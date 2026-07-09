"""
경매 정보 REST API (FastAPI).

    uvicorn api.main:app --reload --port 4000
    문서: http://127.0.0.1:4000/docs

엔드포인트
  GET  /                      서비스 정보
  GET  /properties           물건 목록/검색 (5종·지역·법원·가격 필터)
  GET  /properties/{case_no} 물건 상세 + 권리분석 + 배당(최저가 가정)
  POST /analyze              임의 등기/임차 입력 → 권리분석(+배당)

시작 시 MockSource 를 SQLite 에 적재한다. (실제 소스 연결 시 ingest 만 교체)
"""

from __future__ import annotations

import os
import re
import threading
from datetime import date
from typing import Optional

import httpx

# ── 전역 httpx 공유 클라이언트(2026-06-29) — SSL 컨텍스트 1회 생성으로 예열 정체 근본해결 ──
# 코드 전반이 httpx.get()을 직접 호출 → 매 요청 새 SSL 컨텍스트(load_verify_locations로 CA 번들을
# 디스크에서 ~0.7s/회 재로딩). 아파트 예열 한 건에 실거래·단지·좌표 호출이 수~수십 회라 SSL 로드만으로
# ~16초/건이 됐음(예열 정체의 진짜 원인). httpx.get/post를 공유 클라이언트 호출로 일괄 치환 →
# SSL 컨텍스트 1회 생성 + keep-alive 커넥션 재사용(httpx.Client는 스레드 안전). 호출부 인자
# (params/headers/timeout/follow_redirects 등)는 요청별 오버라이드로 그대로 전달됨.
_HTTPX_SHARED = httpx.Client(timeout=30,
                             transport=httpx.HTTPTransport(retries=2,
                                 limits=httpx.Limits(max_keepalive_connections=50, max_connections=100)))
def _shared_httpx_get(url, **kw):
    return _HTTPX_SHARED.get(url, **kw)
def _shared_httpx_post(url, **kw):
    return _HTTPX_SHARED.post(url, **kw)
httpx.get = _shared_httpx_get
httpx.post = _shared_httpx_post

from fastapi import (FastAPI, HTTPException, Query, Depends, Response, Cookie, Body,
                     Request, UploadFile, File, Form)
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, RedirectResponse, HTMLResponse
from pydantic import BaseModel, Field

from auction_analysis import (
    MockSource, ListingStore, ingest,
    ResidentialType, Region,
    Right, RightType, Tenant, AuctionProperty,
    analyze, calculate_distribution,
)
from auction_analysis.auth import UserStore
from auction_analysis.supabase_source import SupabaseSource, normalize_address, _SIDO_VARIANTS
from .serializers import (
    listing_summary, analysis_to_dict, distribution_to_dict, compute_stats,
)

app = FastAPI(title="경매 정보 API", version="0.1.0")


@app.middleware("http")
async def _no_cache_static(request, call_next):
    """정적 화면(HTML/JS) + 동적 지도 API는 항상 최신을 받도록 캐시 비활성화. 브라우저가 옛 trade_area/배후세대 캐싱 방지."""
    response = await call_next(request)
    path = request.url.path
    if path == "/" or path.startswith("/static/") or path.startswith("/auction"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"   # /auction* + /auctions(검색) 포함
    return response


store = ListingStore(":memory:")

_ROOT = os.path.dirname(os.path.dirname(__file__))
_STATIC_DIR = os.path.join(_ROOT, "static")


def _load_dotenv() -> None:
    """프로젝트 루트 .env 의 KEY=VALUE 를 환경변수로 로드(기존 env 우선)."""
    p = os.path.join(_ROOT, ".env")
    if not os.path.exists(p):
        return
    with open(p, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


_load_dotenv()

# 스피드옥션 제휴 데이터(Supabase + R2). anon키·R2 URL은 공개 가능 값(DATA_ACCESS.md).
# 개인정보 마스킹은 MASK_PERSONAL_INFO=1 로 켤 수 있음(기본 표시).
auction_db = SupabaseSource(
    url=os.environ.get("SUPABASE_URL", "https://jakwbngokvlzehpjiozh.supabase.co"),
    key=(os.environ.get("SUPABASE_KEY") or os.environ.get("SUPABASE_ANON_KEY")
         or "sb_publishable_OAKI_mJcm8v9M4n1WLRotQ_wF0sl5p-"),
    r2_url=os.environ.get("R2_PUBLIC_URL", "https://pub-edb1dd3fca454e75b710b61210fb9cbe.r2.dev"),
)
# 회원 계정은 PostgreSQL(Supabase)에 영속 저장. UserStore가 SUPABASE_DB_URL로 직접 연결한다.
# (인자는 하위호환용으로 받되 무시됨. AUCTION_AUTH_DB/파일경로는 더 이상 사용 안 함.)
user_store = UserStore(os.environ.get("AUCTION_AUTH_DB"))
# 관리자 지정: 기본 '진혁' + 환경변수 ADMIN_NAMES/ADMIN_EMAILS(쉼표구분).
_ADMIN_NAMES = set(["진혁"] + [x.strip() for x in os.environ.get("ADMIN_NAMES", "").split(",") if x.strip()])
_ADMIN_EMAILS = set(x.strip().lower() for x in os.environ.get("ADMIN_EMAILS", "").split(",") if x.strip())


def _maybe_make_admin(user: Optional[dict]) -> Optional[dict]:
    """이름/이메일이 관리자 목록에 들면 admin으로 승격. 로그인·가입 시 호출(서버 재시작 후 새 계정도 자동 적용)."""
    if not user or user.get("role") == "admin":
        return user
    nm = user.get("name") or ""
    em = (user.get("email") or "").lower()
    if nm in _ADMIN_NAMES or em in _ADMIN_EMAILS:
        try:
            if nm in _ADMIN_NAMES:
                user_store.set_admin(name=nm)
            if em in _ADMIN_EMAILS:
                user_store.set_admin(email=em)
            return user_store.get_user(user["id"]) or user
        except Exception:
            pass
    return user


# 기존 회원에도 즉시 적용(시작 시 1회)
for _nm in _ADMIN_NAMES:
    try:
        user_store.set_admin(name=_nm)
    except Exception:
        pass
for _em in _ADMIN_EMAILS:
    try:
        user_store.set_admin(email=_em)
    except Exception:
        pass


def current_user(sid: Optional[str] = Cookie(None)) -> Optional[dict]:
    """쿠키 세션으로 현재 사용자 조회(없으면 None)."""
    return user_store.get_user_by_session(sid)


def require_user(sid: Optional[str] = Cookie(None)) -> dict:
    """로그인 필수 엔드포인트용 의존성."""
    u = user_store.get_user_by_session(sid)
    if not u:
        raise HTTPException(401, "로그인이 필요합니다.")
    return u


def require_admin(sid: Optional[str] = Cookie(None)) -> dict:
    """관리자 전용 의존성."""
    u = user_store.get_user_by_session(sid)
    if not u:
        raise HTTPException(401, "로그인이 필요합니다.")
    if u.get("role") != "admin":
        raise HTTPException(403, "관리자 권한이 필요합니다.")
    return u


def require_paid_user(user: dict = Depends(require_user)) -> dict:
    """유료회원(paid_until 유효) 전용. 관리자는 통과. 부동산DB(KB) 등 유료기능 게이트."""
    if user.get("role") == "admin":
        return user
    import datetime as _dt
    pu = user.get("paid_until")
    ok = False
    if pu:
        try:
            ok = _dt.date.fromisoformat(str(pu)[:10]) >= _dt.date.today()
        except Exception:
            ok = False
    if not ok:
        raise HTTPException(403, "유료 이용권이 필요합니다. 요금결제 후 이용해주세요.")
    return user


_NATIONAL_MIN_RANK = 10   # '전국' 등급의 rank(무료=0). 이 이상만 물건 상세 열람 허용.
_PREMIUM_MIN_RANK = 20    # '프리미엄' 등급의 rank. 이 이상만 유형별 필터·정렬순서 사용(프런트 게이트).


def _user_grade_rank(user: Optional[dict]) -> int:
    """유저 등급의 rank(무료=0, 미로그인/조회실패=0). grades 테이블 조회."""
    try:
        g = user_store._grade_by_name((user or {}).get("grade") or "")
        return int((g or {}).get("rank") or 0)
    except Exception:
        return 0


def require_national_user(user: dict = Depends(require_user)) -> dict:
    """전국 등급 이상(무료 제외) 또는 관리자만 — 물건 상세 열람 게이트.
    결제 시 grade가 '전국'으로 승격되므로 rank 게이트로 결제자도 포함(paid_until은 폴백)."""
    if user.get("role") == "admin":
        return user
    if _user_grade_rank(user) >= _NATIONAL_MIN_RANK:
        return user
    import datetime as _dt          # grade 미승격 결제 엣지 폴백
    pu = user.get("paid_until")
    if pu:
        try:
            if _dt.date.fromisoformat(str(pu)[:10]) >= _dt.date.today():
                return user
        except Exception:
            pass
    raise HTTPException(403, "물건 상세는 전국 등급 이상 회원만 이용할 수 있습니다.")


@app.on_event("startup")
def _load() -> None:
    report = ingest(MockSource(), store)
    print(report.summary())


@app.get("/")
def home() -> FileResponse:
    """메인 홈 화면."""
    return FileResponse(os.path.join(_STATIC_DIR, "index.html"))


@app.get("/{page}.html")
def _bare_html_redirect(page: str):
    """정적 화면은 /static/에만 마운트돼 있어 /auctions.html 같은 '맨' 경로는 404.
    직접 URL 접속·북마크 대비: 파일이 있으면 /static/<page>.html 로 리다이렉트(없으면 404)."""
    if os.path.exists(os.path.join(_STATIC_DIR, page + ".html")):
        return RedirectResponse("/static/" + page + ".html", status_code=307)
    raise HTTPException(404, "not found")


@app.get("/api")
def api_info() -> dict:
    return {
        "service": "부동산 경매 정보 API",
        "count": store.count(),
        "types": [t.value for t in ResidentialType],
        "docs": "/docs",
    }


# ---------- 경매물건 (스피드옥션 제휴 데이터) ----------

_veh_usage_cache: dict = {"ts": 0.0, "buckets": None}


def _vehicle_usage_buckets() -> dict:
    """차량외 usage_name(모델명) 전체를 차체유형(SUV/승용자동차)으로 분류 → {유형: set(usage_name)}. 30분 캐시."""
    import time as _t
    from auction_analysis.body_type import body_type
    if _veh_usage_cache["buckets"] and (_t.time() - _veh_usage_cache["ts"] < 1800):
        return _veh_usage_cache["buckets"]
    buckets = {"SUV": set(), "승용자동차": set()}
    try:
        r = auction_db._get("items", {"select": "usage_name", "search_group": "eq.차량외",
                                       "limit": "5000"})
        for row in r.json():
            u = row.get("usage_name")
            if u:
                buckets[body_type(u)].add(u)
    except Exception:
        pass
    _veh_usage_cache["buckets"] = buckets
    _veh_usage_cache["ts"] = _t.time()
    return buckets


def _expand_vehicle_usages(usage):
    """usage 필터에 'SUV'/'승용자동차'(차체유형)가 있으면 실제 usage_name(모델명) 집합으로 치환.
    크롤 usage_name이 모델명이라, 차체유형 분류로 매칭시킨다."""
    if not usage:
        return usage
    body = [u for u in usage if u in ("SUV", "승용자동차")]
    if not body:
        return usage
    buckets = _vehicle_usage_buckets()
    expanded = {u for u in usage if u not in ("SUV", "승용자동차")}
    for b in body:
        expanded |= buckets.get(b, set())
    return list(expanded) if expanded else usage


# ───────────── 유형별 필터(분석 캐시 기반 item_key 집합) ─────────────
_type_filter_cache: dict = {}          # name -> (ts, set(item_key))
_APT_EXCL_TENANTS = ("한국토지주택공사", "주택도시보증공사", "서울보증보험")


def _apt_deposit_unknown_keys() -> set:
    """아파트 물건 중 '대항력 + 확정없음 + 배당없음 + 보증금미상(0/없음)' 임차인이 있고,
    제외기관(LH·HUG·SGI) 임차인이 없으며, 인수사항(임차인·등기 status에 '인수')이 없는 물건.
    소스 = item_tenants/item_rights 테이블(크롤러 구조화·정확).
    ※ analysis: api_cache는 PDF폴백이라 확정/배당/보증금 필드가 누락돼 오탐 발생 → 사용하지 않음."""
    # 1) 아파트 item_key 집합(현황+과거 매각완료 — 유형필터를 과거 물건에도 적용, 주인님 요청)
    apts: set = set()
    off = 0
    while True:
        r = auction_db._get("items", {"select": "item_key", "order": "item_key",
                                       "usage_name": "ilike.*아파트*",
                                       "limit": "1000", "offset": str(off)})
        rows = r.json() if r.status_code in (200, 206) else []
        for x in rows:
            if x.get("item_key"):
                apts.add(x["item_key"])
        if len(rows) < 1000:
            break
        off += 1000
    if not apts:
        return set()
    # 2) 조건 임차인(대항력 + 확정없음 + 배당없음 + 보증금 null/0) 보유 item_key
    cand: set = set()
    off = 0
    while True:
        r = auction_db._get("item_tenants",
                            {"select": "item_key", "has_opposing_power": "eq.true",
                             "fixed_date": "is.null", "dividend_date": "is.null",
                             "or": "(deposit.is.null,deposit.eq.0)",
                             "limit": "1000", "offset": str(off)})
        rows = r.json() if r.status_code in (200, 206) else []
        for x in rows:
            k = x.get("item_key")
            if k in apts:
                cand.add(k)
        if len(rows) < 1000:
            break
        off += 1000
    if not cand:
        return set()
    # 3) 제외기관 임차인 또는 인수사항(임차인/등기 status에 '인수') 있는 물건 제외
    cand_list = list(cand)
    excl: set = set()
    for i in range(0, len(cand_list), 100):
        ch = cand_list[i:i + 100]
        inlist = ",".join('"' + str(k).replace('"', '') + '"' for k in ch)
        try:                                                       # 임차인 전체(이름/상태)
            tr = auction_db._get("item_tenants",
                                 {"select": "item_key,name,status", "item_key": f"in.({inlist})",
                                  "limit": "5000"})
            for x in (tr.json() if tr.status_code in (200, 206) else []):
                nm = x.get("name") or ""
                if any(e in nm for e in _APT_EXCL_TENANTS) or ("인수" in (x.get("status") or "")):
                    excl.add(x.get("item_key"))
        except Exception:
            pass
        try:                                                       # 등기 권리 상태(인수)
            rr = auction_db._get("item_rights",
                                 {"select": "item_key,status", "item_key": f"in.({inlist})",
                                  "limit": "5000"})
            for x in (rr.json() if rr.status_code in (200, 206) else []):
                if "인수" in (x.get("status") or ""):
                    excl.add(x.get("item_key"))
        except Exception:
            pass
    # 4) 차익(시세 − 기준가) ≥ 3,000만원만 (보증금미상 중에서도 시세 여유 큰 것만)
    keys = list(cand - excl)
    if not keys:
        return set()
    base_of: dict = {}; dc_of: dict = {}   # 기준가 + data_class(과거 매각완료는 차익 미적용 — 주인님 지시)
    for i in range(0, len(keys), 100):
        ch = keys[i:i + 100]
        inlist = ",".join('"' + str(k).replace('"', '') + '"' for k in ch)
        try:
            rr = auction_db._get("items", {"select": "item_key,min_price,sale_price,result,data_class",
                                           "item_key": f"in.({inlist})", "limit": "5000"})
            for x in (rr.json() if rr.status_code in (200, 206) else []):
                res = x.get("result") or ""
                mn = _to_int(x.get("min_price")); sp = _to_int(x.get("sale_price"))
                if ("매각" in res) and ("재매각" not in res):
                    base = sp or mn
                elif (("재매각" in res) or ("재진행" in res)) and sp:
                    base = sp
                else:
                    base = mn
                if base:
                    base_of[x.get("item_key")] = base
                    dc_of[x.get("item_key")] = x.get("data_class")
        except Exception:
            pass
    match: set = set()
    match.update(k for k in keys if dc_of.get(k) and dc_of.get(k) != "현황")   # 과거(매각완료)는 차익 미적용=속성만 — 주인님 지시
    cur = [k for k in keys if dc_of.get(k) == "현황"]                          # 현황만 차익(시세−기준가) 적용
    for i in range(0, len(cur), 150):
        ch = [k for k in cur[i:i + 150] if k in base_of]
        if not ch:
            continue
        try:
            ests = auction_apt_ests(",".join(ch), compute=False)
        except Exception:
            ests = {}
        for k in ch:
            v = ests.get(k)
            if isinstance(v, dict) and v.get("price") and (v["price"] - base_of[k]) >= _DEPOSIT_MIN_PROFIT:
                match.add(k)
    return match


_DEPOSIT_MIN_PROFIT = 30_000_000   # 보증금미상 필터: 차익(시세−기준가) 최소 3,000만원
_REAUC_MIN_PROFIT = 20_000_000   # 재매각/재진행 필터: 차익(시세−이전낙찰가) 최소 2,000만원


def _apt_reauction_profit_keys() -> set:
    """아파트 재매각/재진행 물건 중 '이전 낙찰가(sale_price>0)'와 '추정시세(est)'가 모두 있고,
    차익(시세 − 이전 낙찰가)이 2,000만원 이상인 물건."""
    # 1) 아파트 재매각/재진행 + 이전 낙찰가 보유 item_key·sale_price 수집
    sale: dict = {}
    off = 0
    while True:
        r = auction_db._get("items", {"select": "item_key,sale_price", "order": "item_key",
                                       "usage_name": "ilike.*아파트*",
                                       "or": "(result.like.재매각*,result.like.재진행*)",
                                       "sale_price": "gt.0",
                                       "limit": "1000", "offset": str(off)})
        rows = r.json() if r.status_code in (200, 206) else []
        for x in rows:
            k = x.get("item_key")
            sp = _to_int(x.get("sale_price"))
            if k and sp:
                sale[k] = sp
        if len(rows) < 1000:
            break
        off += 1000
    keys = list(sale.keys())
    if not keys:
        return set()
    # 2) 추정시세(est) 보유 + 차익(est − 이전낙찰가) ≥ 2,000만원 — apt 캐시 조회(계산 안 함)
    match: set = set()
    for i in range(0, len(keys), 150):
        ch = keys[i:i + 150]
        try:
            ests = auction_apt_ests(",".join(ch), compute=False)
        except Exception:
            ests = {}
        for k in ch:
            v = ests.get(k)
            if isinstance(v, dict) and v.get("price"):
                if (v["price"] - sale[k]) >= _REAUC_MIN_PROFIT:
                    match.add(k)
    return match


_OVER85_MIN_PROFIT = 30_000_000   # 85초과 필터: 차익(시세−기준가) 최소 3,000만원


_area_idx: dict = {"ts": 0.0, "map": None}   # item_key -> 전용면적(㎡). building_area 컬럼이 NULL이라 area_text 파싱.


def _area_index(force: bool = False) -> dict:
    """전 물건의 전용면적(㎡) 색인 — area_text의 '전용 NN㎡' 파싱. 30분 캐시.
    (건물면적 필터용: building_area 컬럼은 대부분 NULL이라 숫자비교 불가)."""
    import time as _t
    m = _area_idx["map"]
    if m is not None and not force and _t.time() - _area_idx["ts"] < 1800:
        return m
    import re as _re
    out: dict = {}
    off = 0
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key,area_text", "order": "item_key",
                                          "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        for x in rows:
            mt = _re.search(r"전용\s*([0-9.]+)", x.get("area_text") or "")
            if mt and x.get("item_key"):
                try:
                    out[x["item_key"]] = float(mt.group(1))
                except ValueError:
                    pass
        if len(rows) < 1000:
            break
        off += 1000
    if out:
        _area_idx["map"] = out
        _area_idx["ts"] = _t.time()
    return _area_idx["map"] or {}


def _barea_filter_keys(bmin, bmax) -> Optional[set]:
    """건물면적(전용면적 ㎡) 범위 필터 → item_key 집합. 둘 다 없으면 None(필터 안 함)."""
    if bmin is None and bmax is None:
        return None
    m = _area_index()
    out: set = set()
    for k, a in m.items():
        if bmin is not None and a < bmin:
            continue
        if bmax is not None and a > bmax:
            continue
        out.add(k)
    return out


def _area_warm() -> None:
    try:
        _area_index(force=True)
    except Exception:
        pass


_invest_idx: dict = {"ts": 0.0, "map": None, "building": False}   # item_key -> 투자금(단기매도 선금=총필요금액−종소세). min_price+area_text 산출.


def _invest_of(min_price, excl_area) -> int:
    """단기매도 '선금' = 낙찰×20% + 취등록세(말소·채권 제외) + 중도상환 + 보유이자 + 법무. 낙찰가=최저가."""
    bid = _to_int(min_price) or 0
    if not bid:
        return 0
    eok = bid / 1e8
    base = 1.0 if eok <= 6 else ((eok * 2 / 3 - 3) if eok <= 9 else 3.0)   # 취득세율(금액·6~9억 선형)
    acq = round(bid * base / 100)
    nong = round(bid * 0.2 / 100) if (excl_area or 0) > 85 else 0          # 농특세=85㎡ 초과만
    edu = round(acq * 0.1)                                                # 지방교육세
    inji = 0 if bid <= 1e8 else (150000 if bid <= 1e9 else 350000)        # 인지세
    reg = acq + nong + edu + inji                                         # 취등록세(말소·채권 제외)
    loan = bid * 0.8
    pre = round(loan * 0.005)                                             # 중도상환 0.5%
    hold = round(loan * 0.05 / 12 * 3)                                    # 보유 3개월 이자
    return round(bid * 0.2) + reg + pre + hold + 200000                   # +법무 20만


def _invest_index(force: bool = False) -> dict:
    """전 물건의 투자금(선금) 색인 — min_price+area_text(전용/건물 ㎡)로 _invest_of 산출.
    30분 메모리캐시 + **Supabase 공유캐시**(클라우드는 25페이지 재계산=14초 없이 즉시 읽기).
    요청 경로는 캐시(메모리→Supabase) 즉답, 없으면 백그라운드 1회 빌드(요청 stall 방지)."""
    import time as _t
    m = _invest_idx["map"]
    if m is not None and not force and _t.time() - _invest_idx["ts"] < 1800:
        return m
    if not force:                                   # 요청 경로: Supabase 공유캐시 우선(클라우드 즉답)
        try:
            cached = (auction_db.cache_get_many(["invest_index"]) or {}).get("invest_index")
            if isinstance(cached, dict) and cached:
                _invest_idx["map"] = cached
                _invest_idx["ts"] = _t.time()
                return cached
        except Exception:
            pass
        if not _invest_idx.get("building"):         # 캐시 없음 → 백그라운드 1회 빌드(14초 블로킹 안 함)
            _invest_idx["building"] = True
            threading.Thread(target=_invest_warm, daemon=True).start()
        return _invest_idx["map"] or {}
    import re as _re
    out: dict = {}
    off = 0
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key,area_text,min_price",
                                          "data_class": "eq.현황", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        for x in rows:
            k = x.get("item_key")
            if not k:
                continue
            mt = _re.search(r"(?:전용|건물)\s*([0-9.]+)", x.get("area_text") or "")
            try:
                excl = float(mt.group(1)) if mt else 0.0
            except ValueError:
                excl = 0.0
            iv = _invest_of(x.get("min_price"), excl)
            if iv > 0:
                out[k] = iv
        if len(rows) < 1000:
            break
        off += 1000
    if out:
        _invest_idx["map"] = out
        _invest_idx["ts"] = _t.time()
        try:
            auction_db.cache_save("invest_index", out)      # 공유 → 클라우드가 재계산 없이 읽음
        except Exception:
            pass
    _invest_idx["building"] = False
    return _invest_idx["map"] or {}


def _invest_filter_keys(imin, imax) -> Optional[set]:
    """투자금(원) 범위 필터 → item_key 집합. 둘 다 없으면 None(필터 안 함)."""
    if imin is None and imax is None:
        return None
    m = _invest_index()
    out: set = set()
    for k, iv in m.items():
        if imin is not None and iv < imin:
            continue
        if imax is not None and iv > imax:
            continue
        out.add(k)
    return out


def _invest_warm() -> None:
    try:
        _invest_index(force=True)
    except Exception:
        pass


_baedang_idx: dict = {"ts": 0.0, "map": None, "building": False}   # item_key -> 배당요구신청 건수(다가구·근린주택)


def _baedang_count(court_docs) -> int:
    """문건접수내역(court_docs)에서 '배당요구신청' 접수 건수(취하·철회 제외)."""
    n = 0
    for x in (court_docs or []):
        c = x.get("content") or ""
        if x.get("gubun") == "접수" and "배당요구신청" in c and "취하" not in c and "철회" not in c:
            n += 1
    return n


def _baedang_rebuild() -> None:
    """배당요구 건수 색인. 출처=**매각물건명세서 임차인현황**(item_tenants.dividend_date=배당요구일).
    ⚠️문건접수내역(court_docs) 아님 — 거기엔 채권자 배당요구·중복 접수가 섞여 가구 수를 초과(예: 15가구에 18건).
    다가구·근린주택의 '배당요구한 가구(호실)' 수를 카운트 — 같은 호실 1·2차(증액) 계약은 1건으로 합침(임차인현황 확인되면 0건도 표기).
    ⚠️배당요구종기(items.dividend_deadline) 이후 신청한 배당요구는 무효라 제외(예: 대전 도마동 108-6 — 종기 5/20 후 6/28 신청 1건 제외 → 14→13, 명세서 일치). 기관(LH/HUG/SGI)도 종기 내면 정상 카운트. 실패 시 기존 캐시 유지."""
    import time as _t
    try:
        keys: set = set()                      # ① 다가구·근린주택 현황 item_key(원룸등 포함 위해 LIKE)
        deadlines: dict = {}                   # item_key -> 배당요구종기(YYYY-MM-DD) — 종기 후 배당요구는 무효라 카운트 제외
        for upat in ("*다가구*", "*근린주택*"):
            off = 0
            while True:
                r = auction_db._get("items", {"select": "item_key,dividend_deadline", "data_class": "eq.현황",
                                              "usage_name": f"like.{upat}", "limit": "1000", "offset": str(off)})
                rows = r.json() if r.status_code in (200, 206) else []
                for x in rows:
                    k = x.get("item_key")
                    if not k:
                        continue
                    keys.add(k)
                    if x.get("dividend_deadline"):
                        deadlines[k] = str(x["dividend_deadline"])[:10]
                if len(rows) < 1000:
                    break
                off += 1000
        has_tenant: set = set()                # ② item_tenants(매각물건명세서 임차인현황) 스캔
        seen: dict = {}                        # item_key -> 배당요구 가구(호실) 집합 — 같은 호실 1·2차 계약 중복 제거
        extra: dict = {}                       # 점유·이름 식별 불가한 배당요구 행은 개별 카운트
        off = 0
        while True:                            # PostgREST 1000행 캡 → 페이징
            r = auction_db._get("item_tenants", {"select": "item_key,name,occupancy,dividend_date",
                                                 "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
            if not rows:
                break
            for x in rows:
                k = x.get("item_key")
                if k not in keys:
                    continue
                has_tenant.add(k)
                dd = x.get("dividend_date")
                if dd and str(dd).strip():            # 배당요구일 있음 = 그 임차인이 배당요구함
                    ddl = deadlines.get(k)
                    if ddl and str(dd)[:10] > ddl:    # 배당요구종기 이후 신청 = 무효 → 명세서 미산입(제외). LH/HUG/SGI 기관은 종기 내면 정상 카운트
                        continue
                    occ = (x.get("occupancy") or "").strip()
                    nm = re.sub(r"\s*\d+차\s*$", "", (x.get("name") or "").strip())   # '배수지2차'→'배수지'
                    uk = ("o:" + occ) if occ else (("n:" + nm) if nm else "")          # 호실 우선, 없으면 임차인명(차수 제거)
                    if uk:
                        seen.setdefault(k, set()).add(uk)                              # 같은 호실(가구)은 1건으로
                    else:
                        extra[k] = extra.get(k, 0) + 1
            if len(rows) < 1000:
                break
            off += 1000
        out = {k: len(seen.get(k, set())) + extra.get(k, 0) for k in has_tenant}   # 임차인현황 확인된 물건은 0건도 표기
        for k in keys:                          # 임차인현황 자체가 없는(item_tenants 0) 다가구·근린 → -1 = '임차내역 없음'(프론트 표기, 무표기 모호성 제거)
            if k not in has_tenant:
                out[k] = -1
        if out or _baedang_idx["map"] is None:
            _baedang_idx["map"] = out
            _baedang_idx["ts"] = _t.time()
    except Exception:
        pass


def _baedang_index(force: bool = False) -> dict:
    """다가구·근린주택 배당요구 건수 색인(출처=매각물건명세서 임차인현황 item_tenants.dividend_date). 1시간 캐시 + 스테일 즉시반환.
    빌드가 무거워(~9초) 요청 경로는 절대 동기 빌드 안 함 — 빌드 중이면 빈값/옛값 즉시반환(비차단)."""
    import time as _t
    m = _baedang_idx["map"]
    if not force and m is not None:
        if _t.time() - _baedang_idx["ts"] < 3600:
            return m                       # 신선(1시간) → 즉시
        if not _baedang_idx["building"]:   # 만료 → 옛값 즉시반환 + 백그라운드 재빌드
            _baedang_idx["building"] = True

            def _bg():
                try:
                    _baedang_rebuild()
                finally:
                    _baedang_idx["building"] = False
            threading.Thread(target=_bg, daemon=True).start()
        return m
    if not force:                          # 최초(map None) 요청 경로: 빌드 중이면 빈값(워밍이 곧 채움)
        if _baedang_idx["building"]:
            return {}
        _baedang_idx["building"] = True
        try:
            _baedang_rebuild()
        finally:
            _baedang_idx["building"] = False
        return _baedang_idx["map"] or {}
    _baedang_rebuild()                     # force(시작 워밍) → 동기 빌드
    return _baedang_idx["map"] or {}


def _baedang_warm() -> None:
    try:
        _baedang_index(force=True)
    except Exception:
        pass


# ── 다가구 목록 색인: 경쟁분산(같은 기일·법원 건수) + 우량(보증금합+청구 ≥ 감정가) ──
_dagagu_idx: dict = {"ts": 0.0, "map": None, "building": False}
_DAGAGU_OR = "(usage_name.ilike.*다가구*,usage_name.ilike.*근린주택*)"


def _dagagu_rebuild() -> None:
    """진행중 다가구·단독·근린주택: 경쟁분산(매각기일+법원 그룹수) + 우량(보증금합+청구금액 ≥ 감정가)."""
    import time as _t
    from collections import Counter
    def _num(s):
        return int(re.sub(r"[^0-9]", "", str(s or "")) or 0)   # '45,700,000원'·int 모두 처리
    try:
        items, off = [], 0
        while off < 40000:
            r = auction_db._get("items", {"select": "item_key,usage_name,sell_date,court_code,appraisal_price,claim_amount",
                                          "data_class": "eq.현황", "or": _DAGAGU_OR,
                                          "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code == 200 else []
            if not rows:
                break
            items += rows
            if len(rows) < 1000:
                break
            off += 1000
        items = [it for it in items if not re.search(r"아파트|오피스텔|다세대|연립|빌라|도시형", it.get("usage_name") or "")]  # 집합건물 제외(순수 다가구·단독·근린주택만, 패널과 일치)
        keys = set(it["item_key"] for it in items if it.get("item_key"))
        grp = Counter()                                   # (매각기일, 법원) 그룹 건수 = 경쟁분산
        for it in items:
            sd = (it.get("sell_date") or "")[:10]
            if sd:
                grp[sd + "|" + (it.get("court_code") or "")] += 1
        depsum, off = {}, 0                               # item_tenants 보증금합(다가구 키만 누적, PostgREST 1000행 캡)
        while off < 400000:
            r = auction_db._get("item_tenants", {"select": "item_key,deposit", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code == 200 else []
            if not rows:
                break
            for x in rows:
                k = x.get("item_key")
                if k in keys:
                    depsum[k] = depsum.get(k, 0) + _num(x.get("deposit"))
            if len(rows) < 1000:
                break
            off += 1000
        m = {}
        for it in items:
            ik = it.get("item_key")
            if not ik:
                continue
            sd = (it.get("sell_date") or "")[:10]
            compete = grp.get(sd + "|" + (it.get("court_code") or ""), 0) if sd else 0
            appr = _num(it.get("appraisal_price"))
            tot = depsum.get(ik, 0) + _num(it.get("claim_amount"))
            m[ik] = {"compete": compete, "woolyang": bool(appr and tot >= appr)}
        if m or _dagagu_idx["map"] is None:
            _dagagu_idx["map"] = m
            _dagagu_idx["ts"] = _t.time()
    except Exception:
        pass


def _dagagu_index(force: bool = False) -> dict:
    """다가구 경쟁분산·우량 색인. 1시간 캐시 + 스테일 즉시반환(요청경로 비차단)."""
    import time as _t
    m = _dagagu_idx["map"]
    if not force and m is not None:
        if _t.time() - _dagagu_idx["ts"] < 3600:
            return m
        if not _dagagu_idx["building"]:
            _dagagu_idx["building"] = True

            def _bg():
                try:
                    _dagagu_rebuild()
                finally:
                    _dagagu_idx["building"] = False
            threading.Thread(target=_bg, daemon=True).start()
        return m
    if not force:
        if _dagagu_idx["building"]:
            return {}
        _dagagu_idx["building"] = True
        try:
            _dagagu_rebuild()
        finally:
            _dagagu_idx["building"] = False
        return _dagagu_idx["map"] or {}
    _dagagu_rebuild()
    return _dagagu_idx["map"] or {}


def _dagagu_woolyang_keys() -> set:
    """우량 다가구 item_key 집합(유형필터용)."""
    return set(k for k, v in (_dagagu_index() or {}).items() if v.get("woolyang"))


def _dagagu_warm() -> None:
    try:
        _dagagu_index(force=True)
    except Exception:
        pass


# ── 경쟁분산 색인(전 유형): 같은 매각기일+법원에서 '같은 부류'끼리 동시진행 건수 ──
_compete_idx: dict = {"ts": 0.0, "map": None, "building": False}


def _compete_category(u: str, sg: str = "") -> str:
    """경쟁분산 묶음 부류 — 같은 부류끼리만 센다(아파트는 아파트끼리, 다세대·연립·빌라·도시형(도생)은 한 묶음)."""
    u = u or ""
    if sg == "차량외" or re.search(r"차량|자동차|중기|건설기계|선박", u):
        return "차량"
    if "아파트형" in u or "지식산업" in u:           # 아파트형공장·지식산업센터 = 상가
        return "상가"
    if "아파트" in u:
        return "아파트"
    if "오피스텔" in u:
        return "오피스텔"
    if re.search(r"다세대|연립|빌라|도시형", u):      # 다세대·연립·빌라·도시형(도생) 한 묶음
        return "빌라"
    if re.search(r"다가구|근린주택", u):              # 다가구·근린주택 한 묶음
        return "다가구"
    if "단독" in u:
        return "단독"
    if re.search(r"상가|점포|근린생활|사무|공장|숙박|판매|업무|창고|주유|병원|교육|종교|위락|운동|문화|의료", u):
        return "상가"
    if re.search(r"토지|대지|임야|과수원|잡종지|농지|목장|도로|구거|유지|광천|염전", u) or (u.strip() in ("전", "답")):
        return "토지"
    return "기타"


def _compete_rebuild() -> None:
    """진행중 전체 물건의 경쟁분산(같은 부류·매각기일·법원 건수). 1건 초과만 저장."""
    import time as _t
    from collections import Counter
    try:
        items, off = [], 0
        while True:
            r = auction_db._get("items", {"select": "item_key,usage_name,search_group,sell_date,court_code",
                                          "data_class": "eq.현황", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
            if not rows:
                break
            items += rows
            if len(rows) < 1000:
                break
            off += 1000

        def _cat(it):
            return _compete_category(it.get("usage_name"), it.get("search_group") or "")
        grp = Counter()                                   # (부류, 매각기일, 법원) 건수 = 경쟁분산
        for it in items:
            sd = (it.get("sell_date") or "")[:10]
            if sd:
                grp[_cat(it) + "|" + sd + "|" + (it.get("court_code") or "")] += 1
        m = {}
        for it in items:
            ik = it.get("item_key")
            sd = (it.get("sell_date") or "")[:10]
            if not ik or not sd:
                continue
            c = grp.get(_cat(it) + "|" + sd + "|" + (it.get("court_code") or ""), 0)
            if c > 1:                                     # 1건(자기뿐)은 분산 의미 없음 → 저장 안 함
                m[ik] = c
        if m or _compete_idx["map"] is None:
            _compete_idx["map"] = m
            _compete_idx["ts"] = _t.time()
    except Exception:
        pass


def _compete_index(force: bool = False) -> dict:
    """경쟁분산 색인(전 유형). 1시간 캐시 + 스테일 즉시반환(요청경로 비차단)."""
    import time as _t
    m = _compete_idx["map"]
    if not force and m is not None:
        if _t.time() - _compete_idx["ts"] < 3600:
            return m
        if not _compete_idx["building"]:
            _compete_idx["building"] = True

            def _bg():
                try:
                    _compete_rebuild()
                finally:
                    _compete_idx["building"] = False
            threading.Thread(target=_bg, daemon=True).start()
        return m
    if not force:
        if _compete_idx["building"]:
            return {}
        _compete_idx["building"] = True
        try:
            _compete_rebuild()
        finally:
            _compete_idx["building"] = False
        return _compete_idx["map"] or {}
    _compete_rebuild()
    return _compete_idx["map"] or {}


def _compete_warm() -> None:
    try:
        _compete_index(force=True)
    except Exception:
        pass


def _apt_over85_keys() -> set:
    """아파트 전용 85㎡ 초과 + 차익(시세 − 기준가) 3,000만원 이상.
    전용면적은 area_text의 '전용 NN㎡' 파싱(building_area 컬럼은 '51.84㎡ (15.68평)' 문자열이라 숫자비교 불가).
    기준가 = 목록 표시와 동일(낙찰=낙찰가, 재매각/재진행=이전낙찰가, 그 외=최저가)."""
    import re as _re
    base_of: dict = {}; dc_of: dict = {}   # item_key -> 기준가(원) + data_class(과거 매각완료는 차익 미적용)
    off = 0
    while True:
        r = auction_db._get("items", {"select": "item_key,area_text,min_price,sale_price,result,data_class", "order": "item_key",
                                       "usage_name": "ilike.*아파트*",
                                       "limit": "1000", "offset": str(off)})
        rows = r.json() if r.status_code in (200, 206) else []
        for x in rows:
            m = _re.search(r"전용\s*([0-9.]+)", x.get("area_text") or "")
            if not m:
                continue
            try:
                if float(m.group(1)) <= 85:
                    continue
            except ValueError:
                continue
            k = x.get("item_key")
            if not k:
                continue
            res = x.get("result") or ""
            mn = _to_int(x.get("min_price")); sp = _to_int(x.get("sale_price"))
            if ("매각" in res) and ("재매각" not in res):          # 낙찰
                base = sp or mn
            elif (("재매각" in res) or ("재진행" in res)) and sp:    # 재매각/재진행
                base = sp
            else:                                                  # 신건/유찰
                base = mn
            if base:
                base_of[k] = base; dc_of[k] = x.get("data_class")
        if len(rows) < 1000:
            break
        off += 1000
    keys = list(base_of.keys())
    if not keys:
        return set()
    # 추정시세(est) 보유 + 차익(est − 기준가) ≥ 3,000만원 (현황만; 과거 매각완료는 차익 미적용=속성만)
    match: set = set()
    match.update(k for k in keys if dc_of.get(k) and dc_of.get(k) != "현황")   # 과거(매각완료)는 차익 미적용
    cur = [k for k in keys if dc_of.get(k) == "현황"]
    for i in range(0, len(cur), 150):
        ch = cur[i:i + 150]
        try:
            ests = auction_apt_ests(",".join(ch), compute=False)
        except Exception:
            ests = {}
        for k in ch:
            v = ests.get(k)
            if isinstance(v, dict) and v.get("price"):
                if (v["price"] - base_of[k]) >= _OVER85_MIN_PROFIT:
                    match.add(k)
    return match


_NEW_MIN_PROFIT = 30_000_000   # 신건 필터: 차익(시세−최저가) 최소 3,000만원


def _apt_new_keys() -> set:
    """아파트 신건(result like 신건*) + 차익(시세 − 최저가) 3,000만원 이상.
    신건은 유찰 전 첫 진행이라 기준가 = 최저가(=감정가)."""
    base_of: dict = {}
    off = 0
    while True:
        r = auction_db._get("items", {"select": "item_key,min_price", "order": "item_key",
                                       "usage_name": "ilike.*아파트*", "result": "like.신건*",
                                       "limit": "1000", "offset": str(off)})
        rows = r.json() if r.status_code in (200, 206) else []
        for x in rows:
            k = x.get("item_key"); mn = _to_int(x.get("min_price"))
            if k and mn:
                base_of[k] = mn
        if len(rows) < 1000:
            break
        off += 1000
    keys = list(base_of.keys())
    if not keys:
        return set()
    match: set = set()
    for i in range(0, len(keys), 150):
        ch = keys[i:i + 150]
        try:
            ests = auction_apt_ests(",".join(ch), compute=False)
        except Exception:
            ests = {}
        for k in ch:
            v = ests.get(k)
            if isinstance(v, dict) and v.get("price"):
                if (v["price"] - base_of[k]) >= _NEW_MIN_PROFIT:
                    match.add(k)
    return match


_SENIOR_LEASE_MIN_GAP = 30_000_000   # 선순위 임차권 필터: 시세 − 보증금 최소 3,000만원


def _apt_senior_lease_keys() -> set:
    """아파트 중 '대항력 있는 선순위 임차인(전입 O + 확정 O + 배당요구 X)'이 있고,
    그 임차인의 보증금이 시세(est)보다 3,000만원 이상 낮은 물건.
    소스 = item_tenants 테이블(크롤러 구조화 데이터; 배당요구일 dividend_date 정확·커버리지 넓음).
    ※ analysis: api_cache는 PDF폴백이라 배당요구일이 누락돼 오탐 발생 → 사용하지 않음."""
    # 1) 아파트 item_key + data_class(현황+과거 — 과거 매각완료는 차익 미적용)
    apts: set = set(); dc_of: dict = {}
    off = 0
    while True:
        r = auction_db._get("items", {"select": "item_key,data_class", "order": "item_key",
                                       "usage_name": "ilike.*아파트*",
                                       "limit": "1000", "offset": str(off)})
        rows = r.json() if r.status_code in (200, 206) else []
        for x in rows:
            if x.get("item_key"):
                apts.add(x["item_key"]); dc_of[x["item_key"]] = x.get("data_class")
        if len(rows) < 1000:
            break
        off += 1000
    if not apts:
        return set()
    # 2) item_tenants: 대항력O + 전입O + 확정O + 배당요구일 없음 → item_key별 최대 보증금(인수 리스크 큰 값)
    cand: dict = {}
    off = 0
    while True:
        r = auction_db._get("item_tenants",
                            {"select": "item_key,deposit,move_in_date,fixed_date,dividend_date",
                             "has_opposing_power": "eq.true",
                             "move_in_date": "not.is.null", "fixed_date": "not.is.null",
                             "dividend_date": "is.null",
                             "limit": "1000", "offset": str(off)})
        rows = r.json() if r.status_code in (200, 206) else []
        for x in rows:
            k = x.get("item_key")
            if k not in apts:
                continue
            dep = _to_int(x.get("deposit"))
            if dep:
                cand[k] = max(cand.get(k, 0), dep)
        if len(rows) < 1000:
            break
        off += 1000
    if not cand:
        return set()
    # 2.5) 임차인 혼재 제외: 한 물건의 임차인들이 전입일도 서로 다르고(2종↑) + 보증금도 서로 다르면(2종↑) 제외.
    #   (작은 보증금 임차인만 보고 '시세보다 한참 낮은 선순위'로 잡히지만, 전입·보증금이 다른 거액 임차인이
    #    섞인 혼재 상황이라 깨끗한 단일 선순위가 아님 → 배제)
    cand_list = list(cand.keys())
    excl: set = set()
    for i in range(0, len(cand_list), 100):
        ch = cand_list[i:i + 100]
        inlist = ",".join('"' + str(k).replace('"', '') + '"' for k in ch)
        per: dict = {}   # item_key -> {"mv": set(전입일), "dp": set(보증금)}
        try:
            tr = auction_db._get("item_tenants",
                                 {"select": "item_key,move_in_date,deposit",
                                  "item_key": f"in.({inlist})", "limit": "5000"})
            for x in (tr.json() if tr.status_code in (200, 206) else []):
                k = x.get("item_key")
                if k is None:
                    continue
                d = per.setdefault(k, {"mv": set(), "dp": set()})
                if x.get("move_in_date"):
                    d["mv"].add(x.get("move_in_date"))
                dp = _to_int(x.get("deposit"))
                if dp:
                    d["dp"].add(dp)
        except Exception:
            pass
        for k, d in per.items():
            if len(d["mv"]) > 1 and len(d["dp"]) > 1:   # 전입 다름 AND 보증금 다름 → 혼재
                excl.add(k)
    for k in excl:
        cand.pop(k, None)
    if not cand:
        return set()
    # 3) 시세(est) − 보증금 ≥ 3,000만원 (현황만; 과거 매각완료는 차익 미적용=속성만)
    keys = list(cand.keys())
    match: set = set()
    match.update(k for k in keys if dc_of.get(k) and dc_of.get(k) != "현황")   # 과거(매각완료)는 차익 미적용
    cur = [k for k in keys if dc_of.get(k) == "현황"]
    for i in range(0, len(cur), 150):
        ch = cur[i:i + 150]
        try:
            ests = auction_apt_ests(",".join(ch), compute=False)
        except Exception:
            ests = {}
        for k in ch:
            v = ests.get(k)
            if isinstance(v, dict) and v.get("price"):
                if (v["price"] - cand[k]) >= _SENIOR_LEASE_MIN_GAP:
                    match.add(k)
    return match


def _apt_expbid_keys() -> set:
    """아파트 중 예상낙찰가(동일건물 매각사례 평균)가 산출된 물건 item_key 집합 — '아파트:백데이터' 필터용."""
    keys, off = [], 0
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key", "data_class": "eq.현황",
                                          "usage_name": "ilike.*아파트*", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        keys += [x["item_key"] for x in rows if x.get("item_key")]
        if len(rows) < 1000:
            break
        off += 1000
    match: set = set()
    for i in range(0, len(keys), 200):
        ch = keys[i:i + 200]
        try:
            cc = auction_db.cache_get_many(["expbid:" + k for k in ch])
        except Exception:
            cc = {}
        for k in ch:
            v = cc.get("expbid:" + k)
            if isinstance(v, dict) and v.get("v") == _EXPBID_V and v.get("available"):
                match.add(k)
    # 시세(추정시세) 없는 물건은 백데이터 필터에서 제외 — 시세 있는 것만(주인님 요청 2026-06-30)
    if match:
        ml = list(match); match = set()
        for i in range(0, len(ml), 150):
            ch = ml[i:i + 150]
            try:
                ests = auction_apt_ests(",".join(ch), compute=False)
            except Exception:
                ests = {}
            for k in ch:
                v = ests.get(k)
                if isinstance(v, dict) and v.get("price"):
                    match.add(k)
    return match


def _villa_expbid_keys() -> set:
    """빌라/도생 중 예상낙찰가(반경1km)가 산출된 물건 item_key 집합 — '빌라:백데이터' 필터용."""
    keys, off = [], 0
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key", "data_class": "eq.현황",
                                          "or": _VILLA_OR, "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        keys += [x["item_key"] for x in rows if x.get("item_key")]
        if len(rows) < 1000:
            break
        off += 1000
    match: set = set()
    for i in range(0, len(keys), 200):
        ch = keys[i:i + 200]
        try:
            cc = auction_db.cache_get_many(["vexpbid:" + k for k in ch])
        except Exception:
            cc = {}
        for k in ch:
            v = cc.get("vexpbid:" + k)
            if isinstance(v, dict) and v.get("v") == _VEXPBID_V and v.get("available"):
                match.add(k)
    # 시세(추정시세) 없는 물건은 백데이터 필터에서 제외 — 시세 있는 것만(주인님 요청 2026-06-30)
    if match:
        ml = list(match); match = set()
        for i in range(0, len(ml), 100):                 # villa_ests는 keys 120개 캡 → 100씩
            ch = ml[i:i + 100]
            try:
                ests = auction_villa_ests(",".join(ch), compute=False)
            except Exception:
                ests = {}
            for k in ch:
                v = ests.get(k)
                if isinstance(v, dict) and v.get("price"):
                    match.add(k)
    return match


def _car_expbid_keys() -> set:
    """차량(차량외) 중 예상낙찰가(백데이터 낙찰사례 중앙값)가 산출된 물건 item_key 집합 — '차량외:백데이터' 필터용.
    주인님 지시: 예상낙찰가 available한 물건만(차량은 시세 조건 없이 예상낙찰가만)."""
    keys, off = [], 0
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key", "data_class": "eq.현황",
                                          "search_group": "eq.차량외", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        keys += [x["item_key"] for x in rows if x.get("item_key")]
        if len(rows) < 1000:
            break
        off += 1000
    match: set = set()
    for i in range(0, len(keys), 200):
        ch = keys[i:i + 200]
        try:
            cc = auction_db.cache_get_many(["carexpbid:" + k for k in ch])
        except Exception:
            cc = {}
        for k in ch:
            v = cc.get("carexpbid:" + k)
            if isinstance(v, dict) and v.get("v") == _CAREXPBID_V and v.get("available"):
                match.add(k)
    return match


_TYPE_FILTER_FNS = {"apt_deposit_unknown": _apt_deposit_unknown_keys,
                    "apt_over85": _apt_over85_keys,
                    "apt_reauction_profit": _apt_reauction_profit_keys,
                    "apt_new": _apt_new_keys,
                    "apt_senior_lease": _apt_senior_lease_keys,
                    "apt_expbid": _apt_expbid_keys,
                    "villa_expbid": _villa_expbid_keys,
                    "car_expbid": _car_expbid_keys,
                    "dagagu_woolyang": _dagagu_woolyang_keys}


_type_filter_building: dict = {}       # name -> bool (백그라운드 재계산 중)


def _type_filter_rebuild(name, fn) -> None:
    """타입필터 1종을 계산해 캐시 갱신(백그라운드 전용). 끝나면 building 플래그 해제."""
    import time as _t
    try:
        keys = fn()
        if keys:                                   # 빈 결과(대개 일시 조회실패)는 캐시 안 함 → 옛 정상값 유지
            _type_filter_cache[name] = (_t.time(), keys)
    except Exception:
        pass
    finally:
        _type_filter_building[name] = False


def _type_filter_warm() -> None:
    """모든 타입필터를 미리 계산(시작 시 1회). 첫 사용 클릭이 13초 멈추는 것 방지."""
    for name, fn in _TYPE_FILTER_FNS.items():
        if not _type_filter_cache.get(name):
            _type_filter_rebuild(name, fn)


def _type_filter_keys(names) -> Optional[set]:
    """유형별 필터 이름들 → 매칭 item_key 집합(교집합). 30분 캐시.
    만료돼도 '스테일 캐시 즉시 반환 + 백그라운드 재계산'으로 요청을 막지 않음(클릭 지연 제거).
    캐시가 아예 없을 때만(최초 1회) 동기 계산. names 없으면 None."""
    import time as _t
    if not names:
        return None
    sets: list[set] = []
    for name in names:
        fn = _TYPE_FILTER_FNS.get(name)
        if fn is None:
            continue
        ent = _type_filter_cache.get(name)
        if ent:
            keys = ent[1]                                  # 있으면 (만료여도) 즉시 반환
            if (_t.time() - ent[0] >= 1800) and not _type_filter_building.get(name):
                _type_filter_building[name] = True         # 만료 → 백그라운드 갱신(논블로킹)
                threading.Thread(target=_type_filter_rebuild, args=(name, fn), daemon=True).start()
        else:                                              # 캐시 전무 → 최초 1회만 동기 계산
            try:
                keys = fn()
            except Exception:
                keys = set()
            if keys:                                       # 빈 결과는 캐시 안 함 → 일시 실패로 0건이 30분 고착되는 것 방지
                _type_filter_cache[name] = (_t.time(), keys)
        sets.append(keys)
    if not sets:
        return None
    out: set = set()                       # 합집합(OR): 선택한 유형 중 하나라도 해당하면 포함(여러 개 체크 시 넓어짐)
    for s in sets:
        out |= s
    return out


# ───────────── 연료별 필터(차량 vehicle_specs.fuel → item_key 집합) ─────────────
_fuel_cache: dict = {"ts": 0.0, "buckets": None}
_FUEL_LABELS = ["휘발유", "경유", "LPG", "하이브리드", "전기"]


def _classify_fuel(fuel: str) -> Optional[str]:
    """원본 연료표기 → 5버킷. 하이브리드 먼저 판정(하이브리드(휘발유+전기) 오분류 방지)."""
    f = re.sub(r"\s+", "", fuel or "")
    if not f:
        return None
    if "하이브리드" in f or "hybrid" in f.lower():
        return "하이브리드"
    if "전기" in f or f.lower() == "ev":
        return "전기"
    if "lpg" in f.lower() or "엘피지" in f:
        return "LPG"
    if "경유" in f or "디젤" in f:
        return "경유"
    if "휘발유" in f or "가솔린" in f:
        return "휘발유"
    return None


def _fuel_buckets() -> dict:
    """vehicle_specs 전체 → {연료라벨: set(item_key)}. 30분 캐시."""
    import time as _t
    if _fuel_cache["buckets"] and (_t.time() - _fuel_cache["ts"] < 1800):
        return _fuel_cache["buckets"]
    buckets = {lab: set() for lab in _FUEL_LABELS}
    off = 0
    while True:
        try:
            r = auction_db._get("vehicle_specs", {"select": "item_key,fuel",
                                                  "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        for x in rows:
            b = _classify_fuel(x.get("fuel") or "")
            if b:
                buckets[b].add(x.get("item_key"))
        if len(rows) < 1000:
            break
        off += 1000
    _fuel_cache["buckets"] = buckets
    _fuel_cache["ts"] = _t.time()
    return buckets


def _fuel_filter_keys(fuels) -> Optional[set]:
    """선택 연료들 → 해당 차량 item_key 합집합. 없으면 None."""
    if not fuels:
        return None
    bk = _fuel_buckets()
    out: set = set()
    for f in fuels:
        out |= bk.get(f, set())
    return out


_buy_cache: dict = {"ts": 0.0, "keys": None}


def _buy_ok_keys() -> set:
    """매수 양호(buy_grade.ok) 차량 item_key 집합. 30분 캐시. (목록 개수 정확 반영용)"""
    import time as _t
    if _buy_cache["keys"] is not None and (_t.time() - _buy_cache["ts"] < 1800):
        return _buy_cache["keys"]
    from auction_analysis.vehicle_parser import buy_grade
    keys: set = set()
    off = 0
    while True:
        try:
            r = auction_db._get("vehicle_specs", {"select": "*", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        for spec in rows:
            g = buy_grade(spec)
            if g and g.get("ok"):
                keys.add(spec.get("item_key"))
        if len(rows) < 1000:
            break
        off += 1000
    _buy_cache["keys"] = keys
    _buy_cache["ts"] = _t.time()
    return keys


# ───────────── 매수판정(AI 위험도 기반: 매수양호/검토/금지) ─────────────
_grade_cache: dict = {"ts": 0.0, "buckets": None}
_GRADE_LABELS = ["매수양호", "매수검토", "매수금지"]

# items.buy_grade 컬럼 가속: 컬럼 있으면 매수판정 필터를 IN-리스트 대신 컬럼 WHERE로(수배 빠름)
_buy_grade: dict = {"col": None, "synced": False}   # col: 컬럼 존재(None=미확인)


def _buy_grade_col_exists() -> bool:
    if _buy_grade["col"] is None:
        try:
            r = auction_db._get("items", [("select", "buy_grade"), ("limit", "1")])
            _buy_grade["col"] = (r.status_code == 200)
        except Exception:
            _buy_grade["col"] = False
    return bool(_buy_grade["col"])


def _buy_grade_ready() -> bool:
    """컬럼이 있고 동기화됐으면 True → 컬럼 WHERE 사용 가능."""
    return bool(_buy_grade["col"]) and _buy_grade["synced"]


def _sync_buy_grade(buckets: dict) -> None:
    """매수판정 버킷 → items.buy_grade 컬럼 일괄 upsert(merge). 컬럼 + service_role 키 있을 때만. 워밍 때 갱신."""
    if not buckets or not _buy_grade_col_exists():
        return
    svc = os.environ.get("SUPABASE_SERVICE_KEY", "")     # items 쓰기는 RLS 우회 필요(서버전용 secret 키)
    if not svc:
        return
    rev = {k: g for g, s in buckets.items() for k in s}
    if not rev:
        return
    rows = [{"item_key": k, "buy_grade": g} for k, g in rev.items()]
    h = {"apikey": svc, "Authorization": f"Bearer {svc}",
         "Content-Type": "application/json",
         "Prefer": "resolution=merge-duplicates,return=minimal"}
    ok = 0
    for i in range(0, len(rows), 500):
        chunk = rows[i:i + 500]
        try:
            r = httpx.post(f"{auction_db.url}/rest/v1/items?on_conflict=item_key",
                           headers=h, json=chunk, timeout=30)
            if r.status_code in (200, 201, 204):
                ok += len(chunk)
        except Exception:
            pass
    if ok:
        _buy_grade["synced"] = True
        print(f"[buy_grade] 컬럼 동기화 {ok}건", flush=True)


def _grade_buckets(force: bool = False) -> dict:
    """전 물건 → 매수판정 집합 {매수양호/매수검토/매수금지}.
    요청 경로는 항상 캐시 즉시 반환(재계산 안 함, ~16초 stall 방지). 재계산은 워밍 스레드(force=True)만 수행."""
    import time as _t
    from collections import defaultdict
    cached = _grade_cache["buckets"]
    if cached is not None and not force:
        return cached                       # 요청: 무조건 캐시 반환(만료돼도 옛값 — 워밍이 갱신)
    if not force:                           # 아직 빌드 전 → 요청은 빈값(배지 비동기 폴백) + 백그라운드 1회 빌드
        if not _grade_cache.get("building"):
            _grade_cache["building"] = True
            threading.Thread(target=lambda: _grade_buckets(force=True), daemon=True).start()
        return {}
    if os.environ.get("CLOUD_READER", "0") in ("1", "true", "True"):
        # 클라우드 안전장치: 어떤 경로로 force가 들어와도 수만 행 페이징 재계산 금지(컬럼 신뢰). 스파이크→OOM 방지.
        _grade_cache["building"] = False
        return _grade_cache.get("buckets") or {}
    db = auction_db
    out = {g: set() for g in _GRADE_LABELS}

    def _page(table, params):
        rows, off = [], 0
        while True:
            try:
                r = db._get(table, {**params, "limit": "1000", "offset": str(off)})
                page = r.json() if r.status_code in (200, 206) else []
            except Exception:
                break
            rows += page
            if len(page) < 1000:
                break
            off += 1000
        return rows

    # ── 주거용(analyzed_at) 위험도 ──
    res_rows = _page("items", {"select": "item_key,usage_name,address",
           "search_group": "eq.주거용", "analyzed_at": "not.is.null"})
    res = {x["item_key"] for x in res_rows}
    res_info = {x["item_key"]: (x.get("usage_name"), x.get("address")) for x in res_rows}
    # 승강기(brief 캐시) — 다세대·도시형 4층↑ 무승강기 매수검토 상향용. 미상(brief 없음)도 트리거(정책).
    elev_map: dict = {}
    vio_map: dict = {}                              # 위반건축물(brief 캐시) — 신규 물건도 컬럼/필터에 자동 반영
    _rk = list(res)
    for _i in range(0, len(_rk), 100):
        _ch = _rk[_i:_i + 100]
        try:
            _got = db.cache_get_many(["brief:" + k for k in _ch]) or {}
        except Exception:
            _got = {}
        for k in _ch:
            _bv = _got.get("brief:" + k)
            elev_map[k] = _bv.get("elevator") if isinstance(_bv, dict) else None
            vio_map[k] = bool(_bv.get("violation")) if isinstance(_bv, dict) else False
    tmap = defaultdict(list)
    for x in _page("item_tenants", {"select": "item_key,has_opposing_power,assume_amount,status,"
                                              "move_in_date,fixed_date,dividend_date"}):
        tmap[x["item_key"]].append(x)
    try:
        assume_right = {x["item_key"] for x in
                        (db._get("item_rights", {"select": "item_key", "status": "like.*인수*",
                                                 "limit": "1000"}).json() or [])}
    except Exception:
        assume_right = set()
    # 인수 면제 조건 → ①확약서(말소동의·포기) ②특별매각조건(보증금 반환청구권/채권 포기).
    #  ⚠️'*반환청구권*포기*'(이중 와일드카드) 전수스캔은 간헐 500 → waiver 셋이 비어 아파트 등이 매수금지로 오분류됨.
    #  → 연속 phrase 단일패턴으로 분리(안정) + 빈결과면 재시도(간헐 500 방어).
    def _wpage(params):
        r = []
        for _ in range(3):
            r = _page("items", {"select": "item_key", **params})
            if r:
                break
        return r
    waiver = {x["item_key"] for x in _wpage({"detail_text": "like.*확약*",
              "or": "(detail_text.like.*포기*,detail_text.like.*말소동의*)"})}
    for _wp in ("*반환청구권을 포기*", "*반환채권을 포기*"):   # 특별매각조건 두 표현
        waiver |= {x["item_key"] for x in _wpage({"detail_text": f"like.{_wp}"})}
    # ③ 보증기관(HUG/SGI/HF) 인수조건변경 태그 = 임차보증금 인수 면제(명세서 본문에 포기문구가 없어도 태그로 인정)
    waiver |= {x["item_key"] for x in _wpage({"tags": "like.*인수조건변경*"})}
    if not waiver:                          # 정상이면 수천 건 → 0이면 조회 실패(고부하) → 재계산 중단·옛 버킷 유지(대량 오분류 방지)
        print("[buy_grade] waiver 0건(조회 실패 추정) — 재계산 중단, 기존 버킷 유지", flush=True)
        _grade_cache["building"] = False
        return _grade_cache.get("buckets") or {}
    for k in res:
        assume, has_opp, danger = 0, False, False
        waived = k in waiver
        for t in tmap.get(k, []):
            opp = t.get("has_opposing_power")
            # ★전액 인수 위험: 대항력 + 전입O + 확정X + 배당X(말소기준 없어 전액 인수). 확약(인수조건변경)은
            #  보증기관 승계분 미배당만 면제 → '별도 대항력 임차인의 전액 인수'(danger)는 확약으로도 안 풀림 → 항상 매수금지.
            if opp and t.get("move_in_date") and not t.get("fixed_date") and not t.get("dividend_date"):
                danger = True
            if waived:
                continue                       # 확약: 미배당 인수(assume)·대항력(has_opp)은 면제 — 단 danger는 위에서 이미 반영
            if "인수" in (t.get("status") or "") and (t.get("assume_amount") or 0):
                assume += t["assume_amount"]
            if opp:
                has_opp = True
        if assume > 0 or k in assume_right or danger:
            out["매수금지"].add(k)
        elif has_opp:
            out["매수검토"].add(k)
        else:
            out["매수양호"].add(k)

    # ── 다세대·도시형 4층↑ 승강기 없음/미상 → 매수양호만 매수검토로 상향(금지·검토는 유지) ──
    from auction_analysis.crawler_analysis import elevator_caution
    _elev_moved = 0
    for k in list(out["매수양호"]):
        u, addr = res_info.get(k, (None, None))
        if elevator_caution(u, addr, elev_map.get(k)):
            out["매수양호"].discard(k)
            out["매수검토"].add(k)
            _elev_moved += 1
    if _elev_moved:
        print(f"[buy_grade] 승강기없음·4층↑ 매수검토 상향 {_elev_moved}건", flush=True)

    # ── 위반건축물(건축물대장 스탬프) → 매수양호만 매수검토로 상향(대출 난항·이행강제금 — 금지·검토는 유지) ──
    _vio_moved = 0
    for k in list(out["매수양호"]):
        if k in _VIOLATION_KEYS or vio_map.get(k):
            out["매수양호"].discard(k)
            out["매수검토"].add(k)
            _vio_moved += 1
    if _vio_moved:
        print(f"[buy_grade] 위반건축물 매수검토 상향 {_vio_moved}건", flush=True)

    # ── 차량외 buy_grade ──
    from auction_analysis.vehicle_parser import buy_grade
    for spec in _page("vehicle_specs", {"select": "*"}):
        g = buy_grade(spec)
        if not g:
            continue
        out["매수양호" if g.get("ok") else "매수금지"].add(spec.get("item_key"))

    _grade_cache["buckets"] = out
    _grade_cache["ts"] = _t.time()
    _grade_cache["building"] = False
    try:
        _sync_buy_grade(out)          # 버킷 → items.buy_grade 컬럼(있으면) 동기화
    except Exception:
        pass
    return out


def _grade_filter_keys(grades) -> Optional[set]:
    """선택 매수판정들 → 해당 item_key 합집합. 없으면 None."""
    if not grades:
        return None
    bk = _grade_buckets()
    if not bk:                      # 아직 빌드 전 → 필터는 정확성 우선, 동기 빌드(최초 1회만 ~16초, 이후 캐시)
        bk = _grade_buckets(force=True)
    out: set = set()
    for g in grades:
        out |= bk.get(g, set())
    return out


@app.get("/auction/grades")
def auction_grades(keys: str) -> dict:
    """목록 배지용: item_key별 매수판정(매수양호/매수검토/매수금지). 버킷 캐시 기반(빠름)."""
    klist = [k for k in keys.split(",") if k][:200]
    bk = _grade_buckets()
    rev: dict = {}
    for g, s in bk.items():
        for k in s:
            rev[k] = g
    return {k: rev.get(k) for k in klist}


# ───────────── 브랜드별 필터(차량 vehicle_specs.manufacturer → item_key 집합) ─────────────
_brand_cache: dict = {"ts": 0.0, "buckets": None}
# 원본 제조사 표기(법인명·영문·변형) → 대표 브랜드. 부분일치(소문자/원문) 우선순위 순.
_BRAND_RULES = [
    ("제네시스", "제네시스"), ("genesis", "제네시스"),
    ("현대", "현대"), ("hyundai", "현대"),
    ("기아", "기아"), ("kia", "기아"),
    ("쌍용", "KGM(쌍용)"), ("kgm", "KGM(쌍용)"), ("케이지엠", "KGM(쌍용)"),
    ("르노", "르노"), ("renault", "르노"), ("삼성", "르노"),
    ("쉐보레", "쉐보레"), ("한국gm", "쉐보레"), ("지엠", "쉐보레"), ("chevrolet", "쉐보레"), ("gm", "쉐보레"),
    ("벤츠", "벤츠"), ("benz", "벤츠"), ("mercedes", "벤츠"), ("메르세데스", "벤츠"),
    ("다임러", "벤츠"), ("daimler", "벤츠"),
    ("비엠더블유", "BMW"), ("bmw", "BMW"),
    ("크라이슬러", "크라이슬러"), ("chrysler", "크라이슬러"),
    ("제너럴", "쉐보레"), ("general", "쉐보레"),
    ("mini", "MINI"), ("미니", "MINI"),
    ("아우디", "아우디"), ("audi", "아우디"),
    ("폭스바겐", "폭스바겐"), ("volkswagen", "폭스바겐"),
    ("포르쉐", "포르쉐"), ("porsche", "포르쉐"),
    ("마세라티", "마세라티"), ("maserati", "마세라티"),
    ("페라리", "페라리"), ("ferrari", "페라리"),
    ("람보르기니", "람보르기니"), ("lamborghini", "람보르기니"),
    ("벤틀리", "벤틀리"), ("bentley", "벤틀리"), ("롤스로이스", "롤스로이스"),
    ("렉서스", "렉서스"), ("lexus", "렉서스"),
    ("토요타", "토요타"), ("도요타", "토요타"), ("toyota", "토요타"),
    ("혼다", "혼다"), ("honda", "혼다"),
    ("닛산", "닛산"), ("nissan", "닛산"), ("인피니티", "인피니티"), ("infiniti", "인피니티"),
    ("볼보", "볼보"), ("volvo", "볼보"),
    ("재규어", "재규어"), ("jaguar", "재규어"),
    ("랜드로버", "랜드로버"), ("로버", "랜드로버"), ("land rover", "랜드로버"),
    ("포드", "포드"), ("ford", "포드"), ("링컨", "링컨"), ("lincoln", "링컨"),
    ("캐딜락", "캐딜락"), ("cadillac", "캐딜락"), ("지프", "지프"), ("jeep", "지프"),
    ("푸조", "푸조"), ("peugeot", "푸조"), ("시트로엥", "시트로엥"),
    ("테슬라", "테슬라"), ("tesla", "테슬라"),
    ("폴스타", "폴스타"), ("polestar", "폴스타"),
]


def _classify_brand(mfr: str) -> Optional[str]:
    """원본 제조사 표기 → 대표 브랜드. 미매칭은 법인 접미사 제거한 원문."""
    m = re.sub(r"\s+", "", mfr or "")
    if not m:
        return None
    ml = m.lower()
    for kw, brand in _BRAND_RULES:
        k = kw.replace(" ", "")
        if k in ml or k in m:
            return brand
    cleaned = re.sub(r"\(주\)|주식회사|자동차|모터스|코리아|motors|korea|inc|corp|ag|co\.?,?ltd",
                     "", m, flags=re.I).strip("()㈜·, ")
    return cleaned or None


def _brand_buckets() -> dict:
    """vehicle_specs 전체 → {브랜드: set(item_key)}. 30분 캐시."""
    import time as _t
    if _brand_cache["buckets"] and (_t.time() - _brand_cache["ts"] < 1800):
        return _brand_cache["buckets"]
    buckets: dict[str, set] = {}
    off = 0
    while True:
        try:
            r = auction_db._get("vehicle_specs", {"select": "item_key,manufacturer",
                                                  "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        for x in rows:
            b = _classify_brand(x.get("manufacturer") or "")
            if b:
                buckets.setdefault(b, set()).add(x.get("item_key"))
        if len(rows) < 1000:
            break
        off += 1000
    if buckets:                       # 빈 결과(DB 일시오류)는 캐시 안 함 → 다음 요청 때 재시도(오염 방지)
        _brand_cache["buckets"] = buckets
        _brand_cache["ts"] = _t.time()
    return buckets


def _brand_filter_keys(brands) -> Optional[set]:
    """선택 브랜드들 → 해당 차량 item_key 합집합. 없으면 None."""
    if not brands:
        return None
    bk = _brand_buckets()
    out: set = set()
    for b in brands:
        out |= bk.get(b, set())
    return out


@app.get("/auction/brands")
def auction_brands() -> dict:
    """차량 브랜드 드롭다운 옵션: [{brand, count}] (건수 내림차순)."""
    bk = _brand_buckets()
    items = sorted(({"brand": b, "count": len(ks)} for b, ks in bk.items()),
                   key=lambda x: -x["count"])
    return {"brands": items}


def _combine_item_keys(*sets) -> Optional[set]:
    """여러 필터의 item_key 집합 교집합(None은 무제한이라 무시)."""
    res = None
    for s in sets:
        if s is None:
            continue
        res = s if res is None else (res & s)
    return res


# 용도지역(주거/준주거/상업/준공업 등) — 단독·다가구·근린주택 목록/상세 표기용. detail_text에서 추출·캐시.
_ZONE_RE = re.compile(
    r"제\s*\d+\s*종\s*(?:일반|전용)?\s*주거지역|준주거지역|"
    r"(?:중심|일반|근린|유통)?\s*상업지역|"
    r"(?:전용|일반|준)?\s*공업지역|"
    r"(?:보전|생산|자연)?\s*녹지지역|"
    r"(?:계획|생산|보전)\s*관리지역|농림지역|자연환경보전지역")
_zone_cache: dict = {}   # item_key -> 용도지역(짧은 라벨) | ""


def _extract_zone(detail_text: str) -> str:
    m = _ZONE_RE.search(detail_text or "")
    return re.sub(r"\s+", "", m.group(0)) if m else ""


def _zone_map(zkeys: list) -> dict:
    """item_key별 용도지역(짧은 라벨). detail_text 1회 파싱 후 메모리 캐시.
    detail_text에 용도지역이 없으면(농촌형 등) V-World 보완 캐시(_lu_cache)로 채워 반환."""
    need = [k for k in zkeys if k not in _zone_cache]
    for i in range(0, len(need), 100):
        ch = need[i:i + 100]
        inlist = ",".join('"' + str(k).replace('"', '') + '"' for k in ch)
        try:
            r = auction_db._get("items", {"select": "item_key,detail_text",
                                          "item_key": f"in.({inlist})", "limit": "200"})
            got = {x["item_key"]: x.get("detail_text")
                   for x in (r.json() if r.status_code in (200, 206) else [])}
        except Exception:
            got = {}
        for k in ch:
            _zone_cache[k] = _extract_zone(got.get(k) or "")
    out = {}
    for k in zkeys:
        z = _zone_cache.get(k, "")
        if not z:                                  # detail_text에 없음 → V-World 보완값(있으면)
            lv = _lu_cache.get(k)                  # 호출시점 조회(prewarm이 채우면 즉시 반영)
            if lv and lv != "NF":
                z = lv
        out[k] = z
    return out


# ───────────── 용도지역별 필터(주거/준주거/상업/준공업/공업/녹지/관리) ─────────────
#  주거지역 = 제1·2·3종 일반/전용 주거지역 전부를 하나로 묶음(준주거 제외).
_ZONE_FILTER_CATS = ["주거지역", "준주거지역", "상업지역", "준공업지역", "공업지역", "녹지지역", "관리지역"]
_zone_bucket_cache: dict = {"ts": 0.0, "buckets": None, "building": False}


_ZONE_APPL_OR = ("(usage_name.ilike.*단독*,usage_name.ilike.*다가구*,"
                 "usage_name.ilike.*근린주택*,usage_name.eq.주택,usage_name.ilike.*농가*)")
_zone_build_lock = threading.Lock()


def _zone_ilike_keys(pattern: str) -> set:
    """용도지역 표기 대상(단독·다가구·근린주택 등)이면서 detail_text에 용도지역 키워드(pattern) 포함 item_key.
    용도필터를 쿼리에 포함 → 스캔 행수 축소(전체 현황 ilike보다 빠르고 안정)."""
    keys: set = set()
    off = 0
    while True:
        r = auction_db._get("items", {"select": "item_key", "data_class": "eq.현황",
                                       "or": _ZONE_APPL_OR,
                                       "detail_text": f"ilike.*{pattern}*",
                                       "limit": "1000", "offset": str(off)})
        rows = r.json() if r.status_code in (200, 206) else []
        for x in rows:
            if x.get("item_key"):
                keys.add(x["item_key"])
        if len(rows) < 1000:
            break
        off += 1000
    return keys


_zone_bkt: dict = {"ts": 0.0, "buckets": None}   # detail_text에서 직접 추출한 카테고리별 키 집합


def _zone_categorize(zlabel: str) -> Optional[str]:
    """용도지역 라벨 → 카테고리. 주거 전종류는 '주거지역'으로 묶음(준주거 별도)."""
    if not zlabel:
        return None
    if "준주거" in zlabel:
        return "준주거지역"
    if "주거지역" in zlabel:
        return "주거지역"
    if "상업지역" in zlabel:
        return "상업지역"
    if "준공업" in zlabel:
        return "준공업지역"
    if "공업지역" in zlabel:
        return "공업지역"
    if "녹지지역" in zlabel:
        return "녹지지역"
    if "관리지역" in zlabel:
        return "관리지역"
    return None


def _zone_build(force: bool = False) -> dict:
    """단독·다가구·근린주택 등의 detail_text를 직접 읽어 용도지역 추출 → 카테고리별 키 집합.
    ilike(불안정) 대신 직접 추출이라 신뢰성 높음. 30분 캐시, Lock 직렬화."""
    import time as _t
    b = _zone_bkt["buckets"]
    if b and not force and _t.time() - _zone_bkt["ts"] < 1800:
        return b
    with _zone_build_lock:
        b = _zone_bkt["buckets"]
        if b and not force and _t.time() - _zone_bkt["ts"] < 1800:
            return b
        appl: list = []
        off = 0
        while True:
            r = auction_db._get("items", {"select": "item_key", "data_class": "eq.현황",
                                           "or": _ZONE_APPL_OR, "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
            appl += [x["item_key"] for x in rows if x.get("item_key")]
            if len(rows) < 1000:
                break
            off += 1000
        zmap = _zone_map(appl)                             # {key: 용도지역 라벨} — detail_text 직접 파싱
        out = {c: set() for c in _ZONE_FILTER_CATS}
        for k, z in zmap.items():
            cat = _zone_categorize(z)
            if cat:
                out[cat].add(k)
        if any(out.values()):                              # 유효 결과만 캐시(빈 clobber 방지)
            _zone_bkt["buckets"] = out
            _zone_bkt["ts"] = _t.time()
    return _zone_bkt["buckets"] or {c: set() for c in _ZONE_FILTER_CATS}


def _zone_filter_keys(zone: Optional[str]) -> Optional[set]:
    if not zone:
        return None
    return set(_zone_build().get(zone, set()))   # 캐시 버킷 보호 위해 사본 반환(호출측 변형 방지)


def _zone_warm() -> None:
    try:
        _zone_build(force=True)
    except Exception:
        pass


_VIOLATION_KEYS: set = set()


def _load_violation_keys() -> None:
    """전수 스캔(_audit_viol.py)이 쓴 _audit_viol_result.json의 viol(건축물대장상 '위반건축물' 키)을 메모리 셋으로 로드.
    표제부 API엔 위반 필드가 없어 건축물대장 문서 스캔 결과를 목록 배지에 사용. 신규 물건은 brief.violation으로도 잡힘."""
    global _VIOLATION_KEYS
    try:
        import json as _json
        p = os.path.join(os.path.dirname(__file__), "..", "_audit_viol_result.json")
        with open(p, encoding="utf-8") as f:
            _VIOLATION_KEYS = set(_json.load(f).get("viol") or [])
    except Exception:
        _VIOLATION_KEYS = set()


_load_violation_keys()


def _enrich_list(items: list) -> None:
    """목록 응답에 '이미 캐시된' 부가정보(준공/세대/승강기·유사거래·시세·매수판정·용도지역)를 붙여 한 프레임에 렌더.
    계산은 하지 않음(캐시만 읽음) → 응답 빠름. 미캐시 항목은 None → 프론트 비동기 폴백이 채움."""
    keys = [it.get("item_key") for it in items if it.get("item_key")]
    if not keys:
        return
    kjoin = ",".join(keys)

    def _safe(fn, default):
        try:
            return fn()
        except Exception:
            return default
    # 캐시 읽기(brief/빌라시세/아파트시세/판정)를 '로컬 캐시만'(remote=False) 병렬로 → Supabase 왕복 0 → ~0.003초.
    #  로컬에 없는 시세는 프론트 fillEstimates가 async로 채움(예열된 것은 즉시 표시).
    # 숙박시설: 여관·생활형숙박시설 세부용도 표시용 brief를 Supabase에서 타깃 로딩(페이지당 소수라 저비용,
    #  로컬 예열 꺼져있거나 클라우드(무예열)서도 목록에 세부용도가 뜨게 함). 이미 캐시된 건 miss 없어 스킵.
    _sukbak_keys = [it.get("item_key") for it in items
                    if "숙박" in (it.get("usage") or "") and it.get("item_key")]
    with _cf.ThreadPoolExecutor(max_workers=5) as ex:
        f_brief = ex.submit(_safe, lambda: _load_briefs_from_db(keys, remote=False), None)
        f_sukbak = ex.submit(_safe, lambda: _load_briefs_from_db(_sukbak_keys, remote=True) if _sukbak_keys else None, None)
        f_villa = ex.submit(_safe, lambda: auction_villa_ests(kjoin, compute=False), {})  # 빌라는 메모리캐시 없어 remote=True(keep-alive로 빠름·캐시도 채움)
        f_apt = ex.submit(_safe, lambda: auction_apt_ests(kjoin, compute=False, remote=False), {})
        f_grade = ex.submit(_safe, _grade_buckets, {})   # 캐시 즉시반환(워밍이 갱신)
        f_brief.result()
        f_sukbak.result()
        villa = f_villa.result() or {}
        apt = f_apt.result() or {}
        grade_rev = {k: g for g, s in (f_grade.result() or {}).items() for k in s}
    _bd = _safe(_baedang_index, {}) or {}        # 다가구·근린주택 배당요구신청 건수(캐시 즉시)
    _dg = _safe(_dagagu_index, {}) or {}          # 다가구 우량(캐시 즉시)
    _cp = _safe(_compete_index, {}) or {}          # 경쟁분산(전 유형, 같은 부류·기일·법원, 캐시 즉시)
    from auction_analysis.crawler_analysis import elevator_caution
    for it in items:
        k = it.get("item_key")
        b = _brief_cache.get(k)
        if isinstance(b, dict) and b.get("available"):
            it["brief"] = b
            # 숙박시설 → 건축물대장 세부용도(여관·생활형숙박시설 등)로 표시 정밀화. usage_name(필터·그룹)은 유지.
            if b.get("usage_detail") and "숙박" in (it.get("usage") or ""):
                it["usage_disp"] = b["usage_detail"]
        if (isinstance(b, dict) and b.get("violation")) or k in _VIOLATION_KEYS:
            it["violation"] = True               # 위반건축물(건축물대장 스탬프 — 전수스캔셋 또는 brief)
        # 다세대·도시형 4층↑ 승강기없음/미상 → 목록 칩(매수검토 사유 표시)
        ev = elevator_caution(it.get("usage"), it.get("address"),
                              b.get("elevator") if isinstance(b, dict) else None)
        if ev:
            it["elev_caution"] = ev.split(" — ")[0].replace(" ", "")   # "승강기없음·5층"
        s = _similar_cache.get(k)
        if s:
            it["similar"] = s
        e = villa.get(k) or apt.get(k)
        if isinstance(e, dict) and e.get("price"):
            it["est"] = e["price"]
            it["est_kind"] = e.get("kind")
            if e.get("trades_3m") is not None:
                it["trades_3m"] = e["trades_3m"]
        g = grade_rev.get(k) or it.get("buy_grade")   # in-메모리 버킷(로컬) 우선 → 없으면 items.buy_grade 컬럼(클라우드)
        if g:
            it["grade"] = g
        if it.get("violation") and it.get("grade") == "매수양호":
            it["grade"] = "매수검토"          # 위반건축물 → 매수검토(신규 물건도 brief.violation으로 즉시 자동 반영)
        bd = _bd.get(k)                          # 다가구·근린주택 배당요구신청 건수(상세 보유면 0도 표기)
        if bd is not None:
            it["baedang"] = bd
        dg = _dg.get(k)                          # 다가구 우량(임차보증금+근저당 ≥ 감정가)
        if dg:
            it["dagagu"] = dg
        cp = _cp.get(k)                          # 경쟁분산(전 유형, 같은 부류·매각기일·법원 동시진행 건수)
        if cp:
            it["compete"] = cp
    # 용도지역: 단독·다가구·근린주택·주택(아파트/다세대/도시형 제외)만 표기
    zkeys = [it.get("item_key") for it in items
             if it.get("item_key")
             and re.search(r"단독|다가구|근린|주택", it.get("usage") or "")
             and not re.search(r"아파트|오피스텔|다세대|연립|도시형", it.get("usage") or "")]
    if zkeys:
        zmap = _safe(lambda: _zone_map(zkeys), {})
        for it in items:
            z = (zmap or {}).get(it.get("item_key"))
            if z:
                it["zone"] = z


@app.get("/auctions")
def auctions(
    group: Optional[list[str]] = Query(None, description="수집그룹(다중)"),
    usage: Optional[list[str]] = Query(None, description="현황용도(다중)"),
    keyword: Optional[str] = None,
    region: Optional[str] = Query(None, description="소재지(구군/동 키워드)"),
    regions: Optional[list[str]] = Query(None, description="소재지 다중(+버튼, 각 '시도 구군 동' 문자열, 지역끼리 OR)"),
    sido: Optional[str] = Query(None, description="시/도 표준명(변형 OR 매칭)"),
    year: Optional[str] = Query(None, description="사건 연도"),
    caseno: Optional[str] = Query(None, description="타경번호(사건번호 숫자부)"),
    court: Optional[str] = Query(None, description="법원(계)"),
    court_code: Optional[list[str]] = Query(None, description="법원코드(다중)"),
    status: Optional[str] = Query(None, description="물건상태(유찰/신건/매각 등)"),
    special: Optional[list[str]] = Query(None, description="특수물건(tags 부분일치, 다중=AND)"),
    type_filter: Optional[list[str]] = Query(None, alias="type", description="유형별 필터(예: apt_deposit_unknown)"),
    fuel: Optional[list[str]] = Query(None, description="연료별(휘발유/경유/LPG/하이브리드/전기)"),
    brand: Optional[list[str]] = Query(None, description="브랜드별(현대/기아/BMW/벤츠 등)"),
    grade: Optional[list[str]] = Query(None, description="매수판정(매수양호/매수검토/매수금지)"),
    zone: Optional[str] = Query(None, description="용도지역(주거지역/준주거지역/상업지역/준공업지역/공업지역/녹지지역/관리지역)"),
    buy_ok: bool = Query(False, description="매수 양호 차량만"),
    sort2: Optional[str] = Query(None, description="2차 정렬"),
    appraisal_min: Optional[int] = None,
    appraisal_max: Optional[int] = None,
    price_min: Optional[int] = None,
    price_max: Optional[int] = None,
    fail_min: Optional[int] = None,
    fail_max: Optional[int] = None,
    barea_min: Optional[float] = None,
    barea_max: Optional[float] = None,
    invest_min: Optional[int] = Query(None, description="투자금(선금) 최소(원)"),
    invest_max: Optional[int] = Query(None, description="투자금(선금) 최대(원)"),
    sell_from: Optional[str] = Query(None, description="매각기일 시작(YYYY-MM-DD)"),
    sell_to: Optional[str] = Query(None, description="매각기일 종료(YYYY-MM-DD)"),
    sort: str = "사건번호",
    simsort: Optional[str] = Query(None, description="유사거래 많은순 전역 정렬(서버, _similar_cache 기준)"),
    limit: int = Query(15, le=200),
    offset: int = 0,
) -> dict:
    usage = _expand_vehicle_usages(usage)         # SUV/승용자동차 → 실제 모델명 usage_name 집합
    if caseno:   # 사건번호 = 특정 물건 직접조회 → 목록 좁히는 필터(등급/유형/특수/지역/법원/상태) 전부 무시.
        grade = type_filter = fuel = brand = zone = special = None   #  매수판정 등이 localStorage로 복원돼 사건번호 검색에
        buy_ok = False                                                #  AND로 걸려, 그 물건 등급/유형이 필터와 다르면 0건 나오던 버그
        barea_min = barea_max = invest_min = invest_max = None        #  (2025타경100006=매수금지 물건이 매수양호 필터에 걸려 사라짐)
        group = usage = keyword = region = regions = sido = court = court_code = status = None
    _use_col = bool(grade) and _buy_grade_ready()  # 컬럼 준비되면 매수판정을 컬럼 WHERE로(IN-리스트 회피)
    item_keys = _combine_item_keys(_type_filter_keys(type_filter), _fuel_filter_keys(fuel),
                                   _brand_filter_keys(brand), _zone_filter_keys(zone),
                                   None if _use_col else _grade_filter_keys(grade),
                                   _buy_ok_keys() if buy_ok else None,
                                   _barea_filter_keys(barea_min, barea_max),  # +건물면적(area_text 전용 파싱)
                                   _invest_filter_keys(invest_min, invest_max))  # +투자금(min_price+면적 산출)
    kw = dict(group=group, usages=usage, keyword=keyword,
              region=region, regions=regions, sido=sido, year=year, caseno=caseno, court=court, court_code=court_code,
              result_prefix=status, special=special, item_keys=item_keys,
              buy_grade=(grade if _use_col else None),
              appraisal_min=appraisal_min, appraisal_max=appraisal_max,
              price_min=price_min, price_max=price_max,
              fail_min=fail_min, fail_max=fail_max,
              barea_min=None, barea_max=None,   # building_area 컬럼은 NULL이라 위 키셋으로 대체(컬럼필터 끔)
              sell_from=sell_from, sell_to=sell_to)
    # 큰 item_keys(용도지역 등) 필터에서 목록·카운트를 병렬 실행하면 청크 IN-리스트 카운트가 0으로
    #  깨지는 경우가 있어, item_keys가 큰 경우는 순차 실행(작은 경우는 병렬 유지로 속도).
    if simsort:
        # 유사거래 많은순: 필터셋 '전체' item_key를 _similar_cache(반경500m 카운트) 기준 전역 정렬 후 페이지.
        # ⚠️ item_keys(유형필터/용도지역 등)를 _filters에 그대로 넣으면 큰 IN-리스트로 URL이 깨져 0건/에러 →
        #    item_keys는 파이썬 교집합으로 처리(다른 필터만 _filters로).
        kw2 = dict(kw); ikset = kw2.pop("item_keys", None)
        if isinstance(ikset, (set, list)) and not any(v not in (None, [], "") for v in kw2.values()):
            all_keys = list(ikset)                              # 유형필터만 → 키셋 그대로(쿼리 불필요)
        else:
            all_keys = auction_db.filtered_item_keys(**kw2)     # 지역·상태 등 다른 필터만 적용
            if isinstance(ikset, (set, list)):
                _s = set(ikset); all_keys = [k for k in all_keys if k in _s]
        all_keys.sort(key=lambda k: _similar_cache.get(k, -1), reverse=True)
        total = len(all_keys)
        items = auction_db.summaries_by_keys(all_keys[offset:offset + limit])
    elif isinstance(item_keys, set) and len(item_keys) > 250:
        items = auction_db.list_auctions(limit=limit, offset=offset, sort=sort, sort2=sort2, **kw)
        total = auction_db.count_filtered(**kw)
    else:
        with _cf.ThreadPoolExecutor(max_workers=2) as ex:   # 목록·총건수 병렬
            f_items = ex.submit(auction_db.list_auctions, limit=limit, offset=offset,
                                sort=sort, sort2=sort2, **kw)
            f_total = ex.submit(auction_db.count_filtered, **kw)
            items = f_items.result()
            total = f_total.result()
    _enrich_list(items)                          # 캐시된 부가정보(준공/시세/유사거래/판정)를 응답에 합침
    return {"total": total, "count": len(items), "offset": offset,
            "limit": limit, "items": items}


@app.get("/auctions/stats")
def auction_stats(
    group: Optional[list[str]] = Query(None),
    usage: Optional[list[str]] = Query(None),
    keyword: Optional[str] = None,
    region: Optional[str] = None,
    sido: Optional[str] = None,
    year: Optional[str] = None,
    caseno: Optional[str] = None,
    court: Optional[str] = None,
    court_code: Optional[list[str]] = Query(None),
    special: Optional[list[str]] = Query(None),
    type_filter: Optional[list[str]] = Query(None, alias="type"),
    fuel: Optional[list[str]] = Query(None),
    brand: Optional[list[str]] = Query(None),
    grade: Optional[list[str]] = Query(None),
    zone: Optional[str] = Query(None),
    buy_ok: bool = Query(False),
    appraisal_min: Optional[int] = None,
    appraisal_max: Optional[int] = None,
    price_min: Optional[int] = None,
    price_max: Optional[int] = None,
    fail_min: Optional[int] = None,
    fail_max: Optional[int] = None,
    barea_min: Optional[float] = None,
    barea_max: Optional[float] = None,
    invest_min: Optional[int] = None,
    invest_max: Optional[int] = None,
    sell_from: Optional[str] = None,
    sell_to: Optional[str] = None,
) -> dict:
    """현재 검색조건 기준 물건 상태별 건수(물건통계)."""
    usage = _expand_vehicle_usages(usage)         # SUV/승용자동차 → 실제 모델명 usage_name 집합
    if caseno:   # 사건번호 = 특정 물건 직접조회 → 목록 좁히는 필터 전부 무시(위 /auctions 와 동일, 물건통계도 0 나오던 버그)
        grade = type_filter = fuel = brand = zone = special = None
        buy_ok = False
        barea_min = barea_max = invest_min = invest_max = None
        group = usage = keyword = region = sido = court = court_code = None
    _use_col = bool(grade) and _buy_grade_ready()  # 컬럼 준비되면 매수판정을 컬럼 WHERE로
    item_keys = _combine_item_keys(_type_filter_keys(type_filter), _fuel_filter_keys(fuel),
                                   _brand_filter_keys(brand), _zone_filter_keys(zone),
                                   None if _use_col else _grade_filter_keys(grade),
                                   _buy_ok_keys() if buy_ok else None,
                                   _barea_filter_keys(barea_min, barea_max),  # +건물면적(area_text 전용 파싱)
                                   _invest_filter_keys(invest_min, invest_max))  # +투자금(min_price+면적 산출)
    counts = auction_db.status_stats(
        group=group, usages=usage, keyword=keyword, region=region, sido=sido, special=special,
        item_keys=item_keys, buy_grade=(grade if _use_col else None),
        year=year, caseno=caseno, court=court, court_code=court_code,
        appraisal_min=appraisal_min, appraisal_max=appraisal_max,
        price_min=price_min, price_max=price_max,
        fail_min=fail_min, fail_max=fail_max,
        barea_min=None, barea_max=None,   # building_area 컬럼 NULL → 위 키셋으로 대체
        sell_from=sell_from, sell_to=sell_to,
    )
    return {"counts": counts}


@app.get("/auctions/facets")
def auction_facets() -> dict:
    return {"groups": auction_db.group_counts()}


@app.get("/auctions/usages")
def auction_usages(group: str = "주거용") -> dict:
    return {"group": group, "usages": auction_db.usages_in_group(group)}


_regions_cache: Optional[dict] = None


@app.get("/auctions/regions")
def auction_regions() -> dict:
    """소재지(시도→구군→동읍) 트리 + 연도 + 법원(계) 목록. 1회 집계 후 캐시."""
    global _regions_cache
    if _regions_cache is None:
        _regions_cache = auction_db.region_facets()
    return _regions_cache


@app.get("/auction")
def auction_detail(item_key: str, user: dict = Depends(require_national_user)) -> dict:
    try:
        d = auction_db.get_auction(item_key)
    except Exception:
        # Supabase 지연(크롤러 과부하 등) → 500/"찾을 수 없음" 오표시 대신 503(프런트가 자동 재시도)
        raise HTTPException(503, "DB 혼잡 — 잠시 후 다시 시도")
    if d is None:
        raise HTTPException(404, f"물건을 찾을 수 없습니다: {item_key}")
    # 숙박시설 → 건축물대장 세부용도(여관·생활형숙박시설 등)를 상세 용도 표시에 반영(숙박만 brief 조회 — 다른 유형 지연 없음)
    try:
        if "숙박" in (d.get("usage") or ""):
            _b = _get_brief(item_key)
            if isinstance(_b, dict) and _b.get("usage_detail"):
                d["usage_disp"] = _b["usage_detail"]
    except Exception:
        pass
    return d


def _cached_doc(prefix: str, item_key: str, compute) -> dict:
    """문서분석(등기·감정평가서·명세서·차량) 결과를 Supabase api_cache(prefix:)에 영구 저장.
    ①DB 있으면 즉시 반환(파싱 X) ②없으면 계산 후 DB 저장. 재시작·재배포에도 유지."""
    ck = prefix + ":" + item_key
    try:
        db = auction_db.cache_get_many([ck]).get(ck)
    except Exception:
        db = None
    if isinstance(db, dict):                # 캐시 있으면 available 여부 무관 즉시 반환
        return db                           #  → 스캔본/서류없음(available=False)도 캐시해 매번 재파싱·블로킹 방지
    out = compute()
    if isinstance(out, dict):
        try:
            auction_db.cache_save(ck, out)  # available=False 결과도 저장(재파싱 안 함)
        except Exception:
            pass
    return out


@app.get("/auction/case_objects")
def case_objects(item_key: str) -> dict:
    """같은 사건(법원|연도|사건일련)의 물번 목록 + 결과 — 상세 물번 선택기용."""
    parts = item_key.split("|")
    if len(parts) < 4:
        return {"objects": []}
    prefix = "|".join(parts[:3]) + "|"
    try:
        r = auction_db._get("items", {"select": "item_key,obj_no,result,status_reason",
                                       "item_key": f"like.{prefix}*", "order": "obj_no.asc"})
        rows = r.json() if r.status_code in (200, 206) else []
    except Exception:
        rows = []
    objs = [{"item_key": x["item_key"], "obj_no": x.get("obj_no"),
             "result": (x.get("result") or x.get("status_reason") or "")} for x in rows]
    return {"objects": objs}


@app.get("/auction/analysis")
def auction_analysis(item_key: str) -> dict:
    """권리분석(말소기준·인수/소멸). ①크롤러 구조화DB(analyzed_at) 우선 ②없으면 등기 PDF 파싱 폴백."""
    from auction_analysis.crawler_analysis import analyze_from_crawler
    res = None
    try:
        res = analyze_from_crawler(auction_db, item_key)   # 크롤러 구조화 데이터 우선(판정 포함)
    except Exception:
        res = None
    if res is None:                                        # 미분석 물건 → 기존 PDF 파싱 폴백
        from auction_analysis.doc_analysis import analyze_registry
        res = _cached_doc("analysis", item_key,
                          lambda: analyze_registry(auction_db, item_key))
    # 환매등기(환매특약)는 소유권이전에 딸린 부기등기 → 소멸분은 별도 권리로 표시 안 함.
    #  (선순위=인수 환매는 실제 중요하므로 유지). 캐시 객체는 보존하고 사본으로 반환.
    if isinstance(res, dict) and isinstance(res.get("rights"), list):
        filtered = [r for r in res["rights"]
                    if not (r.get("type") == "환매등기" and r.get("status") == "소멸")]
        if len(filtered) != len(res["rights"]):
            res = {**res, "rights": filtered}
    # 차임(월세) 병합 — item_tenants엔 차임 컬럼이 없어 매각물건명세서(tenant_rents)에서 이름 매칭으로 채움
    if isinstance(res, dict) and res.get("tenants"):
        try:
            from auction_analysis.doc_analysis import analyze_doc_summary
            ds = analyze_doc_summary(auction_db, item_key)
            _nm = lambda s: re.sub(r"\([^)]*\)|\s", "", s or "")   # 괄호(호수·별지)·공백 제거 후 이름 비교
            rmap = {_nm(tr.get("name")): tr
                    for tr in (ds.get("tenant_rents") or []) if tr.get("name")}
            if rmap:
                ts = [dict(t) for t in res["tenants"]]
                for t in ts:
                    src = rmap.get(_nm(t.get("name")))
                    if not src:
                        continue
                    if src.get("rent"):
                        t["rent"] = src["rent"]
                    # 전입일·확정일: item_tenants(현황조사)가 미상일 때만 명세서 값으로 보완(현황조사 우선)
                    if src.get("move_in") and not t.get("move_in"):
                        t["move_in"] = src["move_in"]
                    if src.get("fixed") and not t.get("fixed"):
                        t["fixed"] = src["fixed"]
                res = {**res, "tenants": ts}
        except Exception:
            pass
    if isinstance(res, dict) and isinstance(res.get("tenants"), list):
        res["rent_total"] = sum((t.get("rent") or 0) for t in res["tenants"])   # 월세 합계(패널 보증금합계 옆 표시)
    # 위반건축물 → AI분석 '검토'(위험도 안전이면 주의 상향 + 경고). 목록 매수검토는 _grade_buckets에서 반영.
    if isinstance(res, dict):
        _b = _brief_cache.get(item_key)
        if item_key in _VIOLATION_KEYS or (isinstance(_b, dict) and _b.get("violation")):
            res = dict(res)
            if res.get("risk_level") == "안전":
                res["risk_level"] = "주의"
            res["needs_expert_review"] = True
            res["warnings"] = list(res.get("warnings") or []) + [
                "건축물대장상 위반건축물 — 대출 난항·이행강제금·1주택 비과세가 깨질 수 있음 (원상복구 가능 여부·이행강제금 확인, 매수 검토 필요)"]
    return res


_DOCVIEW_CSS = """
<style id="jh-docview">
  html,body{background:#f4f5f7 !important;}
  body{font-family:-apple-system,"Segoe UI","Malgun Gothic",sans-serif !important;
    color:#1c1c1a !important;max-width:920px;margin:0 auto !important;padding:18px !important;line-height:1.65;}
  .jh-hd{background:#1f4fa3;color:#fff;padding:13px 18px;border-radius:10px;font-size:18px;font-weight:700;
    margin-bottom:16px;display:flex;justify-content:space-between;align-items:center;}
  .jh-hd a{color:#cfe0ff;font-size:13px;text-decoration:none;font-weight:500;}
  table{border-collapse:collapse !important;width:100% !important;margin:12px 0 !important;background:#fff;
    border:1px solid #dfe3ea;border-radius:8px;overflow:hidden;}
  th,td{border:1px solid #e3e7ee !important;padding:9px 12px !important;font-size:13px !important;
    text-align:left;vertical-align:top;color:#1c1c1a;}
  th{background:#eef2fb !important;color:#2c5db0 !important;font-weight:600;white-space:nowrap;}
  h1,h2,h3,h4{color:#1f4fa3 !important;margin:18px 0 8px;}
  b,strong{color:#1f4fa3;}
  img{max-width:100%;height:auto;}
  hr{border:none;border-top:1px solid #dfe3ea;margin:14px 0;}
  a{color:#1f4fa3;}
  .dis-none,.zoom_que_modal,#zoom_mask{display:none !important;}
  [style*="margin-top:-"]{margin-top:0 !important;}   /* 원본 음수마진(margin-top:-80px 등) 핵 제거 — 제목·경고 겹침 방지 */
  .vd_pdftop,.print_cnt,.print_button,.prt_btn,.print{display:none !important;}   /* 원본 인쇄 툴바·버튼(빈 흰 박스) 제거 */
</style>
"""


@app.get("/docview")
def docview(item_key: str, kind: str = "현황조사서") -> HTMLResponse:
    """HTML 서류(부동산표시·사건내역·기일내역·문건접수송달·현황조사서)를 우리 스타일로 입혀 렌더. 원본 /pub/ 깨진 이미지·외부 CSS/JS 제거."""
    url = auction_db.media_url(item_key, kind)
    if not url:
        return HTMLResponse(f'<div style="padding:30px;font-family:sans-serif">{kind} 문서가 없습니다.</div>')
    try:
        html = httpx.get(url, timeout=30, follow_redirects=True).text
    except Exception:
        return HTMLResponse('<div style="padding:30px">문서를 불러오지 못했습니다.</div>')
    # 스피드옥션 상대경로(/pub/...) 외부 CSS·JS·로고는 우리 도메인에서 깨지므로 제거
    html = re.sub(r'<link\b[^>]*href=["\']?/pub/[^>]*>', '', html, flags=re.I)
    html = re.sub(r'<script\b[^>]*src=["\']?/pub/[^>]*>\s*</script>', '', html, flags=re.I)
    # 인쇄 버튼(우리 뷰어에서 불필요) + 깨진 /pub/ 이미지(인쇄아이콘 등) 제거
    html = re.sub(r'<button\b[^>]*id=["\']?print_btn["\']?[^>]*>.*?</button>', '', html, flags=re.I | re.S)
    html = re.sub(r'<img\b[^>]*src=["\']?/pub/[^>]*>', '', html, flags=re.I)
    # 원본 헤더(제목/도움말)·서류목록 내비는 우리 상단바·탭바와 중복 → 제거(내용만 남김)
    html = re.sub(r'<header\b[^>]*>.*?</header>', '', html, flags=re.I | re.S)
    html = re.sub(r'<nav\b[^>]*>.*?</nav>', '', html, flags=re.I | re.S)
    header = (f'<div class="jh-hd">{kind}'
              f'<a href="javascript:history.back()">← 돌아가기</a></div>')
    css = _DOCVIEW_CSS
    if re.search(r'</head>', html, re.I):
        html = re.sub(r'</head>', css + '</head>', html, count=1, flags=re.I)
    elif re.search(r'<html[^>]*>', html, re.I):
        html = re.sub(r'(<html[^>]*>)', r'\1' + css, html, count=1, flags=re.I)
    else:
        html = css + html
    if re.search(r'<body[^>]*>', html, re.I):
        html = re.sub(r'(<body[^>]*>)', r'\1' + header, html, count=1, flags=re.I)
    else:
        html = css + header + html
    return HTMLResponse(html)


@app.get("/auction/appraisal")
def auction_appraisal(item_key: str) -> dict:
    """감정평가서 PDF 파싱 → 물건현황(요항표)·감정평가현황. DB 우선 → 계산 후 DB 저장."""
    from auction_analysis.doc_analysis import analyze_appraisal
    return _cached_doc("appraisal", item_key,
                       lambda: analyze_appraisal(auction_db, item_key))


import json as _json                                          # noqa: E402
import concurrent.futures as _cf                               # noqa: E402
_BRIEF_FILE = os.path.join(_ROOT, "brief_cache.json")
_SIMILAR_FILE = os.path.join(_ROOT, "similar_cache.json")


def _load_brief_cache() -> dict:
    try:
        with open(_BRIEF_FILE, encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


_brief_cache: dict[str, dict] = _load_brief_cache()   # 디스크 영구캐시(available=True만 저장)
_brief_dirty = False
_BRIEF_NEG_TTL = 3 * 86400   # 미등재(건축물대장·문서 없음) 네거티브 캐시 수명: 3일(만료 후 예열·조회가 재시도)


def _load_similar_cache() -> dict:
    try:
        with open(_SIMILAR_FILE, encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


_similar_cache: dict[str, int] = _load_similar_cache()   # item_key -> 유사거래 건수
_similar_dirty = False
_sim_warming = {"on": False}    # simsort 빠진 카운트 백그라운드 보충 중복 스폰 방지


def _save_similar_cache() -> None:
    global _similar_dirty
    if not _similar_dirty:
        return
    try:
        with open(_SIMILAR_FILE, "w", encoding="utf-8") as f:
            _json.dump(_similar_cache, f, ensure_ascii=False)
        _similar_dirty = False
    except Exception:
        pass


def _save_brief_cache() -> None:
    global _brief_dirty
    if not _brief_dirty:
        return
    try:
        # available=True만 디스크 저장. False(주소변환·건축물대장 실패)는 저장 안 함
        #  → resolve_bjd 폴백 개선/데이터 보강 시 재시작하면 자동 재시도(영구 누락 방지).
        keep = {k: v for k, v in _brief_cache.items() if isinstance(v, dict) and v.get("available")}
        with open(_BRIEF_FILE, "w", encoding="utf-8") as f:
            _json.dump(keep, f, ensure_ascii=False)
        _brief_dirty = False
    except Exception:
        pass


_rh_cache: dict[str, list] = {}   # 시군구 연립다세대 실거래(12개월) 풀 캐시


def _pool_from_db(prefix: str, lawd: str) -> Optional[list]:
    """시군구 실거래 풀을 DB(api_cache)에서 로딩(7일 이내만 — 실거래 신선도)."""
    import time as _t
    try:
        d = auction_db.cache_get_many([prefix + lawd]).get(prefix + lawd)
        if isinstance(d, dict) and d.get("trades") is not None and (_t.time() - d.get("ts", 0) < 7 * 86400):
            return d["trades"]
    except Exception:
        pass
    return None


def _pool_to_db(prefix: str, lawd: str, trades: list) -> None:
    import time as _t
    try:
        auction_db.cache_save(prefix + lawd, {"ts": _t.time(), "trades": trades})
    except Exception:
        pass


def _rh_trades(lawd: str) -> list:
    """시군구 12개월 연립다세대 실거래 풀. ①메모리 ②DB(rhpool:, 7일) ③molit 계산+DB저장.
    ※ 아파트 실거래는 미포함: 일반 아파트가 섞여 오염되고 조회가 3배 느려짐."""
    if lawd in _rh_cache:
        return _rh_cache[lawd]
    db = _pool_from_db("rhpool:", lawd)
    if db is not None:
        _rh_cache[lawd] = db
        return db
    try:
        from auction_analysis.molit_source import MolitSource
        res = MolitSource().recent_trades(lawd, months=12) or {}
        tr = res.get("trades") or []
        if res.get("error"):                  # 할당량/오류 → 캐시 안 함(리셋 후 자동 재조회)
            return tr
    except Exception:
        return []
    _rh_cache[lawd] = tr
    _pool_to_db("rhpool:", lawd, tr)          # DB 영구 저장(7일 후 자동 갱신)
    return tr


_shrent_cache: dict[str, list] = {}   # 시군구 단독·다가구 전월세(12개월) 캐시


def _shrent_trades(lawd: str) -> list:
    """시군구 12개월 단독·다가구 전월세 풀. ①메모리 ②DB(shrentpool:, 7일) ③molit 계산+DB저장."""
    if lawd in _shrent_cache:
        return _shrent_cache[lawd]
    db = _pool_from_db("shrentpool:", lawd)
    if db is not None:
        _shrent_cache[lawd] = db
        return db
    try:
        from auction_analysis.molit_source import MolitSource
        res = MolitSource().sh_rent_recent(lawd, months=12) or {}
        tr = res.get("trades") or []
        if res.get("error"):                  # 할당량/오류 → 캐시 안 함
            return tr
    except Exception:
        return []
    _shrent_cache[lawd] = tr
    _pool_to_db("shrentpool:", lawd, tr)
    return tr


@app.get("/auction/dagagu_market")
def auction_dagagu_market(item_key: str) -> dict:
    """다가구·근린주택 주변 단독·다가구 전월세 시세(국토부 — 강의 손품[다방/직방] 대체, 원천데이터)."""
    try:
        r = auction_db._get("items", {"select": "usage_name,address", "item_key": "eq." + item_key, "limit": "1"})
        rows = r.json() if r.status_code == 200 else []
    except Exception:
        rows = []
    if not rows or not _is_dagagu_usage(rows[0].get("usage_name")):
        return {"available": False}
    addr = rows[0].get("address") or ""
    lawd = resolve_lawd(addr)
    if not lawd:
        return {"available": False, "reason": "시군구 코드 없음"}
    trades = _shrent_trades(lawd)
    if not trades:
        return {"available": False, "reason": "주변 단독·다가구 전월세 실거래 없음(또는 할당량 초과)"}
    mj = re.search(r"([가-힣]+(?:동|읍|면|리|가))", addr)
    umd = mj.group(1) if mj else None
    loc = [t for t in trades if umd and t.get("umd") == umd]
    use = loc if len(loc) >= 5 else trades
    scope = ("법정동 " + umd) if len(loc) >= 5 else "시군구 전체"

    def _med(xs):
        xs = sorted(xs)
        return xs[len(xs) // 2] if xs else 0

    def _agg(src):
        wol = [t for t in src if t.get("rent")]
        jeon = [t for t in src if not t.get("rent")]
        return {"wolse_n": len(wol), "wolse_deposit": _med([t["deposit"] for t in wol]),
                "wolse_rent": _med([t["rent"] for t in wol]),
                "jeonse_n": len(jeon), "jeonse_deposit": _med([t["deposit"] for t in jeon])}

    oneroom = [t for t in use if 0 < (t.get("area") or 0) <= 40]   # 원룸급(≤40㎡) — 강의 기준 판정용
    _BR = [(0, 20, "~20㎡"), (20, 30, "20~30㎡"), (30, 40, "30~40㎡"),
           (40, 50, "40~50㎡"), (50, 70, "50~70㎡"), (70, 10 ** 9, "70㎡~")]
    brackets = []
    for lo, hi, lab in _BR:
        sub = [t for t in use if lo < (t.get("area") or 0) <= hi]
        if sub:
            brackets.append({"label": lab, "agg": _agg(sub)})
    return {"available": True, "scope": scope, "months": 12, "total": len(use),
            "all": _agg(use), "oneroom": _agg(oneroom), "brackets": brackets}


_apt_trades_cache: dict[str, list] = {}   # 시군구 아파트 실거래(12개월) 캐시


def _apt_trades(lawd: str, months: int = 12) -> list:
    """시군구 아파트 실거래 풀. ①메모리 ②DB(aptpool:, 7일) ③molit 계산+DB저장."""
    if lawd in _apt_trades_cache:
        return _apt_trades_cache[lawd]
    db = _pool_from_db("aptpool:", lawd)
    if db is not None:
        _apt_trades_cache[lawd] = db
        return db
    try:
        from auction_analysis.molit_source import MolitSource
        tr = (MolitSource().apt_recent_trades(lawd, months=months) or {}).get("trades") or []
    except Exception:
        tr = []
    if tr:                                     # 빈 결과(할당량 등)는 캐시 안 함
        _apt_trades_cache[lawd] = tr
        _pool_to_db("aptpool:", lawd, tr)
    return tr


# ── 서버측 지오코딩(V-World) + 디스크 캐시: 목록/상세 1km 필터 공용 ──
from auction_analysis.geocode_source import VGeocoder, haversine_m  # noqa: E402
_geocoder = VGeocoder()
_GEO_FILE = os.path.join(_ROOT, "geocode_cache.json")


def _load_geo_cache() -> dict:
    try:
        with open(_GEO_FILE, encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


_geo_cache: dict[str, object] = _load_geo_cache()   # 주소 -> [lng,lat] or None
_geo_dirty = False


def _save_geo_cache() -> None:
    global _geo_dirty
    if not _geo_dirty:
        return
    try:
        with open(_GEO_FILE, "w", encoding="utf-8") as f:
            _json.dump(_geo_cache, f, ensure_ascii=False)
        _geo_dirty = False
    except Exception:
        pass


def _geo_preload(addrs) -> None:
    """주소들의 좌표를 Supabase api_cache(geo:)에서 일괄 로딩(메모리에 없는 것만). 대량 지오코딩 전 호출."""
    miss = [a for a in dict.fromkeys(addrs) if a and a not in _geo_cache]
    for i in range(0, len(miss), 100):
        try:
            rows = auction_db.cache_get_many(["geo:" + a for a in miss[i:i + 100]])
            for a in miss[i:i + 100]:
                dd = rows.get("geo:" + a)
                if isinstance(dd, dict) and dd.get("ll"):
                    _geo_cache[a] = dd["ll"]
        except Exception:
            pass


def _geocode(addr: str):
    """주소 → (lng,lat). 성공만 캐시(메모리+로컬+DB). 실패는 캐시 안 함(일시 실패 재시도)."""
    global _geo_dirty
    if not addr:
        return None
    v = _geo_cache.get(addr)
    if v:                          # 성공 캐시만 사용(None/없음은 재조회)
        return tuple(v)
    ll = _geocoder.coord(addr)
    if ll:                         # 성공 시에만 캐시(실패는 저장 안 함)
        _geo_cache[addr] = list(ll)
        _geo_dirty = True
        try:
            auction_db.cache_save("geo:" + addr, {"ll": list(ll)})   # DB 영구 저장(재배포에도 유지)
        except Exception:
            pass
    return ll


# ── 용도지역(V-World 토지이용계획) 보완: detail_text에 용도지역 없는 물건용. 디스크+DB 캐시 ──
#   경매 원천데이터에 용도지역이 없는 농촌형 물건(~21%)을 좌표→V-World LT_C_UQ111로 채운다.
from auction_analysis.landuse_source import LandUseSource  # noqa: E402
_landuse_src = LandUseSource()
_LU_FILE = os.path.join(_ROOT, "landuse_cache.json")


def _load_lu_cache() -> dict:
    try:
        with open(_LU_FILE, encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


_lu_cache: dict = _load_lu_cache()   # item_key -> 용도지역 라벨 | "NF"(조회했으나 없음)
_lu_dirty = False


def _save_lu_cache() -> None:
    global _lu_dirty
    if not _lu_dirty:
        return
    try:
        with open(_LU_FILE, "w", encoding="utf-8") as f:
            _json.dump(_lu_cache, f, ensure_ascii=False)
        _lu_dirty = False
    except Exception:
        pass


def _clean_addr_for_geo(addr: str) -> str:
    """지오코딩용 주소 정리: '외 N필지'·괄호(건물명 등) 제거 → 기본 지번만."""
    addr = re.sub(r"\s*외\s*\d+\s*필지.*$", "", addr or "")
    addr = re.sub(r"\s*\([^)]*\)", "", addr)
    return addr.strip()


def _landuse_lookup(item_key: str, address: str) -> str:
    """item_key의 용도지역(V-World). 캐시 우선. 미캐시면 1회 좌표→V-World 조회 후 캐시.
    성공=라벨, 좌표는 되나 용도지역 없음='NF'(재조회 방지) → 둘 다 표시는 ''로 반환."""
    global _lu_dirty
    v = _lu_cache.get(item_key)
    if v is not None:
        return "" if v == "NF" else v
    ll = _geocode(_clean_addr_for_geo(address)) if address else None
    label, status = _landuse_src.zone_by_coord(*ll) if ll else (None, "ERROR")
    if label:
        _lu_cache[item_key] = label
        _lu_dirty = True
        try:
            auction_db.cache_save("lu:" + item_key, {"z": label})
        except Exception:
            pass
        return label
    if status == "NOT_FOUND":       # '진짜 없음'만 NF 캐시(재조회 방지). 에러/쿼터는 캐시 안 함→재시도
        _lu_cache[item_key] = "NF"
        _lu_dirty = True
        try:
            auction_db.cache_save("lu:" + item_key, {"z": "NF"})
        except Exception:
            pass
    return ""                      # 좌표실패/일시오류는 캐시 안 함(추후 재시도 여지)


def _landuse_prewarm(limit: Optional[int] = None, sleep: float = 0.4) -> dict:
    """detail_text에 용도지역 없는 표기대상 물건을 V-World로 채워 캐시.
    순차+간격(V-World 쿼터 보호). 완료 시 용도지역 버킷 재빌드."""
    todo: list = []          # [(item_key, address)] — 미충족·미조회만
    off = 0
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key,address,detail_text",
                                          "data_class": "eq.현황", "or": _ZONE_APPL_OR,
                                          "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            rows = []
        for x in rows:
            k = x.get("item_key")
            if not k or k in _lu_cache:
                continue
            if _extract_zone(x.get("detail_text") or ""):   # detail_text에 이미 있음
                continue
            todo.append((k, x.get("address") or ""))
        if len(rows) < 1000:
            break
        off += 1000
    total = len(todo)
    work = todo[:limit] if limit else todo
    filled = 0
    done = 0
    #  병렬(4워커)로 좌표+V-World 조회 — 순차 ~9분을 ~1.5분으로. dict 갱신은 GIL로 안전.
    with _cf.ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(_landuse_lookup, k, addr): k for k, addr in work}
        for fu in _cf.as_completed(futs):
            done += 1
            try:
                if fu.result():
                    filled += 1
            except Exception:
                pass
            if done % 50 == 0:          # 주기적 디스크 저장(중단돼도 진행 보존)
                _save_lu_cache()
    _save_lu_cache()
    if filled:
        for k in list(_zone_cache):      # detail_text 추출 ""였던 키 무효화 → 보완값 반영
            if not _zone_cache.get(k):
                _zone_cache.pop(k, None)
        try:
            _zone_build(force=True)
        except Exception:
            pass
    return {"checked": len(work), "filled": filled, "remaining": max(0, total - len(work))}


def _landuse_warm() -> None:
    """백그라운드: 미조회 용도지역을 천천히 채움(서버 시작 시 1회)."""
    try:
        _landuse_prewarm()
    except Exception:
        pass


def _nearby_filtered(d: dict, geocode: bool = True) -> Optional[dict]:
    """면적±10㎡·층±1 + (geocode 시) 반경 1km 필터된 유사거래.
    geocode=False면 지오코딩 생략하고 시군구 면적/층 후보만(목록 카운트용 — V-World 무관·빠름).
    반환 dict: sigungu_prefix·addr_prefix·addr_jibun·prop_area·prop_floor·prop_lng/lat·geo_ok·trades[](lng,lat 포함)."""
    usage = d.get("usage") or ""
    addr = d.get("address") or ""
    if not re.search(r"다세대|연립|빌라|도시형", usage):
        return None
    lawd = resolve_lawd(addr)
    if not lawd:
        return None
    prop_area = _area_num(d.get("building_area"), d.get("area_text"))
    fm = re.search(r"(\d+)\s*층", addr)
    prop_floor = int(fm.group(1)) if fm else None
    # 시군구(시도+시/군/구) 추출 — 지번주소·도로명주소 모두 대응. 동/읍/면/리·도로명(로/길)·지번 토큰에서 중단.
    #  (기존 '동 지번' 가정은 도로명주소에서 sigungu_prefix에 도로명이 섞여 후보 지오코딩이 전부 물건좌표를 반환 → 마커 겹침·공시가격 깨짐)
    _sgg_toks = []
    for _tk in addr.split():
        if re.search(r"(동|읍|면|리|로|길|가|번길)$", _tk) or re.match(r"^\d", _tk):
            break
        _sgg_toks.append(_tk)
    sigungu_prefix = " ".join(_sgg_toks) or re.sub(r"\s+\S+$", "", addr)
    pm = re.match(r"^(.*?(?:동|읍|면|리))\s+\d", addr)
    jm = re.search(r"(?:동|읍|면|리)\s+(\d+(?:-\d+)?)", addr)
    addr_jibun = jm.group(1) if jm else ""
    addr_prefix = pm.group(1) if pm else sigungu_prefix       # 지번주소는 '…동', 도로명은 시군구
    base = {"sigungu_prefix": sigungu_prefix, "addr_prefix": addr_prefix,
            "addr_jibun": addr_jibun, "prop_area": prop_area, "prop_floor": prop_floor}
    # 연립다세대는 면적·위치·연식이 시세를 좌우하고 층 영향이 작음. 층 하드필터는 고층 빌라(예: 7층)에서
    # 대부분(1~5층) 거래를 통째로 잘라내므로 제거 → 면적±10㎡ + 반경1km만 적용. 층 정밀비교는 프런트 '가장 유사' 그룹.
    cand = []
    for t in _rh_trades(lawd):
        a = t.get("area") or 0
        if prop_area and abs(a - prop_area) > 10:
            continue
        cand.append(t)
    if not geocode:                             # 목록 카운트용: 지오코딩 생략(시군구 면적/층 수)
        base["geo_ok"] = False
        base["trades"] = cand[:200]
        return base
    _geo_preload([(addr_prefix + " " + addr_jibun).strip(), addr])   # DB 좌표 우선
    # 도로명주소(지번 없음)는 시군구만으론 물건 좌표가 안 나오므로 전체주소로 지오코딩.
    pc = (_geocode((addr_prefix + " " + addr_jibun).strip()) if addr_jibun else None) or _geocode(addr)
    base["prop_lng"], base["prop_lat"] = (pc[0], pc[1]) if pc else (None, None)
    if not pc:                                  # 물건 지오코딩 실패 → 1km 불가, 시군구 후보 폴백
        base["geo_ok"] = False
        base["trades"] = cand[:120]
        return base
    blds: dict[str, list] = {}
    for t in cand:
        blds.setdefault(f"{t.get('umd')} {t.get('jibun')}", []).append(t)
    keys = list(blds.keys())[:600]
    coords: dict[str, object] = {}
    if keys:
        _geo_preload([sigungu_prefix + " " + k for k in keys])   # 후보 건물 좌표 DB에서 일괄(재배포에도 유지)
        with _cf.ThreadPoolExecutor(max_workers=8) as ex:
            for k, ll in ex.map(lambda k: (k, _geocode(sigungu_prefix + " " + k)), keys):
                coords[k] = ll
        _save_geo_cache()
    trades = []
    for k in keys:
        ll = coords.get(k)
        if not ll or haversine_m(pc[0], pc[1], ll[0], ll[1]) > 1000:
            continue
        for t in blds[k]:
            tt = dict(t)
            tt["lng"], tt["lat"] = ll[0], ll[1]
            trades.append(tt)
    # 후보 건물 지오코딩이 대부분 실패(V-World 할당량 등)하면 반경결과 불신뢰 → geo_ok=False(캐시/집계 제외, 추후 재시도)
    resolved = sum(1 for k in keys if coords.get(k))
    base["geo_ok"] = (not keys) or (resolved >= max(3, int(len(keys) * 0.3)))
    base["trades"] = trades
    return base


# (제거됨) _similar_count: 시군구(구/군) 전체 카운트였으나 호출처 0의 죽은 코드.
#  목록 "유사거래 N건"은 _get_similar→_compute_similar→auction_nearby_trades(반경 500m 트림)로 산출됨(상세와 동일).


def _pdf_text_pages(url: str, max_pages: int, max_bytes: int = 0) -> str:
    """PDF 앞 max_pages 페이지만 텍스트 추출. max_bytes>0이면 그보다 큰 PDF는 스킵('')
    — 대용량 스캔본 감정평가서(파싱 수십초)가 목록을 막지 않게."""
    import io
    import pdfplumber
    data = httpx.get(url, timeout=40, follow_redirects=True).content
    if max_bytes and len(data) > max_bytes:
        return ""
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages[:max_pages])


def _doc_building_brief(item_key: str, want_fields: bool = True) -> dict:
    """저장 문서에서 준공·세대·승강기 + **위반건축물** 추출(API 쿼터 미사용). 결과는 디스크 캐시(_compute_brief)로 1회만.
    want_fields=False면 위반 플래그만 필요(준공/세대/승강기 폴백용 감정평가서 파싱 생략 — 위반은 건축물대장 문서에만 있음).
    ① 건축물대장(작음): 표제부=전부, 전유부=준공.
    ② 부족분은 감정평가서 앞 14쪽: 'N개호'(세대수)·'사용승인'(준공)·'승강기설비'(승강기). (~1.6초)"""
    from auction_analysis.building_doc_parser import (
        parse_bldg_doc, parse_appraisal_bldg, merge_doc_brief)
    bldg = appr = None
    url = auction_db.media_url(item_key, "건축물대장")
    if url:
        try:
            bldg = parse_bldg_doc(_pdf_text_pages(url, 4))
        except Exception:
            bldg = None
    need = want_fields and ((not bldg) or (not bldg.get("build_year")) or (not bldg.get("units"))
                            or (bldg.get("elevator") is None))
    if need:
        au = auction_db.media_url(item_key, "감정평가서")
        if au:
            try:
                # 4MB 초과(스캔 위주 대용량)는 스킵 → 파싱 수십초로 목록 막힘 방지
                appr = parse_appraisal_bldg(_pdf_text_pages(au, 14, max_bytes=4_000_000))
            except Exception:
                appr = None
    return merge_doc_brief(bldg, appr)


def _compute_brief(item_key: str) -> dict:
    """목록용 경량. 주거용=준공·세대·승강기, 차량외=연식·주행거리.
    준공/세대/승강기는 ①저장문서(건축물대장·감정평가서) → ②건축물대장 API 순(쿼터 절약)."""
    out = {"available": False}
    d = auction_db.get_auction(item_key)
    if d:
        usage = d.get("usage") or ""
        addr = d.get("address") or ""
        if (d.get("group") or "") == "차량외":
            from auction_analysis.doc_analysis import analyze_vehicle
            v = analyze_vehicle(auction_db, item_key)
            if v.get("available"):
                fm = {f["label"]: f["value"] for f in v.get("fields", [])}
                if fm.get("연식") or fm.get("주행거리"):
                    out = {"available": True, "kind": "vehicle",
                           "year": fm.get("연식"), "mileage": fm.get("주행거리"),
                           "fuel": _classify_fuel(fm.get("사용연료") or ""),   # 5버킷 라벨(필터와 일치)
                           "grade": v.get("grade")}
            return out
        if re.search(r"아파트|오피스텔", usage):
            lawd = resolve_lawd(addr)
            name = _apt_name_from_addr(addr)
            if lawd and name:
                b = kapt.brief(lawd, name)
                if b and (b.get("build_year") or b.get("households")):
                    out = {"available": True, "unit_label": "세대", **b}
        if not out.get("available") and re.search(r"아파트|오피스텔|다세대|연립|빌라|도시형|다가구|단독|주택|숙박", usage):
            by = un = ul = ev = None
            used_api = used_doc = False
            # ① 건축물대장 API 우선(표제부 — 정확. 트래픽 증가됨). 결과는 DB 저장돼 1회만.
            try:
                bi = building.info(addr)
            except Exception:
                bi = None
            purpose = ""
            if bi:
                by = bi.get("build_year")
                purpose = bi.get("purpose") or ""
                # 집합건물(아파트·오피스텔·다세대·연립·도시형)인데 unit_label='호'면, 지번 조회로
                #  잘못된 동(부속/전유)을 읽은 것 → 세대수·승강기 신뢰 안 함(준공년도만 사용).
                _collective = bool(re.search(r"아파트|오피스텔|다세대|연립|빌라|도시형", usage))
                _is_ho = (bi.get("unit_label") == "호")
                if bi.get("units") and not (_collective and _is_ho):
                    un, ul = bi.get("units"), bi.get("unit_label")
                if bi.get("elevator") is not None and not (_collective and _is_ho):
                    ev = int(bi.get("elevator") or 0) > 0
                used_api = bool(by or un or (ev is not None))
            # ② 저장문서: 위반건축물(문서 스탬프에만 존재 — 표제부 API엔 없음)은 항상 확인 + API 미충족 항목 보완
            _want = (not by) or (not un) or (ev is None)
            doc = _doc_building_brief(item_key, want_fields=_want)
            vio = bool(doc.get("violation"))
            if _want:
                if not by:
                    by = doc.get("build_year")
                if not un and doc.get("units"):
                    # 집합건물(아파트·다세대·연립·도시형)의 '호' 라벨은 전유부(그 호실=1호)일 수 있어 건물 세대수로 쓰지 않음(표제부/API 세대수만)
                    if not (doc.get("unit_label") == "호" and re.search(r"아파트|오피스텔|다세대|연립|빌라|도시형", usage)):
                        un, ul = doc.get("units"), doc.get("unit_label")
                if ev is None and doc.get("elevator") is not None:
                    ev = doc.get("elevator")
            used_doc = bool(doc.get("build_year") or doc.get("units")
                            or doc.get("elevator") is not None or vio)
            # 숙박 세부용도(여관·생활숙박 등): 전유부 문서 우선(그 호실 용도) → 표제부 API 기타용도 폴백.
            _sub = doc.get("sukbak_sub") or (bi.get("sukbak_sub") if bi else None)
            # 단독주택(다가구 아닌 단독)인데 가구수 못 구하면 '단독'으로 표기(빈칸/오추출 방지)
            hh_disp = str(un) if un else None
            hh_label = ul or "세대"
            if not un and ("단독주택" in purpose or "단독주택" in usage) and "다가구" not in (purpose + usage):
                hh_disp, hh_label = "단독", ""
            if by or un or (ev is not None) or hh_disp or vio or _sub:
                out = {"available": True, "build_year": by,
                       "households": hh_disp,
                       "unit_label": hh_label,
                       "elevator": (("1" if ev else "0") if ev is not None else None),
                       "floors": (bi.get("floors") if bi else None),    # 지상 층수(다가구 3요건)
                       "purpose": (purpose or None),                     # 주용도(상가주택·위반 판별)
                       "usage_detail": _sub,                             # 숙박 세부용도(여관·생활형숙박시설 등) → 목록/상세 표시
                       "violation": vio,                                  # 위반건축물(건축물대장 스탬프)
                       "source": ("api+doc" if (used_api and used_doc) else
                                  "api" if used_api else "doc")}
    return out


def _load_briefs_from_db(keys: list, remote: bool = True) -> None:
    """미캐시 key들을 Supabase api_cache(brief:)에서 일괄 로딩 → _brief_cache 채움(계산 생략).
    remote=False면 Supabase 왕복 생략(이미 메모리에 로드된 것만 사용 — 목록 enrich 고속용)."""
    miss = [k for k in keys if k not in _brief_cache]
    if not miss or not remote:
        return
    try:
        rows = auction_db.cache_get_many(["brief:" + k for k in miss])
    except Exception:
        rows = {}
    import time as _t
    now = _t.time()
    for k in miss:
        d = rows.get("brief:" + k)
        if not isinstance(d, dict):
            continue
        if d.get("available"):
            _brief_cache[k] = d
        elif d.get("neg") and (now - (d.get("ts") or 0) < _BRIEF_NEG_TTL):
            _brief_cache[k] = {"available": False}   # TTL 내 미등재 → 재계산 생략(즉시 '정보없음')
        # 만료된 네거티브는 로딩 안 함 → todo에 포함되어 재계산(self-heal)


def _get_brief(item_key: str) -> dict:
    """캐시 우선(메모리/디스크). 없으면 계산 → 캐시 + 주거 building 정보는 Supabase에 영구 저장."""
    global _brief_dirty
    if item_key in _brief_cache:
        return _brief_cache[item_key]
    out = _compute_brief(item_key)
    _brief_cache[item_key] = out
    _brief_dirty = True
    try:
        if out.get("available"):
            auction_db.cache_save("brief:" + item_key, out)   # 성공분: 영구 저장
        else:
            import time as _t
            # 미등재(건축물대장·문서 모두 없음) → 짧은 TTL 네거티브 캐시(매번 4초 재호출 방지, 만료 후 재시도)
            auction_db.cache_save("brief:" + item_key, {"available": False, "neg": True, "ts": _t.time()})
    except Exception:
        pass
    return out


# ---------- 다가구·근린주택 분석(강의 기준 — 다가구주택·근린주택만 적용) ----------
_DAGAGU_RE = re.compile(r"다가구|근린주택")


def _is_dagagu_usage(usage: str) -> bool:
    """**다가구주택·근린주택만**(다가구 투자반 강의 대상). 단독·(일반)주택·농가주택·도시형 등 기타 전부 제외."""
    return bool(_DAGAGU_RE.search(usage or ""))


def _tenant_deposit_sum(item_key: str):
    """item_tenants 보증금 합계 + 임차인 수."""
    try:
        r = auction_db._get("item_tenants",
                            {"select": "deposit", "item_key": "eq." + item_key, "limit": "100"})
        rows = r.json() if r.status_code == 200 else []
    except Exception:
        rows = []
    dep = sum(int(x.get("deposit") or 0) for x in rows if x.get("deposit"))
    return dep, len(rows)


@app.get("/auction/dagagu_analysis")
def auction_dagagu_analysis(item_key: str) -> dict:
    """다가구·단독·근린주택 분석 — 3요건 체크리스트·우량판별·임대수요/명도·위반(강의 기준)."""
    try:
        r = auction_db._get("items",
                            {"select": "usage_name,address,appraisal_price,area_text,claim_amount,detail_text,tags",
                             "item_key": "eq." + item_key, "limit": "1"})
        rows = r.json() if r.status_code == 200 else []
    except Exception:
        rows = []
    if not rows:
        return {"available": False}
    it = rows[0]
    if not _is_dagagu_usage(it.get("usage_name")):
        return {"available": False, "reason": "다가구·단독·근린주택 전용 분석"}
    br = _get_brief(item_key) or {}
    units, floors, purpose = br.get("households"), br.get("floors"), br.get("purpose")
    if floors is None or not purpose:                 # 옛 brief엔 층수/주용도 없음 → 건축물대장 보강
        try:
            bi = building.info(it.get("address") or "")
        except Exception:
            bi = None
        if bi:
            if floors is None:
                floors = bi.get("floors")
            if not purpose:
                purpose = bi.get("purpose")
            if not units:
                units = bi.get("units")
    dep_sum, tcount = _tenant_deposit_sum(item_key)
    baedang = (_baedang_idx.get("map") or {}).get(item_key, 0)
    item = {"appraisal_price": it.get("appraisal_price"), "area_text": it.get("area_text"),
            "claim_amount": it.get("claim_amount"), "detail_text": it.get("detail_text"),
            "tags": it.get("tags")}
    res = dagagu_analysis.analyze(item, {"units": units, "floors": floors, "purpose": purpose},
                                  deposit_sum=dep_sum, baedang_count=baedang, tenant_count=tcount)
    return {"available": True, "usage": it.get("usage_name"), **res}


def _prewarm_briefs() -> None:
    """모든 주거·차량 물건 brief(준공/세대/승강기)를 DB(api_cache, brief:)에 자동 예열(미캐시만).
    저부하: 배치마다 잠깐 양보 → 라이브 요청 보호."""
    import time as _t
    try:
        keys = []
        for grp in ("주거용", "차량외"):
            off = 0
            while True:   # 현황 전체 페이징(if not items/len<limit로 종료) — 옛 12000 상한은 DB 커지며 현황 누락 유발해 제거
                items = auction_db.list_auctions(limit=200, offset=off, group=[grp])
                if not items:
                    break
                keys += [it.get("item_key") for it in items if it.get("item_key")]
                if len(items) < 200:
                    break
                off += 200
        # 숙박시설(상가 그룹) — 여관/생활형숙박시설 세부용도 표기용 brief(문서 우선=무쿼터, 없으면 표제부 API 1회)
        off = 0
        while True:
            items = auction_db.list_auctions(limit=200, offset=off, group=["상가"], usages=["숙박시설"])
            if not items:
                break
            keys += [it.get("item_key") for it in items if it.get("item_key")]
            if len(items) < 200:
                break
            off += 200
        _load_briefs_from_db(keys)                 # ① DB에 이미 있으면 로딩(계산 생략)
        todo = [k for k in keys if k not in _brief_cache]
        for i in range(0, len(todo), 40):          # ② 고속 병렬 계산 → DB 저장
            batch = todo[i:i + 40]
            with _cf.ThreadPoolExecutor(max_workers=12) as ex:
                list(ex.map(_get_brief, batch))
            _save_brief_cache()
    except Exception:
        pass


def _prewarm_apt_info() -> None:
    """아파트/오피스텔 apt_info(단지정보·실거래)를 DB(api_cache, apt:)에 자동 예열(미캐시만). 저부하."""
    import time as _t
    try:
        keys = []
        off = 0
        while True:   # 현황 전체 페이징(if not items/len<limit로 종료) — 옛 12000 상한은 DB 커지며 현황 누락 유발해 제거
            items = auction_db.list_auctions(limit=200, offset=off)
            if not items:
                break
            for d in items:
                u = d.get("usage") or ""
                if ("아파트" in u or "오피스텔" in u) and d.get("item_key"):
                    keys.append(d["item_key"])
            if len(items) < 200:
                break
            off += 200
        # DB에 이미 있는 것 스킵. 단, complex_detail 누락분은 따로 모아 보충(매 요청 동기호출 방지).
        have = set()
        need_detail = []
        for i in range(0, len(keys), 100):
            try:
                rows = auction_db.cache_get_many(["apt:" + k for k in keys[i:i + 100]])
                for ck, data in rows.items():
                    k = ck.split("apt:", 1)[1]
                    have.add(k)
                    if isinstance(data, dict) and data.get("available") and not data.get("complex_detail"):
                        need_detail.append((k, data))
            except Exception:
                pass
        todo = [k for k in keys if k not in have and k not in _apt_cache]

        def one(k):
            try:
                out = _apt_info_compute(k, 12)
                _apt_cache.remember(k, out)
                if out.get("available"):
                    auction_db.cache_save("apt:" + k, out)
            except Exception:
                pass

        def fill(kv):                            # 기존 캐시 entry에 상세 단지정보만 보충
            k, data = kv
            try:
                det = _complex_detail_for(data)
                if det:
                    data["complex_detail"] = det
                    _apt_cache.remember(k, data)
                    auction_db.cache_save("apt:" + k, data)
            except Exception:
                pass
        if todo:
            with _cf.ThreadPoolExecutor(max_workers=6) as ex:
                list(ex.map(one, todo))
        if need_detail:
            with _cf.ThreadPoolExecutor(max_workers=6) as ex:
                list(ex.map(fill, need_detail))
    except Exception:
        pass


def _prewarm_loop() -> None:
    """brief(준공/세대/승강기) 자동 예열 → DB. 시작 90초 후 + 12시간마다."""
    import time as _t
    _t.sleep(90)
    while True:
        _prewarm_briefs()
        _flush_all_caches()
        _t.sleep(12 * 3600)


def _grade_warm_loop() -> None:
    """매수판정 버킷(목록 배지·필터용)을 항상 따뜻하게 유지(시작 직후 빌드 + 25분마다 재산출, 사용자 대기 방지).
    ⚠️ CLOUD_READER=1(클라우드 얇은 리더)이면 재계산 안 함 — 로컬 워머가 채운 items.buy_grade 컬럼을 신뢰.
       (25분마다 수만 행 페이징 재계산이 1GB 인스턴스에서 주기적 OOM→503을 유발하던 것을 제거)"""
    import time as _t
    _t.sleep(3)          # 재기동 직후 곧바로 빌드 → 필터·배지 즉시 정상(첫 필터 16초 stall 방지)
    if os.environ.get("CLOUD_READER", "0") in ("1", "true", "True"):
        # 클라우드: 재계산 생략. 필터=items.buy_grade 컬럼 WHERE, 배지=컬럼 SELECT, 히어로=Supabase 캐시(로컬이 채움).
        try:
            if _buy_grade_col_exists():
                _buy_grade["synced"] = True     # 컬럼 신뢰 → 매수판정 필터가 컬럼 WHERE 사용(버킷 불필요)
                _grade_cache["buckets"] = {}    # 요청 경로가 무거운 재계산 스레드를 스폰하지 않도록(배지는 컬럼 SELECT로)
                print("[buy_grade] CLOUD_READER=1 → 재계산 생략, items.buy_grade 컬럼 신뢰", flush=True)
            else:
                print("[buy_grade] CLOUD_READER=1이나 buy_grade 컬럼 없음 → 배지/필터는 컬럼 채워질 때까지 제한", flush=True)
        except Exception:
            pass
        return
    while True:
        try:
            _grade_buckets(force=True)          # 워밍 스레드만 재계산(캐시는 유지 → 요청 stall 없음)
        except Exception:
            pass
        try:
            _hero_picks_build()                 # 매수양호 버킷 갱신 직후 홈 히어로 추천(차익 상위12) 재계산
        except Exception:
            pass
        _t.sleep(25 * 60)


def _prewarm_apt_loop() -> None:
    """apt_info(아파트정보/실거래) 자동 예열 → DB. brief와 병렬(별도 스레드). 12시간마다."""
    import time as _t
    _t.sleep(120)
    while True:
        _prewarm_apt_info()
        _t.sleep(12 * 3600)


def _prewarm_pools_loop() -> None:
    """B: 시군구 실거래 풀(_rh_cache·_apt_trades_cache) 저부하 예열 → apt_info/주변실거래 첫 조회 가속.
    molit 보호 위해 동시성 1, 시군구당 간격을 둠. 이미 캐시된 시군구는 건너뜀."""
    import time as _t
    _t.sleep(150)                      # 브리프 예열 우선, 서버 안정 후 시작
    try:
        lawds = []
        seen = set()
        for off in range(0, 12000, 200):
            items = auction_db.list_auctions(limit=200, offset=off)
            if not items:
                break
            for d in items:
                u = d.get("usage") or ""
                lw = resolve_lawd(d.get("address") or "")
                if lw and lw not in seen:
                    seen.add(lw)
                    lawds.append((lw, "아파트" in u or "오피스텔" in u,
                                  bool(re.search(r"다세대|연립|빌라|도시형", u))))
            if len(items) < 200:
                break
    except Exception:
        return
    for lw, is_apt, is_villa in lawds:
        try:
            if is_villa and lw not in _rh_cache:
                _rh_trades(lw)
                _t.sleep(0.6)
            if is_apt and lw not in _apt_trades_cache:
                _apt_trades(lw)
                _t.sleep(0.6)
        except Exception:
            pass


def _prewarm_nearby() -> None:
    """주변 유사 실거래(nearby:) 자동 예열 → DB(api_cache). 이미 DB에 있으면 스킵.
    V-World 지오코딩 부하 보호: 순차 + 간격(0.5s). 좌표는 공유 캐시라 같은 시군구 빨라짐."""
    import time as _t
    try:
        keys = []
        off = 0
        while True:   # 현황 전체 페이징(if not items/len<limit로 종료) — 옛 12000 상한은 DB 커지며 현황 누락 유발해 제거
            items = auction_db.list_auctions(limit=200, offset=off)
            if not items:
                break
            for d in items:
                if re.search(r"다세대|연립|빌라|도시형", d.get("usage") or "") and d.get("item_key"):
                    keys.append(d["item_key"])
            if len(items) < 200:
                break
            off += 200
        have = set()
        for i in range(0, len(keys), 100):
            try:
                rows = auction_db.cache_get_many(["nearby:" + k for k in keys[i:i + 100]])
                have |= {ck.split("nearby:", 1)[1] for ck in rows}
            except Exception:
                pass
        todo = [k for k in keys if k not in have and k not in _nearby_cache]

        def one(k):
            try:
                auction_nearby_trades(k)        # 계산 + DB 저장(geo_ok 결과만)
            except Exception:
                pass
        if todo:
            with _cf.ThreadPoolExecutor(max_workers=4) as ex:   # V-World 지오코딩 적정 동시성
                list(ex.map(one, todo))
        # 유사거래 카운트(_similar_cache)도 예열 → 목록 enrich가 즉시 embed(첫 페인트에 표시, async 동기계산 제거)
        sim_todo = [k for k in keys if k not in _similar_cache]
        if sim_todo:
            with _cf.ThreadPoolExecutor(max_workers=8) as ex:   # nearby 캐시 읽어 카운트만 → 가벼움
                list(ex.map(_get_similar, sim_todo))
            _save_similar_cache()
    except Exception:
        pass


def _prewarm_nearby_loop() -> None:
    """주변 유사 실거래 자동 예열 → DB. 다른 예열 이후 시작, 12시간마다."""
    import time as _t
    _t.sleep(240)
    while True:
        _prewarm_nearby()
        _t.sleep(12 * 3600)


def _docs_shard(todo: list) -> list:
    """멀티프로세스 샤딩: env DOCS_SHARD='i/N'면 안정 해시(md5)로 1/N만 처리 — 프로세스마다 다른 i를 주면
    GIL(파이썬 PDF 파싱 1코어 제약) 회피 병렬 + 중복 없음. 미설정이면 전체(단일 프로세스 동작 그대로)."""
    sh = os.environ.get("DOCS_SHARD")
    if not sh:
        return todo
    try:
        import hashlib
        i, n = (int(x) for x in sh.split("/"))
        return [k for k in todo if int(hashlib.md5(k.encode()).hexdigest(), 16) % n == i]
    except Exception:
        return todo


def _prewarm_docs() -> None:
    """물건현황(감정평가서)·권리분석(등기)·명세서요약·차량 → DB 자동 예열(이미 DB에 있으면 스킵).
    PDF 파싱 무거움 → 순차+간격(0.2s). 결과는 api_cache(analysis:/appraisal:/docsummary:/vehicle:)."""
    import time as _t
    from auction_analysis.doc_analysis import (
        analyze_registry, analyze_appraisal, analyze_doc_summary, analyze_vehicle)
    try:
        residential, vehicles = [], []
        off = 0
        while True:   # 현황 전체 페이징(if not items/len<limit로 종료) — 옛 12000 상한은 DB 커지며 현황 누락 유발해 제거
            items = auction_db.list_auctions(limit=200, offset=off)
            if not items:
                break
            for d in items:
                ik = d.get("item_key")
                if not ik:
                    continue
                if (d.get("group") or "") == "차량외" or "차량" in (d.get("usage") or ""):
                    vehicles.append(ik)
                else:
                    residential.append(ik)
            if len(items) < 200:
                break
            off += 200

        def warm(prefix, keys, analyze):
            have = set()
            for i in range(0, len(keys), 100):
                try:
                    rows = auction_db.cache_get_many([prefix + ":" + k for k in keys[i:i + 100]])
                    have |= {ck.split(prefix + ":", 1)[1] for ck in rows}
                except Exception:
                    pass
            todo = _docs_shard([k for k in keys if k not in have])   # 멀티프로세스 샤딩(DOCS_SHARD=i/N) — GIL 회피 병렬

            def one(k):
                try:
                    _cached_doc(prefix, k, lambda: analyze(auction_db, k))
                except Exception:
                    pass
            if todo:
                with _cf.ThreadPoolExecutor(max_workers=4) as ex:   # PDF 파싱 병렬
                    list(ex.map(one, todo))
        # 엔카 매칭 예열(차량외): vehicle_specs DB + Neon만 쓰므로 PDF 파싱과 무관.
        #  → docs(감정평가서 ~수h) join에 묶지 말고 '독립 병렬 스레드'로 빨리 캐시 채움.
        def _warm_prefix(prefix, compute):       # 차량외 공통 예열(미캐시만, 병렬)
            done = set()
            for i in range(0, len(vehicles), 100):
                try:
                    rows = auction_db.cache_get_many([prefix + ":" + k for k in vehicles[i:i + 100]])
                    done |= {ck.split(prefix + ":", 1)[1] for ck in rows}
                except Exception:
                    pass
            todo = [k for k in vehicles if k not in done]
            if todo:
                with _cf.ThreadPoolExecutor(max_workers=6) as ex:
                    list(ex.map(lambda k: _cached_doc(prefix, k, lambda kk=k: compute(kk)), todo))

        def warm_encar():
            _warm_prefix("encar2", _compute_encar)

        def warm_review():                       # 차종 후기(Neon reviews)
            _warm_prefix("review", _compute_review)
        # analysis·appraisal·docsummary·vehicle·encar 를 병렬 스레드로(순차 ~11h → 병렬)
        import threading as _th
        jobs = [("analysis", residential, analyze_registry),
                ("appraisal", residential, analyze_appraisal),
                ("docsummary", residential, analyze_doc_summary),
                ("vehicle2", vehicles, analyze_vehicle)]   # vehicle2: vehicle_specs DB 기반(구 PDF캐시 우회)
        ths = [_th.Thread(target=warm, args=j, daemon=True) for j in jobs]
        ths.append(_th.Thread(target=warm_encar, daemon=True))    # encar 독립(docs 안 기다림)
        ths.append(_th.Thread(target=warm_review, daemon=True))   # review 독립
        for th in ths:
            th.start()
        for th in ths:
            th.join()
    except Exception:
        pass


def _prewarm_docs_loop() -> None:
    """문서분석 자동 예열 → DB. 가장 무거워 후순위 시작, 12시간마다."""
    import time as _t
    _t.sleep(300)
    while True:
        _prewarm_docs()
        _flush_all_caches()
        _t.sleep(12 * 3600)


def _flush_all_caches() -> None:
    """모든 디스크 캐시 강제 저장(종료/주기)."""
    try:
        from auction_analysis.doc_analysis import flush_caches
        flush_caches()
    except Exception:
        pass
    for c in (_apt_cache,):
        try:
            c.flush()
        except Exception:
            pass
    _save_brief_cache()
    _save_similar_cache()
    _save_geo_cache()


_CACHE_PREFIXES = ("brief", "apt", "nearby", "analysis", "appraisal",
                   "docsummary", "vehicle2", "encar", "encar2", "review")


def _invalidate_item_caches(keys: list) -> None:
    """크롤러 갱신된 물건의 모든 캐시(DB api_cache + 메모리/디스크) 제거 → 다음 조회 시 재계산."""
    if not keys:
        return
    global _brief_dirty
    try:
        cks = [p + ":" + k for k in keys for p in _CACHE_PREFIXES]
        auction_db.cache_delete_many(cks)
    except Exception:
        pass
    from auction_analysis.doc_analysis import evict_item
    for k in keys:
        _brief_cache.pop(k, None)
        try:
            _apt_cache.pop(k, None)
        except Exception:
            pass
        _nearby_cache.pop(k, None)
        try:
            evict_item(k)
        except Exception:
            pass
    _brief_dirty = True


def _freshness_loop() -> None:
    """크롤러가 물건을 갱신(updated_at 변경)하면 그 물건 캐시를 자동 무효화. 90초 주기.
    서버 시작 시점의 최신 updated_at 이후 변경분만 감지(기존 캐시는 보존)."""
    import time as _t
    last = auction_db.max_updated_at() or ""
    while True:
        _t.sleep(90)
        try:
            if not last:
                last = auction_db.max_updated_at() or ""
                continue
            keys, newest = auction_db.items_updated_since(last)
            if keys:
                _invalidate_item_caches(keys)
                _save_brief_cache()
                last = newest
                print(f"[freshness] {len(keys)}건 갱신 감지 → 캐시 무효화(재계산 유도)", flush=True)
        except Exception:
            pass


@app.on_event("startup")
def _start_prewarm() -> None:
    import threading
    # CLOUD_READER=1(클라우드 얇은 리더) = 로컬 워머가 Supabase에 채운 캐시만 읽는 인스턴스 → 무거운 예열/워밍 전부 끔.
    _cloud = os.environ.get("CLOUD_READER", "0") in ("1", "true", "True")
    # DISABLE_PREWARM=1 또는 CLOUD_READER=1 이면 무거운 백그라운드 예열(브리프·apt·풀·주변·문서) 끔 → OOM/크래시루프 방지.
    #  끄더라도 캐시는 조회 시 on-demand로 Supabase api_cache에서 로딩되어 기능엔 지장 없음.
    if _cloud or os.environ.get("DISABLE_PREWARM", "0") in ("1", "true", "True"):
        print(f"[prewarm] 백그라운드 예열 비활성(CLOUD_READER={_cloud}) → on-demand Supabase 로딩, freshness 유지", flush=True)
    else:
        threading.Thread(target=_prewarm_loop, daemon=True).start()         # brief → DB
        threading.Thread(target=_prewarm_apt_loop, daemon=True).start()     # apt_info → DB (병렬)
        threading.Thread(target=_prewarm_pools_loop, daemon=True).start()   # 시군구 풀
        threading.Thread(target=_prewarm_nearby_loop, daemon=True).start()  # 주변 유사 실거래 → DB
        threading.Thread(target=_prewarm_docs_loop, daemon=True).start()    # 물건현황·권리분석·명세서 → DB
    # 매수판정 버킷 워밍·freshness는 DISABLE_PREWARM과 무관하게 항상 실행(경량·목록 배지/필터 필수)
    threading.Thread(target=_grade_warm_loop, daemon=True).start()      # 매수판정 버킷 워밍(목록 배지·필터)
    threading.Thread(target=_freshness_loop, daemon=True).start()       # 크롤러 갱신 감지 → 캐시 자동 무효화(경량, 항상 유지)
    threading.Thread(target=_type_filter_warm, daemon=True).start()     # 유형별 필터 미리 데움(첫 클릭 지연 방지)
    threading.Thread(target=_zone_warm, daemon=True).start()  # 용도지역 카테고리 미리 데움
    threading.Thread(target=_area_warm, daemon=True).start()  # 전용면적 색인 미리 데움(건물면적 필터)
    threading.Thread(target=_invest_warm, daemon=True).start()  # 투자금(선금) 색인 미리 데움(투자금 필터)
    threading.Thread(target=_baedang_warm, daemon=True).start()  # 배당요구신청 건수 색인 미리 데움(다가구·근린주택)
    threading.Thread(target=_dagagu_warm, daemon=True).start()    # 다가구 우량 색인 미리 데움
    threading.Thread(target=_compete_warm, daemon=True).start()   # 경쟁분산 색인(전 유형) 미리 데움
    if not _cloud:
        # V-World를 수분간 순차 호출 → 클라우드(얇은 리더)선 스킵(로컬 워머가 Supabase landuse_cache에 채움).
        threading.Thread(target=_landuse_warm, daemon=True).start()  # 원천 누락 용도지역 V-World 보완(순차, ~수분)
    try:
        _kb_apply_token()   # Supabase 공유 토큰(api_cache kb:auth) 로드 — 시작 시 1회, 가벼움(브라우저 없음)
    except Exception:
        pass
    if os.environ.get("KB_AUTH_ENABLE", "0") in ("1", "true", "True"):   # 기본 꺼짐(안정화 게이트) — 카카오 발급은 켤 때만
        threading.Thread(target=_kb_auth_loop, daemon=True).start()      # KB 부동산DB 토큰 관리(로컬 카카오 발급/서버 소비)
    # 카카오 자동발송 스케줄러 — 카카오톡 로그인된 로컬 PC(4011)에서만(KAKAO_BROADCAST=1). 클라우드엔 미설정 → 미기동.
    if os.environ.get("KAKAO_BROADCAST", "0") in ("1", "true", "True"):
        threading.Thread(target=_kakao_scheduler_loop, daemon=True).start()


@app.on_event("shutdown")
def _flush_on_shutdown() -> None:
    _flush_all_caches()


# ---------- 관리자: 수동 예열(DB 적재) ----------
_prewarm_running: dict[str, bool] = {}
_PREWARM_FNS = {
    "brief": "_prewarm_briefs", "apt": "_prewarm_apt_info",
    "nearby": "_prewarm_nearby", "docs": "_prewarm_docs",
}


def _run_prewarm_bg(kind: str) -> None:
    import threading

    def job():
        kinds = ["brief", "apt", "nearby", "docs"] if kind == "all" else [kind]
        for k in kinds:
            fn = globals().get(_PREWARM_FNS.get(k, ""))
            if not fn:
                continue
            _prewarm_running[k] = True
            try:
                fn()
            except Exception:
                pass
            finally:
                _prewarm_running[k] = False
        _flush_all_caches()
    threading.Thread(target=job, daemon=True).start()


@app.get("/admin/me")
def admin_me(admin: dict = Depends(require_admin)) -> dict:
    return {"id": admin["id"], "name": admin["name"], "email": admin["email"],
            "role": admin["role"]}


# ───────────── 관리자: 회원관리 / 마일리지 / 쿠폰 ─────────────
class MileageIn(BaseModel):
    user_id: int
    amount: int                      # +적립 / -차감
    reason: str = ""


class GradeIn(BaseModel):
    user_id: int
    grade: str


class CouponIn(BaseModel):
    name: str = "쿠폰"
    amount: int                      # 사용 시 적립 마일리지
    user_id: Optional[int] = None    # 지정 회원(없으면 범용 코드)
    count: int = 1                   # 발급 장수(범용 코드용)
    expires_at: Optional[str] = None # YYYY-MM-DD


class RedeemIn(BaseModel):
    code: str


@app.get("/admin/users")
def admin_users(q: str = "", admin: dict = Depends(require_admin)) -> dict:
    """회원 목록(검색 q: 이름/이메일/연락처)."""
    return {"users": user_store.list_users(q=q)}


@app.get("/admin/users/mileage_log")
def admin_mileage_log(user_id: int, admin: dict = Depends(require_admin)) -> dict:
    return {"log": user_store.mileage_log(user_id)}


@app.post("/admin/users/mileage")
def admin_adjust_mileage(body: MileageIn, admin: dict = Depends(require_admin)) -> dict:
    try:
        u = user_store.adjust_mileage(body.user_id, body.amount, body.reason)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "user": u}


@app.post("/admin/users/grade")
def admin_set_grade(body: GradeIn, admin: dict = Depends(require_admin)) -> dict:
    u = user_store.set_grade(body.user_id, body.grade)
    if not u:
        raise HTTPException(404, "회원을 찾을 수 없습니다.")
    return {"ok": True, "user": u}


@app.get("/admin/coupons")
def admin_list_coupons(admin: dict = Depends(require_admin)) -> dict:
    return {"coupons": user_store.list_coupons()}


@app.post("/admin/coupons")
def admin_create_coupons(body: CouponIn, admin: dict = Depends(require_admin)) -> dict:
    try:
        coupons = user_store.create_coupons(
            body.name, body.amount, user_id=body.user_id,
            count=body.count, expires_at=body.expires_at)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "coupons": coupons}


@app.post("/admin/coupons/delete")
def admin_delete_coupon(body: dict = Body(...), admin: dict = Depends(require_admin)) -> dict:
    ok = user_store.delete_coupon(int(body.get("id", 0)))
    if not ok:
        raise HTTPException(400, "미사용 쿠폰만 회수할 수 있습니다.")
    return {"ok": True}


@app.post("/admin/kakao_send")
def admin_kakao_send(body: dict = Body(...), admin: dict = Depends(require_admin)) -> dict:
    """관리자 카카오톡 발송 — 서버(이 PC)의 카카오톡 PC버전을 제어해 지정 방에 메시지 전송.
    send_now=True면 즉시 전송(수 초간 마우스/키보드 자동조작), False면 입력창에 넣어만 둠(수동 Enter).
    ⚠️ 4011 서버가 카카오톡 로그인된 PC에서 돌 때만 동작(클라우드 배포 시 불가)."""
    chat = (body.get("chat_name") or "").strip()
    msg = body.get("message") or ""
    send_now = bool(body.get("send_now"))
    if not chat or not msg.strip():
        raise HTTPException(400, "방 이름과 메시지를 입력하세요.")
    try:
        import sys as _sys
        _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # api/ → 프로젝트 루트
        if _root not in _sys.path:
            _sys.path.insert(0, _root)
        from kakao_sender import send_kakao_message, KakaoTalkControlError
    except Exception as e:
        raise HTTPException(500, f"kakao_sender 로드 실패(pywin32·pyperclip 필요): {e}")
    try:
        send_kakao_message(chat, msg, send_now=send_now)
    except KakaoTalkControlError as e:
        raise HTTPException(500, f"카카오톡 제어 실패(방 이름·로그인 상태 확인): {e}")
    except Exception as e:
        raise HTTPException(500, f"전송 실패: {e}")
    return {"ok": True, "sent": send_now, "chat": chat}


# ───────── 카카오 자동발송(뉴스/매각예정/매각완료) — 로컬 4011 전용 ─────────
_KAKAO_SEND_LOCK = threading.Lock()
_KAKAO_LAST_RESULT: dict = {}      # kind -> 마지막 '지금 발송' 결과({ok,msg,at}) 또는 None(진행 중)


def _kb():
    """kakao_broadcast 모듈(프로젝트 루트) 로드. api/ 하위라 루트 path 보장 후 import."""
    import sys as _sys
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _root not in _sys.path:
        _sys.path.insert(0, _root)
    import kakao_broadcast as kb
    return kb


def _kakao_do_send(room, payload):
    """지정 방(들)에 실제 전송. payload=str이면 텍스트, list이면 시퀀스(사진→정보). 실패 시 예외 전파."""
    import sys as _sys
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _root not in _sys.path:
        _sys.path.insert(0, _root)
    if isinstance(payload, list):
        from kakao_sender import send_kakao_sequence
        seq, tmp = _kakao_materialize(payload)
        try:
            send_kakao_sequence(room, seq)
        finally:
            for p in tmp:
                try:
                    os.remove(p)
                except Exception:
                    pass
    else:
        from kakao_sender import send_kakao_message
        send_kakao_message(room, payload, send_now=True)


def _kakao_materialize(items):
    """시퀀스의 image url → 임시파일 다운로드. 반환 (seq, 임시파일목록). 실패 이미지는 사진만 스킵(정보는 유지)."""
    import tempfile
    import hashlib
    import httpx as _httpx
    seq, tmp = [], []
    for it in items:
        if it.get("type") == "image" and it.get("url"):
            try:
                r = _httpx.get(it["url"], timeout=15, follow_redirects=True)
                if r.status_code == 200 and r.content:
                    p = os.path.join(tempfile.gettempdir(),
                                     "kbimg_" + hashlib.md5(it["url"].encode()).hexdigest()[:12] + ".img")
                    with open(p, "wb") as f:
                        f.write(r.content)
                    seq.append({"type": "image", "path": p})
                    tmp.append(p)
                    continue
            except Exception:
                pass
        else:
            seq.append(it)
    return seq, tmp


def _kakao_run(kind, force=False):
    """kind=news/upcoming/sold 콘텐츠 생성 → 중복 아니면 발송 → 이력 저장. dict 반환.
    중복방지: upcoming/sold는 sent_date(같은 매각일 재발송 차단·주말 반복 방지), news는 sent_links."""
    kb = _kb()
    st = kb.load_state()
    c = st.get(kind, {})
    room = (c.get("room") or "").strip()
    if not room:
        return {"ok": False, "msg": "카카오 방 이름이 설정되지 않았습니다."}
    date = None
    new_links = []
    if kind == "news":
        used_sent = [] if force else c.get("sent_links", [])       # 지금 발송(force)은 이미 보낸 기사도 포함
        payload, new_links = kb.build_news(used_sent, c.get("openai_key", ""), c.get("openai_model", "gpt-4o-mini"))
        if not payload:
            return {"ok": False, "msg": ("뉴스 데이터가 없습니다." if force else "발송할 새 뉴스가 없습니다(이미 모두 발송).")}
    elif kind == "upcoming":
        date, payload = kb.build_upcoming()
        if not date:
            return {"ok": False, "msg": "다음 매각기일을 찾지 못했습니다."}
        if not payload:
            return {"ok": False, "msg": f"{date} 예상낙찰가 있는 대상 물건이 없습니다."}
        if not force and c.get("sent_date") == date:
            return {"ok": False, "msg": f"{date} 물건은 이미 발송했습니다(중복 방지)."}
    elif kind == "sold":
        date, payload = kb.build_sold()
        if not date:
            return {"ok": False, "msg": "직전 매각일을 찾지 못했습니다."}
        if not payload:
            return {"ok": False, "msg": f"{date} 매각완료 물건이 없습니다."}
        if not force and c.get("sent_date") == date:
            return {"ok": False, "msg": f"{date} 매각완료는 이미 발송했습니다(중복 방지)."}
    else:
        return {"ok": False, "msg": f"알 수 없는 종류: {kind}"}
    try:
        _kakao_do_send(room, payload)
    except Exception as e:
        return {"ok": False, "msg": f"카카오 전송 실패(방 이름·로그인 확인): {e}"}
    import datetime as _dt
    now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    if kind == "news":
        merged = list(c.get("sent_links", [])) + new_links
        c["sent_links"] = list(dict.fromkeys(merged))[-800:]   # 중복 제거 + 최근 800개만 유지
    else:
        c["sent_date"] = date   # upcoming/sold: 매각기일(내용 중복방지용) — 달력날짜 아님(스케줄러는 auto_date 사용)
    c["last"] = now
    st[kind] = c
    kb.save_state(st)
    cnt = len(payload) if isinstance(payload, list) else len(payload)
    unit = "개 항목" if isinstance(payload, list) else "자"
    return {"ok": True, "msg": f"발송 완료({cnt}{unit})", "date": date}


_KAKAO_FIRED = {}   # kind -> "YYYY-MM-DD HH:MM"(같은 분 재발화 방지)


def _kakao_scheduler_loop():
    """매 30초 확인 → '발송시각 경과(+2시간 창) & 오늘 미처리(auto_date)'면 _kakao_run. (KAKAO_BROADCAST=1일 때만)
    정확한 분 매칭이 아니라 '시각 지났고 오늘 아직 자동발송 안 함'으로 판정 → 서버 재시작·일시적 실패(매각기일
    조회 등)에도 그날 안에 확실히 발송(일시실패는 매 분 재시도, 처리완료 시 auto_date=오늘 달력날짜로 재발화 방지).
    ※ auto_date=자동발송한 '달력날짜' / sent_date=발송한 '매각기일'(내용 중복방지) — 서로 다른 축이라 분리."""
    import time as _t, datetime as _dt
    print("[kakao] 자동발송 스케줄러 시작(로컬 전용)", flush=True)
    while True:
        try:
            kb = _kb()
            now = _dt.datetime.now()
            stamp = now.strftime("%Y-%m-%d %H:%M")
            today = now.strftime("%Y-%m-%d")
            st = kb.load_state()
            for kind in ("news", "upcoming", "sold"):
                c = st.get(kind, {})
                t = (c.get("time") or "").strip()
                if not (c.get("on") and (c.get("room") or "").strip() and len(t) == 5 and t[2] == ":"):
                    continue
                if c.get("auto_date") == today:
                    continue                       # 오늘 이미 자동발송 처리함(성공/정상스킵) — 재발화 방지
                try:
                    sched = now.replace(hour=int(t[:2]), minute=int(t[3:5]), second=0, microsecond=0)
                except Exception:
                    continue
                elapsed = (now - sched).total_seconds()
                if not (0 <= elapsed < 2 * 3600):  # 발송시각~+2시간 창(놓친 것 따라잡기·과도한 지각발송 방지)
                    continue
                if _KAKAO_FIRED.get(kind) == stamp:
                    continue                       # 같은 분 중복 방지(실패 시 다음 분 재시도)
                _KAKAO_FIRED[kind] = stamp
                r = _kakao_run(kind)
                print(f"[kakao] 정기발송 {kind} @{stamp}: {r.get('msg')}", flush=True)
                # 일시적 읽기실패('찾지 못')만 다음 분 재시도. 그 외(발송성공·이미발송·대상없음)는 오늘 처리완료로 마킹.
                if "찾지 못" not in (r.get("msg") or ""):
                    try:
                        st2 = kb.load_state(); st2.setdefault(kind, {})["auto_date"] = today; kb.save_state(st2)
                    except Exception:
                        pass
        except Exception as e:
            print(f"[kakao] 스케줄러 오류: {e}", flush=True)
        _t.sleep(30)


@app.get("/admin/kakao/config")
def admin_kakao_config_get(admin: dict = Depends(require_admin)) -> dict:
    """3메뉴(뉴스/매각예정/매각완료) 설정·이력 조회."""
    kb = _kb()
    st = kb.load_state()
    out = {}
    for kind in ("news", "upcoming", "sold"):
        c = st.get(kind, {})
        out[kind] = {"room": c.get("room", ""), "time": c.get("time", ""),
                     "on": bool(c.get("on")), "last": c.get("last", ""),
                     "sent_date": c.get("sent_date", ""),
                     "sent_count": len(c.get("sent_links", [])),
                     "has_openai_key": bool(c.get("openai_key")),
                     "openai_key": c.get("openai_key", ""),          # 로컬 전용 — 화면에 그대로 표시
                     "openai_model": c.get("openai_model", "gpt-4o-mini"),
                     "last_result": _KAKAO_LAST_RESULT.get(kind)}     # 지금발송 결과(None=진행중)
    return {"ok": True, "config": out}


@app.post("/admin/kakao/config")
def admin_kakao_config_set(body: dict = Body(...), admin: dict = Depends(require_admin)) -> dict:
    """방이름·발송시각(HH:MM)·자동발송 on/off 저장. 이력(sent_date/sent_links)은 보존."""
    kb = _kb()
    st = kb.load_state()
    import re as _re
    for kind in ("news", "upcoming", "sold"):
        if kind in body and isinstance(body[kind], dict):
            c = st.setdefault(kind, {})
            b = body[kind]
            if "room" in b:
                c["room"] = (b.get("room") or "").strip()
            if "time" in b:
                t = (b.get("time") or "").strip()
                if t and not _re.match(r"^\d{2}:\d{2}$", t):
                    raise HTTPException(400, f"{kind} 시각 형식은 HH:MM 이어야 합니다.")
                c["time"] = t
            if "on" in b:
                c["on"] = bool(b.get("on"))
            if "openai_key" in b:
                kv = (b.get("openai_key") or "").strip()
                if kv:
                    c["openai_key"] = kv          # 빈칸이면 기존 키 유지(재입력 불필요)
            if "openai_model" in b:
                c["openai_model"] = (b.get("openai_model") or "gpt-4o-mini").strip()
    kb.save_state(st)
    return {"ok": True}


@app.post("/admin/kakao/send_now")
def admin_kakao_send_now(kind: str, admin: dict = Depends(require_admin)) -> dict:
    """즉시 발송 — 카카오톡 GUI 자동조작은 수십 초~수 분 걸리므로 백그라운드 실행 후 즉시 응답.
    결과는 config의 last_result 로 폴링(프론트)."""
    if kind not in ("news", "upcoming", "sold"):
        raise HTTPException(400, "kind는 news/upcoming/sold 중 하나여야 합니다.")
    if _KAKAO_SEND_LOCK.locked():
        return {"ok": False, "msg": "다른 발송이 진행 중입니다. 완료 후 다시 시도하세요."}
    _KAKAO_LAST_RESULT[kind] = None      # 진행 중 표시(폴링이 None이면 대기)

    def _job():
        with _KAKAO_SEND_LOCK:
            import datetime as _dt
            try:
                r = _kakao_run(kind, force=True)
            except Exception as e:
                r = {"ok": False, "msg": f"발송 오류: {e}"}
            _KAKAO_LAST_RESULT[kind] = {"ok": bool(r.get("ok")), "msg": r.get("msg", ""),
                                        "at": _dt.datetime.now().strftime("%H:%M:%S")}
            print(f"[kakao] 지금발송 {kind}: {r.get('msg')}", flush=True)

    threading.Thread(target=_job, daemon=True).start()
    return {"ok": True, "started": True,
            "msg": "발송을 시작했습니다(카카오톡 자동 조작, 최대 1~2분). 그동안 PC를 건드리지 마세요."}


@app.get("/admin/kakao/preview")
def admin_kakao_preview(kind: str, admin: dict = Depends(require_admin)) -> dict:
    """발송하지 않고 콘텐츠만 미리보기."""
    kb = _kb()
    if kind == "news":
        st = kb.load_state()
        n = st.get("news", {})
        text, links = kb.build_news(n.get("sent_links", []), n.get("openai_key", ""), n.get("openai_model", "gpt-4o-mini"))
        return {"ok": True, "text": text or "", "count": len(links)}
    if kind == "upcoming":
        date, items = kb.build_upcoming()
        return {"ok": True, "text": kb.seq_to_text(items), "date": date}
    if kind == "sold":
        date, items = kb.build_sold()
        return {"ok": True, "text": kb.seq_to_text(items), "date": date}
    raise HTTPException(400, "kind는 news/upcoming/sold 중 하나여야 합니다.")


@app.post("/coupons/redeem")
def redeem_coupon(body: RedeemIn, user: dict = Depends(require_user)) -> dict:
    """회원이 쿠폰코드 사용 → 마일리지 적립."""
    try:
        res = user_store.redeem_coupon(body.code, user["id"])
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, **res}


# ───────────── 요금제 결제(포트원/아임포트 v1) ─────────────
#  요금제 단일 출처. 금액은 서버가 확정(프론트 변조 방지). 스피드옥션 전국이용 기준.
_PAY_PLANS: dict = {
    "m1":  {"months": 1,  "amount": 84000,  "name": "전국이용 1개월",  "list": 84000},
    "m3":  {"months": 3,  "amount": 214000, "name": "전국이용 3개월",  "list": 252000},
    "m6":  {"months": 6,  "amount": 378000, "name": "전국이용 6개월",  "list": 504000},
    "m12": {"months": 12, "amount": 650000, "name": "전국이용 12개월", "list": 1008000},
    "p1":  {"months": 1,  "amount": 150000,  "name": "프리미엄 1개월",  "list": 150000,  "grade": "프리미엄"},
    "p3":  {"months": 3,  "amount": 383000,  "name": "프리미엄 3개월",  "list": 450000,  "grade": "프리미엄"},
    "p6":  {"months": 6,  "amount": 675000,  "name": "프리미엄 6개월",  "list": 900000,  "grade": "프리미엄"},
    "p12": {"months": 12, "amount": 1170000, "name": "프리미엄 12개월", "list": 1800000, "grade": "프리미엄"},
}
_PAY_GRADE = "전국"


def _portone_token() -> Optional[str]:
    key = os.environ.get("PORTONE_API_KEY")
    sec = os.environ.get("PORTONE_API_SECRET")
    if not (key and sec):
        return None
    try:
        r = httpx.post("https://api.iamport.kr/users/getToken",
                       json={"imp_key": key, "imp_secret": sec}, timeout=10)
        return ((r.json() or {}).get("response") or {}).get("access_token")
    except Exception:
        return None


def _portone_prepare(merchant_uid: str, amount: int) -> None:
    """포트원에 (주문번호,금액) 사전등록 — 결제창 금액 변조 방지."""
    t = _portone_token()
    if not t:
        return
    try:
        httpx.post("https://api.iamport.kr/payments/prepare",
                   json={"merchant_uid": merchant_uid, "amount": int(amount)},
                   headers={"Authorization": t}, timeout=10)
    except Exception:
        pass


def _portone_get_payment(imp_uid: str) -> Optional[dict]:
    t = _portone_token()
    if not t:
        return None
    try:
        r = httpx.get(f"https://api.iamport.kr/payments/{imp_uid}",
                      headers={"Authorization": t}, timeout=10)
        return (r.json() or {}).get("response") or None
    except Exception:
        return None


def _portone_cancel(imp_uid: str, reason: str) -> None:
    t = _portone_token()
    if not t:
        return
    try:
        httpx.post("https://api.iamport.kr/payments/cancel",
                   json={"imp_uid": imp_uid, "reason": reason},
                   headers={"Authorization": t}, timeout=10)
    except Exception:
        pass


class PayPrepareIn(BaseModel):
    plan: str


class PayCompleteIn(BaseModel):
    imp_uid: str
    merchant_uid: str


@app.get("/payment/config")
def payment_config(user: Optional[dict] = Depends(current_user)) -> dict:
    """요금제 목록 + 포트원 가맹점코드 + (로그인 시)현재 이용권. 결제설정 여부 안내."""
    configured = bool(os.environ.get("PORTONE_IMP_CODE")
                      and os.environ.get("PORTONE_API_KEY")
                      and os.environ.get("PORTONE_API_SECRET"))
    plans = [{"code": k, **v} for k, v in _PAY_PLANS.items()]
    mem = None
    if user:
        mem = {"grade": user.get("grade"), "paid_until": user.get("paid_until"),
               "name": user.get("name") or user.get("email")}
    return {"imp_code": os.environ.get("PORTONE_IMP_CODE", ""),
            "configured": configured, "plans": plans, "membership": mem,
            "logged_in": bool(user)}


@app.post("/payment/prepare")
def payment_prepare(body: PayPrepareIn, user: dict = Depends(require_user)) -> dict:
    """결제 전 주문 생성 + 포트원 금액 사전등록. 주문번호·금액(서버확정) 반환."""
    plan = _PAY_PLANS.get(body.plan)
    if not plan:
        raise HTTPException(400, "잘못된 요금제입니다.")
    import secrets as _s
    merchant_uid = f"jh_{user['id']}_{_s.token_hex(8)}"
    user_store.create_payment(user["id"], merchant_uid, body.plan,
                              plan["months"], plan["amount"], plan.get("grade", _PAY_GRADE))
    _portone_prepare(merchant_uid, plan["amount"])
    return {"merchant_uid": merchant_uid, "amount": plan["amount"],
            "name": plan["name"], "imp_code": os.environ.get("PORTONE_IMP_CODE", ""),
            "buyer_name": user.get("name") or "", "buyer_email": user.get("email") or "",
            "buyer_tel": user.get("phone") or ""}


@app.post("/payment/complete")
def payment_complete(body: PayCompleteIn, user: dict = Depends(require_user)) -> dict:
    """결제창 성공 콜백 후: 포트원 REST로 실제 결제 재검증 → 금액 대조 → 이용권 부여."""
    pay = _portone_get_payment(body.imp_uid)
    if not pay:
        raise HTTPException(400, "결제 정보를 조회할 수 없습니다(포트원 설정 확인).")
    if pay.get("status") != "paid":
        raise HTTPException(400, "결제가 완료되지 않았습니다.")
    if pay.get("merchant_uid") != body.merchant_uid:
        raise HTTPException(400, "주문번호가 일치하지 않습니다.")
    order = user_store.get_payment(body.merchant_uid)
    if not order or order["user_id"] != user["id"]:
        raise HTTPException(403, "본인 주문이 아닙니다.")
    try:
        res = user_store.complete_payment(body.merchant_uid, body.imp_uid,
                                          int(pay.get("amount") or 0))
    except ValueError as e:
        _portone_cancel(body.imp_uid, str(e))     # 금액불일치 등 → 자동 결제취소(환불)
        raise HTTPException(400, str(e))
    u = res["user"]
    return {"ok": True, "already": res.get("already", False),
            "paid_until": u.get("paid_until"), "grade": u.get("grade"),
            "plan_name": _PAY_PLANS.get(order["plan"], {}).get("name", "")}


@app.get("/admin/payments")
def admin_payments(admin: dict = Depends(require_admin)) -> dict:
    """결제 내역(관리자)."""
    return {"payments": user_store.list_payments()}


class GrantIn(BaseModel):
    user_id: int
    months: int = 1
    grade: str = "전국"


@app.post("/admin/grant_membership")
def admin_grant_membership(body: GrantIn, admin: dict = Depends(require_admin)) -> dict:
    """관리자 수동 이용권 부여(무통장입금 확인 등)."""
    try:
        u = user_store.grant_membership_admin(body.user_id, body.months, body.grade)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "user": u}


# ── 회원 등급 생성·관리(관리자) ──
class GradeCreateIn(BaseModel):
    name: str
    rank: int = 0
    lecture: bool = False
    color: str = ""
    comment: str = ""


class GradeUpdateIn(BaseModel):
    name: Optional[str] = None
    rank: Optional[int] = None
    lecture: Optional[bool] = None
    color: Optional[str] = None
    comment: Optional[str] = None


@app.get("/admin/grades")
def admin_grades(admin: dict = Depends(require_admin)) -> dict:
    """등급 목록 + 각 등급 회원 수."""
    grades = user_store.list_grades()
    counts = user_store.grade_counts() if hasattr(user_store, "grade_counts") else {}
    for g in grades:
        g["users"] = counts.get(g["name"], 0)
    return {"grades": grades}


@app.post("/admin/grades")
def admin_grade_create(body: GradeCreateIn, admin: dict = Depends(require_admin)) -> dict:
    try:
        g = user_store.create_grade(body.name, body.rank, body.lecture, body.color, body.comment)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "grade": g}


@app.put("/admin/grades/{gid}")
def admin_grade_update(gid: int, body: GradeUpdateIn,
                       admin: dict = Depends(require_admin)) -> dict:
    try:
        g = user_store.update_grade(gid, rank=body.rank, lecture=body.lecture,
                                    color=body.color, name=body.name, comment=body.comment)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "grade": g}


@app.delete("/admin/grades/{gid}")
def admin_grade_delete(gid: int, admin: dict = Depends(require_admin)) -> dict:
    try:
        res = user_store.delete_grade(gid)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, **res}


# ───────────── 게시판(예정/강의/커뮤니티/경매자료실) ─────────────
#  권한은 board_settings(DB)에서 관리자가 설정. _BOARDS는 게시판 키·설명·기본권한(시드값).
#  read:  all(전체공개) | user(로그인) | grade(등급회원=grades.lecture)
#  write: admin(관리자) | user(로그인회원) | grade(등급회원)
#  comment: off(불가) | user(로그인회원) | grade(등급회원)
_BOARDS: dict = {
    "yejung":    {"title": "경매예정", "desc": "경매 예정 물건·일정을 안내합니다.",
                  "read": "all",   "write": "admin", "comment": "user"},
    "video":     {"title": "강의",     "desc": "경매 강의입니다(등급 회원 전용).",
                  "read": "grade", "write": "admin", "comment": "grade"},
    "community": {"title": "커뮤니티", "desc": "회원 자유게시판입니다.",
                  "read": "all",   "write": "user",  "comment": "user"},
    "data":      {"title": "경매자료실", "desc": "경매 자료·서식을 공유합니다.",
                  "read": "all",   "write": "admin", "comment": "user"},
    "support":   {"title": "고객센터", "desc": "문의·건의 사항을 남겨주세요(회원 작성).",
                  "read": "all",   "write": "user",  "comment": "user"},
}
_PERM_LABEL = {"all": "전체공개", "user": "로그인 회원", "grade": "등급 회원",
               "admin": "관리자", "off": "사용 안 함"}

# ── 게시글 본문(WYSIWYG HTML) 새니타이즈 — XSS 방지(nh3/ammonia, 허용 태그·속성만) ──
import nh3                                                        # noqa: E402
_HTML_TAGS = {"p", "br", "span", "div", "b", "strong", "i", "em", "u", "s", "strike",
              "del", "sub", "sup", "h1", "h2", "h3", "h4", "h5", "h6", "blockquote",
              "pre", "code", "hr", "mark", "ul", "ol", "li", "a", "img",
              "table", "thead", "tbody", "tfoot", "tr", "th", "td", "figure", "figcaption"}
_HTML_ATTRS = {
    "*": {"style", "class", "title"},
    "a": {"href", "target", "name"},      # rel은 link_rel이 자동 관리(중복 지정 금지)
    "img": {"src", "alt", "width", "height", "data-align", "data-proportion",
            "data-size", "data-rotate", "data-file-name", "data-file-size", "data-origin"},
    "td": {"colspan", "rowspan"}, "th": {"colspan", "rowspan"},
}


def _sanitize_html(html: str) -> str:
    """WYSIWYG 본문 HTML을 허용 태그·속성만 남기고 정리(script/onerror/javascript: 제거)."""
    if not html:
        return ""
    return nh3.clean(html[:300000], tags=_HTML_TAGS, attributes=_HTML_ATTRS,
                     url_schemes={"http", "https", "mailto", "tel"},
                     link_rel="noopener noreferrer")


_UPLOAD_DIR = os.path.join(_STATIC_DIR, "uploads")
os.makedirs(_UPLOAD_DIR, exist_ok=True)
_IMG_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
_IMG_MAX = 8 * 1024 * 1024


@app.post("/board/upload")
async def board_upload(request: Request, user: dict = Depends(require_user)) -> dict:
    """에디터 이미지 업로드 — 회원만. SunEditor 응답형식 {result:[{url,name,size}]}."""
    import secrets as _s
    form = await request.form()
    files = [v for v in form.values() if hasattr(v, "filename") and hasattr(v, "read")]
    results = []
    for f in files:
        ext = os.path.splitext(f.filename or "")[1].lower()
        if ext not in _IMG_EXT:
            continue
        data = await f.read()
        if len(data) > _IMG_MAX:
            raise HTTPException(400, "이미지는 8MB 이하만 업로드할 수 있습니다.")
        name = _s.token_hex(12) + ext
        with open(os.path.join(_UPLOAD_DIR, name), "wb") as out:
            out.write(data)
        results.append({"url": "/static/uploads/" + name,
                        "name": f.filename, "size": len(data)})
    if not results:
        raise HTTPException(400, "이미지 파일만 업로드할 수 있습니다.")
    return {"result": results}


class PostIn(BaseModel):
    title: str
    content: str = ""


class CommentIn(BaseModel):
    content: str


def _board_cfg(b: str) -> Optional[dict]:
    """_BOARDS(키·설명·시드권한) + board_settings(DB·관리자설정) 병합. 미시드면 시드."""
    base = _BOARDS.get(b)
    if not base:
        return None
    s = user_store.get_board_setting(b)
    if not s:
        user_store.ensure_board(b, base["title"], base["read"], base["write"], base["comment"])
        s = user_store.get_board_setting(b)
    return {"board": b, "desc": base["desc"], "title": s.get("title") or base["title"],
            "read": s["read_perm"], "write": s["write_perm"], "comment": s["comment_perm"]}


def _board_cfg_or_404(b: str) -> dict:
    cfg = _board_cfg(b)
    if not cfg:
        raise HTTPException(404, "존재하지 않는 게시판입니다.")
    return cfg


def _grade_label() -> str:
    names = [g["name"] for g in user_store.list_grades() if g.get("lecture")]
    return ", ".join(names) if names else "지정 등급"


def _is_admin(user: Optional[dict]) -> bool:
    return bool(user and user.get("role") == "admin")


def _perm_read_ok(cfg: dict, user: Optional[dict]) -> bool:
    if _is_admin(user):
        return True
    lv = cfg["read"]
    if lv == "all":
        return True
    if lv == "user":
        return bool(user)
    return bool(user) and user_store.grade_can_lecture(user.get("grade"))   # grade


def _perm_write_ok(cfg: dict, user: Optional[dict]) -> bool:
    if not user:
        return False
    if _is_admin(user):
        return True
    lv = cfg["write"]
    if lv == "admin":
        return False
    if lv == "user":
        return True
    return user_store.grade_can_lecture(user.get("grade"))                   # grade


def _perm_comment_ok(cfg: dict, user: Optional[dict]) -> bool:
    if not user:
        return False
    lv = cfg["comment"]
    if lv == "off":
        return False
    if lv == "user":
        return True
    return _is_admin(user) or user_store.grade_can_lecture(user.get("grade"))  # grade


def _read_reason(cfg: dict, user: Optional[dict]) -> str:
    if cfg["read"] == "grade":
        return ("로그인이 필요합니다." if not user
                else f"이 게시판은 {_grade_label()} 등급 회원만 이용할 수 있습니다.")
    return "로그인이 필요합니다."


@app.get("/board/{b}")
def board_list(b: str, page: int = 1, user: Optional[dict] = Depends(current_user)) -> dict:
    """게시판 글목록 + 메타. 읽기 권한 미달 시 locked 응답."""
    cfg = _board_cfg_or_404(b)
    base = {"board": b, "title": cfg["title"], "desc": cfg["desc"],
            "read": cfg["read"], "write": cfg["write"], "comment": cfg["comment"],
            "is_admin": _is_admin(user), "uid": (user["id"] if user else None)}
    if not _perm_read_ok(cfg, user):
        return {**base, "locked": True, "can_write": False, "items": [], "total": 0,
                "page": 1, "size": 15, "reason": _read_reason(cfg, user)}
    data = user_store.list_posts(b, page)
    return {**base, "locked": False, "can_write": _perm_write_ok(cfg, user), **data}


@app.get("/board/{b}/post/{pid}")
def board_view(b: str, pid: int, user: Optional[dict] = Depends(current_user)) -> dict:
    """게시글 보기(조회수 +1). 읽기 권한 필요."""
    cfg = _board_cfg_or_404(b)
    if not _perm_read_ok(cfg, user):
        raise HTTPException(403, _read_reason(cfg, user))
    p = user_store.get_post(pid)
    if not p or p["board"] != b:
        raise HTTPException(404, "게시글을 찾을 수 없습니다.")
    can_edit = bool(user) and (_is_admin(user) or user["id"] == p.get("author_id"))
    return {**p, "can_edit": can_edit,
            "comments": user_store.list_comments(pid),
            "can_comment": _perm_comment_ok(cfg, user),
            "comment_off": cfg["comment"] == "off"}


@app.post("/board/{b}")
def board_create(b: str, body: PostIn, user: dict = Depends(require_user)) -> dict:
    """글쓰기 — 게시판 쓰기 권한에 따름."""
    cfg = _board_cfg_or_404(b)
    if not _perm_write_ok(cfg, user):
        raise HTTPException(403, f"이 게시판은 '{_PERM_LABEL.get(cfg['write'], cfg['write'])}'만 글을 등록할 수 있습니다.")
    content = _sanitize_html(body.content)        # 에디터 HTML → XSS 제거 후 저장
    try:
        p = user_store.create_post(b, body.title, content, user["id"],
                                   user.get("name") or user.get("email") or "회원")
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "id": p["id"]}


@app.delete("/board/{b}/post/{pid}")
def board_delete(b: str, pid: int, user: dict = Depends(require_user)) -> dict:
    """글삭제(작성자 본인 또는 관리자)."""
    _board_cfg_or_404(b)
    p = user_store.get_post(pid, bump=False)
    if not p or p["board"] != b:
        raise HTTPException(404, "게시글을 찾을 수 없습니다.")
    if not _is_admin(user) and user["id"] != p.get("author_id"):
        raise HTTPException(403, "본인 글만 삭제할 수 있습니다.")
    user_store.delete_post(pid)
    return {"ok": True}


@app.post("/board/{b}/post/{pid}/comments")
def comment_create(b: str, pid: int, body: CommentIn,
                   user: dict = Depends(require_user)) -> dict:
    """댓글 작성 — 읽기 권한 + 댓글 권한 충족 시."""
    cfg = _board_cfg_or_404(b)
    if not _perm_read_ok(cfg, user):
        raise HTTPException(403, _read_reason(cfg, user))
    if not _perm_comment_ok(cfg, user):
        raise HTTPException(403, ("이 게시판은 댓글을 받지 않습니다." if cfg["comment"] == "off"
                                  else f"댓글은 '{_PERM_LABEL.get(cfg['comment'], cfg['comment'])}'만 작성할 수 있습니다."))
    p = user_store.get_post(pid, bump=False)
    if not p or p["board"] != b:
        raise HTTPException(404, "게시글을 찾을 수 없습니다.")
    try:
        c = user_store.create_comment(pid, user["id"],
                                      user.get("name") or user.get("email") or "회원",
                                      body.content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "comment": c}


@app.delete("/board/{b}/post/{pid}/comments/{cid}")
def comment_delete(b: str, pid: int, cid: int,
                   user: dict = Depends(require_user)) -> dict:
    """댓글 삭제(작성자 본인 또는 관리자)."""
    _board_cfg_or_404(b)
    c = user_store.get_comment(cid)
    if not c or c["post_id"] != pid:
        raise HTTPException(404, "댓글을 찾을 수 없습니다.")
    if not _is_admin(user) and user["id"] != c.get("author_id"):
        raise HTTPException(403, "본인 댓글만 삭제할 수 있습니다.")
    user_store.delete_comment(cid)
    return {"ok": True}


# ── 게시판 권한 설정(관리자) ──
class BoardSettingIn(BaseModel):
    title: Optional[str] = None
    read_perm: Optional[str] = None
    write_perm: Optional[str] = None
    comment_perm: Optional[str] = None


@app.get("/admin/boards")
def admin_boards(admin: dict = Depends(require_admin)) -> dict:
    """게시판별 권한 설정 목록."""
    out = []
    for b in _BOARDS:
        cfg = _board_cfg(b)        # 미시드면 시드
        cnt = user_store.list_posts(b, 1).get("total", 0)
        out.append({"board": b, "title": cfg["title"], "desc": cfg["desc"],
                    "read_perm": cfg["read"], "write_perm": cfg["write"],
                    "comment_perm": cfg["comment"], "posts": cnt})
    return {"boards": out}


@app.put("/admin/boards/{b}")
def admin_board_update(b: str, body: BoardSettingIn,
                       admin: dict = Depends(require_admin)) -> dict:
    """게시판 권한 설정 변경."""
    _board_cfg_or_404(b)           # 존재 확인 + 시드 보장
    try:
        s = user_store.update_board_setting(b, title=body.title, read_perm=body.read_perm,
                                             write_perm=body.write_perm,
                                             comment_perm=body.comment_perm)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "setting": s}


_prewarm_proc: dict = {"p": None}   # 통합 예열 스크립트 프로세스(버튼/나이트 동일 _prewarm_selfheal.py)


@app.post("/admin/prewarm")
def admin_prewarm(only: str = Query("", pattern="^[a-z_,]*$"),
                  admin: dict = Depends(require_admin)) -> dict:
    """수동 예열: 통합 자가치유 스크립트를 백그라운드 프로세스로 실행.
    버튼·나이트 자동예열이 '같은 스크립트'를 돌림. only: ''(전체)|docs|villa_est|apt|car|nearby."""
    import subprocess as _sp
    import sys as _sys
    p = _prewarm_proc["p"]
    if p is not None and p.poll() is None:
        return {"started": False, "reason": "이미 실행 중"}
    args = [_sys.executable, "_prewarm_selfheal.py"]
    if only:
        args += ["--only", only]
    try:
        logf = open(os.path.join(_ROOT, "_selfheal.log"), "a", encoding="utf-8")
        env = dict(os.environ, PYTHONIOENCODING="utf-8")
        _prewarm_proc["p"] = _sp.Popen(args, cwd=_ROOT, stdout=logf,
                                       stderr=_sp.STDOUT, env=env)
    except Exception as e:
        raise HTTPException(500, f"예열 시작 실패: {e}")
    return {"started": True, "only": only or "all"}


@app.get("/admin/prewarm/status")
def admin_prewarm_status(admin: dict = Depends(require_admin)) -> dict:
    """예열 적재 현황: 종류별 DB 행수 + 실행 중 여부 + 최근 로그."""
    counts = {k: auction_db.cache_count(k) for k in ("brief", "apt", "nearby",
                                                     "analysis", "appraisal",
                                                     "docsummary", "vehicle2")}
    p = _prewarm_proc["p"]
    running = bool(p is not None and p.poll() is None)
    tail = []
    try:
        with open(os.path.join(_ROOT, "_selfheal.log"), encoding="utf-8") as f:
            tail = [ln.rstrip() for ln in f.readlines()[-16:] if ln.strip()]
    except Exception:
        pass
    return {"counts": counts, "total": sum(counts.values()),
            "running": running, "log": tail}


_landuse_proc = {"running": False, "last": None}


@app.post("/admin/landuse_prewarm")
def admin_landuse_prewarm(limit: int = Query(0, ge=0, le=2000),
                          admin: dict = Depends(require_admin)) -> dict:
    """원천 누락 용도지역을 V-World로 보완(백그라운드). limit=0이면 전체."""
    if _landuse_proc["running"]:
        return {"started": False, "reason": "이미 실행 중"}

    def job():
        _landuse_proc["running"] = True
        try:
            _landuse_proc["last"] = _landuse_prewarm(limit=limit or None)
        except Exception as e:
            _landuse_proc["last"] = {"error": f"{type(e).__name__}: {e}"}
        finally:
            _landuse_proc["running"] = False

    threading.Thread(target=job, daemon=True).start()
    return {"started": True, "limit": limit or "all"}


@app.get("/admin/landuse_prewarm/status")
def admin_landuse_status(admin: dict = Depends(require_admin)) -> dict:
    """용도지역 보완 현황: 캐시 적재(라벨/NF) + 실행 여부 + 마지막 결과."""
    labeled = sum(1 for v in _lu_cache.values() if v and v != "NF")
    nf = sum(1 for v in _lu_cache.values() if v == "NF")
    return {"running": _landuse_proc["running"], "last": _landuse_proc["last"],
            "vworld_filled": labeled, "vworld_notfound": nf, "cache_total": len(_lu_cache)}


@app.get("/auction/apt_brief")
def auction_apt_brief(item_key: str) -> dict:
    """단건(상세페이지용). 모든 주거용 준공·세대·승강기."""
    r = _get_brief(item_key)
    _save_brief_cache()
    return r


@app.get("/auction/briefs")
def auction_briefs(keys: str) -> dict:
    """검색목록용 배치: 여러 item_key를 서버에서 동시 조회해 한 번에 반환(순차호출 제거)."""
    klist = [k for k in keys.split(",") if k][:80]
    _load_briefs_from_db(klist)                    # ① Supabase building_brief 일괄 로딩
    todo = [k for k in klist if k not in _brief_cache]
    if todo:                                        # ② 없는 것만 계산(계산 시 DB 저장)
        with _cf.ThreadPoolExecutor(max_workers=8) as ex:
            list(ex.map(_get_brief, todo))
        _save_brief_cache()
    return {k: _brief_cache.get(k, {"available": False}) for k in klist}


def _compute_similar(item_key: str):
    """빌라/연립/도시형: 반경 1km 유사거래 건수(상세와 '동일' — nearby 캐시 공유). 무거움, 캐시."""
    try:
        nb = auction_nearby_trades(item_key)        # 상세와 같은 반경1km 계산+캐시(nearby:)
        if isinstance(nb, dict) and nb.get("available") and nb.get("geo_ok"):
            return len(nb.get("trades") or [])      # 지오코딩 신뢰 결과만(불신뢰는 None→재시도)
    except Exception:
        pass
    return None


def _get_similar(item_key: str):
    global _similar_dirty
    if item_key in _similar_cache:
        return _similar_cache[item_key]
    c = _compute_similar(item_key)
    if c:                                 # 양수만 캐시(0/불명은 캐시 안 함 → 할당량 리셋 후 자동 재조회)
        _similar_cache[item_key] = c
        _similar_dirty = True
    return c


@app.get("/auction/similar")
def auction_similar(keys: str) -> dict:
    """목록용 유사거래 건수 배치(비동기). brief와 분리해 준공/세대/승강기 지연 방지."""
    klist = [k for k in keys.split(",") if k][:80]
    todo = [k for k in klist if k not in _similar_cache]
    if todo:
        with _cf.ThreadPoolExecutor(max_workers=8) as ex:
            list(ex.map(_get_similar, todo))
        _save_similar_cache()
    return {k: _similar_cache[k] for k in klist if k in _similar_cache}


@app.get("/auction/docsummary")
def auction_docsummary(item_key: str) -> dict:
    """명세서 요약사항(최선순위·소멸되지않는권리·지상권·주의사항) + 법원문건접수/송달. DB 우선→계산 후 DB저장."""
    from auction_analysis.doc_analysis import analyze_doc_summary
    from auction_analysis.sale_statement_parser import clean_summary
    res = _cached_doc("docsummary", item_key,
                      lambda: analyze_doc_summary(auction_db, item_key))
    # 캐시된 명세서 요약 정리: 다운로드 워터마크 제거 + 글자 2배 추출(doubled) 복원
    if isinstance(res, dict):
        for f in ("surviving_rights", "ground_rights", "notes"):
            if isinstance(res.get(f), str):
                res[f] = clean_summary(res[f])
    return res


@app.get("/auction/vehicle")
def auction_vehicle(item_key: str) -> dict:
    """차량외(자동차·중기): 매각물건명세서 '자동차의 표시' → 차량/중기현황. DB 우선→계산 후 DB저장."""
    from auction_analysis.doc_analysis import analyze_vehicle, _mark_multi
    r = _cached_doc("vehicle2", item_key,
                    lambda: analyze_vehicle(auction_db, item_key))
    return _mark_multi(r, auction_db, item_key) if isinstance(r, dict) else r   # 캐시된 옛값도 일괄매각 플래그 보장(사건에 타물번 차량 있으면 오판 해제)


def _match_fields(v: dict) -> dict:
    """차량 결과 → 매칭용 필드. 세부등급(model_grade)이 있으면 차명에 합쳐 정밀 매칭(예: BMW→BMW 528i)."""
    fm = {f["label"]: f["value"] for f in v.get("fields", [])}
    grade = v.get("model_grade") or fm.get("세부등급")
    if grade and str(grade) not in (fm.get("차명") or ""):
        fm["차명"] = ((fm.get("차명") or "") + " " + str(grade)).strip()
    return fm


def _compute_encar(item_key: str) -> dict:
    """차량외 매물 → 엔카 동일차량 매칭(브랜드·모델·연식·연료·배기량·주행거리±5천)."""
    from auction_analysis.doc_analysis import analyze_vehicle
    from auction_analysis.encar_match import match_vehicle
    v = analyze_vehicle(auction_db, item_key)
    if not v.get("available"):
        return {"available": False, "reason": "차량 정보 없음"}
    if v.get("multi_vehicle"):   # 일괄매각(여러 대) → 여러 대가 한 필드로 섞여 단일 시세가 오도 → 미제공
        return {"available": False, "reason": "일괄매각(여러 대) — 개별 시세 미제공", "multi_vehicle": True}
    return match_vehicle(_match_fields(v))


@app.get("/auction/encar_matches")
def auction_encar_matches(item_key: str) -> dict:
    """경매/공매 차량외 매물에 해당하는 엔카 중고차 매물(가격순+평균). DB 우선→계산 후 DB저장.
    보험이력(내차/상대차 피해·용도변경)은 doc 캐시 바깥에서 매 요청 부착(insurance_cache 기반)."""
    res = _cached_doc("encar2", item_key, lambda: _compute_encar(item_key))
    if isinstance(res, dict) and res.get("cars"):
        try:
            from auction_analysis.encar_match import _attach_insurance
            _attach_insurance(res["cars"])
        except Exception:
            pass
    return res


def _apt_est_value(d) -> Optional[dict]:
    """apt_info dict → 목록 시세값. 추정시세 우선, 없으면 최근 실거래가 폴백.
    trades_3m = 동일평형 최근 3개월(92일) 실거래 건수(deal_date 기준)."""
    if not isinstance(d, dict):
        return None
    t3 = None
    tr = d.get("trades")
    if isinstance(tr, list):
        from datetime import date as _date, timedelta as _td
        cut = (_date.today() - _td(days=92)).isoformat()
        t3 = sum(1 for t in tr if (t.get("deal_date") or "") >= cut)
    est = d.get("est")
    if isinstance(est, dict) and est.get("price"):
        return {"price": est["price"], "kind": "est", "trades_3m": t3}   # 추정시세
    summ = d.get("summary")                                       # 산출불가 → 최근 실거래가 폴백
    rec = summ.get("recent") if isinstance(summ, dict) else None
    return {"price": rec, "kind": "recent", "trades_3m": t3} if rec else None


@app.get("/auction/apt_ests")
def auction_apt_ests(keys: str, compute: bool = True, remote: bool = True) -> dict:
    """목록용: 아파트 추정시세(원). ①캐시 우선 ②미캐시는 즉석 계산(상세와 동일)+캐시 저장.
    remote=False면 Supabase 왕복 생략(로컬 캐시만) — 목록 enrich 고속용(미캐시는 프론트가 async로 채움)."""
    klist = [k for k in keys.split(",") if k][:200]
    out: dict = {}
    need_remote: list[str] = []
    for k in klist:                                               # ① 로컬 메모리 캐시 우선(원격 왕복 0)
        d = _apt_cache.get(k)
        if isinstance(d, dict) and d.get("available") and d.get("v", 0) >= APT_VER:
            out[k] = _apt_est_value(d)
        else:
            need_remote.append(k)                                # 옛버전(v<APT_VER)도 재계산 대상
    miss: list[str] = []
    if need_remote and remote:                                   # ② 메모리에 없으면 Supabase
        try:
            rows = auction_db.cache_get_many(["apt:" + k for k in need_remote])
        except Exception:
            rows = {}
        for k in need_remote:
            d = rows.get("apt:" + k)
            if isinstance(d, dict) and d.get("v", 0) >= APT_VER:
                out[k] = _apt_est_value(d)
                _apt_cache.remember(k, d)                         # 메모리에도 올려 다음엔 즉시
            else:
                out[k] = None
                miss.append(k)                                    # ③ 미캐시/옛버전 → 계산 대상
    if compute and miss:
        def one(k):
            try:
                d = _apt_info_compute(k, 12)
                if d.get("available"):                            # 실패(molit 등)는 캐시 안 함
                    _apt_cache.remember(k, d)
                    try:
                        auction_db.cache_save("apt:" + k, d)
                    except Exception:
                        pass
                return k, _apt_est_value(d)
            except Exception:
                return k, None
        with _cf.ThreadPoolExecutor(max_workers=6) as ex:
            for k, v in ex.map(one, miss):
                out[k] = v
    return out


def _villa_est_value(item_key: str) -> Optional[dict]:
    """빌라/연립/도시형 목록 시세 = 이 빌라 공시가격 × 유사 실거래 평균 요율.
    가장 유사(준공±5년) 그룹 우선 — 상세 화면(loadNearbyMap)의 'sim' 그룹과 동일 기준.
    거래별 공시가격은 auction_gongsi(gongsi: 캐시 공유)로 조회 → 요율 평균."""
    try:
        nb = auction_nearby_trades(item_key)
    except Exception:
        return None
    if not (isinstance(nb, dict) and nb.get("available") and nb.get("geo_ok")):
        return None
    trades = nb.get("trades") or []
    if not trades:
        return None                              # 주변 유사거래 0 → 못 구함(원천 없음)
    py = nb.get("prop_build_year")
    sgg = nb.get("sigungu_prefix") or ""
    # molit 거래의 build_year/amount는 문자열일 수 있어 정수로 변환(프론트 JS는 자동변환).
    best = [t for t in trades
            if py and _to_int(t.get("build_year")) is not None
            and abs(_to_int(t["build_year"]) - py) <= 5] or trades
    best = sorted(best, key=lambda t: t.get("deal_date", ""), reverse=True)[:14]
    # ① 이 빌라 공시가격 있으면 = 공시가격 × 유사실거래 평균요율(정확)
    pg = nb.get("prop_gongsi")
    if isinstance(pg, dict) and pg.get("price"):
        keys = [f"{(sgg + ' ' + (t.get('umd') or '') + ' ' + (t.get('jibun') or '')).strip()}"
                f"|{t.get('area') or ''}|{t.get('floor') or ''}" for t in best]
        try:
            gmap = auction_gongsi(";".join(keys))   # DB/V-World 공시가격 캐시 공유
        except Exception:
            gmap = {}
        yuls = []
        for t, gk in zip(best, keys):
            amt = _to_int(t.get("amount"))
            g = gmap.get(gk)
            if amt and isinstance(g, dict) and g.get("price"):
                yuls.append(amt / g["price"])
        pgp = _to_int(pg.get("price"))
        if yuls and pgp:
            yul = sum(yuls) / len(yuls)
            return {"price": round(pgp * yul), "yul": round(yul, 4), "gongsi": pgp, "kind": "est"}
    # ② 공시가격 없음/요율 못구함 → 유사 실거래 평균가 폴백(가장 유사 준공±5년 그룹)
    amts = [_to_int(t.get("amount")) for t in best if _to_int(t.get("amount"))]
    if amts:
        return {"price": round(sum(amts) / len(amts)), "kind": "trade_avg", "n": len(amts)}
    return None


_villa_cache: dict = {}   # item_key -> {"price","kind"} | None  (메모리 캐시 — 목록 enrich 고속화)


@app.get("/auction/villa_ests")
def auction_villa_ests(keys: str, compute: bool = True, remote: bool = True) -> dict:
    """목록용: 빌라/연립/도시형 추정시세(원, 공시가격×요율). ①메모리 ②Supabase(villaest:) ③미캐시 즉석.
    remote=False면 Supabase 왕복 생략(메모리만) — 목록 enrich 고속용."""
    _all = [k for k in keys.split(",") if k]
    if len(_all) > 120:                          # ★120 초과 호출도 안전(舊 [:120] 조용한 잘라냄이 예열 covered_set 오판·허수 미처리 유발) — 120씩 재귀 처리
        merged: dict = {}
        for _i in range(0, len(_all), 120):
            merged.update(auction_villa_ests(",".join(_all[_i:_i + 120]), compute, remote))
        return merged
    klist = _all
    out: dict = {}
    need: list[str] = []
    for k in klist:                               # ① 메모리 캐시 우선(원격 왕복 0)
        if k in _villa_cache:
            out[k] = _villa_cache[k]
        else:
            need.append(k)
    rows = {}
    if need and remote:                           # ② 메모리에 없으면 Supabase
        try:
            rows = auction_db.cache_get_many(["villaest:" + k for k in need])
        except Exception:
            rows = {}
    miss: list[str] = []
    for k in need:
        d = rows.get("villaest:" + k)
        if isinstance(d, dict) and d.get("price"):
            v = {"price": d["price"], "kind": "est"}
            out[k] = v; _villa_cache[k] = v       # 메모리에도 올려 다음엔 즉시
        else:
            out[k] = None
            if remote:
                miss.append(k)                   # 미캐시 → 계산 대상(remote일 때만)
    if compute and miss:
        def one(k):
            try:
                v = _villa_est_value(k)
                if v and v.get("price"):
                    try:
                        auction_db.cache_save("villaest:" + k, v)
                    except Exception:
                        pass
                    return k, {"price": v["price"], "kind": "est"}
            except Exception:
                pass
            return k, None
        with _cf.ThreadPoolExecutor(max_workers=4) as ex:
            for k, v in ex.map(one, miss):
                out[k] = v
    return out


@app.get("/auction/encar_avgs")
def auction_encar_avgs(keys: str, compute: bool = True) -> dict:
    """목록용: 여러 차량외 매물의 엔카 동일중고차 '평균가(만원)'. ①캐시 우선 ②미캐시는 즉석 계산(상세와 동일)+캐시 저장."""
    klist = [k for k in keys.split(",") if k][:200]
    try:
        rows = auction_db.cache_get_many(["encar2:" + k for k in klist])
    except Exception:
        rows = {}
    out, miss = {}, []
    for k in klist:
        d = rows.get("encar2:" + k)
        if isinstance(d, dict):
            out[k] = d.get("avg_price") if d.get("count") else None
        else:
            out[k] = None
            miss.append(k)                 # 미캐시 → 계산 대상
    if compute and miss:
        def one(k):
            try:
                d = _cached_doc("encar2", k, lambda kk=k: _compute_encar(kk))
                return k, (d.get("avg_price") if isinstance(d, dict) and d.get("count") else None)
            except Exception:
                return k, None
        with _cf.ThreadPoolExecutor(max_workers=6) as ex:
            for k, v in ex.map(one, miss):
                out[k] = v
    return out


def _compute_review(item_key: str) -> dict:
    """차량외 매물 → 해당 차종 전문가/사용자 후기(Neon reviews, 연식 근접)."""
    from auction_analysis.doc_analysis import analyze_vehicle
    from auction_analysis.review_match import match_reviews
    v = analyze_vehicle(auction_db, item_key)
    if not v.get("available"):
        return {"available": False, "reason": "차량 정보 없음"}
    return match_reviews(_match_fields(v))


@app.get("/auction/reviews")
def auction_reviews(item_key: str) -> dict:
    """차량외 매물의 해당 차종 후기(전문가 평가·실사용자 댓글·FAQ). DB 우선→계산 후 DB저장."""
    return _cached_doc("review", item_key, lambda: _compute_review(item_key))


# ---------- 공매(온비드) ----------
from auction_analysis.onbid_source import OnbidSource  # noqa: E402
onbid = OnbidSource()


def _gm_q(v: str) -> str:
    """PostgREST and=() 내부 값 이스케이프(따옴표/콤마 제거)."""
    return '"' + str(v).replace('"', "").replace(",", " ").strip() + '"'


_GM_SORTS = {   # 정렬 키 → PostgREST order (감정가·최저가는 nullslast로 비공개(NULL) 뒤로)
    "bid_close": "bid_close.asc", "appr": "appraisal_price.desc.nullslast",
    "low": "min_price.asc.nullslast",
    # 프리미엄 정렬(전역) — 종전엔 매핑 부재로 bid_close.asc 폴백 = 정렬 미작동 버그였음(주인님 2026-07-08).
    "profit_desc": "profit.desc.nullslast",     # 차익 높은순
    "nb_desc": "nb_count.desc.nullslast",       # 유사거래 많은순(빌라) / 실거래 3개월(아파트) — nb_count 공용
    "trade_desc": "nb_count.desc.nullslast",    # 실거래 많은순(아파트) — nb_count 공용
}


def _gm_esc_like(v: str) -> str:
    """PostgREST or=()/and=() ilike 값에 들어갈 소재지·용도 토큰 이스케이프.
    or/and 그룹 문법( , ( ) )을 깨는 문자를 제거 — 값에 콤마·괄호가 오면 그룹이 오해석되므로."""
    return str(v).replace(",", " ").replace("(", " ").replace(")", " ").replace('"', "").strip()


def _gm_region_usage_filters(*, regions=None, sido=None, sgg=None, usages=None,
                             usage=None) -> list[tuple]:
    """공매 소재지(복수)·용도(복수) 필터를 (key,value) 튜플 리스트로.
    **경매 supabase_source._filters 의 regions/sido/usages 로직을 그대로 이식**:
      · regions(여러 소재지) = 지역끼리 OR, 각 지역 안 토큰(구군+동)은 AND
        → or=(and(address.ilike.*구군*,address.ilike.*동*),and(...))
        · normalize_address 로 도로명주소도 (시도,구군,동) 토큰화 → 비연속 AND 로 매칭
      · sido(시/도만) = **전방일치**(대구*) + 표기변형(_SIDO_VARIANTS) OR
        → '%대구%'가 부산 '해운대구'에 오매칭되던 버그 방지(전방일치가 핵심)
      · usages(여러 용도) = ilike OR
    호출부(gongmae_list)에서 여러 or 그룹은 and=(or(),or()) 로 병합(PostgREST root or 1개 제한).
    legacy sgg/usage(단일, 저장검색 호환)도 처리."""
    f: list[tuple] = []

    # ── 소재지 복수(+버튼) : 지역끼리 OR, 각 지역 토큰 AND ──
    reg_list = list(regions) if isinstance(regions, (list, tuple)) else ([regions] if regions else [])
    # legacy 단일(sido+sgg): regions 없을 때만 하나의 지역으로 취급(구군은 토큰 AND, 시도는 아래 전방일치)
    _legacy_region = None
    if not reg_list and sgg and sgg not in ("구/군", "구/군 전체"):
        _legacy_region = ((sido + " ") if sido else "") + sgg  # "대구광역시 수성구" → 토큰 AND
    if _legacy_region:
        reg_list = [_legacy_region]
        sido = None   # 구군까지 지정되면 시도 전방일치는 불필요(토큰 AND에 시도 포함)
    if reg_list:
        _rgroups = []
        for _reg in reg_list:
            # normalize_address 로 시도표준화+도로명 보정 → 토큰. 실패해도 원문 split 로 폴백.
            _sd, _gu, _dong = normalize_address(str(_reg))
            _toks = [t for t in (_sd, _gu, _dong) if t]
            if not _toks:                                   # normalize 실패 → 원문 토큰
                _toks = [t for t in str(_reg).split() if t]
            _toks = [_gm_esc_like(t) for t in _toks if _gm_esc_like(t)]
            if not _toks:
                continue
            if len(_toks) == 1:
                _rgroups.append(f"address.ilike.*{_toks[0]}*")
            else:                                           # 비연속 AND(지번·도로명 모두 매칭)
                _inner = ",".join(f"address.ilike.*{t}*" for t in _toks)
                _rgroups.append(f"and({_inner})")
        if _rgroups:
            f.append(("or", f"({','.join(_rgroups)})"))

    # ── 시/도만 : 전방일치 + 표기변형 OR (해운대구 오매칭 방지) ──
    if sido:
        variants = _SIDO_VARIANTS.get(sido, [sido])         # 병합라벨(전남광주통합특별시)은 키 없음 → self 전방일치(DB 저장형이 그 자체)
        ors = ",".join(f"address.ilike.{_gm_esc_like(v)}*" for v in variants)
        f.append(("or", f"({ors})"))

    # ── 용도 복수 : ilike OR (세부용도/대분류 문자열 부분일치) ──
    usg_list = list(usages) if isinstance(usages, (list, tuple)) else ([usages] if usages else [])
    if not usg_list and usage:                              # legacy 단일 용도
        usg_list = [usage]
    usg_list = [u for u in (_gm_esc_like(x) for x in usg_list) if u]
    if usg_list:
        if len(usg_list) == 1:
            f.append(("usage", f"ilike.*{usg_list[0]}*"))
        else:
            ors = ",".join(f"usage.ilike.*{u}*" for u in usg_list)
            f.append(("or", f"({ors})"))
    return f


@app.get("/gongmae")
def gongmae_list(page: int = 1, rows: int = Query(20, le=100),
                 prop: Optional[str] = "압류재산", dpsl_mtd: Optional[str] = None,
                 usg_lcls: Optional[str] = None, goods: Optional[str] = None,
                 sido: Optional[str] = None, sgg: Optional[str] = None,
                 usage: Optional[str] = None,
                 regions: Optional[list[str]] = Query(None, description="소재지 다중(+버튼, 각 '시도 구군 동' 문자열, 지역끼리 OR·토큰 AND)"),
                 usages: Optional[list[str]] = Query(None, description="용도 다중(+버튼, usage 부분일치 OR)"),
                 appr_min: Optional[int] = None, appr_max: Optional[int] = None,
                 low_min: Optional[int] = None, low_max: Optional[int] = None,
                 grade: Optional[str] = None,
                 sort: Optional[str] = None, sort2: Optional[str] = None) -> dict:
    """온비드 공매물건 목록 — 우리 DB(gongmae_items) 소재지/재산유형/명칭/감정가·최저가/정렬 필터.
    온비드 API가 소재지 검색을 지원하지 않아 전량 적재분을 우리가 필터한다.
    소재지·용도는 **복수 선택**(경매 supabase_source._filters regions/sido/usages 로직 이식):
      · regions = 지역끼리 OR, 각 지역 토큰 AND / sido = 전방일치+표기변형 OR(해운대구 오매칭 방지)
      · usages = ilike OR / normalize_address 로 도로명주소도 매칭
    DB 실패 시 온비드 라이브 API로 폴백. (특수물건·유찰수는 공매 데이터 없어 제외)"""
    try:
        # ① 스칼라(단일 컬럼) 조건 → and=(...) 그룹
        conds = []   # (col, op, val)
        if prop:
            conds.append(("prop_type", "eq", prop))
        if goods:
            conds.append(("name", "ilike", f"*{goods}*"))
        if dpsl_mtd:
            conds.append(("disposal", "ilike", f"*{dpsl_mtd}*"))
        if appr_min is not None:
            conds.append(("appraisal_price", "gte", str(int(appr_min))))
        if appr_max is not None:
            conds.append(("appraisal_price", "lte", str(int(appr_max))))
        if low_min is not None:
            conds.append(("min_price", "gte", str(int(low_min))))
        if low_max is not None:
            conds.append(("min_price", "lte", str(int(low_max))))
        if grade:                                           # 매수판정 서버필터(매수양호/검토/금지) — 100건 채운 해당등급만(클라 필터는 현재페이지만 걸러 <10건 문제)
            conds.append(("buy_grade", "eq", grade))
        # ② 소재지(복수)·용도(복수) → or 그룹(들) (경매 검증 로직 이식). legacy sido/sgg/usage 도 흡수.
        ru = _gm_region_usage_filters(regions=regions, sido=sido, sgg=sgg,
                                      usages=usages, usage=usage)
        # 전역 정렬: 1차(sort) + 2차(sort2, 동점 시) + bid_close 안정 타이브레이커(페이지네이션 결정성 보장).
        _o1 = _GM_SORTS.get(sort or "", None)
        _o2 = _GM_SORTS.get(sort2 or "", None)
        order = ",".join(dict.fromkeys([p for p in (_o1, _o2, "bid_close.asc") if p]))
        # 사전계산(warm_gongmae_grade.py) 컬럼(buy_grade·sise·profit·grade_reason·nb_count)도 select해
        # item에 병합. 목록이 행별 라이브 호출 없이 배지·시세·차익·유사거래건수를 즉시 렌더.
        params = {"select": "data,bid_close,buy_grade,sise,profit,grade_reason,nb_count,apt_hoga,recent_trade_price,recent_trade_date",
                  "order": order, "offset": str(max(0, (page - 1) * rows)), "limit": str(rows)}
        # ③ and(스칼라) + or(지역/용도) 병합.
        #    PostgREST root or= 는 1개만 허용 → or 그룹이 2개↑(복수지역 + 용도복수, 또는 시도 + 용도)면
        #    and=(or(X),or(Y)) 로 묶는다(경매 _filters 말미 로직과 동일).
        or_groups = [v for (k, v) in ru if k == "or"]
        scalar_and = list(conds)   # (col,op,val)
        # ru 중 or 아닌 것(단일 usage ilike 등)은 scalar_and 로 편입
        for (k, v) in ru:
            if k != "or":
                # v = "ilike.*xxx*" 형태 → (col=k, op·val 통째)
                _op, _val = v.split(".", 1)
                scalar_and.append((k, _op, _val))

        def _v(op, val):
            return val if op in ("gte", "lte") else _gm_q(val)
        and_parts = [f"{c}.{op}.{_v(op, val)}" for c, op, val in scalar_and]
        if len(or_groups) >= 2:
            # 여러 or 그룹 → and 안에 or(...) 들로 결합(+스칼라 조건도 같은 and 에)
            merged = and_parts + [f"or{g}" for g in or_groups]
            params["and"] = "(" + ",".join(merged) + ")"
        elif len(or_groups) == 1:
            if and_parts:
                params["and"] = "(" + ",".join(and_parts) + ")"
            params["or"] = or_groups[0]   # root or 1개는 그대로
        else:
            if and_parts:
                params["and"] = "(" + ",".join(and_parts) + ")"
        resp = auction_db._get("gongmae_items", params, count=True)
        if resp.status_code >= 400:
            raise RuntimeError(f"db {resp.status_code}")
        data_rows = resp.json()
        cr = resp.headers.get("content-range", "")
        total = int(cr.split("/")[-1]) if "/" in cr and cr.split("/")[-1].isdigit() else len(data_rows)
        items = []
        for r in data_rows:
            d = r.get("data")
            if not d:
                continue
            # 사전계산 컬럼을 item에 병합(warm된 것만 값 존재, 미워밍은 None→프론트가 폴백 소수 호출)
            d["buy_grade"] = r.get("buy_grade")
            d["sise"] = r.get("sise")
            d["profit"] = r.get("profit")
            d["grade_reason"] = r.get("grade_reason")
            d["nb_count"] = r.get("nb_count")   # 유사(주변) 실거래 건수(빌라) / 아파트·오피스텔은 실거래 3개월 건수
            d["apt_hoga"] = r.get("apt_hoga")   # 아파트·오피스텔 KB 호가(같은 단지·평형 매매매물) 건수(precompute_apt_chips.py)
            d["recent_trade_price"] = r.get("recent_trade_price")   # 최근 실거래가(추정시세 없을 때 fallback 기준·목록 시세칸 표시)
            d["recent_trade_date"] = r.get("recent_trade_date")     # 그 체결일(YYYY-MM-DD)
            items.append(d)
        return {"items": items, "total": total, "page": page, "source": "db"}
    except Exception as e:
        # DB 미가용 시 라이브 폴백(소재지 필터는 불가)
        out = onbid.list_items(page=page, rows=rows, prop=prop or "압류재산",
                               dpsl_mtd=dpsl_mtd, usg_lcls=usg_lcls, goods=goods)
        out["source"] = f"live({type(e).__name__})"
        return out


@app.get("/gongmae/enrich")
def gongmae_enrich(mng: str = Query(..., description="물건관리번호 cltrMngNo"),
                   cdtn: Optional[str] = Query(None, description="공매조건번호 pbctCdtnNo")) -> dict:
    """공매 물건상세 보강 — 온비드 상세 API로 면적·전체주소(도로명/지번)·PNU·사진·임대차 조회.
    경매 재사용 기능(유사거래·시세·예상낙찰가·아파트정보·경쟁매물·단기매도계산기)에 필요한
    {전용면적·전체주소·감정가}를 프론트에 제공한다.
    **온디맨드 캐시**: 결과를 api_cache('gm_enrich:*' 영구저장)에 저장 → 재조회는 온비드 호출 없이 즉답
    (예열 아님·요청 시 1회만 온비드 호출). 물건상세(면적·주소)는 경매 진행중 불변이라 영구캐시 안전."""
    ckey = "gm_enrich:" + mng + ((":" + cdtn) if cdtn else "")
    try:
        hit = auction_db.cache_get_many([ckey]).get(ckey)
        # notice_info 키 없으면 공고문·첨부·공고정보 추가 전 옛 캐시 → stale로 재조회(스키마 진화)
        if hit and "notice_info" in hit:
            return {**hit, "_cache": "hit"}
    except Exception:
        pass
    try:
        d = onbid.detail(mng, cdtn) or {}
    except Exception as e:
        return {"error": f"{type(e).__name__}"}
    # 성공(면적 또는 주소 확보)만 캐시 — 에러·빈결과는 캐시 안 함(다음에 재시도)
    if d and not d.get("error") and (d.get("bld_area") or d.get("addr_jibun")):
        try:
            auction_db.cache_save(ckey, d)
        except Exception:
            pass
    return d


# ---------- 공매 물건상세: 경매기능 재사용(온디맨드·gm_* 캐시·경매코드 무수정) ----------
_gm_nearby_cache: dict = {}


def _gm_cur(mng: str, cdtn: Optional[str] = None):
    """공매 enrich → 경매 내부함수용 합성 dict.
    주소에 '층' 부착(예낙/유사거래 층필터가 주소 정규식 의존)·usage=세부용도·building_area='N㎡'(전용)."""
    e = gongmae_enrich(mng, cdtn)
    if not e or e.get("error"):
        return None, None
    addr = (e.get("addr_jibun") or "").strip()
    if e.get("floor"):
        addr = f"{addr} {e['floor']}층"
    cur = {
        "address": addr,
        "usage": e.get("usage_scls") or "",
        "building_area": (f"{e['bld_area']}㎡" if e.get("bld_area") else None),
        "area_text": None,
        "appraisal_price": e.get("appraisal_price"),
        "land_area": e.get("land_area"),
        "item_key": None,
    }
    return e, cur


def _gm_attach_gongsi(result: dict, mng: str, cdtn: Optional[str], top: int = 20) -> dict:
    """공매 유사실거래 결과에 공시가격·요율 부여(경매 _villa_est_value와 동일 방식).
    ①물건 자체 공시가격(prop_gongsi) ②상위 top개 거래에 gongsi(공시가격)·yul(거래가/공시가격) 부착.
    V-World 쿼터/도로명 실패는 해당 항목만 공란(크래시 금지). base 'nearby:' 캐시는 오염 안 됨(별도 gm_gongsi: 캐시)."""
    if not (isinstance(result, dict) and result.get("available") and result.get("trades")):
        return result
    sgg = result.get("sigungu_prefix") or ""
    trades = result["trades"]
    picked = sorted(trades, key=lambda t: t.get("deal_date", ""), reverse=True)[:top]
    # ① 물건 자체 공시가격(gm_gongsi_prop:<mng> 캐시 — 성공분만 영구)
    pgk = "gm_gongsi_prop:" + mng
    prop_g = None
    try:
        hit = auction_db.cache_get_many([pgk]).get(pgk)
        if isinstance(hit, dict) and hit.get("price"):
            prop_g = hit
    except Exception:
        prop_g = None
    if prop_g is None:
        # 물건 전체 지번주소(끝 'N층' 제거 — resolve_bjd 지번 파싱 방해 방지). addr_jibun은 지번 숫자만이라 부적합.
        paddr = re.sub(r"\s*\d+\s*층\s*$", "", result.get("address") or "").strip()
        try:
            prop_g = _compute_prop_gongsi(None, paddr,
                                          result.get("prop_area"), result.get("prop_floor"))
        except Exception:
            prop_g = None
        if isinstance(prop_g, dict) and prop_g.get("price"):
            try:
                auction_db.cache_save(pgk, prop_g)
            except Exception:
                pass
    if isinstance(prop_g, dict) and prop_g.get("price"):
        result["prop_gongsi"] = prop_g
    # ② 거래별 공시가격·요율(auction_gongsi 캐시 공유)
    keys = [f"{(sgg + ' ' + (t.get('umd') or '') + ' ' + (t.get('jibun') or '')).strip()}"
            f"|{t.get('area') or ''}|{t.get('floor') or ''}" for t in picked]
    try:
        gmap = auction_gongsi(";".join(keys))
    except Exception:
        gmap = {}
    for t, gk in zip(picked, keys):
        g = gmap.get(gk)
        if isinstance(g, dict) and g.get("price"):
            t["gongsi"] = int(g["price"])
            amt = _to_int(t.get("amount"))
            if amt:
                t["yul"] = round(amt / g["price"], 4)
    return result


@app.get("/gongmae/nearby_trades")
def gongmae_nearby_trades(mng: str, cdtn: Optional[str] = None,
                          months: int = Query(12, le=24)) -> dict:
    """공매 빌라/도생 주변 유사실거래 — 경매 `_nearby_filtered` 그대로 재사용(온디맨드·gm_nearby: 캐시).
    반환 각 거래에 공시가격(gongsi)·요율(yul=거래가/공시가격) 부여(경매 표와 동일, V-World 실패분은 공란)."""
    if not isinstance(months, int):   # 내부에서 함수로 호출 시 Query 기본값 객체가 넘어옴 → 결과·캐시 오염(직렬화 500) 방지
        months = 12
    ck = "gm_nearby:" + mng
    if ck in _gm_nearby_cache:
        return _gm_attach_gongsi(_trim_to_radius(_gm_nearby_cache[ck]), mng, cdtn)
    try:
        db = auction_db.cache_get_many([ck]).get(ck)
    except Exception:
        db = None
    if isinstance(db, dict) and db.get("available"):
        _gm_nearby_cache[ck] = db
        return _gm_attach_gongsi(_trim_to_radius(db), mng, cdtn)
    e, cur = _gm_cur(mng, cdtn)
    if not cur:
        return {"available": False}
    if not re.search(r"다세대|연립|빌라|도시형", cur["usage"]):
        return {"available": False, "reason": "연립다세대/도시형 대상 아님"}
    addr = cur["address"]

    def _bi():
        try:
            b = building.info(addr)
            return _to_int(b.get("build_year")) if b else None
        except Exception:
            return None
    with _cf.ThreadPoolExecutor(max_workers=2) as ex:
        f_r = ex.submit(_nearby_filtered, cur)
        f_by = ex.submit(_bi)
        r = f_r.result()
        prop_build_year = f_by.result()
    if r is None:
        return {"available": False, "reason": "법정동코드 변환 실패", "address": addr}
    picked = sorted(r["trades"], key=lambda t: t.get("deal_date", ""), reverse=True)[:500]
    result = {
        "available": True, "v": 2, "address": addr, "addr_prefix": r["addr_prefix"],
        "sigungu_prefix": r["sigungu_prefix"], "addr_jibun": r["addr_jibun"],
        "prop_area": r["prop_area"], "prop_floor": r["prop_floor"],
        "prop_build_year": prop_build_year, "prop_gongsi": None, "months": months,
        "geo_ok": r.get("geo_ok", False), "prop_lng": r.get("prop_lng"), "prop_lat": r.get("prop_lat"),
        "trades": [{"amount": t["amount"], "area": t["area"], "build_year": t.get("build_year"),
                    "floor": t.get("floor"), "jibun": t.get("jibun"), "name": t.get("name"),
                    "umd": t.get("umd"), "deal_date": t.get("deal_date"),
                    "lng": t.get("lng"), "lat": t.get("lat")} for t in picked],
    }
    if result.get("geo_ok"):
        _gm_nearby_cache[ck] = result
        try:
            auction_db.cache_save(ck, result)
        except Exception:
            pass
    return _gm_attach_gongsi(_trim_to_radius(result), mng, cdtn)


@app.get("/gongmae/building_brief")
def gongmae_building_brief(mng: str, cdtn: Optional[str] = None) -> dict:
    """공매 건물정보(준공년도·세대·승강기) — 경매 building.info(주소) 재사용. gm_brief: 캐시."""
    ck = "gm_brief:" + mng
    try:
        hit = auction_db.cache_get_many([ck]).get(ck)
        if hit is not None:
            return hit
    except Exception:
        pass
    e, _cur = _gm_cur(mng, cdtn)
    out = {"available": False}
    if e:
        try:
            b = building.info(e.get("addr_jibun") or "")
            if b:
                out = {"available": True, **b}
        except Exception:
            pass
    if out.get("available"):
        try:
            auction_db.cache_save(ck, out)
        except Exception:
            pass
    return out


@app.get("/gongmae/villa_est")
def gongmae_villa_est(mng: str, cdtn: Optional[str] = None) -> dict:
    """공매 빌라 시세 = 공시가격×유사실거래 평균요율(없으면 유사거래 평균). 경매 _villa_est_value 로직 복제(nb는 공매)."""
    try:
        nb = gongmae_nearby_trades(mng, cdtn)
    except Exception:
        return {"available": False}
    if not (isinstance(nb, dict) and nb.get("available") and nb.get("geo_ok")):
        return {"available": False}
    trades = nb.get("trades") or []
    if not trades:
        return {"available": False}
    py = nb.get("prop_build_year")
    sgg = nb.get("sigungu_prefix") or ""
    best = [t for t in trades if py and _to_int(t.get("build_year")) is not None
            and abs(_to_int(t["build_year"]) - py) <= 5] or trades
    best = sorted(best, key=lambda t: t.get("deal_date", ""), reverse=True)[:14]
    pg = nb.get("prop_gongsi")
    if isinstance(pg, dict) and pg.get("price"):
        keys = [f"{(sgg + ' ' + (t.get('umd') or '') + ' ' + (t.get('jibun') or '')).strip()}"
                f"|{t.get('area') or ''}|{t.get('floor') or ''}" for t in best]
        try:
            gmap = auction_gongsi(";".join(keys))
        except Exception:
            gmap = {}
        yuls = []
        for t, gk in zip(best, keys):
            amt = _to_int(t.get("amount"))
            g = gmap.get(gk)
            if amt and isinstance(g, dict) and g.get("price"):
                yuls.append(amt / g["price"])
        pgp = _to_int(pg.get("price"))
        if yuls and pgp:
            yul = sum(yuls) / len(yuls)
            return {"available": True, "price": round(pgp * yul), "yul": round(yul, 4), "gongsi": pgp, "kind": "est"}
    amts = [_to_int(t.get("amount")) for t in best if _to_int(t.get("amount"))]
    if amts:
        return {"available": True, "price": round(sum(amts) / len(amts)), "kind": "trade_avg", "n": len(amts)}
    return {"available": False}


@app.get("/gongmae/villa_expected_bid")
def gongmae_villa_expected_bid(mng: str, cdtn: Optional[str] = None, sid: Optional[str] = Cookie(None)) -> dict:
    """공매 빌라/도생 예상낙찰가 — 경매 백데이터(빌라 매각사례) 참조·eb.compute_villa 무수정 재사용.
    반경1km·전용±6㎡·층±1·감정가±1500만·낙찰가율<100%·3년. gm_vexpbid: 캐시(온디맨드)."""
    from auction_analysis import expected_bid as eb
    ck = "gm_vexpbid:" + mng
    try:
        hit = auction_db.cache_get_many([ck]).get(ck)
        if isinstance(hit, dict):
            return _expbid_gate(hit, sid)   # cases_used(참조 백데이터)는 관리자만
    except Exception:
        pass
    e, cur = _gm_cur(mng, cdtn)
    if not cur:
        return {"available": False}
    if not _is_villa_usage(cur.get("usage")):
        return {"available": False, "reason": "빌라/도생 아님"}
    ll = _geocode(eb.geo_addr(cur.get("address")))
    if not ll:
        return {"available": False, "reason": "좌표 없음"}
    region = _villa_region_prefix(cur.get("address"))
    cases = []
    if region:
        try:
            rr = auction_db._get("items", {
                "select": "item_key,address,area_text,building_area,appraisal_price,sale_price,sale_rate,sell_date,bid_count,sale_2nd_price",
                "or": _VILLA_OR, "address": f"ilike.*{region}*",
                "sale_price": "gt.0", "order": "item_key", "limit": "2000"})
            cases = rr.json() if rr.status_code in (200, 206) else []
        except Exception:
            cases = []
    _geo_preload([eb.geo_addr(c.get("address")) for c in cases])   # ⚡Supabase geo: 배치 프리로드 → 2000건 개별 라이브 지오코딩 회피(클라우드 52초→수초)
    for c in cases:
        c["ll"] = _geocode(eb.geo_addr(c.get("address")))   # 프리로드로 대부분 메모리 적중
    est = None
    try:
        ev = gongmae_villa_est(mng, cdtn)
        if isinstance(ev, dict) and ev.get("available"):
            est = ev.get("price")
    except Exception:
        pass
    r = eb.compute_villa(cur, ll, cases, est_price=est)
    r["v"] = _VEXPBID_V
    if r.get("available"):
        try:
            auction_db.cache_save(ck, r)
        except Exception:
            pass
    return _expbid_gate(r, sid)   # cases_used는 관리자만(캐시엔 원본 저장, 반환만 게이트)


@app.get("/gongmae/expected_bid")
def gongmae_expected_bid(mng: str, cdtn: Optional[str] = None, sid: Optional[str] = Cookie(None)) -> dict:
    """공매 아파트 예상낙찰가 — 경매 백데이터(동일건물 아파트 매각사례)·eb.compute 무수정 재사용. gm_expbid: 캐시."""
    from auction_analysis import expected_bid as eb
    ck = "gm_expbid:" + mng
    try:
        hit = auction_db.cache_get_many([ck]).get(ck)
        if isinstance(hit, dict):
            return _expbid_gate(hit, sid)   # cases_used(참조 백데이터)는 관리자만
    except Exception:
        pass
    e, cur = _gm_cur(mng, cdtn)
    if not cur:
        return {"available": False}
    if "아파트" not in (cur.get("usage") or ""):
        return {"available": False, "reason": "아파트 아님"}
    pre, bunji = eb.building_key(cur.get("address"))
    cases = []
    if bunji:
        try:
            rr = auction_db._get("items", {
                "select": "item_key,address,area_text,building_area,appraisal_price,sale_price,sale_rate,sell_date,result,bid_count,sale_2nd_price",
                "usage_name": "ilike.*아파트*", "address": f"ilike.*{bunji}*",
                "or": "(result.like.매각*,result.like.잔금납부*,result.like.배당종결*)",
                "sale_price": "gt.0", "limit": "1000"})
            cases = rr.json() if rr.status_code in (200, 206) else []
        except Exception:
            cases = []
    est = None   # 시세(차익용) = 아파트정보 추정시세 연동(gongmae_apt_info.est.price)
    try:
        ai = gongmae_apt_info(mng, cdtn)
        if isinstance(ai, dict) and isinstance(ai.get("est"), dict) and ai["est"].get("price"):
            est = ai["est"]["price"]
    except Exception:
        pass
    r = eb.compute(cur, cases, est_price=est)
    r["v"] = _EXPBID_V
    if r.get("available"):
        try:
            auction_db.cache_save(ck, r)
        except Exception:
            pass
    return _expbid_gate(r, sid)   # cases_used는 관리자만(캐시엔 원본 저장, 반환만 게이트)


@app.get("/gongmae/apt_info")
def gongmae_apt_info(mng: str, cdtn: Optional[str] = None,
                     months: int = Query(12, le=24)) -> dict:
    """공매 아파트/오피스텔 정보·실거래 — 경매 `_apt_info_compute` 로직을 cur(합성 dict)로 복제.
    resolve_lawd·_apt_trades(시군구캐시)·match_apt(같은평형±5%)·_estimate_price·_complex_detail_for 재사용.
    item_key 없는 공매라 _brief_as_detail 폴백 대신 building.info(주소)로 단지 최소정보 보완. gm_apt: 캐시(온디맨드)."""
    if not isinstance(months, int):   # 내부 함수호출 시 Query 기본값 객체 유입 → 결과·캐시 오염 방지
        months = 12
    ck = "gm_apt:" + mng
    try:
        hit = auction_db.cache_get_many([ck]).get(ck)
        if isinstance(hit, dict) and hit.get("v", 0) >= APT_VER:
            return hit
    except Exception:
        pass
    e, cur = _gm_cur(mng, cdtn)
    if not cur:
        return {"available": False, "reason": "물건 없음", "v": APT_VER}
    usage = cur.get("usage") or ""
    if "아파트" not in usage and "오피스텔" not in usage:
        return {"available": False, "reason": "아파트/오피스텔 물건이 아님", "usage": usage}
    address = cur.get("address") or ""
    lawd = resolve_lawd(address)
    if not lawd:
        return {"available": False, "reason": "주소에서 법정동코드를 찾지 못함", "address": address}
    area = _area_num(cur.get("building_area"), cur.get("area_text"))

    def _brief_detail():
        """공매엔 item_key가 없으므로 건축물대장(building.info) 주소조회로 준공·세대·승강기 최소 단지정보 구성."""
        try:
            b = building.info(address)
        except Exception:
            b = None
        if not (isinstance(b, dict) and (b.get("build_year") or b.get("households") or b.get("elevator") is not None)):
            return None
        return {"name": _apt_name_from_addr(address) or "", "households": b.get("households"),
                "approved": (str(b.get("build_year")) if b.get("build_year") else None),
                "elevator": b.get("elevator"), "_src": "건축물대장"}

    trades = _apt_trades(lawd, months)
    if not trades:
        nm = _apt_name_from_addr(address)
        cd = None
        if nm:
            try:
                cd = kapt.complex_detail(lawd, nm)
            except Exception:
                cd = None
        cd = cd or _brief_detail()
        out = {"available": False, "reason": "해당 시군구 아파트 실거래 없음",
               "lawd_cd": lawd, "address": address, "area": area,
               "complex": nm or "", "complex_detail": cd, "v": APT_VER}
        if cd:                                   # 단지정보라도 확보되면 캐시(실거래는 다음에 재시도되도록 available=False여도 저장)
            try:
                auction_db.cache_save(ck, out)
            except Exception:
                pass
        return out
    mt = match_apt(trades, address, area=area, area_pct=0.05)   # 같은 평형 ±5%
    same = mt["same_area"] if mt["area_matched"] else []
    from datetime import timedelta
    cutoff = (date.today() - timedelta(days=183)).isoformat()
    pool6 = mt["same_area"] if mt["area_matched"] else mt["trades"]
    c6 = sum(1 for t in pool6 if t.get("deal_date", "") >= cutoff)
    demand = {"count6": c6, "per_month": round(c6 / 6, 1),
              "status": "양호" if c6 >= 3 else "검토",
              "scope": "같은평형" if mt["area_matched"] else "단지전체"}
    fm = re.search(r"(\d+)\s*층", address)
    auction_floor = int(fm.group(1)) if fm else None
    est = _estimate_price(same, auction_floor)
    amounts = [t["amount"] for t in same if t.get("amount")]
    summary = None
    if amounts:
        summary = {"count": len(same), "recent": same[0]["amount"],
                   "recent_date": same[0]["deal_date"],
                   "min": min(amounts), "max": max(amounts),
                   "avg": round(sum(amounts) / len(amounts))}
    out = {
        "available": bool(mt["trades"]),
        "lawd_cd": lawd, "address": address, "area": area,
        "complex": mt["complex"] or "", "build_year": mt["build_year"] or "",
        "summary": summary, "trades": same[:100],
        "complex_trades_total": len(mt["trades"]),
        "area_matched": mt["area_matched"], "demand": demand, "est": est,
        "months": months, "v": APT_VER,
    }
    out["complex_detail"] = _complex_detail_for(out) or _brief_detail()
    if not out.get("complex"):
        out["complex"] = (out.get("complex_detail") or {}).get("name") or _apt_name_from_addr(address) or ""
    if out.get("available"):
        try:
            auction_db.cache_save(ck, out)
        except Exception:
            pass
    return out


@app.get("/gongmae/competing_listings")
def gongmae_competing_listings(mng: str, cdtn: Optional[str] = None) -> dict:
    """공매 아파트 경쟁매물보기 — kb_crawler.match_address(주소)로 KB 단지를 찾고 kb_listing 동일평형(전용±3㎡) 매매를 조회.
    ⚠️KB 인증(KB_SITE_TOKEN) 미설정/무효 시 match_address가 예외 → {matched:False} 우아하게 반환(크래시 금지). gm_kbmatch: 캐시."""
    e, cur = _gm_cur(mng, cdtn)
    if not cur:
        return {"matched": False, "count": 0, "listings": []}
    if "아파트" not in (cur.get("usage") or "") and "오피스텔" not in (cur.get("usage") or ""):
        return {"matched": False, "count": 0, "listings": []}
    # KB 단지 매칭(지번주소 우선, 실패 시 도로명). 결과(complex_no·region_ok)만 캐시.
    ck = "gm_kbmatch:" + mng
    cno = None
    m = None
    try:
        m = auction_db.cache_get_many([ck]).get(ck)
    except Exception:
        m = None
    # 옛 버그(도로명 우선)로 저장된 미매칭 캐시({complex_no:None})는 무시하고 재매칭(근본수정 효과 반영)
    if isinstance(m, dict) and not m.get("complex_no"):
        m = None
    if not isinstance(m, dict):
        # ★근본수정: 지번주소 우선. 도로명주소는 단지명이 괄호 안(예: '평산6길 22 (평산동, 장원하이드파크)')이라
        #   kb_crawler.extract_complex_name이 괄호를 먼저 제거 → 단지명추출실패 → 미매칭. 지번주소는 단지명이 인라인.
        addr_jibun = (e.get("addr_jibun") or cur.get("address") or "")
        addr_road = (e.get("addr_road") or "")
        try:
            import kb_crawler as _kb
            m = _kb.match_address(addr_jibun) if addr_jibun else None
            # 지번으로 미매칭이면 도로명으로 재시도(지번주소 자체가 없거나 단지명 인라인 실패한 경우 대비)
            if (not isinstance(m, dict) or not m.get("complex_no")) and addr_road:
                m2 = _kb.match_address(addr_road)
                if isinstance(m2, dict) and m2.get("complex_no"):
                    m = m2
        except Exception as ex:
            # KB 미가용(토큰 없음/무효/네트워크) — 우아하게 미매칭 반환(재시도 위해 캐시 안 함)
            return {"matched": False, "count": 0, "listings": [], "reason": f"KB 미가용({type(ex).__name__})"}
        if isinstance(m, dict):
            try:
                auction_db.cache_save(ck, {"complex_no": m.get("complex_no"),
                                           "region_ok": m.get("region_ok"),
                                           "kb_name": m.get("kb_name")})
            except Exception:
                pass
    cno = m.get("complex_no") if isinstance(m, dict) else None
    if not cno or (isinstance(m, dict) and m.get("region_ok") is False):
        return {"matched": False, "count": 0, "listings": []}
    # kb_listing 동일평형(전용±3㎡) 매매 — auction_competing_listings 쿼리 복제
    area = _area_num(cur.get("building_area"))
    params = [("select", "listing_id,area_excl,price,floor,dong,ho,unit_price,"
                         "direction,room_cnt,bath_cnt,feature,agent_name,confirm_date"),
              ("complex_no", f"eq.{cno}"), ("trade_type", "eq.매매"),
              ("order", "price.asc"), ("limit", "300")]
    if area:
        params += [("area_excl", f"gte.{round(area - 3, 2)}"),
                   ("area_excl", f"lte.{round(area + 3, 2)}")]
    try:
        lr = auction_db._get("kb_listing", params)
        listings = lr.json() if lr.status_code in (200, 206) else []
    except Exception:
        listings = []
    cname = (m.get("kb_name") if isinstance(m, dict) else None)
    if not cname:
        try:
            cr = auction_db._get("kb_complex", [("select", "name"),
                                                ("complex_no", f"eq.{cno}"), ("limit", "1")])
            cj = cr.json() if cr.status_code in (200, 206) else []
            cname = cj[0].get("name") if cj else None
        except Exception:
            pass
    photo_map: dict = {}
    ids = [x.get("listing_id") for x in listings if x.get("listing_id")]
    for i in range(0, len(ids), 100):
        chunk = ",".join(str(v) for v in ids[i:i + 100])
        try:
            pr = auction_db._get("kb_listing_photo",
                                 [("select", "listing_id,url,title"),
                                  ("listing_id", f"in.({chunk})"), ("order", "seq.asc")])
            for p in (pr.json() if pr.status_code in (200, 206) else []):
                photo_map.setdefault(p["listing_id"], []).append(
                    {"url": p.get("url"), "title": p.get("title")})
        except Exception:
            pass
    return {"matched": True, "count": len(listings), "area": area,
            "complex_no": cno, "complex_name": cname,
            "listings": [{"area_excl": x.get("area_excl"), "price": x.get("price"),
                          "floor": x.get("floor"), "dong": x.get("dong"),
                          "ho": x.get("ho"), "unit_price": x.get("unit_price"),
                          "direction": x.get("direction"), "room_cnt": x.get("room_cnt"),
                          "bath_cnt": x.get("bath_cnt"), "feature": x.get("feature"),
                          "agent_name": x.get("agent_name"), "confirm_date": x.get("confirm_date"),
                          "photos": photo_map.get(x.get("listing_id"), [])}
                         for x in listings]}


@app.get("/gongmae/bid_schedule")
def gongmae_bid_schedule(mng: str = Query(..., description="물건관리번호 cltrMngNo")) -> dict:
    """공매 입찰일정 및 장소 — 회차별(유찰 저감) 입찰기간·최저입찰가·개찰. onbid.bid_schedule 래퍼. gm_bidsch: 캐시.
    ⚠️진행중 물건은 회차가 남아 캐시하되, 회차표는 예정일정이라 사실상 불변(온디맨드·1회 온비드 호출)."""
    ck = "gm_bidsch:" + mng
    try:
        hit = auction_db.cache_get_many([ck]).get(ck)
        if isinstance(hit, dict) and hit.get("available"):
            return {**hit, "_cache": "hit"}
    except Exception:
        pass
    try:
        out = onbid.bid_schedule(mng)
    except Exception as e:
        return {"available": False, "reason": type(e).__name__, "rounds": []}
    if isinstance(out, dict) and out.get("available"):
        try:
            auction_db.cache_save(ck, out)
        except Exception:
            pass
    return out


def _gm_last_min(mng: str) -> Optional[int]:
    """입찰일정(bid_schedule) 마지막(최고 회차) 회차의 최저입찰가(원). 유찰 저감된 현재 최저가."""
    try:
        bs = gongmae_bid_schedule(mng)
    except Exception:
        return None
    rounds = (bs or {}).get("rounds") or []
    if not rounds:
        return None
    # rounds는 회차 오름차순 정렬 → 마지막 = 최고 회차(가장 저감된 최저입찰가)
    lp = rounds[-1].get("min_price")
    return _to_int(lp) if lp is not None else None


def _gm_cur_min(mng: str, cdtn: Optional[str]) -> Optional[int]:
    """이 회차(cdtn)의 최저입찰가 = 현재 회차 최저가(목록 행에 표시되는 값과 일치).
    bid_schedule에서 cdtn_no 일치 회차 → 미발견/미지정 시 첫 회차(가장 안 저감된 현재) 폴백.
    (enrich에는 최저입찰가가 없어 bid_schedule을 정본으로 사용.)"""
    try:
        bs = gongmae_bid_schedule(mng)
    except Exception:
        return None
    rounds = (bs or {}).get("rounds") or []
    if not rounds:
        return None
    if cdtn:
        for r in rounds:
            if str(r.get("cdtn_no")) == str(cdtn):
                m = r.get("min_price")
                if m is not None:
                    return _to_int(m)
    m = rounds[0].get("min_price")   # 폴백: 첫 회차(마지막 저감회차 아님 — fantasy 차익 방지)
    return _to_int(m) if m is not None else None


def _gm_stored_recent(mng: str, cdtn: Optional[str]) -> tuple:
    """저장된 최근 실거래가/체결일(bulk precompute_gm_recent 결과) 조회.
    라이브 buy_grade가 목록과 동일 값을 쓰게 해 리스트=상세 일치 보장
    (apt_info 결과캐시 gm_apt: 가 풀 예열 전 stale=available:False 로 굳어 라이브만 recent None 되는 불일치 회피)."""
    try:
        params = {"manage_no": f"eq.{mng}", "select": "recent_trade_price,recent_trade_date", "limit": "1"}
        if cdtn:
            params["data->>pbct_cdtn_no"] = f"eq.{cdtn}"
        r = auction_db._get("gongmae_items", params)
        rows = r.json() if r.status_code < 400 else []
        if rows:
            return _to_int(rows[0].get("recent_trade_price")), rows[0].get("recent_trade_date")
    except Exception:
        pass
    return None, None


@app.get("/gongmae/buy_grade")
def gongmae_buy_grade(mng: str, cdtn: Optional[str] = None) -> dict:
    """공매 매수판정(빌라·다세대·도시형생활주택 + 아파트) — 기준가(base) vs 추정시세.
    · 시세 소스: 빌라류=villa_est.price / 아파트·오피스텔=apt_info.est.price
    · 기준가 base = 예상낙찰가(villa_expected_bid/expected_bid, 있으면) / 없으면 현재 회차 최저입찰가(e.min_price)
    · 시세 산출불가 → 매수금지("수요 없음 · 매수세 없음")
    · 기준가 > 시세 → 매수금지
    · 차익(시세−기준가) ≥ 3,000만 → 매수양호 · 0~3,000만 → 매수검토
    gm_grade: 캐시(온디맨드, v3). 대상 아닌 용도(토지·상가·단독 등)는 applicable:False."""
    ck = "gm_grade:" + mng + ((":" + cdtn) if cdtn else "")
    try:
        hit = auction_db.cache_get_many([ck]).get(ck)
        # v>=5: 추정시세 없을 때 최근 실거래가 fallback 반영(주인님 2026-07-08). v<5는 재계산.
        if isinstance(hit, dict) and hit.get("v", 0) >= 6:
            return {**hit, "_cache": "hit"}
    except Exception:
        pass
    e, cur = _gm_cur(mng, cdtn)
    if not cur:
        return {"applicable": False, "reason": "물건 없음", "v": 2}
    usage = cur.get("usage") or ""
    is_villa = _is_villa_usage(usage)
    is_apt = ("아파트" in usage) or ("오피스텔" in usage)
    if not (is_villa or is_apt):
        return {"applicable": False, "reason": "매수판정 대상 아님(빌라류·아파트만)",
                "usage": usage, "v": 2}
    # ① 추정 시세(용도별 소스 분기) + 유사(주변) 실거래 건수(nb_count) + 최근 실거래가(추정시세 없을 때 fallback)
    sise = None
    nb_count = None   # 빌라=nearby_trades 건수 / 아파트=같은평형 매칭 실거래 수
    recent_price = None   # 최근 실거래가 — 추정시세 산출불가 시 기준가로 사용(아파트=같은평형 최신, 빌라=유사거래 최신)
    recent_date = None    # 그 실거래 체결일(YYYY-MM-DD) — 물건상세 표시용
    try:
        if is_villa:
            ev = gongmae_villa_est(mng, cdtn)
            if isinstance(ev, dict) and ev.get("available"):
                sise = _to_int(ev.get("price"))
            # 유사거래 건수 = nearby_trades 의 trades 수(villa_est가 이미 조회 → 캐시 히트, 추가호출 저렴)
            try:
                nb = gongmae_nearby_trades(mng, cdtn)
                if isinstance(nb, dict) and nb.get("available"):
                    _tr = nb.get("trades") or []
                    nb_count = len(_tr)
                    if _tr:   # 최근 실거래가 = 가장 최근 유사거래(nearby_trades는 deal_date desc 정렬)
                        recent_price = _to_int(_tr[0].get("amount")); recent_date = _tr[0].get("deal_date")
            except Exception:
                nb_count = None
        else:
            ai = gongmae_apt_info(mng, cdtn)
            if isinstance(ai, dict):
                if isinstance(ai.get("est"), dict) and ai["est"].get("price"):
                    sise = _to_int(ai["est"]["price"])
                # 아파트 유사거래 = 같은평형 매칭 실거래(summary.count), 없으면 trades 길이
                _sm = ai.get("summary")
                if isinstance(_sm, dict) and _sm.get("count") is not None:
                    nb_count = _to_int(_sm.get("count"))
                elif isinstance(ai.get("trades"), list):
                    nb_count = len(ai["trades"])
                # 최근 실거래가 = summary.recent(같은평형 최신 실거래 금액)+recent_date — est(3/6개월 창)가 None이어도 존재
                if isinstance(_sm, dict) and _sm.get("recent"):
                    recent_price = _to_int(_sm.get("recent")); recent_date = _sm.get("recent_date")
    except Exception:
        sise = None
    # 추정시세 없을 때 최근 실거래가는 **저장 컬럼 우선**(bulk precompute_gm_recent) — 목록과 동일 소스로
    #   리스트=상세 일치. 저장값 없으면 위에서 apt_info.summary/nearby 로 잡은 값(라이브) 폴백.
    if not sise:
        _sp, _sd = _gm_stored_recent(mng, cdtn)
        if _sp:
            recent_price, recent_date = _sp, _sd
    # ② 판정 기준가(base) = 예상낙찰가(있으면) / 없으면 현재 회차 최저입찰가.  (주인님 지정 2026-07-08)
    #    · 예상낙찰가: villa_expected_bid(빌라)/expected_bid(아파트) — 인근 경매 낙찰사례(표본상 ~20% 물건 존재).
    #    · 현재 최저입찰가: 이 회차(cdtn)의 최저입찰가(e.min_price) = 목록 행에 표시되는 값과 일치.
    #    (구 '마지막회차 최저가(_gm_last_min)' 방식은 목록에 100% 최저가를 보이며 차익은 가장 저감된
    #     마지막회차로 계산 → 최저입찰가와 차익 기준 불일치 버그. 회차 무관하게 현재 회차 기준으로 통일.)
    exp_bid = None
    try:
        xb = gongmae_villa_expected_bid(mng, cdtn) if is_villa else gongmae_expected_bid(mng, cdtn)
        if isinstance(xb, dict) and xb.get("available"):
            exp_bid = _to_int(xb.get("expected_bid"))
    except Exception:
        exp_bid = None
    cur_min = _gm_cur_min(mng, cdtn)   # 현재 회차 최저입찰가(bid_schedule cdtn 매칭) — 목록 표시값과 일치
    # base = 예상낙찰가와 현재 최저입찰가 중 큰 값. 공매는 최저입찰가 미만 낙찰 불가 →
    #   예상낙찰가(경매 백데이터)가 최저보다 낮으면 그 값은 실현 불가(허수 차익 유발) → 최저가 실질 하한.
    if exp_bid and cur_min:
        base = max(exp_bid, cur_min)
        base_src = "예상낙찰가" if exp_bid >= cur_min else "현재 최저입찰가"
    elif exp_bid:
        base, base_src = exp_bid, "예상낙찰가"
    else:
        base, base_src = cur_min, "현재 최저입찰가"
    kind = "villa" if is_villa else "apt"
    # ③ 판정(주인님 지정 규칙) — 기준가 base 사용
    if not sise:
        # 추정시세 산출불가 → 최근 실거래가 fallback(주인님 지정 2026-07-08):
        #   차익 = 최근실거래가 − base. ≥3천만 → 매수검토(추정시세 없어 '매수양호'는 안 줌).
        #   그 외/기준가 부재 → 매수금지. 최근 실거래가 자체가 없으면 매수금지.
        if recent_price and base:
            pf = recent_price - base
            if pf >= 30000000:
                grade, reason = "매수검토", "추정시세 없음 · 최근 실거래가 기준 차익 3천만원 이상"
            else:
                grade, reason = "매수금지", "추정시세 없음 · 최근 실거래가 기준 차익 3천만원 미만"
            out = {"applicable": True, "grade": grade, "kind": kind, "reason": reason,
                   "sise": None, "sise_src": "최근실거래가",
                   "recent_price": recent_price, "recent_date": recent_date,
                   "base": base, "base_src": base_src, "expected_bid": exp_bid,
                   "cur_min": cur_min, "last_min": base, "profit": pf, "nb_count": nb_count, "v": 6}
        else:
            out = {"applicable": True, "grade": "매수금지", "kind": kind,
                   "reason": "수요 없음 · 매수세 없음(실거래 없음)",
                   "sise": None, "sise_src": None, "recent_price": None, "recent_date": None,
                   "base": base, "base_src": base_src, "expected_bid": exp_bid,
                   "cur_min": cur_min, "last_min": base, "profit": None, "nb_count": nb_count, "v": 6}
    elif base is None:
        return {"applicable": False, "reason": "기준가(예상낙찰가/최저입찰가) 확인 불가",
                "sise": sise, "kind": kind, "nb_count": nb_count, "v": 6}
    elif base > sise:
        out = {"applicable": True, "grade": "매수금지", "kind": kind,
               "reason": base_src + "가 추정시세보다 높음",
               "sise": sise, "sise_src": "추정시세", "base": base, "base_src": base_src, "expected_bid": exp_bid,
               "cur_min": cur_min, "last_min": base, "profit": sise - base,
               "nb_count": nb_count, "v": 6}
    else:
        profit = sise - base
        if profit >= 30000000:
            grade, reason = "매수양호", "차익 3천만원 이상"
        else:
            grade, reason = "매수검토", "차익 3천만원 미만"
        out = {"applicable": True, "grade": grade, "kind": kind, "reason": reason,
               "sise": sise, "sise_src": "추정시세", "base": base, "base_src": base_src, "expected_bid": exp_bid,
               "cur_min": cur_min, "last_min": base, "profit": profit,
               "nb_count": nb_count, "v": 6}
    try:
        auction_db.cache_save(ck, out)
    except Exception:
        pass
    return out


# ---------- 실거래(국토부 연립다세대 매매) ----------
from auction_analysis.molit_source import MolitSource, filter_similar  # noqa: E402
molit = MolitSource()


@app.get("/realprice")
def realprice(lawd_cd: str = Query(..., description="시군구 법정동코드 5자리"),
              umd: Optional[str] = Query(None, description="법정동(동) 한정"),
              area: Optional[float] = Query(None, description="전용면적(㎡) 유사기준"),
              area_pct: float = 0.20, months: int = Query(12, le=24)) -> dict:
    """주변 연립다세대 매매 실거래(최근 N개월) + 유사조건 필터 결과."""
    res = molit.recent_trades(lawd_cd, months=months)
    if res.get("error"):
        return {"error": res["error"], "trades": [], "count": 0}
    sim = filter_similar(res["trades"], umd=umd, area=area, area_pct=area_pct)
    return {"total": res["count"], "count": len(sim), "trades": sim, "months": months}


def _to_int(s):
    try:
        return int(re.sub(r"[^0-9]", "", str(s)) or "")
    except Exception:
        return None


def _road_addr_to_pnu(item_key: str, addr: str):
    """도로명주소(지번 없음) → PNU. 주소 괄호의 법정동('(우만동)') + detail_text의 지번으로 재구성.
    (resolve_bjd는 '동 지번' 형식만 처리 → 도로명 물건의 공시가격이 안 나오던 문제 보완)"""
    m = re.search(r"\(([^),]*?[동읍면리])", addr or "")
    umd = m.group(1) if m else None
    if not umd:
        return None
    try:
        row = auction_db._get("items", {"select": "detail_text",
                                        "item_key": f"eq.{item_key}", "limit": "1"}).json()
    except Exception:
        return None
    dt = (row[0].get("detail_text") if row else "") or ""
    jm = re.search(re.escape(umd) + r"\s+(\d+(?:-\d+)?)", dt)
    if not jm:
        return None
    toks = []
    for tk in (addr or "").split():
        if re.search(r"(동|읍|면|리|로|길|가|번길)$", tk) or re.match(r"^\d", tk):
            break
        toks.append(tk)
    return _addr_to_pnu(f"{' '.join(toks)} {umd} {jm.group(1)}")


def _compute_prop_gongsi(item_key: str, addr: str, area, floor):
    """물건 자체 공시가격: 지번주소→PNU, 실패 시 도로명→detail_text 지번으로 PNU 재구성."""
    try:
        pnu = _addr_to_pnu(addr) or _road_addr_to_pnu(item_key, addr)
        return gongsi.price(pnu, area=area, floor=floor) if pnu else None
    except Exception:
        return None


_nearby_cache: dict[str, dict] = {}   # 주변 유사 실거래 결과 메모리 캐시(+DB api_cache 'nearby:')
NEARBY_RADIUS_M = 500                 # 주변 유사거래 반경(m). 캐시는 1km 슈퍼셋, 읽을 때 이 반경으로 트림(재예열 불필요).


def _dedup_trades(trades: list) -> list:
    """같은 실거래(거래일+법정동+지번+면적+층+금액+건물명 동일)가 중복 적재된 것 1건으로. 국토부 API 중복 방어."""
    seen = set()
    out = []
    for t in trades:
        key = (t.get("deal_date"), t.get("umd"), t.get("jibun"),
               t.get("area"), t.get("floor"), t.get("amount"), t.get("name"))
        if key in seen:
            continue
        seen.add(key)
        out.append(t)
    return out


def _trim_to_radius(d: dict) -> dict:
    """캐시된 1km 결과를 현재 반경(NEARBY_RADIUS_M)으로 좌표 필터 + 중복 제거. 반경 바꿔도 재예열 불필요."""
    if not isinstance(d, dict):
        return d
    pla, plo = d.get("prop_lat"), d.get("prop_lng")
    trades = _dedup_trades(d.get("trades") or [])      # 중복 거래 1건으로(목록 카운트도 이걸 씀 → 자동 일치)
    if not trades:
        return d if not (d.get("trades")) else {**d, "trades": []}
    if pla is None or plo is None:
        # 물건 좌표가 없으면 반경 검증 불가 → 1km 통째 노출 금지(500m 보장). 거래 비움.
        return {**d, "trades": []}
    # ⚠️haversine_m 시그니처는 (lng1,lat1,lng2,lat2) — 경도 먼저. 위도부터 넘기면 거리 왜곡(800m도 500m로 통과)!
    kept = [t for t in trades if t.get("lat") is not None and t.get("lng") is not None
            and haversine_m(plo, pla, t["lng"], t["lat"]) <= NEARBY_RADIUS_M]
    return {**d, "trades": kept}


def _warm_similar_from_nearby(keys: list) -> None:
    """유사거래 카운트(_similar_cache)를 'nearby:' 캐시에서 즉석 계산해 채움(지오코딩 없이 빠름).
    simsort가 빈/부분 캐시로 정렬해 고건수 물건이 누락되는 것 방지(자가치유). nearby 캐시 없는 키는 건너뜀(예열 몫)."""
    global _similar_dirty
    if not keys:
        return
    chunks = [keys[i:i + 150] for i in range(0, len(keys), 150)]

    def _one(chunk):
        try:
            got = auction_db.cache_get_many(["nearby:" + k for k in chunk]) or {}
        except Exception:
            return []
        out = []
        for k in chunk:
            nb = got.get("nearby:" + k)
            if isinstance(nb, dict) and nb.get("available") and nb.get("geo_ok"):
                c = len(_trim_to_radius(nb).get("trades") or [])   # 중복제거·500m트림된 정확한 카운트
                if c:
                    out.append((k, c))
        return out
    changed = False
    with _cf.ThreadPoolExecutor(max_workers=8) as ex:   # nearby 캐시 일괄조회 병렬(순차 ~수십초 → ~수초)
        for res in ex.map(_one, chunks):
            for k, c in res:
                _similar_cache[k] = c
                changed = True
    if changed:
        _similar_dirty = True
        try:
            _save_similar_cache()
        except Exception:
            pass


@app.get("/auction/nearby_trades")
def auction_nearby_trades(item_key: str, months: int = Query(12, le=24)) -> dict:
    """빌라/도시형: 시군구 연립다세대 실거래(도시형생활주택 포함, 12개월) 중 ①면적±10㎡ ②층±1
    ③반경 1km(서버 V-World 지오코딩) 필터. 좌표 포함 반환 → 프론트는 지오코딩 없이 즉시 표시.
    결과는 ①메모리 →②DB(api_cache 'nearby:') →③계산 후 DB저장 (무거운 지오코딩 1회만)."""
    if not isinstance(months, int):                # 예열이 직접 호출 시 Query 기본값 방어
        months = 12
    if item_key in _nearby_cache:                  # ① 메모리(1km 슈퍼셋 저장 → 현재 반경으로 트림)
        return _trim_to_radius(_nearby_cache[item_key])
    try:                                            # ② DB
        db = auction_db.cache_get_many(["nearby:" + item_key]).get("nearby:" + item_key)
    except Exception:
        db = None
    if isinstance(db, dict) and db.get("available") and db.get("v", 0) >= 2:
        # 캐시에 물건 공시가격이 비었으면(도로명/쿼터 등) 그때그때 보강 시도 → 추정시세 복구. 성공 시 캐시 갱신.
        pgg = db.get("prop_gongsi")
        if not (isinstance(pgg, dict) and pgg.get("price")):
            v = _compute_prop_gongsi(item_key, db.get("address") or "",
                                     db.get("prop_area"), db.get("prop_floor"))
            if v:
                db["prop_gongsi"] = v
                try:
                    auction_db.cache_save("nearby:" + item_key, db)
                except Exception:
                    pass
        _nearby_cache[item_key] = db               # v<2(구버전: 도로명주소 좌표버그)는 무시하고 재계산
        return _trim_to_radius(db)                 # 캐시(1km) → 현재 반경으로 트림
    d = auction_db.get_auction(item_key)            # ③ 계산
    if not d:
        return {"available": False}
    addr = d.get("address") or ""
    if not re.search(r"다세대|연립|빌라|도시형", d.get("usage") or ""):
        return {"available": False, "reason": "연립다세대/도시형 대상 아님"}
    # 무거운 ①반경 지오코딩 ②준공년도(건축물대장 ~4s) ③물건 공시가격(V-World)은 서로 독립 →
    # 순차로 합산되던 것을 병렬 실행(준공·공시가격이 지오코딩 뒤에서 더해지던 ~5초 제거).
    prop_area = _area_num(d.get("building_area"), d.get("area_text"))
    _fm = re.search(r"(\d+)\s*층", addr)
    prop_floor0 = int(_fm.group(1)) if _fm else None

    def _bi():
        try:
            b = building.info(addr)
            return _to_int(b.get("build_year")) if b else None
        except Exception:
            return None

    def _pg():
        return _compute_prop_gongsi(item_key, addr, prop_area, prop_floor0)
    with _cf.ThreadPoolExecutor(max_workers=3) as ex:
        f_r = ex.submit(_nearby_filtered, d)
        f_by = ex.submit(_bi)
        f_pg = ex.submit(_pg)
        r = f_r.result()
        prop_build_year = f_by.result()
        prop_gongsi = f_pg.result()
    if r is None:
        return {"available": False, "reason": "법정동코드 변환 실패", "address": addr}
    prop_area, prop_floor = r["prop_area"], r["prop_floor"]
    picked = sorted(r["trades"], key=lambda t: t.get("deal_date", ""), reverse=True)[:500]
    result = {
        "available": True, "v": 2, "address": addr, "addr_prefix": r["addr_prefix"],
        "sigungu_prefix": r["sigungu_prefix"], "addr_jibun": r["addr_jibun"],
        "prop_area": prop_area, "prop_floor": prop_floor, "prop_build_year": prop_build_year,
        "prop_gongsi": prop_gongsi, "months": months, "geo_ok": r.get("geo_ok", False),
        "prop_lng": r.get("prop_lng"), "prop_lat": r.get("prop_lat"),
        "trades": [{"amount": t["amount"], "area": t["area"], "build_year": t.get("build_year"),
                    "floor": t.get("floor"), "jibun": t.get("jibun"), "name": t.get("name"),
                    "umd": t.get("umd"), "deal_date": t.get("deal_date"),
                    "lng": t.get("lng"), "lat": t.get("lat")} for t in picked],
    }
    # 지오코딩 성공(1km 확정) 결과만 캐시(실패는 재시도). 메모리 + DB(api_cache).
    if result.get("geo_ok"):
        _nearby_cache[item_key] = result
        try:
            auction_db.cache_save("nearby:" + item_key, result)   # 1km 슈퍼셋 저장
        except Exception:
            pass
    return _trim_to_radius(result)                 # 현재 반경으로 트림 반환


@app.get("/auction/usage_zones")
def auction_usage_zones(item_key: str) -> dict:
    """상가: 경매물건 중심 반경 1.5km 상업 필지(빨강)·1km 주택 필지(파랑) 병합 폴리곤(GeoJSON).
    건축물대장 주용도 기준. DB캐시(usagezone:) 우선 → 없으면 계산 후 저장(사전계산 대상)."""
    try:
        c = auction_db.cache_get_many(["usagezone:" + item_key]).get("usagezone:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == 11:
        return c
    # 캐시미스: 쿼드트리 계산이 ~30초라 동기 응답 불가 → 백그라운드 1회 계산 + 'pending' 반환(프런트 자동 재시도)
    with _uz_lock:
        already = item_key in _uz_computing
        if not already:
            _uz_computing.add(item_key)
    if not already:
        threading.Thread(target=_uz_compute_bg, args=(item_key,), daemon=True).start()
    return {"available": False, "pending": True, "reason": "용도 분포 계산 중"}


_uz_computing: set = set()
_uz_lock = threading.Lock()


def _uz_geocode(addr: str):
    """상세 지도용 좌표 — 정밀 실패 시 시도 보강 + 동·단지명 단계축소 폴백."""
    ll = _geocode(_clean_addr_for_geo(addr))
    if ll:
        return ll
    _a = ("경" + addr) if addr.startswith("기도 ") else addr
    _cands = []
    _m = re.search(r"^(.*?[동읍면리])\s+([가-힣A-Za-z]{2,})", _a)
    if _m:
        _cands.append(_m.group(1) + " " + _m.group(2))   # ...동 + 단지명(숫자 전까지)
    _m2 = re.search(r"^(.*?[동읍면리])", _a)
    if _m2:
        _cands.append(_m2.group(1))                       # ...동
    for _c in dict.fromkeys(_cands):
        ll = _geocode(_c)
        if ll:
            return ll
    return None


def _uz_compute_bg(item_key: str):
    """usage_zones 백그라운드 계산 → usagezone: 캐시 저장(v=2). 라이브 캐시미스·사전계산 공용."""
    from auction_analysis.usage_zones import compute_zones
    try:
        d = auction_db.get_auction(item_key)
        if not d:
            return
        ll = _uz_geocode(d.get("address") or "")
        if not ll:
            res = {"available": False, "reason": "좌표 변환 실패", "v": 11}
        else:
            res = compute_zones(ll[1], ll[0], db=auction_db)   # (_geocode=(lng,lat))
            res["v"] = 11
        try:
            auction_db.cache_save("usagezone:" + item_key, res)
        except Exception:
            pass
    except Exception:
        pass
    finally:
        with _uz_lock:
            _uz_computing.discard(item_key)


def _sgg_from_coords(lng: float, lat: float) -> str:
    """좌표 → 시군구코드 5자리(Kakao 좌표→행정코드, Supabase 무관). 실패 시 ''."""
    if not (lng and lat and KAKAO_REST_KEY):
        return ""
    try:
        d = httpx.get("https://dapi.kakao.com/v2/local/geo/coord2regioncode.json",
                      params={"x": str(lng), "y": str(lat)},
                      headers={"Authorization": "KakaoAK " + KAKAO_REST_KEY}, timeout=10).json()
        for doc in d.get("documents", []) or []:
            if doc.get("region_type") == "B":        # 법정동(B) 10자리 → 앞 5자리=시군구
                return (doc.get("code") or "")[:5]
    except Exception:
        return ""
    return ""


_apthh_mem: dict = {}                                # 프로세스 메모리 캐시(Supabase 부하 무관 즉시 재사용)


@app.get("/auction/apt_households")
def auction_apt_households(item_key: str, lat: float = 0.0, lng: float = 0.0) -> dict:
    """상가 용도지도용: 주변 1.5km 아파트 단지(Kakao POI) + 실제 세대수(kapt). usage_zones와 별도(가벼움).
    프런트가 좌표(lat,lng)를 주면 get_auction/지오코딩 생략(예열 부하 내성). 시군구는 좌표→행정코드로 구함."""
    mem = _apthh_mem.get(item_key)
    if isinstance(mem, dict):
        return mem
    try:
        c = auction_db.cache_get_many(["apthh:" + item_key]).get("apthh:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == 12:
        _apthh_mem[item_key] = c
        return c
    addr = ""
    if not (lat and lng):                            # 좌표 미전달 시에만 DB/지오코딩(폴백)
        d = None
        try:
            d = auction_db.get_auction(item_key)
        except Exception:
            d = None
        if not d:
            return {"available": False, "reason": "좌표 없음"}
        addr = d.get("address") or ""
        try:
            ll = _uz_geocode(addr)
        except Exception:
            ll = None
        if not ll:
            return {"available": False, "reason": "좌표 변환 실패"}
        lng, lat = ll[0], ll[1]
    lawd = _sgg_from_coords(lng, lat) or (resolve_lawd(addr) if addr else "")
    from auction_analysis.usage_zones import _apartment_complexes
    try:
        comps, entrances = _apartment_complexes(lat, lng)   # (lat,lng) → 단지 + 입출구
    except Exception:
        comps, entrances = [], []
    seen, uniq = set(), []
    for cm in comps:
        nm = (cm.get("name") or "").strip()
        if re.search(r"예정|입주\s*20\d\d|준공\s*20\d\d", nm):   # 준공예정(미준공) 아파트 = 배후세대 아님 → 제외
            continue
        if nm and nm not in seen:
            seen.add(nm)
            uniq.append(cm)

    def _one(cm):
        nm = cm["name"].strip()
        hh = None
        if lawd:
            try:
                b = kapt.brief(lawd, nm)
                hv = (b or {}).get("households")
                hh = int(hv) if hv and str(hv).isdigit() else None
            except Exception:
                hh = None
        return {"name": nm, "households": hh, "center": [cm["lng"], cm["lat"]]}
    try:
        with _cf.ThreadPoolExecutor(max_workers=8) as ex:   # kapt 콜드콜 병렬(66초→~12초)
            out = list(ex.map(_one, uniq))
    except Exception:
        out = [_one(cm) for cm in uniq]
    # 누적 병합: 이전 캐시 단지와 union(이름) — Kakao 격자검색이 예열부하·레이트리밋 때 일시적으로 단지를 덜 잡아
    # 캐시가 8개로 굳는 것 방지(단지는 사라지지 않으므로 합집합이 정답). 준공예정은 재필터.
    try:
        _pc = auction_db.cache_get_many(["apthh:" + item_key]).get("apthh:" + item_key)
        _prev = _pc.get("complexes") if isinstance(_pc, dict) else []
    except Exception:
        _prev = []
    _seen = {}
    for cm in (list(out) + list(_prev or [])):
        nm = (cm.get("name") or "").strip()
        if not nm or not cm.get("center"):
            continue
        if re.search(r"예정|입주\s*20\d\d|준공\s*20\d\d", nm):
            continue
        if nm not in _seen or (cm.get("households") and not _seen[nm].get("households")):
            _seen[nm] = cm                              # 세대수 있는 것 우선
    out = list(_seen.values())
    res = {"available": True, "complexes": out, "entrances": entrances, "v": 12}
    _apthh_mem[item_key] = res                       # 메모리 우선 저장(즉시 재사용)
    try:
        auction_db.cache_save("apthh:" + item_key, res)
    except Exception:
        pass
    return res


_catch_mem: dict = {}


@app.get("/auction/catchment")
def auction_catchment(item_key: str) -> dict:
    """배후세대 동선: 주변 아파트 입구→경매 상가 도로 최단경로(다른 상업서 절단). usagezone(apt/red) 재사용.
    무거우니 메모리+Supabase 캐시(catchment:). 상가 물건만."""
    mem = _catch_mem.get(item_key)
    if isinstance(mem, dict):
        return mem
    try:
        cc = auction_db.cache_get_many(["catchment:" + item_key]).get("catchment:" + item_key)
    except Exception:
        cc = None
    if isinstance(cc, dict) and cc.get("v") == 7:
        _catch_mem[item_key] = cc
        return cc
    try:
        z = auction_db.cache_get_many(["usagezone:" + item_key]).get("usagezone:" + item_key)
    except Exception:
        z = None
    if not (isinstance(z, dict) and z.get("v") == 11 and z.get("center")):
        return {"available": False, "pending": True, "reason": "용도분포 먼저 계산 필요"}
    ctr = z.get("center")                            # [lng,lat]
    from auction_analysis.catchment import compute_catchment
    try:                                             # 입구=Kakao단지 대지에 닿는 도로필지(진입로)
        r = compute_catchment(ctr[1], ctr[0], z.get("red"))
    except Exception as e:
        return {"available": False, "reason": "동선 계산 오류:" + type(e).__name__}
    r["v"] = 7
    _catch_mem[item_key] = r
    try:
        auction_db.cache_save("catchment:" + item_key, r)
    except Exception:
        pass
    return r


_flow_mem: dict = {}


@app.get("/auction/market_flow")
def auction_market_flow(item_key: str) -> dict:
    """상권분석 주동선: 배후세대(아파트 kapt + 택지 면적환산)→목적지(학교·생활상권·역) 보행경로를
    세대수 실어 누적한 것이 주동선. 경매물건은 그 위에 얹어 '걸쳤나'만 채점(물건 중심 아님).
    usagezone(center/blue_regions)+apt_households 재사용. 메모리+Supabase 캐시(mflow:)."""
    mem = _flow_mem.get(item_key)
    if isinstance(mem, dict):
        return mem
    try:
        c = auction_db.cache_get_many(["mflow:" + item_key]).get("mflow:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == 11:
        _flow_mem[item_key] = c
        return c
    try:
        z = auction_db.cache_get_many(["usagezone:" + item_key]).get("usagezone:" + item_key)
    except Exception:
        z = None
    if not (isinstance(z, dict) and z.get("v") == 11 and z.get("center")):
        try:
            _uz_compute_bg(item_key)                     # 배후세대(아파트·택지) 먼저 계산 트리거
        except Exception:
            pass
        return {"available": False, "pending": True, "reason": "배후세대 계산 중"}
    ctr = z.get("center")                                # [lng,lat]
    origins = []
    _TAKJI_F = 0.7                                        # 면적환산 보정: 도로·공터 비주거 섞임 보수보정(정밀=표제부세대 TODO)
    for reg in z.get("blue_regions", []):               # 택지(비아파트 주거) 블록 = 면적환산 세대수
        if reg.get("center") and reg.get("households"):
            hh = int(reg["households"] * _TAKJI_F)
            if hh >= 150:                                # 자투리 택지(가장자리·강건너 면적환산 노이즈) 제외
                origins.append([reg["center"][0], reg["center"][1], hh, "택지"])
    try:                                                 # 아파트 단지 = kapt 실세대수
        ah = auction_apt_households(item_key, ctr[1], ctr[0])
        for cm in (ah.get("complexes") or []):
            if cm.get("center") and cm.get("households"):
                origins.append([cm["center"][0], cm["center"][1], int(cm["households"]), "아파트"])
    except Exception:
        pass
    from auction_analysis.market_flow import collect_destinations, compute_market_flow
    try:
        dests = collect_destinations(ctr[0], ctr[1])
        r = compute_market_flow(ctr[1], ctr[0], origins, dests, z.get("red"))  # red=상권(입지등급 제한)
    except Exception as e:
        return {"available": False, "reason": "주동선 계산 오류:" + type(e).__name__}
    r["v"] = 13
    if r.get("route_count", 0) > 0:          # 라우팅 실패(0경로)면 캐시 안 함 → 자동 재시도
        _flow_mem[item_key] = r
        try:
            auction_db.cache_save("mflow:" + item_key, r)
        except Exception:
            pass
    return r


_manal_mem: dict = {}


@app.get("/auction/market_analysis")
def auction_market_analysis(item_key: str) -> dict:
    """상가 계량분석(숫자 계산기): 배후세대(질보정)·가능업종·평당임대료→월세→적정시세→감정가대비·허가권·경고."""
    mem = _manal_mem.get(item_key)
    if isinstance(mem, dict):
        return mem
    try:
        c = auction_db.cache_get_many(["manal:" + item_key]).get("manal:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == 6:
        _manal_mem[item_key] = c
        return c
    try:                                                 # Supabase 지연 시 500 대신 pending(프런트 재시도)
        item = auction_db.get_auction(item_key) or {}
    except Exception:
        return {"available": False, "pending": True, "reason": "DB 지연 — 재시도"}
    if not item:
        return {"available": False, "reason": "물건 없음"}
    try:
        z = auction_db.cache_get_many(["usagezone:" + item_key]).get("usagezone:" + item_key)
    except Exception:
        z = None
    if isinstance(z, dict) and z.get("v") == 11 and z.get("center"):
        ctr = z["center"]
        blue = z.get("blue_regions", []) or []
    else:
        try:
            _uz_compute_bg(item_key)             # 배후세대(택지) 계산 트리거
        except Exception:
            pass
        ll = None
        try:
            ll = _uz_geocode(item.get("address") or "")
        except Exception:
            ll = None
        if not ll:
            return {"available": False, "pending": True, "reason": "좌표·배후세대 계산 중"}
        ctr, blue = ll, []
    lng, lat = ctr[0], ctr[1]
    backers = []
    for reg in blue:                             # 택지(면적환산 ×0.7)
        if reg.get("center") and reg.get("households"):
            backers.append({"lng": reg["center"][0], "lat": reg["center"][1],
                            "hh": int(reg["households"] * 0.7), "name": "택지"})
    try:                                         # 아파트(kapt 실세대수, 단지명 → 임대 판별)
        ah = auction_apt_households(item_key, lat, lng)
        for cm in (ah.get("complexes") or []):
            if cm.get("center") and cm.get("households"):
                backers.append({"lng": cm["center"][0], "lat": cm["center"][1],
                                "hh": int(cm["households"]), "name": cm.get("name", "")})
    except Exception:
        pass
    from auction_analysis.market_flow import _kakao_cat
    from auction_analysis.market_analysis import _hav as _h, compute_analysis
    schools = [(x, y, n, _h(lat, lng, y, x)) for (x, y, n) in _kakao_cat("SC4", lng, lat, 1200)
               if any(t in n for t in ("초등학교", "중학교", "고등학교"))]
    cvs = [(x, y, n, _h(lat, lng, y, x)) for (x, y, n) in _kakao_cat("CS2", lng, lat, 300)]
    in_comm = None                               # 용도지역 상업 여부(숙박·유흥 신규)
    try:
        from auction_analysis.usage_zones import _commercial_zone_polys
        from shapely.geometry import shape as _shape, Point as _Point
        pt = _Point(lng, lat)
        in_comm = False
        for zg in (_commercial_zone_polys(lat, lng, 300) or []):
            try:
                g = _shape(zg) if isinstance(zg, dict) else zg
                if g.buffer(0).contains(pt):
                    in_comm = True
                    break
            except Exception:
                pass
    except Exception:
        in_comm = None
    resi_dominant = sum(b["hh"] for b in backers) > 500
    try:
        r = compute_analysis(item, lat, lng, backers, schools, cvs, in_comm, resi_dominant)
    except Exception as e:
        return {"available": False, "reason": "분석 오류:" + type(e).__name__}
    r["center"] = [lng, lat]                              # 지도용: 물건 위치
    r["schools_pts"] = [{"lng": x, "lat": y, "name": n} for (x, y, n, d) in schools]
    r["cvs_pts"] = [{"lng": x, "lat": y, "name": n} for (x, y, n, d) in cvs]
    r["v"] = 6
    _manal_mem[item_key] = r
    try:
        auction_db.cache_save("manal:" + item_key, r)
    except Exception:
        pass
    return r


_tarea_mem: dict = {}
_tarea_computing: set = set()
_tarea_lock = threading.Lock()


@app.get("/auction/walk_route")
def auction_walk_route(sx: float, sy: float, ex: float, ey: float) -> dict:
    """도보 경로(Tmap 보행) — 지도 도구 '이동경로'용. sx,sy=출발 lng,lat / ex,ey=도착 lng,lat.
    자동차 길찾기 절대 미사용(보행 전용 API). 사용자 클릭 기반이라 호출량 적음(Tmap 일일쿼터 안전)."""
    try:
        from auction_analysis.market_flow import _tmap
        path = _tmap(sx, sy, ex, ey)                          # [[lng,lat],...] or None
    except Exception:
        path = None
    if not path or len(path) < 2:
        return {"available": False}
    import math
    d = 0.0
    for i in range(1, len(path)):
        x1, y1 = path[i - 1]; x2, y2 = path[i]
        mx = math.cos((y1 + y2) / 2 * math.pi / 180) * 111320.0
        d += math.hypot((x2 - x1) * mx, (y2 - y1) * 110540.0)
    return {"available": True, "path": path, "dist_m": int(d)}


@app.get("/auction/trade_area")
def auction_trade_area(item_key: str) -> dict:
    """상권 영역(필지): 물건이 속한 연속 상가필지 클러스터 + 근린 코어(300m) GeoJSON.
    get_auction(Supabase)·필지계산이 느려 동기 응답 불가 → 백그라운드 1회 계산 + pending(프런트 재시도)."""
    mem = _tarea_mem.get(item_key)
    if isinstance(mem, dict):
        return mem
    try:
        c = auction_db.cache_get_many(["tarea:" + item_key]).get("tarea:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == 15:
        _tarea_mem[item_key] = c
        return c
    with _tarea_lock:                                    # 캐시미스 → 백그라운드(요청경로엔 Supabase 읽기 없음 = 500 방지)
        already = item_key in _tarea_computing
        if not already:
            _tarea_computing.add(item_key)
    if not already:
        threading.Thread(target=_tarea_compute_bg, args=(item_key,), daemon=True).start()
    return {"available": False, "pending": True, "reason": "상권 영역 계산 중"}


def _tarea_compute_bg(item_key: str):
    """trade_area 백그라운드 계산 → tarea: 캐시(v=2). get_auction 타임아웃도 여기서 흡수(엔드포인트 500 방지)."""
    from auction_analysis.trade_area import compute_trade_area
    try:
        item = auction_db.get_auction(item_key) or {}
        if not item:
            return
        ctr = None
        z = None
        try:
            z = auction_db.cache_get_many(["usagezone:" + item_key]).get("usagezone:" + item_key)
            if isinstance(z, dict) and z.get("center"):
                ctr = z["center"]
        except Exception:
            pass
        if not ctr:
            try:
                ctr = _uz_geocode(item.get("address") or "")
            except Exception:
                ctr = None
        if not ctr:
            return
        apt_geo = z.get("apt") if isinstance(z, dict) else None   # 근접 아파트영역 → 상권/주택단지서 제외(v10 깨끗한 상권)
        resi_meta = (z.get("blue_regions") if isinstance(z, dict) else None)  # 주거영역(블록별 geo+세대+중심)
        apt_list = []                                         # apt_households 단지(1.5km 전체) → 아파트 대지 판정용
        try:
            ah = auction_apt_households(item_key, ctr[1], ctr[0])
            for cm in (ah.get("complexes") or []):
                if cm.get("center") and cm.get("households"):
                    apt_list.append({"center": cm["center"], "name": cm.get("name"), "hh": cm["households"]})
        except Exception:
            pass
        r = compute_trade_area(ctr[1], ctr[0], radius=560, area_radius=1500,
                               apt_list=apt_list, exclude_geo=apt_geo, resi_regions_meta=resi_meta)
        r["v"] = 15
        if r.get("available"):
            # 다운그레이드 금지: 이전 캐시에 아파트 영역이 더 많으면 union(이름)으로 합침
            # (apt_households 일시 누락으로 trade_area 아파트가 8개로 굳는 것 방지)
            try:
                _pv = _tarea_mem.get(item_key) or auction_db.cache_get_many(["tarea:" + item_key]).get("tarea:" + item_key)
                if isinstance(_pv, dict) and _pv.get("apt_regions"):
                    _m = {}
                    for a in list(r.get("apt_regions") or []) + list(_pv.get("apt_regions") or []):
                        nm = a.get("name") or ""
                        if nm and (nm not in _m or (a.get("geo") and not _m[nm].get("geo"))):
                            _m[nm] = a                       # 영역(geo) 있는 것 우선
                    r["apt_regions"] = list(_m.values())
            except Exception:
                pass
            _tarea_mem[item_key] = r
            try:
                auction_db.cache_save("tarea:" + item_key, r)
            except Exception:
                pass
    except Exception:
        pass
    finally:
        with _tarea_lock:
            _tarea_computing.discard(item_key)


# ───────────── 예상낙찰가(아파트, 동일건물 과거 매각사례 평균) ─────────────
_expbid_mem: dict = {}
_expbid_computing: set = set()
_expbid_lock = threading.Lock()
_EXPBID_V = 6    # v6: cases_used에 입찰인원(bid_count)·2등입찰가(sale_2nd) 추가


def _expbid_gate(d: dict, sid) -> dict:
    """참조 백데이터(cases_used)는 관리자만. 비관리자는 제거하고 사례수만 노출."""
    if not isinstance(d, dict) or d.get("cases_used") is None:
        return d
    is_admin = False
    try:
        u = current_user(sid)
        is_admin = bool(u and u.get("role") == "admin")
    except Exception:
        is_admin = False
    if is_admin:
        return d
    return {k: v for k, v in d.items() if k != "cases_used"}


@app.get("/auction/expected_bid")
def auction_expected_bid(item_key: str, sid: Optional[str] = Cookie(None)) -> dict:
    """아파트 예상낙찰가 = 동일건물 과거 매각사례(3년·전용±6㎡·감정가±2500만·층군·낙찰가율>60%) 낙찰가 단순평균.
    + 차익(추정시세−예상낙찰가). cases_used(참조 백데이터)는 관리자만. 백그라운드 계산+캐시(expbid:), 미완료 시 pending."""
    mem = _expbid_mem.get(item_key)
    if isinstance(mem, dict):
        return _expbid_gate(mem, sid)
    try:
        c = auction_db.cache_get_many(["expbid:" + item_key]).get("expbid:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == _EXPBID_V:
        _expbid_mem[item_key] = c
        return _expbid_gate(c, sid)
    with _expbid_lock:
        already = item_key in _expbid_computing
        if not already:
            _expbid_computing.add(item_key)
    if not already:
        threading.Thread(target=_expbid_compute_bg, args=(item_key,), daemon=True).start()
    return {"available": False, "pending": True, "reason": "예상낙찰가 계산 중"}


def _expbid_compute_bg(item_key: str):
    from auction_analysis import expected_bid as eb
    try:
        cur = auction_db.get_auction(item_key) or {}
        if not cur:
            return
        r = {"available": False, "reason": "아파트 아님", "v": _EXPBID_V}
        if "아파트" in (cur.get("usage") or ""):
            pre, bunji = eb.building_key(cur.get("address"))
            cases = []
            if bunji:                                          # 동일 법정동+지번 매각사례(넓게 받아 파이썬서 건물 prefix 일치 확인)
                try:
                    rr = auction_db._get("items", {
                        "select": "item_key,address,area_text,building_area,appraisal_price,sale_price,sale_rate,sell_date,result,bid_count,sale_2nd_price",
                        "usage_name": "ilike.*아파트*", "address": f"ilike.*{bunji}*",
                        "or": "(result.like.매각*,result.like.잔금납부*,result.like.배당종결*)", "sale_price": "gt.0", "limit": "1000"})
                    cases = rr.json() if rr.status_code in (200, 206) else []
                except Exception:
                    cases = []
            est = None                                         # 추정시세(차익용) — 캐시 우선(계산 안 함)
            try:
                ev = auction_apt_ests(item_key, compute=False)   # 차익은 화면 시세로 클라가 계산 → est 강제계산 불필요(빠르게)
                v = ev.get(item_key) if isinstance(ev, dict) else None
                if isinstance(v, dict):
                    est = v.get("price")
            except Exception:
                pass
            r = eb.compute(cur, cases, est_price=est)
            r["v"] = _EXPBID_V
        _expbid_mem[item_key] = r
        try:
            auction_db.cache_save("expbid:" + item_key, r)
        except Exception:
            pass
    except Exception:
        pass
    finally:
        with _expbid_lock:
            _expbid_computing.discard(item_key)


@app.get("/auction/expbid_batch")
def auction_expbid_batch(keys: str) -> dict:
    """목록용: 여러 item_key의 예상낙찰가/차익을 캐시에서만 일괄 반환(계산 안 함). cases_used는 미포함."""
    ks = [k for k in (keys or "").split(",") if k]
    out: dict = {}
    miss = []
    for k in ks:
        m = _expbid_mem.get(k)
        if isinstance(m, dict):
            if m.get("available"):
                out[k] = {"expected_bid": m.get("expected_bid"), "profit": m.get("profit"), "count": m.get("count")}
            else:
                out[k] = {"unavailable": True}          # 계산완료·산출불가 → '예상낙찰가 산출불가' 명시
        else:
            miss.append(k)
    if miss:
        try:
            cc = auction_db.cache_get_many(["expbid:" + k for k in miss])
        except Exception:
            cc = {}
        for k in miss:
            v = cc.get("expbid:" + k)
            if isinstance(v, dict) and v.get("v") == _EXPBID_V:
                _expbid_mem[k] = v
                if v.get("available"):
                    out[k] = {"expected_bid": v.get("expected_bid"), "profit": v.get("profit"), "count": v.get("count")}
                else:
                    out[k] = {"unavailable": True}       # 계산완료·산출불가
    return out


def _prewarm_expbid(keys=None):
    """예상낙찰가 예열(배치) — 매각사례 전체 1회 로드→건물(prefix)별 그룹 → 진행중 아파트마다 메모리 계산.
    건당 get_auction/번지쿼리 없음. est(차익)는 _apt_cache 일괄(compute=False). 저장은 로컬캐시 배치(synced=0, 02시 flush가 Supabase 동기화)."""
    from auction_analysis import expected_bid as eb
    cases_by, off = {}, 0
    while True:                                          # 1) 매각완료 아파트 사례 → 건물별 그룹
        try:
            r = auction_db._get("items", {"select": "item_key,address,area_text,building_area,appraisal_price,sale_price,sale_rate,sell_date,bid_count,sale_2nd_price",
                                          "usage_name": "ilike.*아파트*", "or": "(result.like.매각*,result.like.잔금납부*,result.like.배당종결*)", "sale_price": "gt.0",
                                          "order": "item_key", "limit": "1000", "offset": str(off)})  # order 고정=offset 페이지네이션 중복/누락 방지
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        for c in rows:
            pre = eb._norm(eb.building_key(c.get("address"))[0] or "")
            if pre:
                cases_by.setdefault(pre, []).append(c)
        if len(rows) < 1000:
            break
        off += 1000
    done, off = 0, 0
    while True:                                          # 2) 진행중 아파트 → 건물 그룹으로 compute(미캐시만)
        try:
            r = auction_db._get("items", {"select": "item_key,address,area_text,building_area,appraisal_price,sell_date",
                                          "data_class": "eq.현황", "usage_name": "ilike.*아파트*",
                                          "order": "item_key", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        kk = [x["item_key"] for x in rows if x.get("item_key")]
        try:
            cc = auction_db.local.get_many(["expbid:" + k for k in kk])
        except Exception:
            cc = {}
        try:
            ests = auction_apt_ests(",".join(kk), compute=False) or {}
        except Exception:
            ests = {}
        saves = []
        for x in rows:
            k = x.get("item_key")
            if not k:
                continue
            v = cc.get("expbid:" + k)
            if isinstance(v, dict) and v.get("v") == _EXPBID_V:
                continue
            pre = eb._norm(eb.building_key(x.get("address"))[0] or "")
            ev = ests.get(k) if isinstance(ests, dict) else None
            est = ev.get("price") if isinstance(ev, dict) else None
            res = eb.compute(x, cases_by.get(pre, []), est_price=est)
            res["v"] = _EXPBID_V
            _expbid_mem[k] = res
            saves.append(("expbid:" + k, res))
            done += 1
        if saves:
            try:
                auction_db.local.put_many(saves, synced=0)   # 로컬 배치쓰기(빠름) → 02시 flush가 Supabase로
            except Exception:
                pass
        if len(rows) < 1000:
            break
        off += 1000
    return done


# ════════ 빌라/도시형생활주택 예상낙찰가(반경 1km, 동일건물 아님) ════════
_VEXPBID_V = 2    # v2: cases_used에 입찰인원·2등입찰가 추가
_vexpbid_mem: dict = {}
_vexpbid_computing: set = set()
_GCELL = 0.012                 # 격자 셀(~1.1km) — 반경 1km 후보를 3×3 셀에서 추림
_VILLA_OR = "(usage_name.ilike.*다세대*,usage_name.ilike.*연립*,usage_name.ilike.*빌라*,usage_name.ilike.*도시형*)"


def _is_villa_usage(u) -> bool:
    u = u or ""
    return any(t in u for t in ("다세대", "연립", "빌라", "도시형"))


def _villa_region_prefix(addr):
    """지역 좁히기 접두(시도+시군구). 구/군 우선, 없으면 시까지."""
    toks = (addr or "").split()
    acc = []
    for t in toks:
        acc.append(t)
        if t.endswith(("구", "군")):
            return " ".join(acc)
    acc = []
    for t in toks:
        acc.append(t)
        if t.endswith("시"):
            return " ".join(acc)
    return " ".join(toks[:2]) if toks else None


@app.get("/auction/villa_expected_bid")
def auction_villa_expected_bid(item_key: str, sid: Optional[str] = Cookie(None)) -> dict:
    """빌라/도생 예상낙찰가 = 반경1km·3년·전용±6㎡·층±1·감정가±1500만·낙찰가율<100% 낙찰가 평균.
    + 차익(추정시세−예상낙찰가). cases_used는 관리자만. 백그라운드 계산+캐시(vexpbid:), 미완료 시 pending."""
    mem = _vexpbid_mem.get(item_key)
    if isinstance(mem, dict):
        return _expbid_gate(mem, sid)
    try:
        c = auction_db.cache_get_many(["vexpbid:" + item_key]).get("vexpbid:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == _VEXPBID_V:
        _vexpbid_mem[item_key] = c
        return _expbid_gate(c, sid)
    with _expbid_lock:
        already = item_key in _vexpbid_computing
        if not already:
            _vexpbid_computing.add(item_key)
    if not already:
        threading.Thread(target=_villa_expbid_compute_bg, args=(item_key,), daemon=True).start()
    return {"available": False, "pending": True, "reason": "예상낙찰가 계산 중"}


@app.get("/auction/vexpbid_batch")
def auction_vexpbid_batch(keys: str) -> dict:
    """목록용: 빌라/도생 예상낙찰가/차익 캐시 일괄 반환(계산 안 함). cases_used 미포함."""
    ks = [k for k in (keys or "").split(",") if k]
    out: dict = {}
    miss = []
    for k in ks:
        m = _vexpbid_mem.get(k)
        if isinstance(m, dict):
            if m.get("available"):
                out[k] = {"expected_bid": m.get("expected_bid"), "profit": m.get("profit"), "count": m.get("count")}
            else:
                out[k] = {"unavailable": True}          # 계산완료·산출불가 → '예상낙찰가 산출불가' 명시
        else:
            miss.append(k)
    if miss:
        try:
            cc = auction_db.cache_get_many(["vexpbid:" + k for k in miss])
        except Exception:
            cc = {}
        for k in miss:
            v = cc.get("vexpbid:" + k)
            if isinstance(v, dict) and v.get("v") == _VEXPBID_V:
                _vexpbid_mem[k] = v
                if v.get("available"):
                    out[k] = {"expected_bid": v.get("expected_bid"), "profit": v.get("profit"), "count": v.get("count")}
                else:
                    out[k] = {"unavailable": True}       # 계산완료·산출불가
    return out


# ───────────── 홈 히어로 추천 물건(매수양호·차익 상위 12) ─────────────
_HERO_V = "v5"
_HERO_OR = ("(usage_name.ilike.*아파트*,usage_name.ilike.*다세대*,usage_name.ilike.*연립*,"
            "usage_name.ilike.*빌라*,usage_name.ilike.*도시형*)")
_hero_picks_cache: dict = {"picks": None}
_hero_building = False
_SIDO_SHORT = {"서울": "서울", "부산": "부산", "대구": "대구", "인천": "인천", "광주": "광주",
               "대전": "대전", "울산": "울산", "세종": "세종", "경기": "경기", "강원": "강원",
               "충청북": "충북", "충청남": "충남", "전라북": "전북", "전라남": "전남",
               "경상북": "경북", "경상남": "경남", "제주": "제주"}


def _region_short(addr: str) -> str:
    """주소 → '시/도(축약) 시·군·구'(예: 경기도 화성시 → 경기 화성시)."""
    a = (addr or "").split("(")[0].strip()
    parts = a.split()
    if len(parts) < 2:
        return a[:12]
    sido = parts[0]
    for k, v in _SIDO_SHORT.items():
        if parts[0].startswith(k):
            sido = v
            break
    return (sido + " " + parts[1])[:14]


def _hero_usage_label(usage: str) -> str:
    u = usage or ""
    if "아파트" in u:
        return "아파트"
    if "다세대" in u:
        return "다세대"
    if "연립" in u:
        return "연립"
    if "도시형" in u:
        return "도시형생활주택"
    if "빌라" in u:
        return "빌라"
    return (u.split()[0] if u.split() else u)


def _hero_picks_compute() -> list:
    """매수양호 ∩ 아파트/다세대/도생 ∩ 진행중(현황) ∩ 사진있음 → 차익 상위 12.
    차익=시세(추정시세 캐시)−예상낙찰가(expbid/vexpbid 캐시, 없으면 최저가). 시세 없으면 제외. 예열 캐시만 읽음(계산X)."""
    bk = _grade_buckets()
    good = bk.get("매수양호") or set()
    if not good:
        bk = _grade_buckets(force=True)    # 콜드(예열 전) → 강제 빌드
        good = bk.get("매수양호") or set()
    if not good:
        return []
    cand: dict = {}
    off = 0
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key,usage_name,address,min_price,appraisal_price,thumb_url",
                                          "data_class": "eq.현황", "or": _HERO_OR,
                                          "thumb_url": "not.is.null", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        for x in rows:
            k = x.get("item_key")
            if k and k in good and x.get("thumb_url") and _to_int(x.get("min_price")):
                cand[k] = x
        if len(rows) < 1000:
            break
        off += 1000
    if not cand:
        return []
    keys = list(cand.keys())
    apt_keys = [k for k in keys if "아파트" in (cand[k].get("usage_name") or "")]
    villa_keys = [k for k in keys if "아파트" not in (cand[k].get("usage_name") or "")]
    est: dict = {}

    def _pull_est(ks, fn, step):
        for i in range(0, len(ks), step):
            ch = ks[i:i + step]
            try:
                ev = fn(",".join(ch), compute=False) or {}
            except Exception:
                ev = {}
            for k in ch:
                v = ev.get(k)
                if isinstance(v, dict) and v.get("price"):
                    est[k] = _to_int(v["price"])
    _pull_est(apt_keys, auction_apt_ests, 120)
    _pull_est(villa_keys, auction_villa_ests, 100)
    expb: dict = {}

    def _pull_expb(ks, prefix, ver):
        for i in range(0, len(ks), 200):
            ch = ks[i:i + 200]
            try:
                cc = auction_db.cache_get_many([prefix + k for k in ch])
            except Exception:
                cc = {}
            for k in ch:
                v = cc.get(prefix + k)
                if isinstance(v, dict) and v.get("v") == ver and v.get("available") and v.get("expected_bid"):
                    expb[k] = _to_int(v["expected_bid"])
    _pull_expb(apt_keys, "expbid:", _EXPBID_V)
    _pull_expb(villa_keys, "vexpbid:", _VEXPBID_V)
    picks = []
    for k in keys:
        sise = est.get(k)
        if not sise:
            continue                       # 시세 없으면 제외
        ap = _to_int(cand[k].get("appraisal_price"))
        if ap and sise > ap * 1.5:
            continue                       # 시세>감정가×1.5 = 추정시세 오류(유사거래 오매칭) 이상치 제외
        bid = expb.get(k)                  # 예상낙찰가만 사용
        if not bid:
            continue                       # ★ 예상낙찰가 없는 물건은 노출 안 함(주인님 지정)
        profit = sise - bid                # 차익 = 시세 − 예상낙찰가
        if profit <= 0:
            continue
        row = cand[k]
        picks.append({"item_key": k, "region": _region_short(row.get("address")),
                      "usage": _hero_usage_label(row.get("usage_name")),
                      "profit": profit, "sise": sise, "bid": bid, "bid_est": True,
                      "thumb": row.get("thumb_url")})
    picks.sort(key=lambda p: p["profit"], reverse=True)
    return picks[:12]


def _hero_picks_build() -> None:
    """계산 후 메모리+디스크 캐시 갱신(예열 루프에서 호출)."""
    global _hero_building
    if _hero_building:
        return
    _hero_building = True
    try:
        picks = _hero_picks_compute()
        if picks:                          # 빈 결과는 캐시 안 함(콜드 레이스 방지)
            _hero_picks_cache["picks"] = picks
            try:
                auction_db.cache_save("hero_picks:" + _HERO_V, {"picks": picks})
            except Exception:
                pass
    finally:
        _hero_building = False


@app.get("/hero_picks")
def hero_picks() -> dict:
    """홈 히어로 캐러셀 — 매수양호+아파트/다세대/도생+진행중 차익 상위 12(예열 캐시)."""
    c = _hero_picks_cache
    if c["picks"] is not None:
        return {"items": c["picks"]}
    try:
        dc = auction_db.cache_get_many(["hero_picks:" + _HERO_V]).get("hero_picks:" + _HERO_V)
        if isinstance(dc, dict) and isinstance(dc.get("picks"), list):
            c["picks"] = dc["picks"]
            return {"items": c["picks"]}
    except Exception:
        pass
    threading.Thread(target=_hero_picks_build, daemon=True).start()
    return {"items": [], "pending": True}


@app.get("/auction/sold_cases")
def auction_sold_cases(item_key: str, mode: str = "bunji") -> dict:
    """새 창 목록용(우리 수집 백데이터=낙찰완료만, 타사 크롤링 X).
    mode=bunji: 해당번지/건물 매각사례 전체 — 아파트=단지(주소 prefix), 빌라/도생=그 건물.
    mode=nearby: 빌라/도생 전용 — 경매물건 중심 반경 1km 매각된 경매물건 전체(빌라/도생만)."""
    from auction_analysis import expected_bid as eb
    cur = auction_db.get_auction(item_key) or {}
    if not cur:
        return {"available": False, "cases": []}
    usage = cur.get("usage") or ""
    is_apt = "아파트" in usage
    is_villa = _is_villa_usage(usage)
    fields = ("item_key,court_name,case_no,obj_no,address,area_text,land_area,usage_name,appraisal_price,"
              "min_price,sale_price,sale_rate,sell_date,bid_count,result,fail_count,hit_count,tags")

    def _final_sale(c):                          # 낙찰완료(재매각·재진행=이전 무산분 제외)
        return (eb._to_int(c.get("sale_price")) or 0) > 0 and not (c.get("result") or "").startswith("재")

    cases = []
    if mode == "nearby":
        if not is_villa:
            return {"available": False, "reason": "인근매각물건은 빌라/도생만", "cases": []}
        ll = _geocode(eb.geo_addr(cur.get("address")))
        if not ll:
            return {"available": False, "reason": "좌표 없음", "cases": []}
        region = _villa_region_prefix(cur.get("address"))
        rows = []
        if region:
            try:
                r = auction_db._get("items", {"select": fields, "or": _VILLA_OR,
                      "address": f"ilike.*{region}*", "sale_price": "gt.0", "limit": "3000"})
                rows = r.json() if r.status_code in (200, 206) else []
            except Exception:
                rows = []
        clng, clat = ll[0], ll[1]
        for c in rows:
            if not _final_sale(c):
                continue
            cl = _geocode(eb.geo_addr(c.get("address")))
            if not cl:
                continue
            d = eb._haversine_m(clng, clat, cl[0], cl[1])
            if d <= 1000:
                c["dist_m"] = round(d)
                cases.append(c)
    else:  # bunji
        if not (is_apt or is_villa):
            return {"available": False, "reason": "백데이터 유형 아님", "cases": []}
        pre, bunji = eb.building_key(cur.get("address"))
        cur_pre = eb._norm(pre)
        rows = []
        if bunji:
            params = {"select": fields, "address": f"ilike.*{bunji}*", "sale_price": "gt.0", "limit": "3000"}
            if is_apt:
                params["usage_name"] = "ilike.*아파트*"
            else:
                params["or"] = _VILLA_OR
            try:
                r = auction_db._get("items", params)
                rows = r.json() if r.status_code in (200, 206) else []
            except Exception:
                rows = []
        for c in rows:
            if _final_sale(c) and eb._norm(eb.building_key(c.get("address"))[0]) == cur_pre:
                cases.append(c)
    cases.sort(key=lambda c: (c.get("sell_date") or ""), reverse=True)
    return {"available": True, "mode": mode, "count": len(cases), "cases": cases,
            "subject": {"address": cur.get("address"), "usage": usage}}


def _villa_expbid_compute_bg(item_key: str):
    from auction_analysis import expected_bid as eb
    try:
        cur = auction_db.get_auction(item_key) or {}
        if not cur:
            return
        r = {"available": False, "reason": "빌라/도생 아님", "v": _VEXPBID_V}
        if _is_villa_usage(cur.get("usage")):
            ll = _geocode(eb.geo_addr(cur.get("address")))
            if not ll:
                r = {"available": False, "reason": "좌표 없음", "v": _VEXPBID_V}
            else:
                region = _villa_region_prefix(cur.get("address"))
                cases = []
                if region:
                    try:
                        rr = auction_db._get("items", {
                            "select": "item_key,address,area_text,building_area,appraisal_price,sale_price,sale_rate,sell_date,bid_count,sale_2nd_price",
                            "or": _VILLA_OR, "address": f"ilike.*{region}*",
                            "sale_price": "gt.0", "order": "item_key", "limit": "2000"})
                        cases = rr.json() if rr.status_code in (200, 206) else []
                    except Exception:
                        cases = []
                for c in cases:
                    c["ll"] = _geocode(eb.geo_addr(c.get("address")))   # 대부분 워밍 캐시 적중
                est = None
                try:
                    ev = auction_villa_ests(item_key, compute=False)
                    vv = ev.get(item_key) if isinstance(ev, dict) else None
                    if isinstance(vv, dict):
                        est = vv.get("price")
                except Exception:
                    pass
                r = eb.compute_villa(cur, ll, cases, est_price=est)
                r["v"] = _VEXPBID_V
        _vexpbid_mem[item_key] = r
        try:
            auction_db.cache_save("vexpbid:" + item_key, r)
        except Exception:
            pass
    except Exception:
        pass
    finally:
        with _expbid_lock:
            _vexpbid_computing.discard(item_key)


def _prewarm_villa_expbid(keys=None):
    """빌라/도생 예상낙찰가 예열(배치) — 낙찰사례 로드+동시 지오코딩 → 격자버킷 → 진행중 빌라마다 반경1km compute.
    저장은 로컬캐시 배치(synced=0, 02시 flush가 Supabase 동기화)."""
    global _geo_dirty
    import concurrent.futures as cf
    from auction_analysis import expected_bid as eb

    def _page(params):
        out, off = [], 0
        while True:
            try:
                r = auction_db._get("items", {**params, "order": "item_key", "limit": "1000", "offset": str(off)})
                rows = r.json() if r.status_code in (200, 206) else []
            except Exception:
                break
            out += rows
            if len(rows) < 1000:
                break
            off += 1000
        return out

    cases = _page({"select": "item_key,address,area_text,building_area,appraisal_price,sale_price,sale_rate,sell_date,bid_count,sale_2nd_price",
                   "or": _VILLA_OR, "sale_price": "gt.0"})
    cur_rows = _page({"select": "item_key,address,area_text,building_area,appraisal_price,sell_date",
                      "data_class": "eq.현황", "or": _VILLA_OR})
    uniq = list({eb.geo_addr(x.get("address")) for x in (cases + cur_rows) if eb.geo_addr(x.get("address"))})
    try:
        _geo_preload(uniq)
    except Exception:
        pass
    todo = [a for a in uniq if a not in _geo_cache]
    print(f"[vexpbid] 사례 {len(cases):,} · 진행중 {len(cur_rows):,} · 지오코딩 {len(todo):,}/{len(uniq):,}", flush=True)

    def _gc1(a):
        try:
            return a, _geocoder.coord(a)
        except Exception:
            return a, None
    if todo:
        with cf.ThreadPoolExecutor(max_workers=12) as ex:
            n = 0
            for a, ll in ex.map(_gc1, todo):
                if ll:
                    _geo_cache[a] = list(ll)
                n += 1
                if n % 1000 == 0:
                    print(f"[vexpbid] 지오코딩 {n:,}/{len(todo):,}", flush=True)
        _geo_dirty = True            # 좌표를 파일에 저장(영구화) — 다음 워밍부터 재지오코딩 안 함
        try:
            _save_geo_cache()
        except Exception:
            pass
    grid: dict = {}
    for c in cases:
        ll = _geo_cache.get(eb.geo_addr(c.get("address")))
        if not ll:
            continue
        c["ll"] = ll
        gk = (int(ll[0] / _GCELL), int(ll[1] / _GCELL))
        grid.setdefault(gk, []).append(c)
    done = 0
    for i in range(0, len(cur_rows), 200):
        chunk = cur_rows[i:i + 200]
        kk = [x["item_key"] for x in chunk if x.get("item_key")]
        try:
            cc = auction_db.local.get_many(["vexpbid:" + k for k in kk])
        except Exception:
            cc = {}
        try:
            ests = auction_villa_ests(",".join(kk), compute=False) or {}
        except Exception:
            ests = {}
        saves = []
        for x in chunk:
            k = x.get("item_key")
            if not k:
                continue
            v = cc.get("vexpbid:" + k)
            if isinstance(v, dict) and v.get("v") == _VEXPBID_V:
                continue
            ll = _geo_cache.get(eb.geo_addr(x.get("address")))
            if not ll:
                res = {"available": False, "reason": "좌표 없음", "v": _VEXPBID_V}
            else:
                gx, gy = int(ll[0] / _GCELL), int(ll[1] / _GCELL)
                cand = []
                for dx in (-1, 0, 1):
                    for dy in (-1, 0, 1):
                        cand += grid.get((gx + dx, gy + dy), [])
                ev = ests.get(k) if isinstance(ests, dict) else None
                est = ev.get("price") if isinstance(ev, dict) else None
                res = eb.compute_villa(x, ll, cand, est_price=est)
                res["v"] = _VEXPBID_V
            _vexpbid_mem[k] = res
            saves.append(("vexpbid:" + k, res))
            done += 1
        if saves:
            try:
                auction_db.local.put_many(saves, synced=0)
            except Exception:
                pass
    return done


# ════════ 차량 예상낙찰가(백데이터 낙찰사례: 같은 제조사+모델·연식±1·주행±20%·최근3년 중앙값) ════════
_CAREXPBID_V = 1
_carexpbid_mem: dict = {}
_carexpbid_computing: set = set()
_carexpbid_lock = threading.Lock()


def _car_specs_many(keys: list) -> dict:
    """item_key 목록 → {item_key: vehicle_specs dict}. 100개씩 배치."""
    out: dict = {}
    for i in range(0, len(keys), 100):
        ch = keys[i:i + 100]
        q = "(" + ",".join('"' + k + '"' for k in ch) + ")"
        try:
            r = auction_db._get("vehicle_specs",
                                 {"select": "item_key,manufacturer,model,model_year,mileage_km", "item_key": f"in.{q}"})
            for row in (r.json() if r.status_code in (200, 206) else []):
                out[row["item_key"]] = row
        except Exception:
            pass
    return out


def _car_expbid_compute_bg(item_key: str):
    from auction_analysis import car_expected_bid as ceb
    import datetime as _dt
    try:
        r = {"available": False, "reason": "차량 스펙 없음", "v": _CAREXPBID_V}
        sp = _car_specs_many([item_key]).get(item_key)
        if sp and sp.get("manufacturer") and sp.get("model_year") and sp.get("mileage_km"):
            mf = sp["manufacturer"]
            try:                                             # 같은 제조사 백데이터 후보(스펙)
                rc = auction_db._get("vehicle_specs",
                                     {"select": "item_key,manufacturer,model,model_year,mileage_km",
                                      "manufacturer": f"eq.{mf}", "limit": "5000"})
                spmap = {x["item_key"]: x for x in (rc.json() if rc.status_code in (200, 206) else [])}
            except Exception:
                spmap = {}
            cases, kk = [], list(spmap.keys())
            for i in range(0, len(kk), 100):                 # 후보 중 백데이터 낙찰건(items) 병합
                ch = kk[i:i + 100]
                q = "(" + ",".join('"' + k + '"' for k in ch) + ")"
                try:
                    ri = auction_db._get("items",
                                         {"select": "item_key,case_no,sale_price,sell_date,bid_count,sale_2nd_price",
                                          "item_key": f"in.{q}", "search_group": "eq.차량외",
                                          "data_class": "eq.백데이터", "sale_price": "gt.0"})
                    for it in (ri.json() if ri.status_code in (200, 206) else []):
                        cases.append({**it, **spmap[it["item_key"]]})
                except Exception:
                    pass
            r = ceb.compute(sp, cases, _dt.date.today().isoformat())
            r["v"] = _CAREXPBID_V
        _carexpbid_mem[item_key] = r
        try:
            auction_db.cache_save("carexpbid:" + item_key, r)
        except Exception:
            pass
    except Exception:
        pass
    finally:
        with _carexpbid_lock:
            _carexpbid_computing.discard(item_key)


@app.get("/auction/car_expected_bid")
def auction_car_expected_bid(item_key: str, sid: Optional[str] = Cookie(None)) -> dict:
    """차량 예상낙찰가 = 백데이터 낙찰사례(같은 제조사+모델·연식±1·주행±20%·최근3년) 낙찰가 중앙값.
    cases_used(참조 백데이터)는 관리자만. 백그라운드 계산+캐시(carexpbid:), 미완료 시 pending."""
    mem = _carexpbid_mem.get(item_key)
    if isinstance(mem, dict):
        return _expbid_gate(mem, sid)
    try:
        c = auction_db.cache_get_many(["carexpbid:" + item_key]).get("carexpbid:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == _CAREXPBID_V:
        _carexpbid_mem[item_key] = c
        return _expbid_gate(c, sid)
    with _carexpbid_lock:
        already = item_key in _carexpbid_computing
        if not already:
            _carexpbid_computing.add(item_key)
    if not already:
        threading.Thread(target=_car_expbid_compute_bg, args=(item_key,), daemon=True).start()
    return {"available": False, "pending": True, "reason": "예상낙찰가 계산 중"}


@app.get("/auction/car_expbid_batch")
def auction_car_expbid_batch(keys: str) -> dict:
    """목록용: 여러 차량 item_key의 예상낙찰가를 캐시에서만 일괄 반환(계산 안 함). cases_used 미포함.
    산출불가(계산완료·available=false)는 {unavailable:True}로 표기 → 목록에 '예상낙찰가 산출불가' 노출.
    미계산(캐시 없음)은 out에서 제외 → 표시 안 함(예열되면 채워짐)."""
    ks = [k for k in (keys or "").split(",") if k]
    out: dict = {}
    miss = []
    for k in ks:
        m = _carexpbid_mem.get(k)
        if isinstance(m, dict):
            if m.get("available"):
                out[k] = {"expected_bid": m.get("expected_bid"), "count": m.get("count")}
            else:
                out[k] = {"unavailable": True}          # 계산완료·산출불가 → '산출불가' 명시
        else:
            miss.append(k)
    if miss:
        try:
            cc = auction_db.cache_get_many(["carexpbid:" + k for k in miss])
        except Exception:
            cc = {}
        for k in miss:
            v = cc.get("carexpbid:" + k)
            if isinstance(v, dict) and v.get("v") == _CAREXPBID_V:
                _carexpbid_mem[k] = v
                if v.get("available"):
                    out[k] = {"expected_bid": v.get("expected_bid"), "count": v.get("count")}
                else:
                    out[k] = {"unavailable": True}       # 계산완료·산출불가
    return out


def _prewarm_car_expbid(keys=None):
    """차량 예상낙찰가 예열 — 백데이터 낙찰사례 전체 1회 로드→(제조사,모델)그룹 → 진행중 차량마다 중앙값 계산."""
    from auction_analysis import car_expected_bid as ceb
    import datetime as _dt
    items, off = [], 0                                       # 1) 백데이터 낙찰차 items 전체
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key,case_no,sale_price,sell_date,bid_count,sale_2nd_price",
                                          "search_group": "eq.차량외", "data_class": "eq.백데이터", "sale_price": "gt.0",
                                          "order": "item_key", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        items += rows
        if len(rows) < 1000:
            break
        off += 1000
    specs = _car_specs_many([x["item_key"] for x in items])
    groups: dict = {}                                        # (manufacturer, model_norm) → cases
    for it in items:
        sp = specs.get(it["item_key"])
        if not sp:
            continue
        mn = ceb.norm_model(sp.get("model"), sp.get("manufacturer"))
        if sp.get("manufacturer") and mn:
            groups.setdefault((sp["manufacturer"], mn), []).append({**it, **sp})
    today = _dt.date.today().isoformat()
    done, off = 0, 0                                         # 2) 진행중 차량 → 그룹으로 compute(미캐시만)
    while True:
        try:
            r = auction_db._get("items", {"select": "item_key", "search_group": "eq.차량외", "data_class": "eq.현황",
                                          "order": "item_key", "limit": "1000", "offset": str(off)})
            rows = r.json() if r.status_code in (200, 206) else []
        except Exception:
            break
        kk = [x["item_key"] for x in rows if x.get("item_key")]
        tspecs = _car_specs_many(kk)
        try:
            cc = auction_db.local.get_many(["carexpbid:" + k for k in kk])
        except Exception:
            cc = {}
        saves = []
        for k in kk:
            v = cc.get("carexpbid:" + k)
            if isinstance(v, dict) and v.get("v") == _CAREXPBID_V:
                continue
            sp = tspecs.get(k)
            if not sp:
                res = {"available": False, "reason": "차량 스펙 없음", "v": _CAREXPBID_V}
            else:
                mn = ceb.norm_model(sp.get("model"), sp.get("manufacturer"))
                res = ceb.compute(sp, groups.get((sp.get("manufacturer"), mn), []), today)
                res["v"] = _CAREXPBID_V
            _carexpbid_mem[k] = res
            saves.append(("carexpbid:" + k, res))
            done += 1
        if saves:
            try:
                auction_db.local.put_many(saves, synced=0)
            except Exception:
                pass
        if len(rows) < 1000:
            break
        off += 1000
    return done


@app.get("/auction/offi_info")
def auction_offi_info(item_key: str, months: int = 12) -> dict:
    """오피스텔: 같은 단지 매매·전월세 실거래(국토부 오피스텔 API) + 추정시세. DB캐시(offi:) 우선."""
    try:
        c = auction_db.cache_get_many(["offi:" + item_key]).get("offi:" + item_key)
    except Exception:
        c = None
    if isinstance(c, dict) and c.get("v") == 2:
        return c
    d = auction_db.get_auction(item_key)
    if not d:
        return {"available": False, "reason": "물건 없음"}
    if "오피스텔" not in (d.get("usage") or ""):
        return {"available": False, "reason": "오피스텔 물건 아님"}
    addr = d.get("address") or ""
    lawd = resolve_lawd(addr)
    if not lawd:
        return {"available": False, "reason": "법정동코드 변환 실패", "address": addr}
    nm = _apt_name_from_addr(addr)
    if not nm:                                   # 폴백: '번지 단지명 N층M호'(동 없는 오피스텔) → 층 앞을 단지명
        _mn = re.search(r"\d+(?:-\d+)?\s+(\S.+?)\s*제?\d+층", addr)
        if _mn:
            nm = _mn.group(1).strip()
    # 법정동+지번 추출(단지명 표기 불일치 대비 1순위 매칭. 예: '에스엠벨리체'↔국토부 'SM벨리체')
    _mj = re.search(r"([가-힣]+(?:[0-9]+가)?[동읍면리])\s+(\d+(?:-\d+)?)", addr)
    umd = _mj.group(1) if _mj else ""
    jibun = _mj.group(2) if _mj else ""
    area = _area_num(d.get("building_area"), d.get("area_text"))
    from auction_analysis.offi_source import offi_deals
    trades = offi_deals(lawd, months, rent=False)
    rents = offi_deals(lawd, months, rent=True)

    def _norm(s):
        return re.sub(r"[\s()]", "", s or "")

    def _match(lst):
        cand = []
        if umd and jibun:                        # 1순위: 같은 법정동+지번(표기 무관, 가장 정확)
            cand = [t for t in lst if (t.get("umd") or "") == umd and (t.get("jibun") or "") == jibun]
        if not cand:                             # 2순위: 단지명
            nn = _norm(nm)
            cand = [t for t in lst if nn and _norm(t.get("name")) and (nn in _norm(t["name"]) or _norm(t["name"]) in nn)]
        if area:
            ca = [t for t in cand if t.get("area") and abs(t["area"] - area) <= area * 0.1]
            cand = ca or cand
        return sorted(cand, key=lambda t: t.get("deal_date", ""), reverse=True)
    mt, mr = _match(trades), _match(rents)
    amts = [t["amount"] for t in mt if t.get("amount")]
    res = {"available": True, "complex": nm, "area": area, "lawd_cd": lawd,
           "est": (round(sum(amts) / len(amts)) if amts else None),
           "trade_count": len(mt), "rent_count": len(mr),
           "trades": mt[:40], "rents": mr[:40],
           "sgg_trade_total": len(trades), "v": 2}
    try:
        auction_db.cache_save("offi:" + item_key, res)
    except Exception:
        pass
    return res


_elev_cache: dict[str, object] = {}


@app.get("/auction/elev")
def auction_elev(addrs: str) -> dict:
    """지번주소 목록(| 구분) → 각 건물 승강기 유무(건축물대장). True/False/None. 동시처리+캐시."""
    lst = [a.strip() for a in addrs.split("|") if a.strip()][:200]
    need = [a for a in lst if a not in _elev_cache]          # ① 메모리에 없는 건 DB(elev:)에서 일괄
    if need:
        try:
            rows = auction_db.cache_get_many(["elev:" + a for a in need])
            for a in need:
                dd = rows.get("elev:" + a)
                if isinstance(dd, dict) and "ev" in dd:
                    _elev_cache[a] = dd["ev"]
        except Exception:
            pass
    todo = [a for a in lst if a not in _elev_cache]          # ② DB에도 없는 것만 건축물대장 조회

    def one(a):
        try:
            bi = building.info(a)
            return a, (int(bi.get("elevator") or 0) > 0 if bi else None)
        except Exception:
            return a, None
    if todo:
        with _cf.ThreadPoolExecutor(max_workers=8) as ex:
            for a, ev in ex.map(one, todo):
                _elev_cache[a] = ev
                if ev is not None:                          # 확정값(True/False)만 DB 저장(불명은 재시도)
                    try:
                        auction_db.cache_save("elev:" + a, {"ev": ev})
                    except Exception:
                        pass
    return {a: _elev_cache.get(a) for a in lst}


from auction_analysis.molit_source import match_apt           # noqa: E402
from auction_analysis.lawd_codes import resolve_lawd          # noqa: E402
from auction_analysis.kapt_source import KaptSource           # noqa: E402
from auction_analysis.building_source import BuildingSource   # noqa: E402
from auction_analysis import dagagu_analysis   # noqa: E402  다가구 분석(3요건·우량·수요·위반)
from auction_analysis.bjd_codes import resolve_bjd            # noqa: E402
from auction_analysis.gongsi_source import GongsiPrice, addr_to_pnu  # noqa: E402
kapt = KaptSource()
building = BuildingSource()
gongsi = GongsiPrice()
_gongsi_cache: dict[str, object] = {}


_pnu_geo_cache: dict = {}   # addr -> PNU(19) : V-World 지오코더 폴백(bjd_codes.tsv 미수록 주소 = 2022+ 신설 법정동 등)


def _addr_to_pnu(addr: str):
    r = resolve_bjd(addr)
    if r:
        sgg, bjd, bun, ji = r
        return addr_to_pnu(sgg, bjd, bun, ji, mountain=bool(re.search(r"\s산\s*\d", addr)))
    # 폴백: TSV가 못 잡는 주소(신설 법정동 등) → V-World 지오코더 level4LC(19자리 PNU) 직접 사용.
    #       성공분만 메모리+DB(pnu_geo:) 영구 캐시. 경매·공매 공용(둘 다 공시가격 보강). 실패=None(오염 방지).
    if not addr:
        return None
    key = addr.strip()
    if key in _pnu_geo_cache:
        return _pnu_geo_cache[key]
    ck = "pnu_geo:" + key
    try:
        hit = auction_db.cache_get_many([ck]).get(ck)
        if isinstance(hit, dict) and hit.get("pnu"):
            _pnu_geo_cache[key] = hit["pnu"]
            return hit["pnu"]
    except Exception:
        pass
    pnu = None
    try:
        vk = os.environ.get("VWORLD_KEY", "")
        if vk:
            rr = httpx.get("https://api.vworld.kr/req/address", params={
                "service": "address", "request": "getcoord", "version": "2.0", "crs": "epsg:4326",
                "address": key, "type": "parcel", "format": "json", "key": vk,
                "domain": os.environ.get("VWORLD_DOMAIN", "http://localhost:4011"), "refine": "true"}, timeout=8)
            js = (rr.json() or {}).get("response", {})
            if js.get("status") == "OK":
                lc = ((js.get("refined", {}) or {}).get("structure", {}) or {}).get("level4LC") or ""
                if len(lc) == 19 and lc.isdigit():
                    pnu = lc
    except Exception:
        pnu = None
    _pnu_geo_cache[key] = pnu   # None도 캐시(같은 주소 반복 지오코딩 방지) — 단 DB엔 성공분만 영구
    if pnu:
        try:
            auction_db.cache_save(ck, {"pnu": pnu})
        except Exception:
            pass
    return pnu


@app.get("/auction/building_vat")
def auction_building_vat(addr: str, area: float = 0, land_area: float = 0, build_year: int = 0) -> dict:
    """[새] 85㎡ 초과 주택 건물분 부가세 자동계산용([9강]/홈택스 방식, 국세청 건물기준시가 산식 추정).
    건물기준시가(추정)=신축가격기준액(82만/㎡)×경과연수잔가율(준공·철콘 내용연수 50년 근사)×연면적(전용×1.3).
    토지기준시가=개별공시지가(V-World)×토지지분 → 매도가를 건물:토지 비율로 안분 후 건물분×10%."""
    out = {"land_price": None, "bldg_std": None, "build_year": None, "year": None}
    try:
        pnu = _addr_to_pnu(addr)
        if not pnu:
            return out
        lp = gongsi.indvd_land_price(pnu)
        if lp and lp.get("price"):
            out["land_price"] = lp["price"]
            out["year"] = lp.get("year")
        by = build_year
        if not by:                       # 준공년도 미전달 → 건축물대장(캐시) 조회
            bi = building.info(addr)
            if bi and bi.get("build_year"):
                try:
                    by = int(bi["build_year"])
                except Exception:
                    by = 0
        out["build_year"] = by or None
        if area and by:
            import datetime as _dtm
            age = max(0, _dtm.datetime.now().year - by)
            resid = max(0.1, 1 - 0.9 * min(age / 50.0, 1.0))   # 경과연수별잔가율(근사)
            out["bldg_std"] = int(820000 * resid * area * 1.3)  # 신축가격기준액×잔가율×연면적(전용×1.3)
    except Exception:
        pass
    return out


@app.get("/auction/gongsi")
def auction_gongsi(items: str) -> dict:
    """'주소|전용면적|층' 묶음(; 구분) → 각 호 공시가격(V-World, 면적·층 매칭). 동시처리+캐시.
    반환 {원본키: {price, year, name}} (실패 시 키 생략)."""
    parts = [p for p in items.split(";") if p.strip()][:120]
    out: dict[str, dict] = {}
    # ① 메모리에 없는 건 Supabase api_cache(gongsi:)에서 일괄 로딩 (한 번 수집한 건 영구 보관)
    need = [p for p in parts if p not in _gongsi_cache]
    if need:
        try:
            rows = auction_db.cache_get_many(["gongsi:" + p for p in need])
            for p in need:
                d = rows.get("gongsi:" + p)
                if isinstance(d, dict) and d.get("price"):
                    _gongsi_cache[p] = d
        except Exception:
            pass
    todo = [p for p in parts if p not in _gongsi_cache]   # ② DB에도 없는 것만 V-World 계산

    def one(p):
        try:
            seg = p.split("|")
            addr = seg[0].strip()
            area = float(seg[1]) if len(seg) > 1 and seg[1] else None
            floor = int(seg[2]) if len(seg) > 2 and seg[2] else None
            pnu = _addr_to_pnu(addr)
            if not pnu:
                return p, None
            return p, gongsi.price(pnu, area=area, floor=floor)
        except Exception:
            return p, None
    if todo:
        with _cf.ThreadPoolExecutor(max_workers=8) as ex:
            for p, v in ex.map(one, todo):
                if v:                       # 실패(V-World 할당량 등)는 캐시 안 함 → 다음에 재시도(오염 방지)
                    _gongsi_cache[p] = v
                    try:
                        auction_db.cache_save("gongsi:" + p, v)   # DB에도 영구 저장
                    except Exception:
                        pass
    for p in parts:
        v = _gongsi_cache.get(p)
        if v:
            out[p] = v
    return out
from auction_analysis.disk_cache import DiskDict  # noqa: E402
_apt_cache: DiskDict = DiskDict(os.path.join(_ROOT, "cache_apt_info.json"))


def _area_num(*vals) -> Optional[float]:
    for v in vals:
        if v is None:
            continue
        s = str(v)
        # ⚠️"대지권 X㎡ / 전용 Y㎡" 형태 → '전용' 면적 우선. 첫 숫자(대지권)를 잡으면
        #   실거래 평형매칭 실패(area_matched=false) → 아파트 시세 없음의 주범.
        m = re.search(r"전용\s*(\d+(?:\.\d+)?)", s)
        if m:
            return float(m.group(1))
        m = re.search(r"(\d+(?:\.\d+)?)", s)
        if m:
            return float(m.group(1))
    return None


@app.get("/auction/apt_info")
def auction_apt_info(item_key: str, months: int = Query(12, le=24)) -> dict:
    """아파트 경매물건: 같은 단지 실거래 + 단지 요약(단지명·건축년도·시세대). 온디맨드+캐시."""
    _mem = _apt_cache.get(item_key)
    if isinstance(_mem, dict) and _mem.get("v", 0) >= APT_VER:
        out = _mem
    else:
        db = None                                  # ① Supabase api_cache(apt:) 우선
        try:
            db = auction_db.cache_get_many(["apt:" + item_key]).get("apt:" + item_key)
        except Exception:
            db = None
        if isinstance(db, dict) and db.get("available") and db.get("v", 0) >= APT_VER:
            out = db
            _apt_cache[item_key] = out
        else:                                       # ② 없음/옛버전 → 계산 → 메모리/디스크 + DB 저장
            out = _apt_info_compute(item_key, months)
            if out.get("available"):            # 실패(molit 할당량 등) 결과는 캐시 안 함 → 다음에 재시도(캐시 오염 방지)
                _apt_cache.remember(item_key, out)
                try:
                    auction_db.cache_save("apt:" + item_key, out)
                except Exception:
                    pass
    # 상세 단지정보(K-apt)가 아직 없으면 → 응답을 막지 말고 '백그라운드'로 채워 다음 방문부터 즉시.
    #  (예전엔 매 요청마다 K-apt를 동기 호출해 ~3초 지연됨)
    if out.get("available") and not out.get("complex_detail"):
        import threading as _th

        def _fill_detail():
            det = _complex_detail_for(out)
            if det:
                out["complex_detail"] = det
                _apt_cache.remember(item_key, out)
                try:
                    auction_db.cache_save("apt:" + item_key, out)
                except Exception:
                    pass
        _th.Thread(target=_fill_detail, daemon=True).start()
    return out


@app.get("/auction/competing_listings")
def auction_competing_listings(item_key: str) -> dict:
    """경매물건과 동일 평형(전용 ±3㎡)의 KB부동산 매매 매물 — '경쟁매물보기'.
    items.kb_complex_no 로 KB 단지를 찾아 kb_listing 에서 같은 전용면적대 매매만 추린다."""
    try:
        r = auction_db._get("items", [("select", "kb_complex_no,area_text"),
                                      ("item_key", f"eq.{item_key}"), ("limit", "1")])
        rows = r.json() if r.status_code in (200, 206) else []
    except Exception:
        rows = []
    if not rows or not rows[0].get("kb_complex_no"):
        return {"matched": False, "count": 0, "listings": []}
    cno = rows[0]["kb_complex_no"]
    m = re.search(r"전용\s*([\d.]+)", rows[0].get("area_text") or "")
    area = round(float(m.group(1)), 2) if m else None
    params = [("select", "listing_id,area_excl,price,floor,dong,ho,unit_price,"
                         "direction,room_cnt,bath_cnt,feature,agent_name,confirm_date"),
              ("complex_no", f"eq.{cno}"), ("trade_type", "eq.매매"),
              ("order", "price.asc"), ("limit", "300")]
    if area:                                       # 동일 평형 = 전용 ±3㎡(같은 ㎡타입만, 59/84 등 구분)
        params += [("area_excl", f"gte.{round(area - 3, 2)}"),
                   ("area_excl", f"lte.{round(area + 3, 2)}")]
    try:
        lr = auction_db._get("kb_listing", params)
        listings = lr.json() if lr.status_code in (200, 206) else []
    except Exception:
        listings = []
    cname = None
    try:
        cr = auction_db._get("kb_complex", [("select", "name"),
                                            ("complex_no", f"eq.{cno}"), ("limit", "1")])
        cj = cr.json() if cr.status_code in (200, 206) else []
        cname = cj[0].get("name") if cj else None
    except Exception:
        pass
    # 매물 사진(kb_listing_photo) 일괄조회 → listing_id별 url 목록(현재 미수집이면 빈 채로 정상)
    photo_map: dict = {}
    ids = [x.get("listing_id") for x in listings if x.get("listing_id")]
    for i in range(0, len(ids), 100):
        chunk = ",".join(str(v) for v in ids[i:i + 100])
        try:
            pr = auction_db._get("kb_listing_photo",
                                 [("select", "listing_id,url,title"),
                                  ("listing_id", f"in.({chunk})"), ("order", "seq.asc")])
            for p in (pr.json() if pr.status_code in (200, 206) else []):
                photo_map.setdefault(p["listing_id"], []).append(
                    {"url": p.get("url"), "title": p.get("title")})
        except Exception:
            pass
    return {"matched": True, "count": len(listings), "area": area,
            "complex_no": cno, "complex_name": cname,
            "listings": [{"area_excl": x.get("area_excl"), "price": x.get("price"),
                          "floor": x.get("floor"), "dong": x.get("dong"),
                          "ho": x.get("ho"), "unit_price": x.get("unit_price"),
                          "direction": x.get("direction"), "room_cnt": x.get("room_cnt"),
                          "bath_cnt": x.get("bath_cnt"), "feature": x.get("feature"),
                          "agent_name": x.get("agent_name"), "confirm_date": x.get("confirm_date"),
                          "photos": photo_map.get(x.get("listing_id"), [])}
                         for x in listings]}


@app.get("/auction/kb_counts")
def auction_kb_counts(keys: str) -> dict:
    """리스트용: 여러 경매물건의 '호가 매물 수'(동일평형 KB 매매 매물) 일괄 → {item_key: count}.
    경쟁매물보기와 같은 기준(items.kb_complex_no + 전용±3㎡)."""
    ks = [k.strip() for k in keys.split(",") if k.strip()][:200]
    if not ks:
        return {}
    inq = "in.(" + ",".join('"' + k + '"' for k in ks) + ")"
    try:
        r = auction_db._get("items", [("select", "item_key,kb_complex_no,area_text"),
                                      ("item_key", inq), ("kb_complex_no", "not.is.null")])
        items = r.json() if r.status_code in (200, 206) else []
    except Exception:
        items = []
    info, complexes = {}, set()
    for it in items:
        cno = it.get("kb_complex_no")
        if not cno:
            continue
        m = re.search(r"전용\s*([\d.]+)", it.get("area_text") or "")   # 전용 미상(area_text 빈 물건)이면 None
        info[it["item_key"]] = (str(cno), float(m.group(1)) if m else None)
        complexes.add(str(cno))
    if not complexes:
        return {}
    cinq = "in.(" + ",".join('"' + c + '"' for c in complexes) + ")"
    by_cx: dict = {}
    off = 0
    while off <= 60000:
        try:
            lr = auction_db._get("kb_listing", [("select", "complex_no,area_excl"),
                                                ("complex_no", cinq), ("trade_type", "eq.매매"),
                                                ("order", "listing_id"),   # ★ 안정 페이징: order 없으면 offset 페이징이 행을 건너뛰어 일부 단지 누락→count 0
                                                ("limit", "1000"), ("offset", str(off))])
            page = lr.json() if lr.status_code in (200, 206) else []
        except Exception:
            page = []
        for x in page:
            by_cx.setdefault(str(x.get("complex_no")), []).append(x.get("area_excl"))
        if len(page) < 1000:
            break
        off += 1000
    return {ik: (len(by_cx.get(cno, [])) if area is None       # 전용 미상 → competing_listings와 동일하게 단지 전체
                 else sum(1 for a in by_cx.get(cno, []) if a is not None and abs(a - area) <= 3))
            for ik, (cno, area) in info.items()}


_geocode_cache: dict[str, dict] = {}


def _geo_candidates(addr: str) -> list[str]:
    """지오코딩 후보(정확도순): ①단지명 ②시군구+법정동(시도접두 제거) ③지번주소.
    OSM은 단지명이 가장 정확하고, '경기도' 같은 시도 접두가 붙으면 매칭 실패가 잦다."""
    cands: list[str] = []
    # ① 단지명(지번 뒤 ~ 동/층/호 앞)
    jm = re.search(r"(?:동|읍|면|리)\s+\d+(?:-\d+)?\s+(.+?)"
                   r"(?=\s*(?:제?\s*\d+\s*동|지하|\d+\s*층|\d+\s*호|외\s*\d)|$)", addr)
    if jm:
        nm = re.sub(r"-", "", jm.group(1)).strip()
        nm = re.sub(r"\s+", " ", nm).replace("이편한세상", "e편한세상")
        if len(nm) >= 2:
            cands.append(nm)
    # ② 시군구 + 법정동 (시도 접두 제거)
    toks = addr.split()
    if toks:
        if re.search(r"(특별시|광역시|특별자치시|특별자치도|도)$", toks[0]):
            toks = toks[1:]
        di = next((i for i, t in enumerate(toks) if re.search(r"(동|읍|면|리)$", t)), -1)
        if di >= 0:
            cands.append(" ".join(toks[:di + 1]))
    # ③ 지번주소
    m = re.match(r"^(.*?(?:동|읍|면|리)\s+\d+(?:-\d+)?)", addr)
    if m:
        cands.append(m.group(1))
    # 중복 제거(순서 유지)
    return list(dict.fromkeys([c for c in cands if c]))


@app.get("/geocode")
def geocode(item_key: str) -> dict:
    """물건 주소 → 좌표(위경도). KB부동산 지도 중심이동용. Nominatim(OSM, 무료·무키)+캐시."""
    d = auction_db.get_auction(item_key)
    if not d:
        return {"available": False}
    addr = d.get("address") or ""
    ck = "geo:" + addr
    if ck in _geocode_cache:
        return _geocode_cache[ck]
    out = {"available": False}
    for q in _geo_candidates(addr):
        try:
            r = httpx.get("https://nominatim.openstreetmap.org/search",
                          params={"q": q, "format": "json", "limit": "1", "countrycodes": "kr"},
                          headers={"User-Agent": "jh-auction-school/1.0"}, timeout=12)
            j = r.json()
            if j:
                out = {"available": True, "lat": float(j[0]["lat"]),
                       "lng": float(j[0]["lon"]), "query": q}
                break
        except Exception:
            continue
    if out.get("available"):                  # 성공만 캐시(전파/일시오류 재시도 허용)
        _geocode_cache[ck] = out
    return out


def _floor_band(f) -> Optional[str]:
    try:
        f = int(f)
    except (TypeError, ValueError):
        return None
    if f <= 6:
        return "1~6층"
    if f <= 15:
        return "7~15층"
    return "16층 이상"


def _price_threshold(p: int) -> int:
    """기준가 구간별 이상치 제외 임계(원). 이 값 '이상' 벌어지면 제외."""
    억, 만 = 100000000, 10000
    if p < 1 * 억:
        return 1000 * 만
    if p < 2 * 억:
        return 1500 * 만
    if p < 3 * 억:
        return 2000 * 만
    if p < 4 * 억:
        return 3000 * 만
    if p < 6 * 억:
        return 3500 * 만
    if p < 10 * 억:
        return 5000 * 만
    return 8000 * 만


def _estimate_price(same_area: list, auction_floor) -> Optional[dict]:
    """추정시세: 같은단지·같은평형 풀에서 ①최근3개월(0건시 6개월) ②유사층수밴드
    ③median(2건이하면 최저가) 기준 가격구간 임계 초과 제외 → 남은 거래 평균."""
    from datetime import date
    band = _floor_band(auction_floor)
    today = date.today()
    for window in (3, 6):
        yy, mm = today.year, today.month - (window - 1)
        while mm <= 0:
            mm += 12
            yy -= 1
        cutoff = f"{yy:04d}-{mm:02d}-01"
        pool1 = [t for t in same_area if t.get("deal_date", "") >= cutoff]
        pool2 = [t for t in pool1 if band is None or _floor_band(t.get("floor")) == band]
        prices = sorted(t["amount"] for t in pool2 if t.get("amount"))
        if not prices:
            continue
        ref = prices[0] if len(prices) <= 2 else prices[(len(prices) - 1) // 2]
        thr = _price_threshold(ref)
        kept = [p for p in prices if abs(p - ref) < thr] or [ref]
        est = round(sum(kept) / len(kept))
        return {"price": est, "count": len(kept), "pool": len(pool2),
                "window": window, "band": band or "-", "floor": auction_floor}
    return None


def _apt_name_from_addr(addr: str) -> str:
    """주소에서 단지명 추정. ①지번 뒤 ~ 'NNN동' 앞(가장 일반적·접미키워드 무관) ②키워드 접미 폴백.
    예: '사수동 833 금호서한이다음 101동 24층2401호' → '금호서한이다음'."""
    # ① 지번(번지) 다음 ~ 건물 'NNN동' 앞 = 단지명. 행정동(사수'동')은 지번 앞이라 안 걸림.
    m = re.search(r"\d+(?:-\d+)?\s+(\S.*?)\s*\d+동(?:\s|\d|호|$)", addr)
    if m:
        nm = m.group(1).strip()
        if len(nm) >= 2:                        # 너무 짧으면(행정동 오인 등) 폴백
            return nm
    # ② 키워드 접미 폴백(동 표기 없는 주소 등)
    m = re.search(r"\d+(?:-\d+)?\s+([^\s]*(?:아파트|오피스텔|마을|캐슬|푸르지오|자이|"
                  r"힐스테이트|더샵|편한세상|위브|파크|시티|타워|팰리스|리슈빌|스카이))", addr)
    return m.group(1) if m else ""


def _complex_detail_for(out: dict):
    """apt_info 결과 → K-apt 상세 단지정보(complex_detail). 실거래 단지명+주소추정 단지명 둘 다 시도. 실패 None."""
    lawd = out.get("lawd_cd")
    if not lawd:
        return None
    cands = [n for n in (out.get("complex"), _apt_name_from_addr(out.get("address", ""))) if n]
    for nm in dict.fromkeys(cands):           # 중복 제거, 순서 유지
        try:
            det = kapt.complex_detail(lawd, nm)
        except Exception:
            det = None
        if det:
            return det
    return None


def _brief_as_detail(item_key: str, name: str):
    """K-apt 단지정보가 없을 때 건축물대장(brief)로 준공·세대·승강기 최소 단지정보 구성(시세없어도 표시용)."""
    try:
        b = _get_brief(item_key)
    except Exception:
        return None
    if not (isinstance(b, dict) and b.get("available")):
        return None
    if not (b.get("build_year") or b.get("households") or b.get("elevator") is not None):
        return None
    return {"name": name or "", "households": b.get("households"),
            "approved": (str(b.get("build_year")) if b.get("build_year") else None),
            "elevator": b.get("elevator"), "_src": "건축물대장"}


APT_VER = 4   # apt 캐시 스키마 버전 — 올리면 옛 캐시는 stale로 재계산(v4: _area_num 전용 우선=대지권 오採用 버그 픽스 반영)


def _apt_info_compute(item_key: str, months: int) -> dict:
    d = auction_db.get_auction(item_key)
    if d is None:
        return {"available": False, "reason": "물건 없음", "v": APT_VER}
    usage = d.get("usage") or ""
    if "아파트" not in usage and "오피스텔" not in usage:
        return {"available": False, "reason": "아파트/오피스텔 물건이 아님", "usage": usage}
    address = d.get("address") or ""
    lawd = resolve_lawd(address)
    if not lawd:
        return {"available": False, "reason": "주소에서 법정동코드를 찾지 못함", "address": address}
    area = _area_num(d.get("building_area"), d.get("area_text"))
    trades = _apt_trades(lawd, months)          # 시군구 캐시(같은 지역 아파트는 즉시)
    if not trades:
        # 실거래가 없어도 단지정보(준공·세대·시공사 등)는 채워서 내려줌
        nm = _apt_name_from_addr(address)
        cd = None
        if nm:
            try:
                cd = kapt.complex_detail(lawd, nm)
            except Exception:
                cd = None
        cd = cd or _brief_as_detail(item_key, nm)   # kapt 없으면 건축물대장 폴백
        return {"available": False, "reason": "해당 시군구 아파트 실거래 없음",
                "lawd_cd": lawd, "address": address, "area": area,
                "complex": nm or "", "complex_detail": cd, "v": APT_VER}
    mt = match_apt(trades, address, area=area, area_pct=0.05)   # 같은 평형 ±5%(다른 평형 혼입 방지)
    # 시세·실거래는 '같은 평형'만 사용 — 없으면 빈 리스트(단지 다른 평형을 시세로 쓰지 않음 → '시세 없음')
    same = mt["same_area"] if mt["area_matched"] else []
    # 매수세: '같은 평형' 최근 6개월(오늘-6개월) 실거래 건수. 같은평형 없으면 단지 전체로.
    from datetime import date, timedelta
    cutoff = (date.today() - timedelta(days=183)).isoformat()
    pool6 = mt["same_area"] if mt["area_matched"] else mt["trades"]
    c6 = sum(1 for t in pool6 if t.get("deal_date", "") >= cutoff)
    demand = {"count6": c6, "per_month": round(c6 / 6, 1),
              "status": "양호" if c6 >= 3 else "검토",
              "scope": "같은평형" if mt["area_matched"] else "단지전체"}
    # 추정시세: 같은단지·같은평형(same) 풀 + 경매물건 층수(주소 추출)
    fm = re.search(r"(\d+)\s*층", address)
    auction_floor = int(fm.group(1)) if fm else None
    est = _estimate_price(same, auction_floor)
    amounts = [t["amount"] for t in same if t.get("amount")]
    summary = None
    if amounts:
        summary = {"count": len(same), "recent": same[0]["amount"],
                   "recent_date": same[0]["deal_date"],
                   "min": min(amounts), "max": max(amounts),
                   "avg": round(sum(amounts) / len(amounts))}
    out = {
        "available": bool(mt["trades"]),
        "lawd_cd": lawd,
        "address": address,
        "area": area,
        "complex": mt["complex"] or "",
        "build_year": mt["build_year"] or "",
        "summary": summary,
        "trades": same[:100],                # 같은 평형(없으면 단지 전체) 최근순 100건(필터용)
        "complex_trades_total": len(mt["trades"]),
        "area_matched": mt["area_matched"],  # True=같은평형 필터 적용, False=단지 전체
        "demand": demand,                    # 매수세(최근 6개월 거래빈도)
        "est": est,                          # 추정시세(3단계 필터 평균)
        "months": months,
        "v": APT_VER,
    }
    # 상세 단지정보는 실거래 유무와 무관하게 시도(시세 없어도 단지정보는 표시)
    out["complex_detail"] = _complex_detail_for(out) or _brief_as_detail(
        item_key, out.get("complex") or _apt_name_from_addr(address))
    if not out.get("complex"):           # 단지 미매칭 시 폴백 이름으로 단지명/링크 채움
        out["complex"] = (out.get("complex_detail") or {}).get("name") or _apt_name_from_addr(address) or ""
    return out


# ---------- 인증 ----------

_COOKIE = "sid"
_COOKIE_MAXAGE = 60 * 60 * 24 * 14  # 14일

# 카카오 OAuth 설정 (개발자센터 발급). 키 없으면 안내 페이지로 유도.
KAKAO_REST_KEY = os.environ.get("KAKAO_REST_KEY", "")
KAKAO_CLIENT_SECRET = os.environ.get("KAKAO_CLIENT_SECRET", "")  # 카카오 콘솔에서 ON이면 필수
KAKAO_REDIRECT_URI = os.environ.get(
    "KAKAO_REDIRECT_URI", "http://localhost:4011/auth/kakao/callback")


# ---------- 휴대폰 본인인증(SMS OTP) ----------
import secrets as _secrets
import threading as _otp_threading
import time as _otp_time

SMS_PROVIDER = os.environ.get("SMS_PROVIDER", "").strip().lower()   # "" → 개발모드(문자 미발송, 코드 노출)
_SMS_DEV = not SMS_PROVIDER
_OTP_TTL = 180            # 인증번호 유효(초)
_OTP_RESEND = 30         # 재발송 최소 간격(초)
_OTP_MAXTRIES = 5        # 코드 검증 최대 시도
_PHONE_TOKEN_TTL = 900   # 인증완료 토큰 유효(가입 완료까지, 초)

_otp_lock = _otp_threading.Lock()
_otp_store: dict = {}     # phone -> {code, exp, tries, last_sent}
_phone_tokens: dict = {}  # token -> {phone, exp}


def _norm_phone(p: str) -> str:
    return re.sub(r"\D", "", p or "")


def _valid_phone(d: str) -> bool:
    return bool(re.match(r"^01[016789]\d{7,8}$", d))


def _otp_purge(now: float) -> None:
    for k in [k for k, v in _otp_store.items() if v.get("exp", 0) < now]:
        _otp_store.pop(k, None)
    for k in [k for k, v in _phone_tokens.items() if v.get("exp", 0) < now]:
        _phone_tokens.pop(k, None)


def _send_sms_solapi(phone: str, text: str) -> bool:
    """Solapi(CoolSMS) 발송. env: SOLAPI_API_KEY, SOLAPI_API_SECRET, SMS_SENDER."""
    import hmac
    import hashlib
    key = os.environ.get("SOLAPI_API_KEY", "")
    sec = os.environ.get("SOLAPI_API_SECRET", "")
    sender = _norm_phone(os.environ.get("SMS_SENDER", ""))
    if not (key and sec and sender):
        print("[SMS] solapi 설정 누락(SOLAPI_API_KEY/SECRET, SMS_SENDER)", flush=True)
        return False
    import datetime as _dt
    dt = _dt.datetime.utcnow().isoformat()
    salt = _secrets.token_hex(16)
    sig = hmac.new(sec.encode(), (dt + salt).encode(), hashlib.sha256).hexdigest()
    auth = f"HMAC-SHA256 apiKey={key}, date={dt}, salt={salt}, signature={sig}"
    try:
        r = httpx.post("https://api.solapi.com/messages/v4/send",
                       headers={"Authorization": auth, "Content-Type": "application/json"},
                       json={"message": {"to": phone, "from": sender, "text": text}}, timeout=10)
        return r.status_code in (200, 201, 202)
    except Exception as e:
        print(f"[SMS] solapi 오류: {e}", flush=True)
        return False


def _send_sms_aligo(phone: str, text: str) -> bool:
    """알리고 발송. env: ALIGO_API_KEY, ALIGO_USER_ID, SMS_SENDER."""
    key = os.environ.get("ALIGO_API_KEY", "")
    uid = os.environ.get("ALIGO_USER_ID", "")
    sender = _norm_phone(os.environ.get("SMS_SENDER", ""))
    if not (key and uid and sender):
        print("[SMS] aligo 설정 누락(ALIGO_API_KEY/USER_ID, SMS_SENDER)", flush=True)
        return False
    try:
        r = httpx.post("https://apis.aligo.in/send/",
                       data={"key": key, "user_id": uid, "sender": sender,
                             "receiver": phone, "msg": text}, timeout=10)
        return r.status_code == 200 and r.json().get("result_code") in (1, "1")
    except Exception as e:
        print(f"[SMS] aligo 오류: {e}", flush=True)
        return False


def _send_sms(phone: str, text: str) -> bool:
    """SMS 발송 어댑터. SMS_PROVIDER 설정 시 실제 발송, 없으면 개발모드(로그만)."""
    if _SMS_DEV:
        print(f"[SMS-DEV] {phone} ← {text}", flush=True)
        return True
    if SMS_PROVIDER == "solapi":
        return _send_sms_solapi(phone, text)
    if SMS_PROVIDER == "aligo":
        return _send_sms_aligo(phone, text)
    print(f"[SMS-?{SMS_PROVIDER}] {phone} ← {text}", flush=True)
    return True


class PhoneSendIn(BaseModel):
    phone: str


@app.post("/auth/phone/send")
def phone_send(body: PhoneSendIn) -> dict:
    """가입용 휴대폰 인증번호 발송. 개발모드면 응답에 dev_code 포함(화면 표시)."""
    d = _norm_phone(body.phone)
    if not _valid_phone(d):
        raise HTTPException(400, "올바른 휴대폰 번호를 입력하세요. (예: 010-1234-5678)")
    if user_store.phone_exists(d):
        raise HTTPException(409, "이미 가입에 사용된 연락처입니다.")
    now = _otp_time.time()
    with _otp_lock:
        _otp_purge(now)
        cur = _otp_store.get(d)
        if cur and now - cur.get("last_sent", 0) < _OTP_RESEND:
            wait = int(_OTP_RESEND - (now - cur["last_sent"])) + 1
            raise HTTPException(429, f"{wait}초 후 다시 시도하세요.")
        code = f"{_secrets.randbelow(1000000):06d}"
        _otp_store[d] = {"code": code, "exp": now + _OTP_TTL, "tries": 0, "last_sent": now}
    ok = _send_sms(d, f"[JH옥션스쿨] 인증번호 [{code}] 를 입력하세요. (3분 내)")
    if not ok:
        raise HTTPException(502, "문자 발송에 실패했습니다. 잠시 후 다시 시도하세요.")
    out = {"sent": True, "ttl": _OTP_TTL, "dev": _SMS_DEV}
    if _SMS_DEV:
        out["dev_code"] = code            # 개발모드 전용: 실제 발송 설정 시 미포함
    return out


class PhoneVerifyIn(BaseModel):
    phone: str
    code: str


@app.post("/auth/phone/verify")
def phone_verify(body: PhoneVerifyIn) -> dict:
    """인증번호 확인 → 가입에 쓸 1회용 phone_token 발급."""
    d = _norm_phone(body.phone)
    now = _otp_time.time()
    with _otp_lock:
        rec = _otp_store.get(d)
        if not rec or now > rec.get("exp", 0):
            raise HTTPException(400, "인증번호가 만료되었습니다. 다시 받아주세요.")
        if rec["tries"] >= _OTP_MAXTRIES:
            _otp_store.pop(d, None)
            raise HTTPException(429, "시도 횟수를 초과했습니다. 다시 받아주세요.")
        rec["tries"] += 1
        if (body.code or "").strip() != rec["code"]:
            raise HTTPException(400, "인증번호가 일치하지 않습니다.")
        _otp_store.pop(d, None)
        token = _secrets.token_urlsafe(24)
        _phone_tokens[token] = {"phone": d, "exp": now + _PHONE_TOKEN_TTL}
    return {"verified": True, "phone": d, "phone_token": token}


class SignupIn(BaseModel):
    email: str
    password: str
    name: str = ""
    phone: str = ""
    phone_token: str = ""


class LoginIn(BaseModel):
    email: str
    password: str


def _set_session_cookie(response: Response, token: str) -> None:
    response.set_cookie(_COOKIE, token, httponly=True, samesite="lax",
                        max_age=_COOKIE_MAXAGE, path="/")


@app.post("/auth/signup")
def signup(body: SignupIn, response: Response) -> dict:
    # 휴대폰 본인인증 필수: phone_verify가 발급한 1회용 토큰이 번호와 일치해야 가입
    d = _norm_phone(body.phone)
    now = _otp_time.time()
    with _otp_lock:
        _otp_purge(now)
        tok = _phone_tokens.get(body.phone_token or "")
        ok = (tok and tok.get("exp", 0) >= now and tok.get("phone") == d and _valid_phone(d))
        if not ok:
            raise HTTPException(400, "휴대폰 본인인증을 완료해주세요.")
    try:
        user = user_store.create_user(body.email, body.password, body.name, phone=d)
    except ValueError as e:
        raise HTTPException(409, str(e))
    user = _maybe_make_admin(user)                   # 관리자 이름/이메일이면 자동 승격
    _phone_tokens.pop(body.phone_token, None)        # 토큰 1회용 소모
    _set_session_cookie(response, user_store.create_session(user["id"]))
    return user


@app.post("/auth/login")
def login(body: LoginIn, response: Response) -> dict:
    user = user_store.authenticate(body.email, body.password)
    if not user:
        raise HTTPException(401, "이메일 또는 비밀번호가 올바르지 않습니다.")
    _set_session_cookie(response, user_store.create_session(user["id"]))
    return user


@app.post("/auth/logout")
def logout(response: Response, sid: Optional[str] = Cookie(None)) -> dict:
    user_store.delete_session(sid)
    response.delete_cookie(_COOKIE, path="/")
    return {"ok": True}


@app.get("/auth/me")
def me(user: Optional[dict] = Depends(current_user)) -> dict:
    if not user:
        raise HTTPException(401, "로그인이 필요합니다.")
    return {**user, "grade_rank": _user_grade_rank(user)}   # 프런트 등급 게이트(유형필터·정렬)용


_INTRO_BODY_CACHE = {"html": None}


@app.get("/intro", response_class=HTMLResponse)
def intro_page(request: Request) -> HTMLResponse:
    """JH옥션스쿨 소개 페이지 — 카카오/SNS 공유 시 'JH옥션스쿨 소개' 미리보기(og 메타).
    og:image·og:url을 요청 host 기준 절대경로로 생성 → 배포 도메인 자동 반영(로컬은 localhost)."""
    if _INTRO_BODY_CACHE["html"] is None:
        try:
            with open(os.path.join(_ROOT, "static", "intro_content.html"), encoding="utf-8") as f:
                _INTRO_BODY_CACHE["html"] = f.read()
        except Exception:
            _INTRO_BODY_CACHE["html"] = "<p style='padding:40px'>소개 페이지를 불러오지 못했습니다.</p>"
    base = str(request.base_url).rstrip("/")
    desc = "경매의 어려움을 AI가 대신 풀어드립니다 — 권리분석·예상낙찰가·시세·수익계산을 한 화면에서."
    og = (
        '<meta property="og:type" content="website">'
        '<meta property="og:site_name" content="JH옥션스쿨">'
        '<meta property="og:title" content="JH옥션스쿨 소개">'
        f'<meta property="og:description" content="{desc}">'
        f'<meta property="og:image" content="{base}/static/og_intro.png">'
        '<meta property="og:image:width" content="1200">'
        '<meta property="og:image:height" content="630">'
        f'<meta property="og:url" content="{base}/intro">'
        '<meta name="twitter:card" content="summary_large_image">'
        '<meta name="twitter:title" content="JH옥션스쿨 소개">'
        f'<meta name="twitter:description" content="{desc}">'
        f'<meta name="twitter:image" content="{base}/static/og_intro.png">'
    )
    html = (
        '<!doctype html><html lang="ko"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        '<title>JH옥션스쿨 소개</title>'
        f'{og}</head><body style="margin:0">{_INTRO_BODY_CACHE["html"]}</body></html>'
    )
    return HTMLResponse(html)


@app.get("/download/sms_app")
def download_sms_app():
    """자동문자 보내기 앱(APK) — 정확한 MIME으로 강제 다운로드(부동산DB 페이지 버튼)."""
    p = os.path.join(_ROOT, "static", "downloads", "리치빌더_자동문자보내기.apk")
    if not os.path.exists(p):
        raise HTTPException(404, "앱 파일을 찾을 수 없습니다.")
    return FileResponse(p, media_type="application/vnd.android.package-archive",
                        filename="리치빌더 자동문자보내기.apk")


# ---------- AI 챗봇(사이트 안내) ----------
_CHAT_SYSTEM = """당신은 'JH옥션스쿨' 부동산 경매·공매 분석 플랫폼의 AI 안내 도우미입니다.
사용자 질문에 친절하고 간결하게(핵심 위주 3~6문장) 한국어로 답하세요.

[사이트] JH옥션스쿨 — 경매 초보자도 쉽게 쓰도록 어려운 판단을 AI가 대신 처리하는 플랫폼.

[핵심 기능]
- AI 권리분석: 등기 자동 분석 → 말소기준·인수/소멸 위험을 '안전·주의·위험' 신호등으로 요약. 분묘기지권·유치권 등 특수물건 경고.
- AI 매수판정: 위험요소 종합 → '매수양호/매수검토/매수금지' 등급.
- 예상 낙찰가: 아파트는 동일단지 매각사례, 빌라·도생은 반경 1km, 차량은 모델·연식 기준 추정.
- 시세·공시가·차익: KB시세·국토부 실거래·공시가 자동 연동, 시세 대비 예상 차익 표시.
- 단기매도 수익 계산기: 취득세·종합소득세(종소세) 자동 반영해 실투자금·순수익 계산.
- 유형별 전문 분석: 다가구 수익률·방공제, 상가 배후세대·업종, 위반건축물·승강기 없음 리스크.
- 서류 통합 뷰어(등기·감정평가서·명세서·건축물대장·사진), 경매+공매(온비드)+부동산DB(KB 실매물) 원스톱.

[회원 등급]
- 무료: 물건 목록·검색.
- 전국: 물건 상세 열람, AI 권리분석·매수판정, 예상낙찰가, 서류, 수익계산기, 부동산DB(이용권 등록 시).
- 프리미엄: 전국 기능 + 유형별 세부 필터, 정렬순서 1·2.

[규칙]
- 사이트 기능·경매 기초·이용법을 안내. 낙찰·수익을 보장하지 않으며 최종 판단은 이용자 책임임을 필요시 안내.
- 사이트와 무관한 질문은 정중히 경매·사이트 관련으로 유도. 구체적 법률자문은 전문가 상담 권유.
- 실제 물건 추천: 사용자가 지역·예산·물건유형으로 '투자 가능한 물건', '어떤 물건 있어?', '추천'을 물으면 반드시 search_properties 도구로 실제 경매 물건을 검색한다. 검색 결과 물건은 사용자 화면에 **카드(사진·시세·차익 포함)로 자동 표시**되므로, 답변 텍스트에는 물건을 일일이 나열하거나 링크를 붙이지 말고 '조건에 맞는 물건 N건을 찾았어요. 아래 카드를 확인해 보세요.'처럼 1~2문장으로 요약한 뒤, 필요하면 무주택·대출 팁만 짧게 덧붙인다. 투자금 3천만원=30000000원. '주택 위주'는 usages에 아파트·다세대 (빌라)·주택·다가구 (원룸등)·도시형생활 주택을 넣고, 안전 우선이면 buy_ok=true. 무주택·대출 관련은 학습자료 규정도 함께 짧게 안내한다. **지역 범위:** 사용자가 '전국', '지방 (전체)', '규제 없는 곳/지역', '어디든', '넓혀서/확대해서'처럼 특정 시/도를 벗어난 넓은 범위를 요청하면 sido를 반드시 비워(생략) 전국을 검색한다. 직전 대화에서 특정 지역(예: 대구)을 봤더라도 그 지역에 갇히지 말고 사용자의 최신 요청 범위를 그대로 반영할 것. **조건 유지(매우 중요):** 대화 중 한 번 유형(usages)·지역(sido)·예산(invest_max)을 정하면, 이후 이어지는 물건 관련 질문에서는 사용자가 **명시적으로 다른 유형/지역/예산을 말하지 않는 한 이전 조건을 그대로 유지**한다. '다른 물건 없어?', '예상낙찰가 있는 거', '차익 큰 거', '더 싼 거', '지방으로 확대' 같은 후속 요청은 정렬·범위·부가조건만 바뀔 뿐 **유형·예산은 그대로**다. 예: '대구 아파트 3천만'으로 보다가 '예상낙찰가 있는 물건 있어?'라고 하면 → usages=아파트·invest_max=30000000을 **반드시 유지**한다(유형을 빌라·다세대 등으로 절대 바꾸지 마라). 사용자가 '빌라도 보여줘', '전국으로', '5천만으로' 처럼 명시적으로 바꿀 때만 해당 항목을 변경한다. **정렬:** 사용자가 '차익 큰/제일 큰/수익 높은 순'을 원하면 order_by=profit, '싼/저렴한/최저가 낮은 순'이면 order_by=cheap을 넣는다(정렬 언급이 없으면 생략). **직전 목록 참조(중요):** 사용자가 '이 중에(서)', '방금/위/추천받은/그 것 중에' 처럼 직전에 보여준 물건을 가리키면, 바로 앞 대화에 남은 '[방금 보여준 물건 N건] …' 목록을 근거로 답한다(그 목록의 사건번호·차익·예상낙찰가 표시를 그대로 활용). 예: '이 중에 예상낙찰가 있는 것'이면 → 직전과 동일한 유형·지역·예산(ctx)에 has_expbid=true를 더해 search_properties를 호출한다(그러면 예상낙찰가가 표시된 카드로 다시 보여진다). **예상낙찰가:** '예상낙찰가 있는/나온 것/만'을 **명시적으로** 원할 때만 has_expbid=true를 넣는다. '단기매매/단기매도 가능한', '단타', '수익 나는 물건' 같은 표현엔 has_expbid를 **절대 넣지 마라**(예상낙찰가가 없어도 시세로 판단 가능하고, 넣으면 예상낙찰가 미산정 물건이 전부 빠져 0건이 된다). **직전 물건 상세 질문(매우 중요):** 사용자가 방금 보여준 물건에 대해 '예상낙찰가 얼마', '시세/감정가/최저가 얼마', '차익 얼마', '안전한가/매수판정 뭐' 등을 물으면 **절대 새로 검색하지 말고** 바로 앞 대화의 '[방금 보여준 물건] …' 목록에 있는 값을 그대로 알려준다(그 목록에 감정가·최저가·시세·차익·예상낙찰가·매수판정이 모두 들어 있다). 예상낙찰가가 '미산정'이면 아직 산정 전이라고 안내한다. 물건이 여러 개라 특정이 안 되면 어느 물건인지 되묻거나 가장 최근(또는 사용자가 가리킨) 물건 기준으로 답한다. '못 찾았어요' 같은 답을 하지 말 것. **결과 개수·이유 질문(매우 중요):** '왜 N건만 나왔어', '왜 이것(이거)밖에 없어', '왜 1건만', '지역에 X건인데 왜 이것뿐이야' 처럼 **직전 검색 결과의 개수·이유**를 물으면 **절대 재검색(search_properties)하지 말고** 직전 결과를 근거로 설명한다(예: '조건에 맞게 산정된 게 이 정도예요. 예산·유형·지역을 넓히면 더 나옵니다'). 재검색하면 이미 보여준 물건이 제외되어 0건이 될 수 있으니 하지 마라. **심층 질문(권리분석·임차인·안전성):** 특정 물건의 권리분석·위험도·말소기준·인수보증금·임차인 현황·위반건축물·안전성('이 물건 권리분석 어때', '임차인 있어?', '안전해?', '2025-32164 어때?')은 get_property_detail 도구로 조회해서, 조회된 위험도·말소기준·인수보증금·임차인 정보를 근거로 이해하기 쉽게 설명한다. 인수보증금이 있으면 매수인이 떠안는 금액임을 분명히 알린다.

[대출 규제지역 사실 — 반드시 이 기준으로 답할 것] 2023년 1월 이후 규제지역(조정대상지역·투기과열지구)은 **서울 강남·서초·송파·용산 4개 구뿐**이다. **대구를 포함한 그 외 전국 모든 지역은 비규제지역**이며, 무주택자 주택담보대출 LTV는 보통 70%다(규제지역만 무주택 50% 등으로 강화). 사용자가 특정 지역의 규제 여부·대출 한도를 물으면, 대구·부산·인천·경기 등은 '비규제지역, LTV 70%'라고 정확히 답하고, 규제지역으로 잘못 단정하지 말 것."""


_CHAT_TOOLS = [{
    "type": "function",
    "function": {
        "name": "search_properties",
        "description": "실제 진행 중인 경매 물건을 조건으로 검색해 추천한다. 사용자가 지역·예산·물건유형으로 '투자 가능한 물건', '어떤 물건 있어?', '추천해줘' 등을 물으면 반드시 호출한다.",
        "parameters": {"type": "object", "properties": {
            "sido": {"type": "string", "description": "시/도. 예: 대구, 서울, 부산, 경기. 사용자가 '전국·지방 전체·규제 없는 곳·어디든·넓혀서'처럼 특정 시/도를 벗어난 넓은 범위를 원하면 이 값을 비워라(생략=전국 검색). 직전 대화에서 특정 지역을 봤더라도 최신 요청 범위를 따른다."},
            "region": {"type": "string", "description": "구/군/동 키워드 1개(선택). 예: 수성구. 여러 지역을 동시에 원하면 이것 대신 regions 배열을 써라."},
            "regions": {"type": "array", "items": {"type": "string"}, "description": "여러 구/군/동을 동시에 검색(OR). 사용자가 말한 각 지역(동·구·군)을 하나씩 원소로 넣어라. 시/도는 sido에 넣고 여기엔 동/구/군 이름만. 예: '부산 온천동, 대연동 빌라' → sido='부산', regions=['온천동','대연동']. '수성구·달서구' → regions=['수성구','달서구']."},
            "usages": {"type": "array", "items": {"type": "string"}, "description": "물건유형(현황용도). 다음에서만 선택: 아파트, 다세대 (빌라), 도시형생활 주택, 주택, 다가구 (원룸등), 근린주택, 농가주택, 오피스텔, 근린상가, 숙박시설"},
            "invest_max": {"type": "integer", "description": "투자금(초기 필요자금) 상한(원). 3천만원=30000000"},
            "price_max": {"type": "integer", "description": "최저매각가 상한(원, 선택)"},
            "buy_ok": {"type": "boolean", "description": "AI 매수판정 '양호'(안전)한 물건만"},
            "order_by": {"type": "string", "enum": ["profit", "cheap"], "description": "정렬 기준. 'profit'=예상차익 큰 순, 'cheap'=최저가 낮은 순. 사용자가 '차익 큰/제일 큰/수익 높은'이면 profit, '싼/저렴한/최저가 낮은'이면 cheap. 특별한 정렬 언급이 없으면 생략."},
            "has_expbid": {"type": "boolean", "description": "예상낙찰가가 산정된 물건만. 사용자가 '예상낙찰가 있는 것/예상낙찰가 나온 것/예상낙찰가만'을 원하면 true."}
        }, "required": []}
    }
}, {
    "type": "function",
    "function": {
        "name": "get_property_detail",
        "description": "특정 물건 하나의 상세 정보(AI 권리분석 위험도·말소기준일·인수보증금·임차인 현황·감정가·최저가·예상낙찰가)를 조회한다. 사용자가 특정 물건(직전에 본 물건 또는 사건번호로 지정)의 '권리분석/안전한지/위험한지/임차인/인수보증금/이 물건 어때' 같은 심층 정보를 물으면 호출한다.",
        "parameters": {"type": "object", "properties": {
            "case_no": {"type": "string", "description": "사건번호(예: 2025-32164). 직전에 본 물건을 가리키면 생략 가능."},
            "item_key": {"type": "string", "description": "물건 내부 키(직전 목록에서 알 수 있으면)."}
        }, "required": []}
    }
}]


def _chat_sotax(diff: int) -> int:
    """차익(원) → 종합소득세(총액, 누진공제 전) — auction.html ssTax와 동일 2023 누진세율표."""
    if diff <= 0:
        return 0
    for cap, rate in ((14000000, 6), (50000000, 15), (88000000, 24), (150000000, 35),
                      (300000000, 38), (500000000, 40), (1000000000, 42)):
        if diff <= cap:
            return round(diff * rate / 100)
    return round(diff * 45 / 100)


def _chat_excl(area_text) -> float:
    """area_text에서 전용/건물 면적(㎡) 파싱 — _invest_index와 동일 규칙."""
    import re as _r
    mt = _r.search(r"(?:전용|건물)\s*([0-9.]+)", area_text or "")
    try:
        return float(mt.group(1)) if mt else 0.0
    except Exception:
        return 0.0


_SELF_BASE = os.environ.get("SELF_BASE", "http://127.0.0.1:4011").rstrip("/")   # 자기 서버 호출 기준 URL. 클라우드 배포 시 포트가 8000 등으로 바뀌므로 SELF_BASE 환경변수로 오버라이드(예: http://127.0.0.1:8000)


def _chat_property_detail(args: dict, last=None) -> dict:
    """특정 물건의 종합 상세(권리분석·말소기준·인수보증금·임차인·예상낙찰가)를 텍스트로 반환.
    '이 물건 권리분석 어때', '2025-32164 임차인 있어?' 등 특정 물건 심층 질문용."""
    import re as _re
    _digits = lambda s: _re.sub(r"[^0-9]", "", s or "")
    ik = (args.get("item_key") or "").strip()
    cno = (args.get("case_no") or "").strip()
    # 1) item_key 확보: 직전 목록(last) 사건번호 매칭 → caseno 검색 → last 단일건
    if not ik and cno and last:
        _cn = _digits(cno.split("-")[-1]) or _digits(cno)
        for x in last:
            if isinstance(x, dict) and _cn and _digits(str(x.get("사건번호"))).endswith(_cn):
                ik = x.get("item_key") or ""
                break
    if not ik and cno:
        try:
            _num = _digits(cno.split("-")[-1]) or _digits(cno)
            r = httpx.get(_SELF_BASE + "/auctions", params=[("caseno", _num), ("limit", "1")], timeout=15).json()
            its = r.get("items") or []
            if its:
                ik = its[0].get("item_key") or ""
        except Exception:
            pass
    if not ik and last and len(last) == 1:
        ik = (last[0] or {}).get("item_key") or ""
    if not ik:
        return {"detail_text": "", "error": "어느 물건인지 특정하지 못했어요. 사건번호를 알려주시거나 목록에서 물건을 먼저 확인해 주세요."}
    # 2) 기본 상세 + 권리분석 조회
    try:
        d = auction_db.get_auction(ik) or {}
    except Exception:
        d = {}
    try:
        an = auction_analysis(ik) or {}
    except Exception:
        an = {}
    L = []
    L.append("사건번호 %s · %s · %s" % (d.get("case_no") or cno or ik,
             d.get("address") or an.get("address") or "", d.get("usage") or an.get("usage") or ""))
    _ap = _to_int(d.get("appraisal_price")); _mp = _to_int(d.get("min_price"))
    if _ap:
        L.append("감정가 %d만원" % round(_ap / 10000))
    if _mp:
        L.append("최저가 %d만원" % round(_mp / 10000))
    if d.get("area_text") or an.get("area_text"):
        L.append("면적 %s" % (d.get("area_text") or an.get("area_text")))
    _rl = an.get("risk_level")
    if _rl:
        L.append("AI 권리분석 위험도: %s" % _rl)
    _bl = an.get("baseline") or {}
    if _bl.get("date"):
        L.append("말소기준일 %s (%s)" % (_bl.get("date"), _bl.get("type") or ""))
    _at = _to_int(an.get("assumed_amount_total"))
    if _at:
        L.append("매수인이 인수하는 보증금 약 %d만원(주의)" % round(_at / 10000))
    _ts = an.get("tenants") or []
    if _ts:
        L.append("임차인 %d명:" % len(_ts))
        for t in _ts[:6]:
            _dep = _to_int(t.get("deposit"))
            L.append("  · %s 보증금%s 전입일%s 대항력%s %s" % (
                t.get("name") or "", ("%d만" % round(_dep / 10000) if _dep else "-"),
                t.get("move_in") or "-", ("있음" if t.get("has_opposing_power") else "없음"),
                t.get("status_label") or ""))
    elif an.get("occupancy_available"):
        L.append("임차인: 없음(현황조사 기준)")
    for w in (an.get("warnings") or [])[:4]:
        L.append("주의: %s" % w)
    try:
        _eb = (auction_expbid_batch(ik) if (d.get("usage") or "") == "아파트" else auction_vexpbid_batch(ik)) or {}
        _ev = _eb.get(ik)
        if isinstance(_ev, dict) and _ev.get("expected_bid"):
            L.append("예상낙찰가 약 %d만원" % round(_ev["expected_bid"] / 10000))
    except Exception:
        pass
    return {"detail_text": "\n".join(L), "item_key": ik}


def _chat_filter_recent(last, args: dict) -> dict:
    """'이 중에서 ~' 요청 — 직전에 보여준 목록(last)에서 조건만 걸러 반환(재검색 안 함).
    last는 직전 카드 배열(예상낙찰가·차익·유형·최저가 포함)."""
    out = [dict(x) for x in (last or []) if isinstance(x, dict)]
    if args.get("has_expbid"):
        out = [x for x in out if x.get("예상낙찰가")]          # 예상낙찰가 산정된 것만
    if args.get("price_max"):
        try:
            _pm = int(args["price_max"])
            out = [x for x in out if (_to_int(x.get("최저가")) or 0) <= _pm]
        except Exception:
            pass
    _order = (args.get("order_by") or "").strip().lower()
    if _order == "profit":
        out.sort(key=lambda x: (x.get("예상차익") or 0), reverse=True)
    elif _order == "cheap":
        out.sort(key=lambda x: (x.get("최저가") or 0))
    return {"총검색건수": len(last or []), "표시건수": len(out[:6]), "물건목록": out[:6]}


def _chat_search_properties(args: dict, exclude_keys=None) -> dict:
    """챗봇 도구 — 자기 서버 /auctions를 호출해 실제 물건 top6 요약을 반환.
    exclude_keys(item_key set): 이미 챗봇에 보여준 물건 — '다른 물건' 요청 시 제외."""
    _order = (args.get("order_by") or "").strip().lower()
    _limit = 150 if _order in ("profit", "cheap") else 60   # 정렬 요청 시 표본 확대(차익·저가 최대 근사)
    params = [("limit", str(_limit)), ("sort", "사건번호")]   # 시세없음·매수금지 제외 후 아래서 정렬→6건
    sido = (args.get("sido") or "").replace("광역시", "").replace("특별시", "").replace("특별자치도", "").strip()
    _regs = args.get("regions")
    if isinstance(_regs, str):
        _regs = [_regs]
    _regs = [str(r).strip() for r in (_regs or []) if str(r).strip()]
    if not _regs and args.get("region"):
        _regs = [str(args["region"]).strip()]
    if _regs:                                  # 여러 동/구 → 각각 '시도 구군동'으로 → 백엔드 regions OR(지역끼리 OR, 토큰 AND)
        for r in _regs:
            params.append(("regions", r if (not sido or sido in r) else sido + " " + r))
    elif sido:
        params.append(("sido", sido))
    _umap = {"다세대": "다세대 (빌라)", "빌라": "다세대 (빌라)", "다세대(빌라)": "다세대 (빌라)",
             "다가구": "다가구 (원룸등)", "원룸": "다가구 (원룸등)", "다가구(원룸등)": "다가구 (원룸등)",
             "도시형생활주택": "도시형생활 주택", "도시형": "도시형생활 주택", "도시형생활": "도시형생활 주택"}
    for u in (args.get("usages") or []):
        if u:
            params.append(("usage", _umap.get(str(u).replace(" ", ""), str(u))))
    for k in ("invest_max", "price_max"):
        if args.get(k):
            try:
                params.append((k, int(args[k])))
            except Exception:
                pass
    # buy_ok(매수양호) 필터는 지역·유형·투자금과 겹치면 0건이 되기 쉬워 검색엔 쓰지 않는다.
    # (안전도는 각 카드의 '매수판정' 배지로 표시)
    try:
        res = httpx.get(_SELF_BASE + "/auctions", params=params, timeout=(40 if _limit > 60 else 25)).json()
    except Exception as e:
        return {"error": "물건 검색에 실패했습니다: %s" % e, "물건목록": []}
    # 결과 0이면 투자금 상한을 풀어 지역·유형만으로 재검색(너무 좁아 0건 방지)
    if not res.get("items") and args.get("invest_max"):
        try:
            res = httpx.get(_SELF_BASE + "/auctions",
                            params=[(k, v) for k, v in params if k != "invest_max"], timeout=(40 if _limit > 60 else 25)).json()
        except Exception:
            pass
    _imax = args.get("invest_max")
    out = []
    _seen = set()                         # 같은 사건번호 중복 카드 방지(여러 회차·호실)
    for it in (res.get("items") or []):
        grade = it.get("buy_grade") or it.get("grade")
        if grade == "매수금지":
            continue                      # 위험(매수금지) 물건은 투자 추천에서 제외
        if exclude_keys and it.get("item_key") in exclude_keys:
            continue                      # 이미 챗봇에 보여준 물건 → '다른 물건' 요청 시 제외
        est = it.get("est"); mp = it.get("min_price")
        if not est:
            continue                      # 추정시세 없으면 차익 계산 불가 → 추천에서 제외(주인님 지시)
        _est = _to_int(est) or 0
        _mp = _to_int(mp) or 0
        if _est <= _mp:
            continue                      # 시세 ≤ 최저가 → 차익 없음/손해 → 추천 제외(마이너스 차익 금지)
        _apx = _to_int(it.get("appraisal_price")) or 0
        if _apx and _est > _apx * 1.5:
            continue                      # 시세 > 감정가×1.5 → 유사거래 오매칭 이상치(허수 차익) 제외(홈 히어로와 동일)
        if _apx and _mp < _apx * 0.2:
            continue                      # 최저가 < 감정가×20% → 비정상 대폭유찰/최저가 파싱오류(허수 차익) 제외
        # 총 필요금액(선금 + 종소세)으로 예산 재검증. /auctions의 invest_max는 '선금'(종소세 제외)이라
        #  실제 총필요금액보다 느슨함 → 여기서 화면 계산기와 동일한 '총 필요금액'으로 상한 초과분을 걸러낸다.
        if _imax:
            _sungeum = _invest_of(_mp, _chat_excl(it.get("area_text")))
            _total_need = _sungeum + _chat_sotax(_est - _mp)
            if _sungeum and _total_need > _imax:
                continue                  # 총 필요금액이 투자금 상한 초과 → 추천에서 제외
        _case = it.get("case_label") or it.get("case_no")
        if _case in _seen:
            continue                      # 같은 사건번호 이미 담음 → 중복 카드 제외
        _seen.add(_case)
        profit = _est - _mp
        out.append({
            "사건번호": it.get("case_label") or it.get("case_no"),
            "소재지": it.get("address"),
            "유형": it.get("usage"),
            "감정가": it.get("appraisal_price"),
            "최저가": mp,
            "시세": est,
            "예상차익": profit,
            "매각기일": (str(it.get("sell_date") or ""))[:10],
            "매수판정": grade,
            "thumb": it.get("thumb_url"),
            "item_key": it.get("item_key"),
            "링크": "/static/auction.html?item_key=" + str(it.get("item_key") or ""),
        })
        if len(out) >= 60:
            break                         # 정렬용 충분 수집(아래서 order_by로 정렬 후 6건)
    # 예상낙찰가(캐시) 일괄 조회 → 각 카드에 부착 + 'has_expbid'(예상낙찰가 있는 것만) 필터
    try:
        _apt_ks = [x["item_key"] for x in out if x.get("유형") == "아파트" and x.get("item_key")]
        _vil_ks = [x["item_key"] for x in out if x.get("유형") in ("다세대 (빌라)", "도시형생활 주택") and x.get("item_key")]
        _exp = {}
        if _apt_ks:
            _exp.update(auction_expbid_batch(",".join(_apt_ks)) or {})
        if _vil_ks:
            _exp.update(auction_vexpbid_batch(",".join(_vil_ks)) or {})
    except Exception:
        _exp = {}
    for x in out:
        _ev = _exp.get(x.get("item_key"))
        x["예상낙찰가"] = (_ev.get("expected_bid") if isinstance(_ev, dict) and _ev.get("expected_bid") else None)
    # 예상낙찰가가 있으면 '실제 차익 = 시세 − 예상낙찰가'로 교체하고, 손해(예상낙찰가 ≥ 시세)는 제외.
    #  (최저가 기준 차익은 허수 — 실제 매수가는 예상낙찰가이므로. 주인님 지시)
    _kept = []
    for x in out:
        _eb = _to_int(x.get("예상낙찰가"))
        _estv = _to_int(x.get("시세")) or 0
        if _eb:
            _real = _estv - _eb
            if _real <= 0:
                continue                  # 예상낙찰가 ≥ 시세 → 사면 손해 → 추천 제외
            x["예상차익"] = _real           # 카드 차익을 실제(시세 − 예상낙찰가)로 교체
        _kept.append(x)
    out = _kept
    if args.get("has_expbid"):
        _we = [x for x in out if x.get("예상낙찰가")]
        if _we and len(_we) * 2 >= len(out):   # 예상낙찰가 산정분이 '과반'이면 그것만(명시적 '예상낙찰가만' 요청 존중).
            out = _we                            #  소수뿐이면(예: 1/9) GPT가 '단기매매' 등에 has_expbid를 잘못 붙인 것으로 보고 전체 유지(1건으로 쪼그라드는 것 방지)
    if _order == "profit":
        out.sort(key=lambda x: (x.get("예상차익") or 0), reverse=True)   # 차익 큰 순
    elif _order == "cheap":
        out.sort(key=lambda x: (x.get("최저가") or 0))                    # 최저가 낮은 순
    out = out[:6]
    return {"총검색건수": res.get("total", 0), "표시건수": len(out), "물건목록": out}


@app.post("/chat")
def chat_bot(body: dict = Body(...), sid: Optional[str] = Cookie(None)) -> dict:
    """사이트 안내 + 실제 물건 검색·추천 AI 챗봇. openai_key 재사용, GPT function calling.
    물건 검색·조회(카드)는 프리미엄 등급 이상(또는 관리자) 전용 — 일반 안내는 누구나."""
    msg = (body.get("message") or "").strip()
    if not msg:
        return {"reply": "무엇을 도와드릴까요? 지역·예산·유형을 알려주시면 실제 물건도 찾아드려요."}
    # 물건 검색·조회 게이트용 등급 판정(비로그인/무료=미허용, 프리미엄 rank≥20 또는 admin=허용)
    try:
        _cu = user_store.get_user_by_session(sid) if sid else None
    except Exception:
        _cu = None
    _is_premium = bool(_cu and (_cu.get("role") == "admin" or _user_grade_rank(_cu) >= _PREMIUM_MIN_RANK))
    _PREMIUM_GATE = {"reply": "AI 물건 검색·조회는 프리미엄 등급 이상 회원 전용이에요. 요금제에서 프리미엄으로 올리시면 "
                              "조건에 맞는 실제 경매 물건을 AI가 바로 찾아드립니다. (사이트 이용법·경매 기초·용어는 지금도 얼마든지 물어보세요!)",
                     "cards": [], "used_args": {}, "list_text": ""}
    key = _openai_key()   # env var(OPENAI_API_KEY) 우선 → 로컬 kakao_broadcast.json 폴백(챗봇이 _openai_key 안 쓰고 파일만 읽던 버그 픽스)
    if not key:
        return {"reply": "죄송합니다. AI 상담 기능이 아직 준비 중입니다(관리자 설정 필요). "
                         "그동안 궁금한 점은 고객센터로 문의해 주세요."}
    _sys = _CHAT_SYSTEM
    _kbtext = _load_ai_kb()
    if _kbtext.strip():
        _sys += ("\n\n[학습 자료 — 아래는 관리자가 등록한 참고자료입니다. 관련 질문(대출·규제·세금·절차 등)은 "
                 "반드시 아래 자료를 근거로 구체적으로 답하세요. 지역·주택 보유수 등 조건을 물으면 자료의 규정을 "
                 "적용해 안내하고, 자료에 없는 세부는 은행·전문가·최신 공고 확인을 권하세요]\n" + _kbtext)
    hdr = {"Authorization": "Bearer " + key, "Content-Type": "application/json"}
    url = "https://api.openai.com/v1/chat/completions"
    _ctx = body.get("ctx") or {}          # 직전 검색 조건(대화 맥락 유지 — 유형·지역·예산)
    if _ctx:
        _cp = []
        if _ctx.get("usages"):
            _cp.append("유형=" + ",".join(str(u) for u in _ctx["usages"]))
        if _ctx.get("sido"):
            _cp.append("지역=" + str(_ctx["sido"]))
        if _ctx.get("invest_max"):
            try:
                _cp.append("투자금상한=%d원" % int(_ctx["invest_max"]))
            except Exception:
                pass
        if _cp:
            _sys += ("\n\n[직전 검색 조건 — 이어지는 물건 질문의 기본값] " + " / ".join(_cp) +
                     ". 사용자가 명시적으로 다른 유형·지역·예산을 말하지 않으면 이 조건을 그대로 사용해 검색하라.")
    msgs = [{"role": "system", "content": _sys}]
    for h in (body.get("history") or [])[-20:]:      # 대화 맥락 넉넉히(6→20턴) — '이 중에서' 등 참조 지원
        role = h.get("role"); content = (h.get("content") or "").strip()
        if role in ("user", "assistant") and content:
            msgs.append({"role": role, "content": content[:2000]})   # 물건 목록 텍스트 보존(1000→2000)
    msgs.append({"role": "user", "content": msg[:1500]})
    _ref = any(_k in msg for _k in ("이 중", "이중", "방금", "그 중", "그중", "위에", "추천받은", "아까", "그것 중", "여기", "이것들", "이 물건들", "이중에"))
    # seen(이미 본 물건 제외)은 '다른/또/말고' 등 명시적으로 '새 물건'을 원할 때만 적용한다.
    #  '차익 큰것 추려줘' 같은 일반 검색·정렬은 전체 대상(이미 본 진짜 top이 빠지면 안 됨).
    _other = any(_k in msg for _k in ("다른", "말고", "그 외", "그외", "이외", "새로운", "새 물건", "또 다른", "또다른", "또 없", "또 보여", "또 추천", "말고 다른"))
    _seen_keys = {str(x) for x in (body.get("seen") or []) if x} if _other else set()
    _last = body.get("last") or []        # 직전에 보여준 카드 목록('이 중에서' 필터 대상)
    _cards = []
    _used_args = {}
    try:
        # 직전에 보여준 물건에 대한 '상세 질문'(예상낙찰가·시세·감정가·안전 등)은 검색하지 말고 대화 목록에서 답한다.
        _deep_q = any(_k in msg for _k in ("권리", "임차", "인수", "안전", "위험", "분석", "어때", "괜찮", "말소", "대항력", "세대수", "준공", "위반", "명세", "보증금"))
        # 필터/추림 의도('예상낙찰가 있는것', '있는건 뭐야', '싼것만', '어떤게') → 카드로 보여줘야 함(텍스트 X)
        _filter_intent = any(_k in msg for _k in ("있는", "없는", "인것", "인건", "인거", "만 ", "만이", "만인", "어떤", "뭐가", "뭐 있", "뭐있", "골라", "추려", "제외"))
        # 단일 값 질문('예상낙찰가 얼마', '몇 층') → 텍스트로 답
        _value_q = any(_k in msg for _k in ("얼마", "몇 층", "몇층", "몇 세대", "면적이", "언제", "며칠", "감정가는", "시세는"))
        _detail_q = any(_k in msg for _k in ("예상낙찰가", "낙찰가", "시세", "감정가", "최저가", "차익", "이거", "이 물건", "매수판정"))
        _search_intent = any(_k in msg for _k in ("추천", "찾아", "매물", "보여", "살만", "살 만", "투자할", "투자 가능", "다른 물건", "이 중", "그 중", "물건 있", "없어", "없나"))
        if _deep_q and not (_search_intent or _filter_intent):
            _tc = {"type": "function", "function": {"name": "get_property_detail"}}   # 특정 물건 심층(권리분석·임차인·안전) → 상세 조회
        elif (_value_q or _detail_q) and not _search_intent and not _filter_intent:
            _tc = "auto"                 # 단일 값 질문(시세·감정가·예상낙찰가 얼마) → 대화 목록에서 텍스트로 답
        else:
            _tc = ({"type": "function", "function": {"name": "search_properties"}}
                   if (_filter_intent or any(_k in msg for _k in ("물건", "추천", "매물", "투자", "구입", "찾아", "살 만", "나와", "보여")))
                   else "auto")
        # [게이트] 물건 검색/조회(도구 강제)를 프리미엄 미만이 요청하면 GPT 호출 없이 안내
        if _tc != "auto" and not _is_premium:
            return _PREMIUM_GATE
        r = httpx.post(url, headers=hdr,
                       json={"model": "gpt-4o-mini", "messages": msgs,
                             "tools": _CHAT_TOOLS, "tool_choice": _tc,
                             "temperature": 0.3, "max_tokens": 600}, timeout=45)
        if r.status_code != 200:
            if r.status_code in (401, 403):   # OpenAI 키 무효/폐기 — 관리자에게 명확히 안내
                return {"reply": "AI 상담 키 인증에 실패했습니다(키 무효/만료). 관리자 페이지 → 카카오 발송 탭에서 "
                                 "유효한 OpenAI 키를 다시 저장해 주세요."}
            if r.status_code == 429:
                return {"reply": "AI 상담 사용량이 한도에 도달했습니다(OpenAI 크레딧/요율). 잠시 후 다시 시도하거나 결제 상태를 확인해 주세요."}
            return {"reply": "죄송합니다. 답변 생성에 실패했습니다. 잠시 후 다시 시도해 주세요."}
        choice = ((r.json().get("choices") or [{}])[0].get("message", {})) or {}
        tcs = choice.get("tool_calls")
        if tcs and not _is_premium:       # [게이트] auto였는데 GPT가 물건 도구를 부른 경우도 차단
            return _PREMIUM_GATE
        if tcs:
            _detail_msgs = []             # get_property_detail 결과(2차 GPT로 자연어화)
            for tc in tcs:
                fn = (tc.get("function") or {}).get("name")
                if fn == "search_properties":
                    try:
                        a = _json.loads((tc.get("function") or {}).get("arguments") or "{}")
                    except Exception:
                        a = {}
                    # 맥락 유지: GPT가 유형·예산·지역을 빠뜨리면 직전 검색 조건(ctx)으로 보완.
                    #  단 '전국·지방 확대·어디든' 요청이면 지역(sido)은 일부러 물려받지 않는다(확대 의도 존중).
                    _expand = any(_k in msg for _k in ("전국", "지방", "확대", "어디든", "다른 지역", "타지역", "전지역", "전국구"))
                    for _ck in (("usages", "invest_max") if _expand else ("usages", "invest_max", "sido", "region", "regions")):
                        if not a.get(_ck) and _ctx.get(_ck):
                            a[_ck] = _ctx[_ck]
                    if _ref and _last:
                        result = _chat_filter_recent(_last, a)     # '이 중에서 ~' — 직전 목록에서만 필터(재검색 안 함)
                    else:
                        result = _chat_search_properties(a, _seen_keys)
                    _used_args = {k: a.get(k) for k in ("usages", "sido", "invest_max", "region", "regions", "price_max") if a.get(k)}
                    _cards = result.get("물건목록", []) or _cards
                elif fn == "get_property_detail":
                    try:
                        _a2 = _json.loads((tc.get("function") or {}).get("arguments") or "{}")
                    except Exception:
                        _a2 = {}
                    _dr = _chat_property_detail(_a2, _last)
                    _detail_msgs.append((tc, _dr.get("detail_text") or _dr.get("error") or "물건 정보를 불러오지 못했어요."))
            if _cards:
                # 물건은 카드로만 표시(GPT 재나열 방지) → 2차 생성 생략, 고정 안내.
                reply = "요청하신 조건에 맞는 물건 %d건이에요. 아래 카드에서 사건번호·감정가·최저가를 확인하시고, 카드를 누르면 물건 상세로 이동합니다." % len(_cards)
            elif _detail_msgs:
                # 조회한 물건 상세를 근거로 2차 GPT 호출 → 자연어 답변
                msgs.append({"role": "assistant", "content": None, "tool_calls": tcs})
                for _tc2, _dt in _detail_msgs:
                    msgs.append({"role": "tool", "tool_call_id": _tc2.get("id"), "content": _dt})
                try:
                    r2 = httpx.post(url, headers=hdr,
                                    json={"model": "gpt-4o-mini", "messages": msgs,
                                          "temperature": 0.3, "max_tokens": 700}, timeout=45)
                    reply = (((r2.json().get("choices") or [{}])[0].get("message", {})) or {}).get("content", "").strip()
                except Exception:
                    reply = ""
                if not reply:
                    reply = _detail_msgs[0][1]   # 2차 실패 시 조회한 상세 텍스트 그대로
            else:
                reply = "조건에 맞는 물건을 찾지 못했어요. 지역이나 예산·유형을 조금 바꿔서 다시 물어봐 주세요."
        else:
            reply = (choice.get("content") or "").strip()
        _list_text = ""
        if _cards:
            _parts = []
            for _i, _c in enumerate(_cards, 1):
                _eb = _c.get("예상낙찰가")
                _parts.append("%d) %s %s %s / 감정가%s만·최저가%s만·시세%s만·차익%s만·예상낙찰가%s·%s" % (
                    _i, _c.get("사건번호") or "", (_c.get("소재지") or "")[:18], _c.get("유형") or "",
                    round((_to_int(_c.get("감정가")) or 0) / 10000),
                    round((_to_int(_c.get("최저가")) or 0) / 10000),
                    round((_to_int(_c.get("시세")) or 0) / 10000),
                    round((_c.get("예상차익") or 0) / 10000),
                    ("%d만" % round(_eb / 10000) if _eb else "미산정"),
                    _c.get("매수판정") or ""))
            _list_text = " [방금 보여준 물건 %d건] %s" % (len(_cards), " / ".join(_parts))
        return {"reply": reply or "죄송합니다. 답변을 만들지 못했습니다. 다시 물어봐 주세요.",
                "cards": _cards, "used_args": _used_args, "list_text": _list_text}
    except Exception:
        return {"reply": "죄송합니다. 일시적 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."}


# ---------- AI 학습자료 관리(관리자) ----------
_AI_KB_DIR = os.path.join(_ROOT, "static", "data", "ai_kb")
_AI_KB_INDEX = os.path.join(_AI_KB_DIR, "_index.json")


def _ai_kb_load_index() -> dict:
    try:
        with open(_AI_KB_INDEX, encoding="utf-8") as f:
            return _json.load(f)
    except FileNotFoundError:
        return {}
    except Exception as e:
        import sys as _s
        print("[ai_kb] index load fail %s: %s %s" % (_AI_KB_INDEX, type(e).__name__, e), file=_s.stderr, flush=True)
        return {}


def _ai_kb_save_index(idx: dict) -> None:
    os.makedirs(_AI_KB_DIR, exist_ok=True)
    with open(_AI_KB_INDEX, "w", encoding="utf-8") as f:
        _json.dump(idx, f, ensure_ascii=False, indent=1)


def _load_ai_kb() -> str:
    """등록된 모든 학습자료를 제목 헤더와 함께 합쳐 반환(챗봇 컨텍스트)."""
    idx = _ai_kb_load_index()
    parts = []
    for slug, meta in idx.items():
        try:
            with open(os.path.join(_AI_KB_DIR, slug + ".txt"), encoding="utf-8") as f:
                txt = f.read().strip()
            if txt:
                parts.append("### [%s]\n%s" % (meta.get("title", slug), txt))
        except Exception:
            continue
    return "\n\n".join(parts)


def _openai_key() -> str:
    k = os.environ.get("OPENAI_API_KEY", "").strip()   # 클라우드=env var 우선(키를 repo에 안 올려 노출·폐기 방지)
    if k:
        return k
    try:
        st = _kb().load_state()                         # 로컬 폴백: kakao_broadcast.json(이제 gitignore됨)
        for kind in ("news", "upcoming", "sold"):
            k = (st.get(kind) or {}).get("openai_key")
            if k:
                return k
    except Exception:
        pass
    return ""


def _pdf_bytes_to_text(data: bytes) -> str:
    """PDF → 텍스트. 텍스트 레이어 있으면 그대로, 스캔본이면 GPT-4o Vision으로 추출."""
    import fitz
    doc = fitz.open(stream=data, filetype="pdf")
    plain = "\n".join((p.get_text() or "") for p in doc)
    if len(plain.strip()) > 60:
        return plain.strip()
    key = _openai_key()
    if not key:
        return ""
    import base64
    imgs = []
    for page in doc:
        pix = page.get_pixmap(dpi=140)
        b64 = base64.b64encode(pix.tobytes("png")).decode()
        imgs.append({"type": "image_url", "image_url": {"url": "data:image/png;base64," + b64, "detail": "high"}})
    if not imgs:
        return ""
    content = [{"type": "text", "text": "이 문서의 모든 내용을 빠짐없이 한국어 텍스트로 정리해줘. "
               "표·수치·조건·예시를 그대로 포함. AI 상담 답변 근거로 쓸 거야."}] + imgs
    try:
        r = httpx.post("https://api.openai.com/v1/chat/completions",
                       headers={"Authorization": "Bearer " + key, "Content-Type": "application/json"},
                       json={"model": "gpt-4o", "messages": [{"role": "user", "content": content}],
                             "temperature": 0.1, "max_tokens": 4000}, timeout=180)
        if r.status_code == 200:
            return (r.json()["choices"][0]["message"]["content"] or "").strip()
    except Exception:
        pass
    return ""


@app.post("/admin/ai_kb/upload")
async def ai_kb_upload(title: str = Form(...), file: UploadFile = File(...),
                       admin: dict = Depends(require_admin)) -> dict:
    data = await file.read()
    fname = (file.filename or "").lower()
    if fname.endswith(".pdf"):
        text = _pdf_bytes_to_text(data)
        if not text:
            raise HTTPException(400, "PDF에서 내용을 추출하지 못했습니다(스캔본은 카카오 설정에 GPT 키가 필요).")
    else:
        try:
            text = data.decode("utf-8")
        except Exception:
            text = data.decode("cp949", errors="ignore")
    text = (text or "").strip()
    if not text:
        raise HTTPException(400, "빈 자료입니다.")
    import re as _re
    base = _re.sub(r"[^0-9A-Za-z가-힣]+", "_", (title or "자료")).strip("_")[:40] or "kb"
    idx = _ai_kb_load_index()
    slug = base
    i = 1
    while slug in idx and idx[slug].get("title") != title:
        slug = "%s_%d" % (base, i); i += 1
    os.makedirs(_AI_KB_DIR, exist_ok=True)
    with open(os.path.join(_AI_KB_DIR, slug + ".txt"), "w", encoding="utf-8") as f:
        f.write(text)
    idx[slug] = {"title": title, "chars": len(text), "src": fname}
    _ai_kb_save_index(idx)
    return {"ok": True, "slug": slug, "chars": len(text)}


@app.get("/admin/ai_kb/list")
def ai_kb_list(admin: dict = Depends(require_admin)) -> dict:
    idx = _ai_kb_load_index()
    items = []
    for slug, meta in idx.items():
        preview = ""
        try:
            with open(os.path.join(_AI_KB_DIR, slug + ".txt"), encoding="utf-8") as f:
                preview = f.read(220)
        except Exception:
            pass
        items.append({"slug": slug, "title": meta.get("title", slug),
                      "chars": meta.get("chars", 0), "preview": preview})
    return {"ok": True, "items": items, "total_chars": sum(x["chars"] for x in items)}


@app.delete("/admin/ai_kb/{slug}")
def ai_kb_delete(slug: str, admin: dict = Depends(require_admin)) -> dict:
    idx = _ai_kb_load_index()
    if slug in idx:
        del idx[slug]
        _ai_kb_save_index(idx)
    try:
        os.remove(os.path.join(_AI_KB_DIR, slug + ".txt"))
    except Exception:
        pass
    return {"ok": True}


@app.get("/admin/ai_kb/{slug}/download")
def ai_kb_download(slug: str, admin: dict = Depends(require_admin)):
    """등록된 학습자료(추출·학습된 텍스트)를 .txt로 다운로드. 원본 PDF는 보관 안 하므로 텍스트 제공."""
    idx = _ai_kb_load_index()
    meta = idx.get(slug)
    if not meta:
        raise HTTPException(404, "자료를 찾을 수 없습니다.")
    try:
        with open(os.path.join(_AI_KB_DIR, slug + ".txt"), encoding="utf-8") as f:
            text = f.read()
    except Exception:
        raise HTTPException(404, "파일이 없습니다.")
    from urllib.parse import quote as _q
    dlname = _q((meta.get("title") or slug) + ".txt")   # 한글 파일명 RFC5987(filename*)
    return Response(content=text, media_type="text/plain; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=\"ai_kb.txt\"; filename*=UTF-8''%s" % dlname})


# ---------- 카카오 간편로그인 (OAuth2) ----------

@app.get("/auth/kakao/login")
def kakao_login():
    """카카오 인증 페이지로 리다이렉트. 키 미설정 시 안내."""
    if not KAKAO_REST_KEY:
        return RedirectResponse("/static/login.html?error=kakao_setup", status_code=303)
    url = (
        "https://kauth.kakao.com/oauth/authorize"
        f"?client_id={KAKAO_REST_KEY}"
        f"&redirect_uri={KAKAO_REDIRECT_URI}"
        "&response_type=code"
        # scope 미지정 → 콘솔에 설정된 동의항목을 그대로 사용(scope ID 불일치/KOE205 방지)
    )
    return RedirectResponse(url, status_code=303)


@app.get("/auth/kakao/callback")
def kakao_callback(code: str):
    """카카오 콜백: code → 토큰 → 회원정보 → 세션 발급 → 홈으로."""
    try:
        token_data = {
            "grant_type": "authorization_code",
            "client_id": KAKAO_REST_KEY,
            "redirect_uri": KAKAO_REDIRECT_URI,
            "code": code,
        }
        if KAKAO_CLIENT_SECRET:                 # 콘솔에서 클라이언트 시크릿 ON이면 포함
            token_data["client_secret"] = KAKAO_CLIENT_SECRET
        tok = httpx.post(
            "https://kauth.kakao.com/oauth/token",
            data=token_data,
            timeout=15,
        )
        tok.raise_for_status()
        access = tok.json()["access_token"]

        prof = httpx.get(
            "https://kapi.kakao.com/v2/user/me",
            headers={"Authorization": f"Bearer {access}"},
            timeout=15,
        )
        prof.raise_for_status()
        p = prof.json()
        kakao_id = str(p["id"])
        account = p.get("kakao_account", {}) or {}
        email = account.get("email")
        nickname = (
            (account.get("profile") or {}).get("nickname")
            or (p.get("properties") or {}).get("nickname")
            or "카카오회원"
        )
    except Exception:
        return RedirectResponse("/static/login.html?error=kakao", status_code=303)

    user = user_store.get_or_create_social_user("kakao", kakao_id, email, nickname)
    user = _maybe_make_admin(user)                    # 관리자 이름/이메일이면 자동 승격
    token = user_store.create_session(user["id"])
    resp = RedirectResponse("/", status_code=303)
    resp.set_cookie(_COOKIE, token, httponly=True, samesite="lax",
                    max_age=_COOKIE_MAXAGE, path="/")
    return resp


# ---------- 관심공매물건(온비드) ----------

class GongmaeFavIn(BaseModel):
    manage_no: str
    data: dict = {}
    folder: str = "기타"
    memo: str = ""


@app.get("/gongmae/favorites")
def list_gongmae_favorites(user: dict = Depends(require_user)) -> dict:
    rows = user_store.list_gongmae_favs(user["id"])
    items = []
    for r in rows:
        try:
            d = _json.loads(r.get("data") or "{}")
        except Exception:
            d = {}
        if not isinstance(d, dict):
            d = {}
        d["manage_no"] = r["manage_no"]
        d["folder"] = r.get("folder") or "기타"
        d["memo"] = r.get("memo") or ""
        d["created_at"] = r.get("created_at")
        items.append(d)
    return {"count": len(items), "items": items}


@app.post("/gongmae/favorites")
def add_gongmae_favorite(body: GongmaeFavIn, user: dict = Depends(require_user)) -> dict:
    if not body.manage_no:
        return {"ok": False, "error": "manage_no 필요"}
    user_store.add_gongmae_fav(user["id"], body.manage_no,
                               _json.dumps(body.data, ensure_ascii=False),
                               body.folder, body.memo)
    return {"ok": True}


@app.delete("/gongmae/favorites/{manage_no}")
def del_gongmae_favorite(manage_no: str, user: dict = Depends(require_user)) -> dict:
    user_store.remove_gongmae_fav(user["id"], manage_no)
    return {"ok": True}


@app.get("/gongmae/favorites/status")
def gongmae_favorite_status(manage_no: str, user: dict = Depends(require_user)) -> dict:
    return {"favorite": user_store.is_gongmae_fav(user["id"], manage_no)}


# ---------- 공매 즐겨쓰는검색(검색조건 저장) ----------

class GongmaeSearchIn(BaseModel):
    name: str = ""
    conditions: dict = {}


@app.get("/gongmae/searches")
def list_gongmae_searches(user: dict = Depends(require_user)) -> dict:
    """로그인 사용자의 저장된 공매 검색조건 목록."""
    out = []
    for r in user_store.list_gongmae_searches(user["id"]):
        try:
            cond = _json.loads(r.get("conditions") or "{}")
        except Exception:
            cond = {}
        if not isinstance(cond, dict):
            cond = {}
        out.append({"id": r["id"], "name": r.get("name") or "",
                    "conditions": cond, "created_at": r.get("created_at")})
    return {"count": len(out), "items": out}


@app.post("/gongmae/searches")
def add_gongmae_search(body: GongmaeSearchIn, user: dict = Depends(require_user)) -> dict:
    """공매 검색조건 저장(로그인 필수). conditions=폼 필터값 dict."""
    cond = body.conditions if isinstance(body.conditions, dict) else {}
    new_id = user_store.add_gongmae_search(
        user["id"], (body.name or "").strip(),
        _json.dumps(cond, ensure_ascii=False))
    return {"ok": True, "id": new_id}


@app.delete("/gongmae/searches/{search_id}")
def del_gongmae_search(search_id: int, user: dict = Depends(require_user)) -> dict:
    user_store.remove_gongmae_search(user["id"], search_id)
    return {"ok": True}


# ---------- 모의입찰(연습) ----------
_MOCKBID_CLOSED = ("매각", "배당", "납부", "취하", "기각", "정지", "미진행",
                   "대금미납", "변경", "불허", "각하", "낙찰", "종국", "취소")


def _mockbid_biddable(result: Optional[str]) -> bool:
    """모의입찰 가능 여부 — 진행중(신건/유찰/재진행/재매각/미정)=가능, 매각·변경 등 종결=불가."""
    r = (result or "").strip()
    if not r:
        return True
    if any(r.startswith(k) for k in ("신건", "유찰", "재진행", "재매각", "진행")):
        return True   # '재매각'을 '매각' 부분일치보다 먼저 판정
    if any(k in r for k in _MOCKBID_CLOSED):
        return False
    return True


def _item_result(item_key: str) -> Optional[str]:
    """items.result 한 건 조회(진행중 서버 재검증용)."""
    try:
        r = auction_db._get("items", [("select", "result"),
                                      ("item_key", f"eq.{item_key}"), ("limit", "1")])
        if r.status_code in (200, 206):
            rows = r.json()
            if rows:
                return rows[0].get("result")
    except Exception:
        pass
    return None


class MockBidIn(BaseModel):
    item_key: str
    bid_amount: int = 0
    data: dict = {}


@app.get("/mock_bids")
def list_mock_bids(user: dict = Depends(require_user)) -> dict:
    """내 모의입찰 목록(마이페이지)."""
    rows = user_store.list_mock_bids(user["id"])
    items = []
    for r in rows:
        try:
            d = _json.loads(r.get("data") or "{}")
        except Exception:
            d = {}
        if not isinstance(d, dict):
            d = {}
        d["item_key"] = r["item_key"]
        d["case_no"] = r.get("case_no") or d.get("case_no")
        d["bid_amount"] = r.get("bid_amount")
        d["created_at"] = r.get("created_at")
        items.append(d)
    return {"count": len(items), "items": items}


@app.post("/mock_bid")
def add_mock_bid(body: MockBidIn, user: dict = Depends(require_user)) -> dict:
    """모의입찰 등록 — 진행 중(매각 예정)인 물건만 허용."""
    if not body.item_key or not body.bid_amount or body.bid_amount <= 0:
        return {"ok": False, "error": "물건과 입찰가가 필요합니다."}
    result = _item_result(body.item_key)
    if not _mockbid_biddable(result):
        return {"ok": False, "error": "not_biddable",
                "message": "이미 매각되었거나 변경된 물건입니다. 매각이 진행 중인 물건만 모의입찰할 수 있습니다."}
    case_no = body.data.get("case_no") or "" if isinstance(body.data, dict) else ""
    user_store.add_mock_bid(user["id"], body.item_key, int(body.bid_amount),
                            case_no, _json.dumps(body.data, ensure_ascii=False))
    return {"ok": True}


@app.delete("/mock_bid")
def del_mock_bid(item_key: str, user: dict = Depends(require_user)) -> dict:
    user_store.remove_mock_bid(user["id"], item_key)
    return {"ok": True}


@app.get("/mock_bid/status")
def mock_bid_status(item_key: str, user: dict = Depends(require_user)) -> dict:
    """상세 페이지용 — 진행중 여부 + 내 기존 입찰가."""
    result = _item_result(item_key)
    return {"biddable": _mockbid_biddable(result), "result": result,
            "bid": user_store.get_mock_bid(user["id"], item_key)}


# ---------- 관심물건 ----------

class FavoriteIn(BaseModel):
    folder: str = "기타"          # 기본/개인 폴더명
    importance: int = 0           # 중요도 별점 0~5
    memo: str = ""
    notify: int = 1               # 알림/입찰달력 표시(1)/제외(0)


@app.get("/favorites")
def list_favorites(user: dict = Depends(require_user)) -> dict:
    rows = user_store.list_favorites_full(user["id"])      # 메타(폴더/중요도/메모/알림) 포함, 저장 역순
    items = auction_db.summaries_by_keys([r["case_no"] for r in rows])  # 실데이터 요약
    meta = {r["case_no"]: r for r in rows}
    for it in items:                                       # 요약에 관심물건 메타 병합(item_key 기준)
        m = meta.get(it.get("item_key"))
        if m:
            it["folder"] = m.get("folder") or "기타"
            it["importance"] = m.get("importance") or 0
            it["memo"] = m.get("memo") or ""
            it["notify"] = 1 if m.get("notify", 1) else 0
    return {"count": len(items), "items": items}


@app.post("/favorites/{case_no}")
def add_favorite(case_no: str, body: Optional[FavoriteIn] = None,
                 user: dict = Depends(require_user)) -> dict:
    # case_no 파라미터는 실제로 item_key. 모달에서 폴더/중요도/메모/알림 함께 저장(바디 없으면 기본값).
    b = body or FavoriteIn()
    user_store.add_favorite(user["id"], case_no, folder=b.folder,
                            importance=b.importance, memo=b.memo, notify=b.notify)
    return {"ok": True, "favorite": True}


@app.delete("/favorites/{case_no}")
def remove_favorite(case_no: str, user: dict = Depends(require_user)) -> dict:
    user_store.remove_favorite(user["id"], case_no)
    return {"ok": True, "favorite": False}


@app.get("/favorites/{case_no}/status")
def favorite_status(case_no: str, user: Optional[dict] = Depends(current_user)) -> dict:
    """이 물건이 관심물건인지(로그인 시) + 저장 메타(모달 채우기용). 비로그인은 favorite=false."""
    if not user:
        return {"logged_in": False, "favorite": False}
    meta = user_store.get_favorite(user["id"], case_no)
    return {"logged_in": True, "favorite": meta is not None, "meta": meta}


# ---------- 개인폴더 ----------

@app.get("/folders")
def list_user_folders(user: dict = Depends(require_user)) -> dict:
    return {"folders": user_store.list_folders(user["id"])}


@app.post("/folders")
def create_user_folder(body: dict = Body(...), user: dict = Depends(require_user)) -> dict:
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="폴더명을 입력하세요.")
    if len(name) > 30:
        raise HTTPException(status_code=400, detail="폴더명은 30자 이내로 입력하세요.")
    user_store.add_folder(user["id"], name)
    return {"ok": True, "folders": user_store.list_folders(user["id"])}


@app.post("/folders/rename")
def rename_user_folder(body: dict = Body(...), user: dict = Depends(require_user)) -> dict:
    old = (body.get("old") or "").strip()
    new = (body.get("new") or "").strip()
    if not new:
        raise HTTPException(status_code=400, detail="새 폴더명을 입력하세요.")
    if len(new) > 30:
        raise HTTPException(status_code=400, detail="폴더명은 30자 이내로 입력하세요.")
    if not user_store.rename_folder(user["id"], old, new):
        raise HTTPException(status_code=400, detail="이미 있는 폴더명이거나 변경할 수 없습니다.")
    return {"ok": True, "folders": user_store.list_folders(user["id"])}


@app.post("/folders/reorder")
def reorder_user_folders(body: dict = Body(...), user: dict = Depends(require_user)) -> dict:
    names = body.get("names")
    if not isinstance(names, list):
        raise HTTPException(status_code=400, detail="names는 배열이어야 합니다.")
    user_store.reorder_folders(user["id"], [str(n) for n in names])
    return {"ok": True, "folders": user_store.list_folders(user["id"])}


@app.delete("/folders/{name}")
def delete_user_folder(name: str, user: dict = Depends(require_user)) -> dict:
    user_store.remove_folder(user["id"], name)            # 그 폴더 물건은 '기타'로 이동
    return {"ok": True, "folders": user_store.list_folders(user["id"])}


@app.get("/properties")
def list_properties(
    type: Optional[list[ResidentialType]] = Query(None, description="현황용도(다중 선택)"),
    region: Optional[Region] = None,
    court: Optional[str] = None,
    keyword: Optional[str] = Query(None, description="소재지/명칭 검색"),
    auction_type: Optional[str] = Query(None, description="임의경매/강제경매"),
    min_price: Optional[int] = None,
    max_price: Optional[int] = None,
    appraisal_min: Optional[int] = None,
    appraisal_max: Optional[int] = None,
    failed_min: Optional[int] = None,
    failed_max: Optional[int] = None,
    building_area_min: Optional[float] = None,
    building_area_max: Optional[float] = None,
    sale_from: Optional[str] = None,
    sale_to: Optional[str] = None,
    sort: str = "매각기일",
    limit: int = Query(100, le=500),
    offset: int = 0,
) -> dict:
    common = dict(
        types=type, region=region, court=court, keyword=keyword,
        auction_type=auction_type, min_price=min_price, max_price=max_price,
        appraisal_min=appraisal_min, appraisal_max=appraisal_max,
        failed_min=failed_min, failed_max=failed_max,
        building_area_min=building_area_min, building_area_max=building_area_max,
        sale_from=sale_from, sale_to=sale_to, sort=sort,
    )
    # 통계는 전체 매칭 기준, 목록은 페이지 단위
    matching = store.search(**common, limit=100000, offset=0)
    page = matching[offset: offset + limit]
    return {
        "count": len(matching),
        "stats": compute_stats(matching),
        "items": [listing_summary(l) for l in page],
    }


@app.get("/properties/{case_no}")
def property_detail(case_no: str) -> dict:
    listing = store.get(case_no)
    if listing is None:
        raise HTTPException(404, f"물건을 찾을 수 없습니다: {case_no}")

    prop = listing.to_property()
    analysis = analyze(prop)
    # 배당은 최저매각가를 가정 매각가로 사용(실낙찰가 미정)
    dist = calculate_distribution(
        prop, sale_price=listing.min_bid_price, region=listing.region,
    )
    return {
        "listing": listing_summary(listing),
        "analysis": analysis_to_dict(analysis),
        "distribution": {
            "assumed_sale_price": listing.min_bid_price,
            "note": "매각가는 최저매각가 가정. 실낙찰가로 재계산 필요.",
            **distribution_to_dict(dist),
        },
    }


@app.get("/properties/{case_no}/distribution")
def property_distribution(
    case_no: str,
    sale_price: int = Query(..., description="가정 매각가(원)"),
    execution_cost: int = 0,
) -> dict:
    """매각가를 바꿔 배당을 재계산(상세 화면 슬라이더용)."""
    listing = store.get(case_no)
    if listing is None:
        raise HTTPException(404, f"물건을 찾을 수 없습니다: {case_no}")
    dist = calculate_distribution(
        listing.to_property(), sale_price=sale_price,
        region=listing.region, execution_cost=execution_cost,
    )
    return {"assumed_sale_price": sale_price, **distribution_to_dict(dist)}


# ---------- POST /analyze ----------

class RightIn(BaseModel):
    type: str = Field(..., description="권리종류(예: 근저당권, 가압류, 소유권이전청구권가등기)")
    reg_date: date
    holder: str = ""
    amount: int = 0
    note: str = ""


class TenantIn(BaseModel):
    name: str = ""
    move_in_date: Optional[date] = None
    fixed_date: Optional[date] = None
    deposit: int = 0
    demanded_distribution: bool = False
    occupying: bool = True


class AnalyzeIn(BaseModel):
    case_no: str = "직접입력"
    rights: list[RightIn] = []
    tenants: list[TenantIn] = []
    sale_price: Optional[int] = None
    region: Region = Region.OTHER
    execution_cost: int = 0


@app.post("/analyze")
def analyze_endpoint(body: AnalyzeIn) -> dict:
    try:
        rights = [
            Right(type=RightType(r.type), reg_date=r.reg_date, holder=r.holder,
                  amount=r.amount, note=r.note)
            for r in body.rights
        ]
    except ValueError as e:
        raise HTTPException(422, f"권리종류 값 오류: {e}")

    tenants = [
        Tenant(name=t.name, move_in_date=t.move_in_date, fixed_date=t.fixed_date,
               deposit=t.deposit, demanded_distribution=t.demanded_distribution,
               occupying=t.occupying)
        for t in body.tenants
    ]
    prop = AuctionProperty(case_no=body.case_no, rights=rights, tenants=tenants)

    out: dict = {"analysis": analysis_to_dict(analyze(prop))}
    if body.sale_price is not None:
        dist = calculate_distribution(
            prop, sale_price=body.sale_price, region=body.region,
            execution_cost=body.execution_cost,
        )
        out["distribution"] = distribution_to_dict(dist)
    return out


# ───────── KB 부동산DB(kb_crawler) 통합 — 유료회원 전용 ─────────
try:
    import sys as _kb_sys
    _kb_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _kb_root not in _kb_sys.path:
        _kb_sys.path.insert(0, _kb_root)
    from kb_crawler import router as _kb_router
    if _kb_router is not None:
        app.include_router(_kb_router, dependencies=[Depends(require_national_user)])
        print("[kb] 부동산DB 라우터 통합 완료(/kb/*, 유료회원 전용)", flush=True)
    else:
        print("[kb] kb_crawler router=None (fastapi 미로드?) — 통합 스킵", flush=True)
except Exception as _kb_e:
    print(f"[kb] 부동산DB 라우터 통합 실패: {_kb_e}", flush=True)


@app.get("/kb/resolve_region")
def kb_resolve_region(address: str, _u: dict = Depends(require_national_user)) -> dict:
    """주소 → 법정동 10자리 코드(kb 지역수집 lawd_code용). 유료회원 전용."""
    try:
        from auction_analysis.bjd_codes import resolve_bjd
        r = resolve_bjd(address)
    except Exception:
        r = None
    if not r:
        return {"address": address, "lawd_code": None, "ok": False}
    return {"address": address, "lawd_code": r[0] + r[1], "ok": True}


# ───────── KB 인증 하이브리드 — 로컬=발급기(카카오 자동로그인), Supabase=공유, 서버=소비 ─────────
def _kb_issue_token() -> bool:
    """발급기(KB_EMAIL/KB_PW 있는 로컬): 별도 프로세스(kb_kakao_login.py)로 카카오 자동로그인
    (playwright sync API를 서버 스레드에서 직접 돌리면 멈추므로 subprocess로 분리) →
    siteToken+캡처헤더 → api_cache 저장 + AUTH 적용. 비번은 로컬에만, 클라우드엔 토큰만 전달됨."""
    try:
        import subprocess
        import sys as _sys
        import json as _json
        import time as _t
        import datetime as _dt
        import kb_crawler as _kb
        _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        helper = os.path.join(_root, "kb_kakao_login.py")
        r = subprocess.run([_sys.executable, helper], capture_output=True, text=True,
                           encoding="utf-8", errors="replace", timeout=180, cwd=_root)
        out = (r.stdout or "").strip()
        line = out.splitlines()[-1] if out else ""
        data = _json.loads(line) if line.startswith("{") else {}
        tok = data.get("token")
        hdrs = data.get("headers") or None
        if not tok:
            print(f"[kb] 토큰 발급 실패(로그인 미완/캡차). stderr: {(r.stderr or '')[-300:]}", flush=True)
            return False
        _kb.AUTH.token = tok
        _kb.AUTH.captured_headers = hdrs
        _kb.AUTH.obtained_at = _t.monotonic()
        try:
            auction_db.cache_save("kb:auth", {"token": tok, "headers": hdrs or {}, "at": _dt.datetime.now().isoformat()})
        except Exception as e:
            print(f"[kb] 토큰 Supabase 저장 실패: {e}", flush=True)
        print("[kb] 토큰 발급·저장 완료(캡처헤더 %s)" % ("있음" if hdrs else "없음"), flush=True)
        return True
    except Exception as e:
        print(f"[kb] 토큰 발급 오류: {e}", flush=True)
        return False


def _kb_apply_token() -> bool:
    """소비자(모든 서버): Supabase api_cache에서 토큰 읽어 AUTH 적용.
    refresh_token이 저장돼 있으면 siteToken 만료 임박(2h 전) 시 카카오 없이 자동 재발급 후 재저장."""
    try:
        import kb_crawler as _kb
        import time as _t
        import datetime as _dt
        d = (auction_db.cache_get_many(["kb:auth"]) or {}).get("kb:auth")
        if not d:
            return False
        rt = d.get("refresh_token")
        if rt:                                   # 만료 임박/토큰없음 → refreshToken으로 자동 재발급
            need = not d.get("token")
            try:
                age = (_dt.datetime.now() - _dt.datetime.fromisoformat(d.get("at"))).total_seconds()
                exp = int(d.get("expires_in") or 0)
                if exp and age > exp - 7200:     # 만료 2시간 전부터 갱신
                    need = True
            except Exception:
                pass
            if need:
                nu = _kb.refresh_site_token(rt)
                if nu:
                    d = {"token": nu["token"], "refresh_token": nu["refresh_token"],
                         "expires_in": nu["expires_in"], "headers": d.get("headers") or {},
                         "at": _dt.datetime.now().isoformat()}
                    try:
                        auction_db.cache_save("kb:auth", d)
                    except Exception:
                        pass
                    print("[kb] refreshToken으로 siteToken 자동 재발급(카카오 불필요)", flush=True)
        if not d.get("token"):
            return False
        _kb.AUTH.token = d["token"]
        _kb.AUTH.captured_headers = d.get("headers") or None
        _kb.AUTH.obtained_at = _t.monotonic()
        return True
    except Exception as e:
        print(f"[kb] 토큰 적용 오류: {e}", flush=True)
        return False


def _kb_auth_loop() -> None:
    """KB_EMAIL/KB_PW 있으면 발급기(30분 주기 재로그인), 없으면 소비자(5분 주기 로드)."""
    import time as _t
    is_issuer = bool(os.environ.get("KB_EMAIL") and os.environ.get("KB_PW"))
    print(f"[kb] 인증 루프 시작: {'발급기(로컬 카카오 자동로그인)' if is_issuer else '소비자(토큰 로드)'}", flush=True)
    while True:
        try:
            if is_issuer:
                _kb_issue_token()
                _t.sleep(30 * 60)
            else:
                _kb_apply_token()
                _t.sleep(5 * 60)
        except Exception as e:
            print(f"[kb] 인증 루프 오류: {e}", flush=True)
            _t.sleep(60)


@app.post("/admin/kb/set_token")
def admin_kb_set_token(body: dict = Body(...), admin: dict = Depends(require_admin)) -> dict:
    """KB siteToken/refreshToken을 Supabase(api_cache kb:auth) 공유 저장 + 즉시 적용.
    refreshToken을 주면 그걸로 새 siteToken을 발급(검증 겸)하고, 이후 만료 임박 시 자동 재발급된다
    (카카오 재로그인 불필요). siteToken만 주면 종전과 동일(자동갱신 없음)."""
    tok = (body.get("token") or "").strip().strip("'\"").strip()   # 콘솔 복사 시 딸려온 따옴표 제거
    rt = (body.get("refresh_token") or "").strip().strip("'\"").strip()
    import kb_crawler as _kb
    import time as _t
    import datetime as _dt
    exp = 0
    if rt:                                   # refreshToken 우선 — 새 siteToken 발급(무효면 즉시 알림)
        nu = _kb.refresh_site_token(rt)
        if not nu:
            raise HTTPException(400, "refreshToken이 무효/만료입니다. kbland.kr 재로그인 후 다시 복사해 주세요.")
        tok = nu["token"]; rt = nu["refresh_token"]; exp = nu["expires_in"]
    if not tok:
        raise HTTPException(400, "siteToken 또는 refreshToken을 입력하세요.")
    _kb.AUTH.token = tok
    _kb.AUTH.obtained_at = _t.monotonic()
    rec = {"token": tok, "headers": {}, "at": _dt.datetime.now().isoformat()}
    if rt:
        rec["refresh_token"] = rt
        rec["expires_in"] = exp
    try:
        auction_db.cache_save("kb:auth", rec)
    except Exception as e:
        raise HTTPException(500, f"Supabase 저장 실패: {e}")
    return {"ok": True, "saved": True, "auto_refresh": bool(rt),
            "expires_hours": round(exp / 3600, 1) if exp else None}


@app.get("/admin/kb/token_status")
def admin_kb_token_status(admin: dict = Depends(require_admin)) -> dict:
    """현재 토큰 상태(저장여부·출처·마지막 갱신·끝 6자리). 값 전체는 노출 안 함."""
    import kb_crawler as _kb
    d = (auction_db.cache_get_many(["kb:auth"]) or {}).get("kb:auth") or {}
    env_tok = os.environ.get("KB_SITE_TOKEN", "")
    cur = _kb.AUTH.token or d.get("token") or env_tok
    src = "Supabase(공유)" if d.get("token") else (".env(로컬)" if env_tok else ("메모리" if _kb.AUTH.token else ""))
    return {"has_token": bool(cur), "source": src, "saved_at": d.get("at", ""), "tail": (cur[-6:] if cur else "")}


@app.post("/admin/kb/test_token")
def admin_kb_test_token(admin: dict = Depends(require_admin)) -> dict:
    """현재 토큰으로 KB 서명 API를 실제 호출해 유효/만료 판정."""
    import kb_crawler as _kb
    try:
        cands = _kb.kb_search("아파트", 5)
        cno = next((c.get("COMPLEX_NO") for c in cands if c.get("COMPLEX_NO")), None)
        if not cno:
            return {"valid": None, "msg": "단지 검색 실패(네트워크 확인)"}
        data = _kb.kb_list_complex(cno, trade_code="1", page=1, size=1)
        ok = bool(data) and (data.get("propertyList") is not None or data.get("페이지개수") is not None)
        return {"valid": bool(ok), "msg": ("유효 (정상 응답)" if ok else "만료/무효 — 새 토큰이 필요합니다")}
    except Exception as e:
        return {"valid": False, "msg": f"만료/오류: {str(e)[:120]}"}


@app.get("/kb/regions")
def kb_regions(sido: str = "", sigungu: str = "", _u: dict = Depends(require_national_user)) -> dict:
    """지역 드롭다운: sido 없으면 시도목록, sido만 있으면 시군구목록, 둘 다면 읍면동(+법정동코드)."""
    from auction_analysis.bjd_codes import _load
    mp = _load()
    if not sido:
        sido_code = {}
        for k, v in mp.items():
            toks = k.split()
            if toks and toks[0] not in sido_code:
                sido_code[toks[0]] = v[:2]        # 시도 법정동코드 앞2자리
        # 서울(11)→광역시(26~31)→세종(36)→경기·도(41~) 순 = 코드 오름차순
        return {"level": "sido", "items": sorted(sido_code, key=lambda s: sido_code[s])}
    if not sigungu:
        base = sido + " "
        sggs = set()
        for k in mp:
            if k.startswith(base):
                sgg = []
                for t in k[len(base):].split():
                    if t.endswith(("시", "군", "구")):
                        sgg.append(t)            # 성남시 분당구처럼 시+구 연속 허용
                    else:
                        break                    # 읍/면/동/리 나오면 시군구 끝
                if sgg:
                    sggs.add(" ".join(sgg))
        return {"level": "sigungu", "items": sorted(sggs)}
    prefix = sido + " " + sigungu + " "
    plen = len(prefix.split())
    emds = {}
    for k, v in mp.items():
        if k.startswith(prefix):
            toks = k.split()
            if len(toks) == plen + 1:        # 읍/면/동 레벨(리 제외)
                emds[toks[-1]] = v
    return {"level": "emd", "items": [{"name": n, "code": c} for n, c in sorted(emds.items())]}


@app.post("/kb/export_brokers")
def kb_export_brokers(body: dict = Body(...), _u: dict = Depends(require_national_user)):
    """중개업소 리스트 → 열 너비 자동 맞춤된 xlsx 다운로드."""
    import io
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    from fastapi.responses import StreamingResponse
    brokers = body.get("brokers") or []
    cols = [("업소명", "중개업소명"), ("대표자", "중개업소대표자명"), ("주소", "중개업소주소"),
            ("전화", "중개업소전화번호"), ("휴대폰", "중개업소대표자휴대폰번호")]

    def vwidth(s):
        return sum(2 if ord(ch) > 0x2000 else 1 for ch in str(s or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "중개업소"
    ws.append([c[0] for c in cols])
    for b in brokers:
        ws.append([str(b.get(c[1]) or "") for c in cols])
    for i, c in enumerate(cols, 1):
        w = vwidth(c[0])
        for b in brokers:
            w = max(w, vwidth(b.get(c[1])))
        ws.column_dimensions[get_column_letter(i)].width = min(w + 2, 60)   # 열 너비 자동
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = quote(f"KB중개업소_{len(brokers)}곳.xlsx")
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fn}"})


@app.post("/kb/collect/region_page")
def kb_collect_region_page(body: dict = Body(...), _u: dict = Depends(require_national_user)) -> dict:
    """지역 매물 '한 페이지(30건)'만 수집 — 프론트가 페이지별로 호출해 진행 표시. page 1에서 좌표 반환→이후 재사용."""
    import kb_crawler as _kb
    address = body.get("address")
    lawd = body.get("lawd_code")
    ptypes = body.get("property_types", "01")
    ttypes = body.get("transaction_types", "1")
    lat = body.get("lat")
    lng = body.get("lng")
    page = int(body.get("page", 1) or 1)
    if not lawd:
        raise HTTPException(400, "lawd_code가 필요합니다.")
    if lat is None or lng is None:
        geo = _kb.kb_geocode(address or "")
        if not geo:
            raise HTTPException(400, "좌표를 찾지 못했습니다(주소 확인).")
        lat, lng = geo
    headers = _kb.AUTH.get_captured_headers() or _kb._signed_headers(_kb.AUTH.get_token())
    payload = _kb._region_payload(lat, lng, lawd, ptypes, ttypes, page, 30)
    r = _kb._api("POST", _kb.PROP_REGION_URL, ctx=f"page:{lawd}:{page}", json=payload, timeout=20, headers=headers)
    data = (_kb._body(r, "page").get("dataBody") or {}).get("data") or {}
    props = data.get("propertyList") or []
    total = _kb._to_int(data.get("총매물건수")) or 0
    return {"lat": lat, "lng": lng, "page": page, "total": total, "count": len(props), "properties": props}


# 정적 화면 서빙: /static/detail.html?case_no=... (상세 권리분석·배당 + 매각가 슬라이더)
app.mount("/static", StaticFiles(directory=_STATIC_DIR, html=True), name="static")
