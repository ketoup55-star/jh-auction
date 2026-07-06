"""KB Land(kbland.kr) 매물 수집기 GUI — 디컴파일(바이트코드 역분석)로 복원한 소스.

원본: kb+부동산+수집.exe (PyInstaller, Python 3.13) 내부 ui_app.pyc
주의: 변수/함수명·문자열 리터럴은 원본 그대로 복원됨. 주석/공백/일부 표현은 추정 재구성.
"""
from __future__ import annotations

import queue
import re
import sys
import threading
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

try:
    from kmong.kb2.fetch import FetchService
    from kmong.kb2.service import Service
except ModuleNotFoundError:
    from fetch import FetchService
    from service import Service


PROJECT_ROOT = Path(__file__).resolve().parents[2]

EXCEL_COLUMNS = [
    '중개업소명',
    '중개업소대표자명',
    '중개업소주소',
    '중개업소전화번호',
    '중개업소대표자휴대폰번호',
]


def get_excel_output_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).resolve().parent
    return PROJECT_ROOT / 'output'


PROPERTY_TYPE_GROUPS = [
    (
        'apartment_officetel',
        '아파트/오피스텔',
        [
            ('01', '아파트'),
            ('05', '아파트분양권'),
            ('41', '아파트재건축'),
            ('04', '오피스텔'),
            ('07', '오피스텔분양권'),
        ],
    ),
    (
        'house',
        '빌라/주택',
        [
            ('08', '빌라'),
            ('38', '다가구주택'),
            ('09', '단독주택'),
            ('10', '전원주택'),
            ('11', '상가주택'),
        ],
    ),
    (
        'room',
        '원룸/투룸',
        [
            ('34', '원룸'),
            ('35', '투룸'),
        ],
    ),
    (
        'redevelopment',
        '재개발',
        [
            ('27', '재개발'),
        ],
    ),
    (
        'etc',
        '기타',
        [
            ('16', '상가'),
            ('19', '사무실'),
            ('21', '지식산업센터'),
            ('23', '건물'),
            ('20', '공장'),
            ('22', '창고'),
            ('28', '토지'),
            ('43', '생활숙박시설'),
        ],
    ),
]

PROPERTY_TYPE_OPTIONS = [
    option
    for _group_code, _group_label, group_options in PROPERTY_TYPE_GROUPS
    for option in group_options
]

TRANSACTION_TYPE_OPTIONS = [
    ('1', '매매'),
    ('2', '전세'),
    ('3', '월세'),
]


@dataclass(frozen=True)
class CollectionJob:
    search_keyword: str
    dong_name: str
    lawd_code: str
    property_type: str
    property_type_label: str
    transaction_type: str
    transaction_type_label: str

    @property
    def key(self) -> tuple[str, str, str]:
        return (self.lawd_code, self.property_type, self.transaction_type)

    @property
    def display_text(self) -> str:
        return f"{self.search_keyword} / {self.property_type_label} / {self.transaction_type_label}"


class ScrollableSelect(ctk.CTkFrame):
    def __init__(
        self,
        parent,
        variable: ctk.StringVar,
        command: Callable[[str], None] | None = None,
        values: list[str] | None = None,
        max_dropdown_height: int = 280,
    ):
        super().__init__(parent, fg_color='transparent')

        self.variable = variable
        self.command = command
        self.values = values or ['']
        self.max_dropdown_height = max_dropdown_height
        self.dropdown = None
        self.enabled = True

        self.grid_columnconfigure(0, weight=1)

        self.button = ctk.CTkButton(
            self,
            text='',
            command=self._toggle_dropdown,
            height=42,
            anchor='w',
            font=ctk.CTkFont(size=14),
        )
        self.button.grid(row=0, column=0, sticky='ew')

        self.arrow_label = ctk.CTkLabel(
            self,
            text='v',
            width=34,
            font=ctk.CTkFont(size=14, weight='bold'),
        )
        self.arrow_label.grid(row=0, column=0, sticky='e', padx=(0, 8))
        self.arrow_label.bind('<Button-1>', lambda _event: self._toggle_dropdown())
        self.arrow_label.lift()

        self.variable.trace_add('write', self._sync_button_text)
        self._sync_button_text()

    def set_values(self, values: list[str]):
        self.values = values or ['']
        self._close_dropdown()

    def set_enabled(self, enabled: bool):
        self.enabled = enabled
        self.button.configure(state='normal' if enabled else 'disabled')
        if not enabled:
            self._close_dropdown()

    def _sync_button_text(self, *_args):
        value = self.variable.get()
        self.button.configure(text=value or '선택 없음')

    def _toggle_dropdown(self):
        if not self.enabled:
            return
        if self.dropdown and self.dropdown.winfo_exists():
            self._close_dropdown()
            return
        self._open_dropdown()

    def _open_dropdown(self):
        self.update_idletasks()

        width = max(self.winfo_width(), 260)
        row_height = 36
        dropdown_height = min(
            self.max_dropdown_height,
            max(48, len(self.values) * row_height + 12),
        )

        x = self.winfo_rootx()
        y = self.winfo_rooty() + self.winfo_height() + 4

        self.dropdown = ctk.CTkToplevel(self)
        self.dropdown.overrideredirect(True)
        self.dropdown.geometry(f"{width}x{dropdown_height}+{x}+{y}")
        self.dropdown.lift()
        self.dropdown.attributes('-topmost', True)
        self.dropdown.after(100, lambda: self.dropdown.attributes('-topmost', False))
        self.dropdown.bind('<Escape>', lambda _event: self._close_dropdown())

        container = ctk.CTkFrame(self.dropdown, corner_radius=8)
        container.pack(fill='both', expand=True)

        scroll_frame = ctk.CTkScrollableFrame(container, corner_radius=8)
        scroll_frame.pack(fill='both', expand=True, padx=4, pady=4)
        scroll_frame.grid_columnconfigure(0, weight=1)

        selected_value = self.variable.get()
        for row, value in enumerate(self.values):
            option_button = ctk.CTkButton(
                scroll_frame,
                text=value or '선택 없음',
                command=lambda selected=value: self._select_value(selected),
                height=32,
                anchor='w',
                font=ctk.CTkFont(size=13),
                fg_color='#3B8ED0' if value == selected_value else 'transparent',
                hover_color=('#36719F', '#144870'),
            )
            option_button.grid(row=row, column=0, sticky='ew', padx=2, pady=1)

        self.dropdown.focus_force()

    def _select_value(self, value: str):
        self.variable.set(value)
        self._close_dropdown()
        if self.command:
            self.command(value)

    def _close_dropdown(self):
        if self.dropdown and self.dropdown.winfo_exists():
            self.dropdown.destroy()
        self.dropdown = None


class LawdSelectApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode('dark')
        ctk.set_default_color_theme('blue')

        self.service = Service()
        self.fetch_service = FetchService()
        self.log_queue = queue.Queue()
        self.is_collecting = False
        self.collection_jobs = []

        self.title('부동산 지역 선택')
        self.geometry('1120x760')
        self.minsize(920, 620)

        self.email_var = ctk.StringVar()
        self.password_var = ctk.StringVar()
        self.browser_visible_var = ctk.BooleanVar(value=False)
        self.merge_excel_var = ctk.BooleanVar(value=False)
        self.sido_var = ctk.StringVar()
        self.sigungu_var = ctk.StringVar()
        self.eupmyeondong_var = ctk.StringVar()
        self.all_eupmyeondong_var = ctk.BooleanVar(value=False)
        self.property_type_group_var = ctk.StringVar(value=PROPERTY_TYPE_GROUPS[0][0])

        initial_property_type_codes = self._get_property_type_group_codes(
            self.property_type_group_var.get()
        )

        self.property_type_vars = {
            code: ctk.BooleanVar(value=code in initial_property_type_codes)
            for code, _label in PROPERTY_TYPE_OPTIONS
        }

        self.property_type_checkboxes = {}

        self.transaction_type_vars = {
            code: ctk.BooleanVar(value=True)
            for code, _label in TRANSACTION_TYPE_OPTIONS
        }

        self._setup_ui()
        self._load_initial_values()
        self._process_log_queue()

    def _setup_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        main_frame = ctk.CTkScrollableFrame(self, corner_radius=8)
        main_frame.grid(row=0, column=0, sticky='nsew', padx=20, pady=20)
        main_frame.grid_columnconfigure(0, weight=1)

        login_frame = ctk.CTkFrame(main_frame)
        login_frame.grid(row=0, column=0, sticky='ew', padx=14, pady=(14, 10))
        login_frame.grid_columnconfigure(1, weight=1)
        login_frame.grid_columnconfigure(3, weight=1)
        login_frame.grid_columnconfigure(4, weight=0)
        login_frame.grid_columnconfigure(5, weight=0)

        self.email_entry = self._create_inline_entry(
            login_frame,
            label_column=0,
            entry_column=1,
            label_text='카카오 이메일',
            variable=self.email_var,
            placeholder_text='카카오 이메일 입력',
        )

        self.password_entry = self._create_inline_entry(
            login_frame,
            label_column=2,
            entry_column=3,
            label_text='카카오 비밀번호',
            variable=self.password_var,
            placeholder_text='카카오 비밀번호 입력',
            show='*',
        )

        self.browser_visible_checkbox = ctk.CTkCheckBox(
            login_frame,
            text='브라우저 보기',
            variable=self.browser_visible_var,
            onvalue=True,
            offvalue=False,
            font=ctk.CTkFont(size=13),
        )
        self.browser_visible_checkbox.grid(
            row=0, column=4, sticky='w', padx=(0, 12), pady=14
        )

        self.merge_excel_checkbox = ctk.CTkCheckBox(
            login_frame,
            text='엑셀 합치기',
            variable=self.merge_excel_var,
            onvalue=True,
            offvalue=False,
            font=ctk.CTkFont(size=13),
        )
        self.merge_excel_checkbox.grid(
            row=0, column=5, sticky='w', padx=(0, 18), pady=14
        )

        settings_frame = ctk.CTkFrame(main_frame, fg_color='transparent')
        settings_frame.grid(row=1, column=0, sticky='ew', padx=14, pady=(0, 10))
        for column in range(2):
            settings_frame.grid_columnconfigure(column, weight=1, uniform='settings')

        region_frame = ctk.CTkFrame(settings_frame)
        region_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 7))
        region_frame.grid_columnconfigure(1, weight=1)

        filter_frame = ctk.CTkFrame(settings_frame)
        filter_frame.grid(row=0, column=1, sticky='nsew', padx=(7, 0))
        filter_frame.grid_columnconfigure(1, weight=1)

        self._create_section_title(region_frame, row=0, text='지역')
        self.sido_select = self._create_select_row(
            region_frame,
            row=1,
            label_text='시도',
            variable=self.sido_var,
            command=self._on_sido_change,
        )
        self.sigungu_select = self._create_select_row(
            region_frame,
            row=2,
            label_text='시군구',
            variable=self.sigungu_var,
            command=self._on_sigungu_change,
        )
        self.eupmyeondong_select = self._create_select_row(
            region_frame,
            row=3,
            label_text='읍면동',
            variable=self.eupmyeondong_var,
            command=self._on_eupmyeondong_change,
        )

        self.all_eupmyeondong_checkbox = ctk.CTkCheckBox(
            region_frame,
            text='읍면동 전체',
            variable=self.all_eupmyeondong_var,
            command=self._on_all_eupmyeondong_toggle,
            onvalue=True,
            offvalue=False,
            font=ctk.CTkFont(size=13),
        )
        self.all_eupmyeondong_checkbox.grid(
            row=4, column=1, sticky='w', padx=(0, 24), pady=(0, 6)
        )

        self.selected_label = ctk.CTkLabel(
            region_frame,
            text='',
            anchor='w',
            font=ctk.CTkFont(size=13),
        )
        self.selected_label.grid(
            row=5, column=0, columnspan=2, sticky='ew', padx=24, pady=(8, 14)
        )

        self._create_section_title(filter_frame, row=0, text='필터')
        self.property_type_frame = self._create_property_type_group(filter_frame, row=1)

        self.transaction_type_frame = self._create_checkbox_group(
            filter_frame,
            row=2,
            label_text='거래유형',
            options=TRANSACTION_TYPE_OPTIONS,
            variables=self.transaction_type_vars,
            columns=3,
        )

        action_frame = ctk.CTkFrame(main_frame, fg_color='transparent')
        action_frame.grid(row=2, column=0, sticky='ew', padx=14, pady=(0, 10))
        action_frame.grid_columnconfigure(0, weight=1)
        action_frame.grid_columnconfigure(1, weight=1)
        action_frame.grid_columnconfigure(2, weight=1)

        self.add_job_button = ctk.CTkButton(
            action_frame,
            text='설정 등록',
            command=self._add_collection_job,
            height=42,
            font=ctk.CTkFont(size=15, weight='bold'),
        )
        self.add_job_button.grid(row=0, column=0, sticky='ew', padx=(0, 6))

        self.clear_jobs_button = ctk.CTkButton(
            action_frame,
            text='목록 비우기',
            command=self._clear_collection_jobs,
            height=42,
            font=ctk.CTkFont(size=15, weight='bold'),
            fg_color='#5A6472',
            hover_color='#49515C',
        )
        self.clear_jobs_button.grid(row=0, column=1, sticky='ew', padx=6)

        self.collect_button = ctk.CTkButton(
            action_frame,
            text='수집 시작',
            command=self._start_collection,
            height=42,
            font=ctk.CTkFont(size=15, weight='bold'),
        )
        self.collect_button.grid(row=0, column=2, sticky='ew', padx=(6, 0))

        self.status_label = ctk.CTkLabel(
            main_frame,
            text='',
            anchor='w',
            font=ctk.CTkFont(size=13),
        )
        self.status_label.grid(row=3, column=0, sticky='ew', padx=16, pady=(0, 8))

        result_frame = ctk.CTkFrame(main_frame, fg_color='transparent')
        result_frame.grid(row=4, column=0, sticky='ew', padx=14, pady=(0, 14))
        for column in range(2):
            result_frame.grid_columnconfigure(column, weight=1, uniform='results')

        job_list_frame = ctk.CTkFrame(result_frame)
        job_list_frame.grid(row=0, column=0, sticky='ew', padx=(0, 7))
        job_list_frame.grid_columnconfigure(0, weight=1)

        job_list_label = ctk.CTkLabel(
            job_list_frame,
            text='등록된 설정',
            anchor='w',
            font=ctk.CTkFont(size=14, weight='bold'),
        )
        job_list_label.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))

        self.job_list_text = ctk.CTkTextbox(
            job_list_frame,
            height=170,
            font=ctk.CTkFont(size=12),
        )
        self.job_list_text.grid(row=1, column=0, sticky='ew', padx=12, pady=(0, 12))
        self.job_list_text.configure(state='disabled')
        self._refresh_job_list_text()

        log_frame = ctk.CTkFrame(result_frame)
        log_frame.grid(row=0, column=1, sticky='ew', padx=(7, 0))
        log_frame.grid_columnconfigure(0, weight=1)

        log_label = ctk.CTkLabel(
            log_frame,
            text='진행 로그',
            anchor='w',
            font=ctk.CTkFont(size=14, weight='bold'),
        )
        log_label.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))

        self.log_text = ctk.CTkTextbox(
            log_frame,
            height=170,
            font=ctk.CTkFont(size=12),
        )
        self.log_text.grid(row=1, column=0, sticky='ew', padx=12, pady=(0, 12))

    def _create_section_title(self, parent: ctk.CTkFrame, row: int, text: str):
        label = ctk.CTkLabel(
            parent,
            text=text,
            anchor='w',
            font=ctk.CTkFont(size=15, weight='bold'),
        )
        label.grid(row=row, column=0, columnspan=2, sticky='ew', padx=24, pady=(14, 4))

    def _create_inline_entry(
        self,
        parent: ctk.CTkFrame,
        label_column: int,
        entry_column: int,
        label_text: str,
        variable: ctk.StringVar,
        placeholder_text: str,
        show: str | None = None,
    ) -> ctk.CTkEntry:
        label = ctk.CTkLabel(
            parent,
            text=label_text,
            width=112,
            anchor='w',
            font=ctk.CTkFont(size=14, weight='bold'),
        )
        label.grid(row=0, column=label_column, sticky='w', padx=(18, 8), pady=14)

        entry = ctk.CTkEntry(
            parent,
            textvariable=variable,
            placeholder_text=placeholder_text,
            show=show,
            height=40,
            font=ctk.CTkFont(size=14),
        )
        entry.grid(row=0, column=entry_column, sticky='ew', padx=(0, 18), pady=14)

        return entry

    def _create_entry_row(
        self,
        parent: ctk.CTkFrame,
        row: int,
        label_text: str,
        variable: ctk.StringVar,
        placeholder_text: str,
        show: str | None = None,
    ) -> ctk.CTkEntry:
        label = ctk.CTkLabel(
            parent,
            text=label_text,
            width=110,
            anchor='w',
            font=ctk.CTkFont(size=15, weight='bold'),
        )
        label.grid(row=row, column=0, sticky='w', padx=(24, 12), pady=8)

        entry = ctk.CTkEntry(
            parent,
            textvariable=variable,
            placeholder_text=placeholder_text,
            show=show,
            height=42,
            font=ctk.CTkFont(size=14),
        )
        entry.grid(row=row, column=1, sticky='ew', padx=(0, 24), pady=8)

        return entry

    def _create_property_type_group(self, parent: ctk.CTkFrame, row: int) -> ctk.CTkFrame:
        label = ctk.CTkLabel(
            parent,
            text='물건종류',
            width=110,
            anchor='nw',
            font=ctk.CTkFont(size=15, weight='bold'),
        )
        label.grid(row=row, column=0, sticky='nw', padx=(24, 12), pady=(14, 8))

        group_frame = ctk.CTkFrame(parent, fg_color='transparent')
        group_frame.grid(row=row, column=1, sticky='ew', padx=(0, 24), pady=8)
        for column in range(3):
            group_frame.grid_columnconfigure(column, weight=1)

        for index, (group_code, group_label, _options) in enumerate(PROPERTY_TYPE_GROUPS):
            radio = ctk.CTkRadioButton(
                group_frame,
                text=group_label,
                variable=self.property_type_group_var,
                value=group_code,
                command=self._on_property_type_group_change,
                font=ctk.CTkFont(size=13, weight='bold'),
            )
            radio.grid(
                row=index // 3,
                column=index % 3,
                sticky='w',
                padx=(0, 10),
                pady=(0, 8),
            )

        self.property_type_checkbox_frame = ctk.CTkFrame(
            group_frame, fg_color='transparent'
        )
        self.property_type_checkbox_frame.grid(
            row=2, column=0, columnspan=3, sticky='ew', pady=(4, 0)
        )
        for column in range(3):
            self.property_type_checkbox_frame.grid_columnconfigure(column, weight=1)

        for code, text in PROPERTY_TYPE_OPTIONS:
            checkbox = ctk.CTkCheckBox(
                self.property_type_checkbox_frame,
                text=text,
                variable=self.property_type_vars[code],
                onvalue=True,
                offvalue=False,
                font=ctk.CTkFont(size=13),
            )
            self.property_type_checkboxes[code] = checkbox

        self._sync_property_type_options(reset_selection=False)
        return group_frame

    def _create_checkbox_group(
        self,
        parent: ctk.CTkFrame,
        row: int,
        label_text: str,
        options: list[tuple[str, str]],
        variables: dict[str, ctk.BooleanVar],
        columns: int,
    ) -> ctk.CTkFrame:
        label = ctk.CTkLabel(
            parent,
            text=label_text,
            width=110,
            anchor='nw',
            font=ctk.CTkFont(size=15, weight='bold'),
        )
        label.grid(row=row, column=0, sticky='nw', padx=(24, 12), pady=(14, 8))

        group_frame = ctk.CTkFrame(parent, fg_color='transparent')
        group_frame.grid(row=row, column=1, sticky='ew', padx=(0, 24), pady=8)
        for column in range(columns):
            group_frame.grid_columnconfigure(column, weight=1)

        for index, (code, text) in enumerate(options):
            checkbox = ctk.CTkCheckBox(
                group_frame,
                text=text,
                variable=variables[code],
                onvalue=True,
                offvalue=False,
                font=ctk.CTkFont(size=13),
            )
            checkbox.grid(
                row=index // columns,
                column=index % columns,
                sticky='w',
                padx=(0, 10),
                pady=5,
            )

        return group_frame

    def _create_select_row(
        self,
        parent: ctk.CTkFrame,
        row: int,
        label_text: str,
        variable: ctk.StringVar,
        command,
    ) -> ScrollableSelect:
        label = ctk.CTkLabel(
            parent,
            text=label_text,
            width=110,
            anchor='w',
            font=ctk.CTkFont(size=15, weight='bold'),
        )
        label.grid(row=row, column=0, sticky='w', padx=(24, 12), pady=8)

        select = ScrollableSelect(
            parent,
            values=[''],
            variable=variable,
            command=command,
            max_dropdown_height=280,
        )
        select.grid(row=row, column=1, sticky='ew', padx=(0, 24), pady=8)

        return select

    def _load_initial_values(self):
        sido_names = self.service.get_sido_names()
        self._set_select_values(self.sido_select, self.sido_var, sido_names)

        if sido_names:
            self._on_sido_change(sido_names[0])

    def _on_sido_change(self, selected_sido: str):
        sigungu_names = self.service.get_sigungu_names(selected_sido)
        self._set_select_values(self.sigungu_select, self.sigungu_var, sigungu_names)

        selected_sigungu = sigungu_names[0] if sigungu_names else ''
        self._on_sigungu_change(selected_sigungu)

    def _on_sigungu_change(self, selected_sigungu: str):
        sido_name = self.sido_var.get()
        eupmyeondong_names = self.service.get_eupmyeondong_names(
            sido_name, selected_sigungu
        )

        self._set_select_values(
            self.eupmyeondong_select,
            self.eupmyeondong_var,
            eupmyeondong_names,
        )

        self._update_selected_label()

    def _on_eupmyeondong_change(self, _selected_eupmyeondong: str):
        self._update_selected_label()

    def _on_all_eupmyeondong_toggle(self):
        self.eupmyeondong_select.set_enabled(not self.all_eupmyeondong_var.get())
        self._update_selected_label()

    def _set_select_values(
        self,
        select: ScrollableSelect,
        variable: ctk.StringVar,
        values: list[str],
    ):
        normalized_values = values or ['']
        select.set_values(normalized_values)
        variable.set(normalized_values[0])

    def _on_property_type_group_change(self):
        self._sync_property_type_options(reset_selection=True)

    def _sync_property_type_options(self, reset_selection: bool):
        selected_codes = self._get_property_type_group_codes(
            self.property_type_group_var.get()
        )

        selected_options = self._get_property_type_group_options(
            self.property_type_group_var.get()
        )

        for code, variable in self.property_type_vars.items():
            if reset_selection:
                variable.set(code in selected_codes)
            elif code not in selected_codes:
                variable.set(False)

            checkbox = self.property_type_checkboxes.get(code)
            if checkbox:
                checkbox.grid_forget()

        for index, (code, _label) in enumerate(selected_options):
            checkbox = self.property_type_checkboxes[code]
            checkbox.grid(
                row=index // 3,
                column=index % 3,
                sticky='w',
                padx=(0, 10),
                pady=5,
            )

    def _get_property_type_group_options(self, group_code: str) -> list[tuple[str, str]]:
        for current_group_code, _group_label, options in PROPERTY_TYPE_GROUPS:
            if current_group_code == group_code:
                return options
        return PROPERTY_TYPE_GROUPS[0][2]

    def _get_property_type_group_codes(self, group_code: str) -> set[str]:
        return {
            code
            for code, _label in self._get_property_type_group_options(group_code)
        }

    def _get_selected_property_type_codes(self) -> str:
        return self._get_selected_codes(
            self._get_property_type_group_options(self.property_type_group_var.get()),
            self.property_type_vars,
        )

    def _get_selected_transaction_type_codes(self) -> str:
        return self._get_selected_codes(
            TRANSACTION_TYPE_OPTIONS, self.transaction_type_vars
        )

    def _get_selected_codes(
        self,
        options: list[tuple[str, str]],
        variables: dict[str, ctk.BooleanVar],
    ) -> str:
        return ','.join(
            code for code, _label in options if variables[code].get()
        )

    def _add_collection_job(self):
        if self.is_collecting:
            return

        jobs = self._make_current_collection_jobs()
        if not jobs:
            return

        existing_keys = {job.key for job in self.collection_jobs}
        added_count = 0
        duplicate_count = 0
        for job in jobs:
            if job.key in existing_keys:
                duplicate_count += 1
                continue

            self.collection_jobs.append(job)
            existing_keys.add(job.key)
            added_count += 1

        if added_count == 0:
            self.status_label.configure(text='이미 등록된 설정입니다.')
            return

        self._refresh_job_list_text()

        status_text = f"설정 {added_count}개 추가, 총 {len(self.collection_jobs)}개 등록됨"
        if duplicate_count:
            status_text += f" ({duplicate_count}개 중복 제외)"
        self.status_label.configure(text=status_text)

    def _make_current_collection_jobs(self) -> list[CollectionJob]:
        sido_name = self.sido_var.get().strip()
        sigungu_name = self.sigungu_var.get().strip()
        property_type = self._get_selected_property_type_codes()
        transaction_type = self._get_selected_transaction_type_codes()

        if not sido_name or not sigungu_name:
            self.status_label.configure(text='지역을 선택해주세요.')
            return []

        if not property_type:
            self.status_label.configure(text='물건종류를 하나 이상 선택해주세요.')
            return []

        if not transaction_type:
            self.status_label.configure(text='거래유형을 하나 이상 선택해주세요.')
            return []

        if self.all_eupmyeondong_var.get():
            eupmyeondong_names = self.service.get_eupmyeondong_names(
                sido_name,
                sigungu_name,
            )

            jobs = [
                job
                for eupmyeondong_name in eupmyeondong_names
                if (
                    job := self._make_collection_job(
                        sido_name,
                        sigungu_name,
                        eupmyeondong_name,
                        property_type,
                        transaction_type,
                    )
                )
            ]

            if not jobs:
                self.status_label.configure(text='등록할 읍면동이 없습니다.')
            return jobs

        eupmyeondong_name = self.eupmyeondong_var.get().strip()
        job = self._make_collection_job(
            sido_name,
            sigungu_name,
            eupmyeondong_name,
            property_type,
            transaction_type,
        )

        if job is None:
            self.status_label.configure(text='읍면동을 선택해주세요.')
            return []

        return [job]

    def _make_collection_job(
        self,
        sido_name: str,
        sigungu_name: str,
        eupmyeondong_name: str,
        property_type: str,
        transaction_type: str,
    ) -> CollectionJob | None:
        eupmyeondong = self.service.get_eupmyeondong(
            sido_name,
            sigungu_name,
            eupmyeondong_name,
        )

        if not eupmyeondong:
            return None

        search_keyword = f"{sido_name} {sigungu_name} {eupmyeondong_name}"

        return CollectionJob(
            search_keyword=search_keyword,
            dong_name=eupmyeondong_name,
            lawd_code=str(eupmyeondong.get('fullCode') or ''),
            property_type=property_type,
            property_type_label=self._get_option_labels(
                property_type, PROPERTY_TYPE_OPTIONS
            ),
            transaction_type=transaction_type,
            transaction_type_label=self._get_option_labels(
                transaction_type, TRANSACTION_TYPE_OPTIONS
            ),
        )

    def _clear_collection_jobs(self):
        if self.is_collecting:
            return

        self.collection_jobs.clear()
        self._refresh_job_list_text()
        self.status_label.configure(text='등록된 설정을 비웠습니다.')

    def _refresh_job_list_text(self):
        if not hasattr(self, 'job_list_text'):
            return

        self.job_list_text.configure(state='normal')
        self.job_list_text.delete('1.0', 'end')

        if not self.collection_jobs:
            self.job_list_text.insert('end', '등록된 설정이 없습니다.\n')
        else:
            for index, job in enumerate(self.collection_jobs, 1):
                self.job_list_text.insert(
                    'end', f"{index}. {job.display_text}\n"
                )

        self.job_list_text.configure(state='disabled')

    def _start_collection(self):
        if self.is_collecting:
            return

        email = self.email_var.get().strip()
        password = self.password_var.get().strip()
        browser_visible = self.browser_visible_var.get()
        merge_excel = self.merge_excel_var.get()
        jobs = list(self.collection_jobs)

        if not email or not password:
            self.status_label.configure(text='카카오 이메일과 비밀번호를 입력해주세요.')
            return

        if not jobs:
            self.status_label.configure(text='수집할 설정을 등록해주세요.')
            return

        self.is_collecting = True
        self.collect_button.configure(text='수집 중...')
        self._set_collection_controls_state('disabled')
        self.status_label.configure(text=f"등록된 설정 {len(jobs)}개 수집 중")
        self.log_text.delete('1.0', 'end')

        worker = threading.Thread(
            target=self._collect_and_save_excel,
            args=(email, password, jobs, browser_visible, merge_excel),
            daemon=True,
        )
        worker.start()

    def _set_collection_controls_state(self, state: str):
        self.add_job_button.configure(state=state)
        self.clear_jobs_button.configure(state=state)
        self.collect_button.configure(state=state)
        self.browser_visible_checkbox.configure(state=state)
        self.merge_excel_checkbox.configure(state=state)
        self.all_eupmyeondong_checkbox.configure(state=state)

    def _collect_and_save_excel(
        self,
        email: str,
        password: str,
        jobs: list[CollectionJob],
        browser_visible: bool,
        merge_excel: bool,
    ):
        try:
            self._log('로그인 중...')
            self.fetch_service.login(email, password, headless=not browser_visible)
        except Exception as error:
            self.log_queue.put(('error', str(error)))
            return

        saved_count = 0
        skipped_count = 0
        error_count = 0
        total_broker_count = 0
        job_count = len(jobs)
        merged_broker_rows = []

        for index, job in enumerate(jobs, 1):
            self.log_queue.put(
                ('status', f"[{index}/{job_count}] {job.search_keyword} 수집 중")
            )

            self._log(f"[{index}/{job_count}] {job.display_text} 수집 시작")

            try:
                self._log('좌표 조회 중...')
                lat, lng = self.fetch_service.get_lat_lng(job.search_keyword)
                self._log(f"좌표 조회 완료: lat={lat}, lng={lng}")

                self._log('전체 매물 수 조회 중...')
                total_count = self.fetch_service.get_property_total_count(
                    lat=lat,
                    lng=lng,
                    property_type=job.property_type,
                    transaction_type=job.transaction_type,
                    lawd_code=job.lawd_code,
                )

                self._log(f"수집 대상 매물 {total_count}건")
                if total_count == 0:
                    skipped_count += 1
                    self._log('매물이 없어 엑셀 저장을 건너뜁니다.')
                    continue

                self._log('전체 매물 조회 중...')
                properties = self.fetch_service.get_all_properties(
                    lat=lat,
                    lng=lng,
                    property_type=job.property_type,
                    transaction_type=job.transaction_type,
                    lawd_code=job.lawd_code,
                    total_count=total_count,
                    progress_callback=self._log_property_page_progress,
                )

                self._log(f"매물 {len(properties)}건 조회 완료")

                if merge_excel:
                    broker_rows = self._make_unique_broker_rows(properties)
                    merged_broker_rows.extend(broker_rows)
                    self._log(f"통합 엑셀에 부동산 {len(broker_rows)}개 추가")
                    continue

                output_path, row_count = self._save_broker_excel(job, properties)
                saved_count += 1
                total_broker_count += row_count
                self._log(f"중복 제거 후 부동산 {row_count}개 수집")
                self._log(f"엑셀 저장 완료: {output_path}")
            except Exception as error:
                error_count += 1
                self._log(f"[{index}/{job_count}] 오류: {error}")

        if merge_excel:
            merged_broker_rows = self._dedupe_broker_rows(merged_broker_rows)
            total_broker_count = len(merged_broker_rows)
            self._log(f"통합 중복 제거 후 부동산 {total_broker_count}개 수집")

            if merged_broker_rows:
                output_path, row_count = self._save_merged_broker_excel(
                    jobs,
                    merged_broker_rows,
                )

                saved_count = 1
                total_broker_count = row_count
                self._log(f"통합 엑셀 저장 완료: {output_path}")
            else:
                self._log('저장할 부동산이 없어 통합 엑셀 저장을 건너뜁니다.')

        self.log_queue.put(
            (
                'done_all',
                job_count,
                saved_count,
                skipped_count,
                error_count,
                total_broker_count,
            )
        )

    def _get_selected_region(self) -> str:
        parts = [
            self.sido_var.get().strip(),
            self.sigungu_var.get().strip(),
            self.eupmyeondong_var.get().strip(),
        ]
        return ' '.join(part for part in parts if part)

    def _log_property_page_progress(
        self,
        page_number: int,
        page_count: int,
        collected_count: int,
        total_count: int,
    ):
        visible_collected_count = min(collected_count, total_count)
        self._log(
            '매물 조회 진행: '
            f"{page_number}/{page_count}페이지, "
            f"{visible_collected_count}/{total_count}건"
        )

    def _get_selected_lawd_code(self) -> str:
        eupmyeondong = self.service.get_eupmyeondong(
            self.sido_var.get(),
            self.sigungu_var.get(),
            self.eupmyeondong_var.get(),
        )

        if not eupmyeondong:
            return ''

        return str(eupmyeondong.get('fullCode') or '')

    def _save_broker_excel(
        self,
        job: CollectionJob,
        properties: list[dict],
    ) -> tuple[Path, int]:
        broker_rows = self._make_unique_broker_rows(properties)

        filename_parts = [
            job.dong_name or job.search_keyword,
            job.property_type_label,
            job.transaction_type_label,
        ]

        return self._save_broker_rows_excel(filename_parts, broker_rows)

    def _save_merged_broker_excel(
        self,
        jobs: list[CollectionJob],
        broker_rows: list[dict[str, str]],
    ) -> tuple[Path, int]:
        filename_parts = ['통합', f"{len(jobs)}개설정"]
        return self._save_broker_rows_excel(filename_parts, broker_rows)

    def _save_broker_rows_excel(
        self,
        filename_parts: list[str],
        broker_rows: list[dict[str, str]],
    ) -> tuple[Path, int]:
        output_dir = get_excel_output_dir()
        output_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_filename = self._make_safe_filename(
            '_'.join([*filename_parts, timestamp])
        )
        output_path = output_dir / f"{safe_filename}.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '중개업소'
        worksheet.append(EXCEL_COLUMNS)

        for row in broker_rows:
            worksheet.append([row.get(column, '') for column in EXCEL_COLUMNS])

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for row_cells in worksheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        worksheet.freeze_panes = 'A2'

        column_widths = {
            '중개업소명': 42,
            '중개업소대표자명': 18,
            '중개업소주소': 70,
            '중개업소전화번호': 18,
            '중개업소대표자휴대폰번호': 24,
        }

        for index, column_name in enumerate(EXCEL_COLUMNS, 1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=index).column_letter
            ].width = column_widths[column_name]

        workbook.save(output_path)

        return output_path, len(broker_rows)

    def _get_option_labels(
        self,
        selected_codes: str,
        options: list[tuple[str, str]],
    ) -> str:
        labels_by_code = dict(options)

        labels = [
            labels_by_code[code]
            for code in selected_codes.split(',')
            if code in labels_by_code
        ]

        if labels:
            return '+'.join(labels)
        return '전체'

    def _make_safe_filename(self, value: str) -> str:
        return re.sub(r'[\\/:*?"<>|]+', '_', value).strip()

    def _make_unique_broker_rows(self, properties: list[dict]) -> list[dict[str, str]]:
        return self._dedupe_broker_rows(
            [self._make_broker_excel_row(property_item) for property_item in properties]
        )

    def _dedupe_broker_rows(
        self,
        broker_rows: list[dict[str, str]],
    ) -> list[dict[str, str]]:
        rows = []
        seen_keys = set()
        for row in broker_rows:
            broker_name = row['중개업소명'].strip()
            broker_address = row['중개업소주소'].strip()
            key = (broker_name, broker_address)
            if not broker_name and not broker_address:
                continue

            if key in seen_keys:
                continue

            seen_keys.add(key)
            rows.append(row)

        return rows

    def _make_broker_excel_row(self, property_item: dict) -> dict[str, str]:
        return {
            '중개업소명': self._string_value(property_item.get('중개업소명')),
            '중개업소대표자명': self._string_value(property_item.get('중개업소대표자명')),
            '중개업소주소': self._string_value(property_item.get('중개업소주소')),
            '중개업소전화번호': self._string_value(property_item.get('중개업소전화번호')),
            '중개업소대표자휴대폰번호': self._string_value(
                property_item.get('중개업소대표자휴대폰번호')
            ),
        }

    def _string_value(self, value) -> str:
        if value is None:
            return ''
        return str(value)

    def _log(self, message: str):
        self.log_queue.put(('log', message))

    def _process_log_queue(self):
        try:
            while True:
                event = self.log_queue.get_nowait()
                event_type = event[0]

                if event_type == 'log':
                    self._append_log(event[1])
                elif event_type == 'status':
                    self.status_label.configure(text=event[1])
                elif event_type == 'done':
                    row_count, output_path = event[1], event[2]
                    self._append_log(f"엑셀 저장 완료: {output_path}")
                    self.status_label.configure(text=f"저장 완료: {row_count}건")
                    self.collect_button.configure(text='수집 시작')
                    self._set_collection_controls_state('normal')
                    self.is_collecting = False
                elif event_type == 'done_all':
                    (
                        _event_type,
                        job_count,
                        saved_count,
                        skipped_count,
                        error_count,
                        total_broker_count,
                    ) = event
                    self.status_label.configure(
                        text=(
                            f"수집 완료: 설정 {job_count}개, 저장 {saved_count}개, "
                            f"부동산 {total_broker_count}개"
                        )
                    )
                    self._append_log(
                        f"전체 수집 완료: 저장 {saved_count}개, "
                        f"매물 없음 {skipped_count}개, 오류 {error_count}개"
                    )
                    self.collect_button.configure(text='수집 시작')
                    self._set_collection_controls_state('normal')
                    self.is_collecting = False
                elif event_type == 'empty':
                    self.status_label.configure(text='저장할 매물이 없습니다.')
                    self.collect_button.configure(text='수집 시작')
                    self._set_collection_controls_state('normal')
                    self.is_collecting = False
                elif event_type == 'error':
                    self._append_log(f"오류: {event[1]}")
                    self.status_label.configure(text='수집 실패')
                    self.collect_button.configure(text='수집 시작')
                    self._set_collection_controls_state('normal')
                    self.is_collecting = False
        except queue.Empty:
            pass

        self.after(100, self._process_log_queue)

    def _append_log(self, message: str):
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.insert('end', f"[{timestamp}] {message}\n")
        self.log_text.see('end')

    def _update_selected_label(self):
        sido_name = self.sido_var.get()
        sigungu_name = self.sigungu_var.get()
        eupmyeondong_name = self.eupmyeondong_var.get()
        eupmyeondong = self.service.get_eupmyeondong(
            sido_name,
            sigungu_name,
            eupmyeondong_name,
        )

        if self.all_eupmyeondong_var.get():
            eupmyeondong_count = len(
                self.service.get_eupmyeondong_names(sido_name, sigungu_name)
            )

            if eupmyeondong_count == 0:
                self.selected_label.configure(text='')
                return

            self.selected_label.configure(
                text=f"{sido_name} {sigungu_name} 전체 읍면동 ({eupmyeondong_count}개)"
            )
            return

        if not eupmyeondong:
            self.selected_label.configure(text='')
            return

        self.selected_label.configure(
            text=f"{sido_name} {sigungu_name} {eupmyeondong_name}"
        )


if __name__ == '__main__':
    app = LawdSelectApp()
    app.mainloop()
